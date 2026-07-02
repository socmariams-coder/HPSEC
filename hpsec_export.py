"""
hpsec_export.py - Mòdul d'exportació HPSEC (Fase 4: EXPORTAR)
==============================================================

FASE 4 del pipeline de 4 fases:
- Generar Excels finals amb estructura estandarditzada
- Fulls: ID (traçabilitat), DOC, DAD, RESULTS
- Usar seleccions de rèpliques de la Fase 3 (Analitzar)

REQUEREIX:
- Fase 3: processed_data amb seleccions DOC/DAD per mostra

Usat per HPSEC_Suite.py
"""

__version__ = "2.2.0"
__version_date__ = "2026-03-03"

import os
import json
import logging
import zipfile
import hashlib
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from scipy.integrate import trapezoid

logger = logging.getLogger(__name__)


def _make_unique_filename(base_name, ext, used_filenames):
    """Genera un nom de fitxer únic, afegint _2, _3... si cal."""
    candidate = f"{base_name}{ext}"
    if candidate not in used_filenames:
        used_filenames.add(candidate)
        return candidate
    n = 2
    while True:
        candidate = f"{base_name}_{n}{ext}"
        if candidate not in used_filenames:
            used_filenames.add(candidate)
            return candidate
        n += 1


def _sample_filename_base(sample_name, mode):
    """Retorna el nom base del fitxer: {sample}_HPSEC_{C|B}."""
    mode_suffix = "C" if mode.upper() == "COLUMN" else "B"
    return f"{sample_name}_HPSEC_{mode_suffix}"


def _get_max_severity(anomalies):
    """Retorna severitat màxima d'una llista d'anomalies (strings o dicts)."""
    try:
        from hpsec_warnings import get_max_anomaly_severity
        return get_max_anomaly_severity(anomalies)
    except ImportError:
        return ""


def _load_export_config():
    """Carrega configuració d'exportació des de hpsec_config.json.

    Retorna dict amb target_wavelengths i time_fractions en format [start, end].
    Fallback a valors hardcoded si la config no està disponible.
    """
    fallback = {
        "target_wavelengths": [220, 252, 254, 272, 290, 362],
        "time_fractions": {
            "BioP": [0, 18],
            "HS": [18, 23],
            "BB": [23, 30],
            "SB": [30, 40],
            "LMW": [40, 70],
        },
    }
    try:
        from hpsec_config import get_config
        cfg = get_config()

        # Wavelengths
        wls = cfg.get("wavelengths", "selected", default=None)
        if wls:
            fallback["target_wavelengths"] = wls

        # Time fractions: config format {"BioP": {"start": 10.8, "end": 18}} → [start, end]
        fractions_raw = cfg.get("time_fractions", default={})
        if fractions_raw:
            fractions = {}
            for name, info in fractions_raw.items():
                if isinstance(info, dict) and "start" in info and "end" in info:
                    fractions[name] = [info["start"], info["end"]]
                elif isinstance(info, (list, tuple)) and len(info) == 2:
                    fractions[name] = list(info)
            if fractions:
                fallback["time_fractions"] = fractions

    except Exception as e:
        logger.warning("Could not load export fractions from config: %s", e)

    return fallback


# Configuració per defecte (carregada dinàmicament)
DEFAULT_EXPORT_CONFIG = _load_export_config()


def write_final_excel(
    out_path: str,
    sample_name: str,
    sample_data: dict,
    calibration_data: dict = None,
    mode: str = "COLUMN",
    config: dict = None,
):
    """
    Escriu Excel final amb estructura estandarditzada.

    Fulls:
        ID: Traçabilitat completa (fitxers, shifts, quantificació)
        DOC: Cromatogrames DOC (final, raw Direct, raw UIB)
        DAD: 6 longituds d'ona seleccionades
        RESULTS: Integracions per fraccions

    Args:
        out_path: Camí del fitxer Excel
        sample_name: Nom de la mostra
        sample_data: Dict amb les dades de la rèplica seleccionada
        calibration_data: Dict amb dades de calibració (opcional)
        mode: "BP" o "COLUMN"
        config: Configuració (opcional)

    Returns:
        dict amb info d'exportació
    """
    config = config or DEFAULT_EXPORT_CONFIG
    calibration_data = calibration_data or {}

    # Extreure dades de la rèplica
    t_doc = sample_data.get("t_doc")
    y_doc_net = sample_data.get("y_doc_net")
    y_doc_raw = sample_data.get("y_doc_raw")
    y_doc_uib_net = sample_data.get("y_doc_uib_net")
    y_doc_uib_raw = sample_data.get("y_doc_uib_raw")
    df_dad = sample_data.get("df_dad")

    # Convertir a numpy arrays
    if t_doc is not None:
        t_doc = np.asarray(t_doc)
    if y_doc_net is not None:
        y_doc_net = np.asarray(y_doc_net)
    if y_doc_raw is not None:
        y_doc_raw = np.asarray(y_doc_raw)
    if y_doc_uib_net is not None:
        y_doc_uib_net = np.asarray(y_doc_uib_net)
    if y_doc_uib_raw is not None:
        y_doc_uib_raw = np.asarray(y_doc_uib_raw)

    is_dual = y_doc_uib_net is not None and len(y_doc_uib_net) > 0

    # === FULL ID: Traçabilitat ===
    id_rows = _build_id_sheet(sample_name, sample_data, calibration_data, mode, is_dual)
    df_id = pd.DataFrame(id_rows, columns=["Field", "Value"])

    # === FULL DOC: Cromatogrames ===
    df_doc = _build_doc_sheet(t_doc, y_doc_net, y_doc_raw, y_doc_uib_net, y_doc_uib_raw, is_dual)

    # === FULL DAD: 6 longituds d'ona ===
    df_dad_export = _build_dad_sheet(df_dad, config)

    # === FULL RESULTS: Integracions ===
    df_results = _build_results_sheet(t_doc, y_doc_net, df_dad, mode, config)

    # Escriure Excel: ID | DOC | DAD | RESULTS
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_id.to_excel(writer, sheet_name="ID", index=False)
        df_doc.to_excel(writer, sheet_name="DOC", index=False)
        if df_dad_export is not None and not df_dad_export.empty:
            df_dad_export.to_excel(writer, sheet_name="DAD", index=False)
        df_results.to_excel(writer, sheet_name="RESULTS", index=False)

    return {
        "success": True,
        "path": out_path,
        "sample": sample_name,
        "n_doc_points": len(t_doc) if t_doc is not None else 0,
        "n_dad_points": len(df_dad) if df_dad is not None else 0,
        "is_dual": is_dual,
        "has_bp": False,
    }


def _build_id_sheet(sample_name, sample_data, calibration_data, mode, is_dual):
    """Construeix les files del full ID (traçabilitat) amb seccions organitzades."""
    rows = []

    # --- 1. EXPORT ---
    from hpsec_version import SUITE_VERSION
    rows.append(("=== EXPORT ===", ""))
    rows.append(("Suite_Version", f"HPSEC Suite v{SUITE_VERSION}"))
    rows.append(("Export_Module", f"hpsec_export v{__version__}"))
    rows.append(("Export_Date", datetime.now().strftime("%Y-%m-%d %H:%M")))
    rows.append(("Encoding", "UTF-8"))

    # --- 2. SAMPLE ---
    rows.append(("=== SAMPLE ===", ""))
    rows.append(("Sample", sample_name))
    rows.append(("Replica_DOC", sample_data.get("replica", "?")))
    rows.append(("Replica_DAD", sample_data.get("replica_dad", sample_data.get("replica", "?"))))
    rows.append(("SEQ", sample_data.get("seq_name", "")))
    rows.append(("Method", mode))
    rows.append(("Sample_Type", sample_data.get("sample_type", "SAMPLE")))

    # --- 3. SOURCE FILES ---
    rows.append(("=== SOURCE FILES ===", ""))
    rows.append(("File_DOC_Direct", sample_data.get("file_doc", "")))
    if is_dual:
        rows.append(("File_DOC_UIB", sample_data.get("file_doc_uib", sample_data.get("file_doc", ""))))
    rows.append(("File_DAD", sample_data.get("file_dad", "")))
    rows.append(("File_MasterFile", sample_data.get("master_file", "")))

    # --- 4. PROCESSING ---
    rows.append(("=== PROCESSING ===", ""))
    shift_direct = sample_data.get("shift_direct")
    shift_uib = sample_data.get("shift_uib")
    if shift_direct is not None:
        rows.append(("Shift_Direct_sec", round(shift_direct * 60, 2)))
    if is_dual and shift_uib is not None:
        rows.append(("Shift_UIB_sec", round(shift_uib * 60, 2)))
    baseline_direct = sample_data.get("baseline_direct")
    baseline_uib = sample_data.get("baseline_uib")
    if baseline_direct is not None:
        rows.append(("Baseline_Direct_mAU", round(float(baseline_direct), 3)))
    if is_dual and baseline_uib is not None:
        rows.append(("Baseline_UIB_mAU", round(float(baseline_uib), 3)))
    # Smoothing & integration config
    try:
        from hpsec_config import get_config
        cfg = get_config()
        rows.append(("Smoothing_Window_SG", cfg.get("smoothing", {}).get("sg_window_points", "")))
        rows.append(("Baseline_Method", cfg.get("baseline", {}).get("method", "mode")))
    except Exception:
        pass
    rows.append(("DOC_Target_DT_min", 0.0667))
    # Integration limits
    peak_limits = sample_data.get("peak_limits", {})
    if peak_limits:
        rows.append(("Integration_T_Start_min", peak_limits.get("t_start", "")))
        rows.append(("Integration_T_End_min", peak_limits.get("t_end", "")))

    # --- 5. SIGNAL QUALITY ---
    rows.append(("=== SIGNAL QUALITY ===", ""))
    snr_info = sample_data.get("snr_info", {})
    if snr_info:
        if snr_info.get("snr_direct") is not None:
            rows.append(("SNR_Direct", round(snr_info["snr_direct"], 1)))
        if snr_info.get("lod_direct") is not None:
            rows.append(("LOD_Direct_mAU", round(snr_info["lod_direct"], 3)))
        if snr_info.get("loq_direct") is not None:
            rows.append(("LOQ_Direct_mAU", round(snr_info["loq_direct"], 3)))
        if is_dual:
            if snr_info.get("snr_uib") is not None:
                rows.append(("SNR_UIB", round(snr_info["snr_uib"], 1)))
            if snr_info.get("lod_uib") is not None:
                rows.append(("LOD_UIB_mAU", round(snr_info["lod_uib"], 3)))

    # --- 6. INJECTION ---
    rows.append(("=== INJECTION ===", ""))
    seq_date = sample_data.get("seq_date", "")
    if seq_date:
        rows.append(("Date", seq_date))
    inj_volume = sample_data.get("inj_volume")
    if inj_volume:
        rows.append(("INJ_Volume_uL", inj_volume))
    inj_idx = sample_data.get("injection_index")
    if inj_idx is not None:
        rows.append(("Injection_Index", inj_idx))
    data_mode = sample_data.get("data_mode", "")
    if data_mode:
        rows.append(("Data_Mode", data_mode))

    # --- 7. QUANTIFICATION ---
    rows.append(("=== QUANTIFICATION ===", ""))
    quantification = sample_data.get("quantification", {})
    if quantification:
        conc = quantification.get("concentration_ppm")
        if conc is not None:
            rows.append(("Concentration_ppm", round(conc, 3)))
        conc_d = quantification.get("concentration_ppm_direct")
        if conc_d is not None and conc_d != conc:
            rows.append(("Concentration_ppm_Direct", round(conc_d, 3)))
        conc_u = quantification.get("concentration_ppm_uib")
        if conc_u is not None:
            rows.append(("Concentration_ppm_UIB", round(conc_u, 3)))
        cr = quantification.get("calibration_ratio")
        if cr is not None:
            rows.append(("Calibration_Ratio", round(cr, 4)))
        area_total = quantification.get("area_total")
        if area_total is not None:
            rows.append(("Area_DOC_total", round(area_total, 2)))
    rows.append(("Quantification_Formula", "ppm = (Area - intercept) * 1000 / (RF * V_uL)"))

    # --- 8. CALIBRATION ---
    rows.append(("=== CALIBRATION ===", ""))
    if calibration_data:
        cal_date = calibration_data.get("date", "")
        cal_khp = calibration_data.get("khp_conc_ppm", "")
        rows.append(("Calibration_Date", cal_date))
        rows.append(("Calibration_KHP_ppm", cal_khp))
        rf_obj = calibration_data.get("rf_mass_cal", {})
        if isinstance(rf_obj, dict):
            for sig_type in ("direct", "uib"):
                sig_rf = rf_obj.get(sig_type, {})
                if isinstance(sig_rf, dict):
                    for mode_key, val in sig_rf.items():
                        rows.append((f"RF_mass_cal_{sig_type}_{mode_key}", val))
        elif rf_obj:
            rows.append(("RF_mass_cal", rf_obj))
        intercept_obj = calibration_data.get("intercept", {})
        if isinstance(intercept_obj, dict):
            for sig_type in ("direct", "uib"):
                sig_int = intercept_obj.get(sig_type, {})
                if isinstance(sig_int, dict):
                    for mode_key, val in sig_int.items():
                        rows.append((f"Intercept_{sig_type}_{mode_key}", val))
        elif intercept_obj:
            rows.append(("Intercept", intercept_obj))
        cal_seq = calibration_data.get("seq_name", "")
        if cal_seq:
            rows.append(("Calibration_SEQ", cal_seq))
    try:
        from hpsec_calibrate import compute_calibration_fingerprint
        cal_fp = compute_calibration_fingerprint()
        rows.append(("Calibration_Fingerprint", cal_fp))
    except Exception:
        pass

    # --- 9. ANOMALIES ---
    rows.append(("=== ANOMALIES ===", ""))
    anomalies = sample_data.get("anomalies", [])
    if anomalies:
        for a in anomalies:
            if isinstance(a, dict):
                code = a.get("code", "")
                sev = a.get("severity", "info")
                rep = " [REPAIRED]" if a.get("repaired") else ""
                rows.append((f"Anomaly_{code}", f"{sev}{rep}"))
            else:
                rows.append(("Anomaly", str(a)))
    else:
        rows.append(("Anomalies", "Cap"))

    # Bi-Gaussian info (BP mode)
    bigaussian = sample_data.get("bigaussian_doc")
    if bigaussian and isinstance(bigaussian, dict):
        r2_bg = bigaussian.get("r2") or bigaussian.get("r_squared")
        if r2_bg is not None:
            rows.append(("Bigaussian_R2", round(r2_bg, 4)))
        asym = bigaussian.get("asymmetry")
        if asym is not None:
            rows.append(("Bigaussian_Asymmetry", round(asym, 3)))

    # Timeout info
    timeout_info = sample_data.get("timeout_info", {})
    if timeout_info.get("n_timeouts", 0) > 0:
        rows.append(("Timeout_Severity", timeout_info.get("severity", "")))
        zone_summary = timeout_info.get("zone_summary", {})
        if zone_summary:
            zones_str = "; ".join(f"{zone}: {count}" for zone, count in zone_summary.items())
            rows.append(("Timeout_Zones", zones_str))
    if timeout_info.get("toc_minute_precision"):
        rows.append(("TOC_Timestamp_Repair", "YES"))

    # Irregular top repair info
    irregular_top_repaired = sample_data.get("irregular_top_direct_repaired",
                                              sample_data.get("batman_direct_repaired"))
    if irregular_top_repaired:
        rows.append(("Irregular_Top_Direct_Repaired", "YES"))
        repair_info = sample_data.get("irregular_top_direct_repair_info",
                                       sample_data.get("batman_direct_repair_info", {}))
        if repair_info:
            rows.append(("Irregular_Top_Y_Max_Original", repair_info.get("y_max_original", "")))
            rows.append(("Irregular_Top_Y_Max_Theoretical", repair_info.get("y_max_theoretical", "")))

    # --- 10. CONFIG ---
    rows.append(("=== CONFIG ===", ""))
    rows.append(("Integration_Method", "tangent_projection"))
    try:
        from hpsec_config import ConfigManager as Config
        cfg_obj = Config()
        rows.append(("Config_Fingerprint", cfg_obj.compute_fingerprint()))
    except Exception:
        pass

    return rows


def _build_doc_sheet(t_doc, y_doc_net, y_doc_raw, y_doc_uib_net, y_doc_uib_raw, is_dual):
    """Construeix el DataFrame del full DOC."""
    if t_doc is None or len(t_doc) == 0:
        return pd.DataFrame({"time (min)": [], "DOC (mAU)": []})

    data = {"time (min)": t_doc}

    # Senyal final (corregit)
    if y_doc_net is not None:
        data["DOC (mAU)"] = y_doc_net
    else:
        data["DOC (mAU)"] = np.zeros_like(t_doc)

    # Raw Direct
    if y_doc_raw is not None and len(y_doc_raw) == len(t_doc):
        data["DOC_Direct_RAW (mAU)"] = y_doc_raw

    # UIB (si DUAL)
    if is_dual:
        if y_doc_uib_net is not None and len(y_doc_uib_net) == len(t_doc):
            data["DOC_UIB (mAU)"] = y_doc_uib_net
        if y_doc_uib_raw is not None and len(y_doc_uib_raw) == len(t_doc):
            data["DOC_UIB_RAW (mAU)"] = y_doc_uib_raw

    return pd.DataFrame(data)


def _build_dad_sheet(df_dad, config):
    """Construeix el DataFrame del full DAD amb 6 longituds d'ona."""
    if df_dad is None or df_dad.empty:
        return None

    target_wls = config.get("target_wavelengths", [220, 252, 254, 272, 290, 362])

    # Columna temps
    if "time (min)" in df_dad.columns:
        result = {"time (min)": df_dad["time (min)"].values}
    elif "Time" in df_dad.columns:
        result = {"time (min)": df_dad["Time"].values}
    else:
        return None

    # Afegir cada longitud d'ona
    for wl in target_wls:
        col_name = str(wl)
        if col_name in df_dad.columns:
            result[f"A{wl}"] = df_dad[col_name].values
        elif f"A{wl}" in df_dad.columns:
            result[f"A{wl}"] = df_dad[f"A{wl}"].values
        else:
            # Buscar columna aproximada
            for col in df_dad.columns:
                try:
                    if abs(float(col) - wl) < 1:
                        result[f"A{wl}"] = df_dad[col].values
                        break
                except (ValueError, TypeError):
                    pass

    return pd.DataFrame(result)


def _build_results_sheet(t_doc, y_doc_net, df_dad, mode, config):
    """Construeix el DataFrame del full RESULTS amb integracions."""
    target_wls = config.get("target_wavelengths", [220, 252, 254, 272, 290, 362])
    fractions = config.get("time_fractions", {})

    # Header
    header = ["Fraction", "Range (min)", "DOC"]
    for wl in target_wls:
        header.append(f"A{wl}")

    rows = []

    # Calcular àrees per cada fracció
    fraction_names = list(fractions.keys())
    if mode == "BP":
        # Per BP o mostres sense fraccions, només total
        fraction_names = ["total"]
    else:
        fraction_names.append("total")

    for frac_name in fraction_names:
        if frac_name == "total":
            t_ini, t_fi = 0, 70
            rang = "0-70"
        else:
            t_ini, t_fi = fractions.get(frac_name, [0, 0])
            rang = f"{t_ini}-{t_fi}"

        row = [frac_name, rang]

        # Àrea DOC
        area_doc = _integrate_fraction(t_doc, y_doc_net, t_ini, t_fi)
        row.append(round(area_doc, 2) if area_doc > 0 else "-")

        # Àrees DAD
        for wl in target_wls:
            area_dad = _integrate_dad_fraction(df_dad, wl, t_ini, t_fi)
            row.append(round(area_dad, 2) if area_dad > 0 else "-")

        rows.append(row)

    return pd.DataFrame(rows, columns=header)


def _integrate_fraction(t, y, t_ini, t_fi):
    """Integra una fracció temporal."""
    if t is None or y is None or len(t) == 0:
        return 0.0

    mask = (t >= t_ini) & (t <= t_fi)
    if not np.any(mask):
        return 0.0

    t_frac = t[mask]
    y_frac = y[mask]

    if len(t_frac) < 2:
        return 0.0

    return float(trapezoid(y_frac, t_frac))


def _integrate_dad_fraction(df_dad, wl, t_ini, t_fi):
    """Integra una fracció temporal per una longitud d'ona DAD."""
    if df_dad is None or df_dad.empty:
        return 0.0

    # Trobar columna temps
    if "time (min)" in df_dad.columns:
        t_col = "time (min)"
    elif "Time" in df_dad.columns:
        t_col = "Time"
    else:
        return 0.0

    # Trobar columna WL
    wl_col = None
    for col in [str(wl), f"A{wl}"]:
        if col in df_dad.columns:
            wl_col = col
            break

    if wl_col is None:
        # Buscar aproximada
        for col in df_dad.columns:
            try:
                if abs(float(col) - wl) < 1:
                    wl_col = col
                    break
            except (ValueError, TypeError):
                pass

    if wl_col is None:
        return 0.0

    t = df_dad[t_col].values
    y = df_dad[wl_col].values

    return _integrate_fraction(t, y, t_ini, t_fi)


# =============================================================================
# EXPORTACIÓ LIGHTWEIGHT PER BLANCS / CONTROLS
# =============================================================================

def write_light_excel(
    out_path: str,
    sample_name: str,
    sample_data: dict,
    config: dict = None,
):
    """
    Escriu Excel lightweight per BLANK/CONTROL (només ID + DOC).

    Args:
        out_path: Camí del fitxer Excel
        sample_name: Nom de la mostra
        sample_data: Dict amb dades de la rèplica light
        config: Configuració (opcional)

    Returns:
        dict amb info d'exportació
    """
    # === FULL ID: Traçabilitat ===
    id_rows = [
        ("Export_Module", f"hpsec_export v{__version__}"),
        ("Export_Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("---", "---"),
        ("Sample", sample_name),
        ("Sample_Type", sample_data.get("sample_type", "BLANK")),
        ("Replica", sample_data.get("replica", "?")),
        ("Analysis_Type", "light"),
        ("---", "---"),
        ("Area_Total_DOC", sample_data.get("area_total", 0)),
        ("SNR", sample_data.get("snr", 0)),
        ("Inj_Volume_uL", sample_data.get("inj_volume", "")),
    ]
    df_id = pd.DataFrame(id_rows, columns=["Field", "Value"])

    # === FULL DOC: Cromatograma ===
    t_doc = sample_data.get("t_doc")
    y_doc_net = sample_data.get("y_doc_net")

    doc_data = {}
    if t_doc is not None:
        doc_data["Time_min"] = np.asarray(t_doc)
    if y_doc_net is not None:
        doc_data["DOC_net"] = np.asarray(y_doc_net)

    df_doc = pd.DataFrame(doc_data) if doc_data else pd.DataFrame()

    # Escriure Excel: ID | DOC
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_id.to_excel(writer, sheet_name="ID", index=False)
        if not df_doc.empty:
            df_doc.to_excel(writer, sheet_name="DOC", index=False)

    return {
        "success": True,
        "path": out_path,
        "sample": sample_name,
        "analysis_type": "light",
        "n_doc_points": len(t_doc) if t_doc is not None else 0,
    }


# =============================================================================
# EXPORTACIÓ DE SEQÜÈNCIES
# =============================================================================

def export_sequence(
    samples_grouped: dict,
    output_dir: str,
    calibration_data: dict = None,
    mode: str = "COLUMN",
    config: dict = None,
    progress_callback=None,
    seq_path: str = None,
    export_raw: bool = False,
    export_processed: bool = False,
    csv_separator: str = ";",
):
    """
    Exporta totes les mostres d'una seqüència.

    Args:
        samples_grouped: Dict amb mostres agrupades
        output_dir: Directori de sortida (RESULTATS/)
        calibration_data: Dades de calibració
        mode: "BP" o "COLUMN"
        config: Configuració
        progress_callback: Funció per reportar progrés (pct, msg)
        seq_path: Path de la seqüència
        export_raw: Exportar CSVs RAW (DOC cru + DAD 101λ) a RAW/
        export_processed: Exportar CSVs PROCESSED (DOC net + fraccions) a PROCESSED/
        csv_separator: Separador pels CSVs (";", ",", "\t")

    Returns:
        dict amb resultats d'exportació
    """
    config = config or DEFAULT_EXPORT_CONFIG
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    results = {
        "success": True,
        "n_exported": 0,
        "n_errors": 0,
        "files": [],
        "errors": [],
        "bp_info": None,
        "raw_files": [],
        "processed_files": [],
    }

    # Tracking de noms per evitar col·lisions (separats per carpeta)
    used_filenames = set()
    used_raw = set()
    used_proc = set()

    # Subcarpetes RAW/ i PROCESSED/
    raw_dir = str(output_path / "RAW")
    proc_dir = str(output_path / "PROCESSED")

    total = len(samples_grouped)
    n_skipped = 0
    for i, (sample_name, sample_info) in enumerate(samples_grouped.items()):
        if progress_callback:
            pct = 5 + int((i / total) * 90)
            progress_callback(pct, f"Exportant {sample_name}...")

        try:
            # Light samples (CONTROL) — exportació lightweight
            if sample_info.get("analysis_type") == "light":
                selected = sample_info.get("selected", {})
                doc_replica = selected.get("doc", "1")
                replicas = sample_info.get("replicas", {})
                doc_data = replicas.get(doc_replica, {})

                sample_type = sample_info.get("sample_type", "BLANK")
                base = _sample_filename_base(sample_name, mode)
                filename = _make_unique_filename(base, ".xlsx", used_filenames)
                filepath = output_path / filename

                result = write_light_excel(
                    str(filepath),
                    sample_name,
                    doc_data,
                    config,
                )

                # CSV RAW per light samples
                if export_raw:
                    try:
                        rf = write_csv_raw(raw_dir, sample_name, doc_data,
                                           mode, csv_separator, used_raw)
                        results["raw_files"].extend(rf)
                    except Exception as e_csv:
                        logger.warning(f"CSV RAW light {sample_name}: {e_csv}")

                # CSV PROCESSED per light samples
                if export_processed:
                    try:
                        pf = write_csv_processed(proc_dir, sample_name, doc_data,
                                                 mode, config, csv_separator, used_proc)
                        results["processed_files"].extend(pf)
                    except Exception as e_csv:
                        logger.warning(f"CSV PROC light {sample_name}: {e_csv}")

                results["files"].append({
                    "sample": sample_name,
                    "path": str(filepath),
                    "doc_replica": doc_replica,
                    "analysis_type": "light",
                })
                results["n_exported"] += 1
                continue

            # Saltar mostres no vàlides o amb selecció "cap"
            selected = sample_info.get("selected", {})
            doc_replica = selected.get("doc", "1")
            dad_replica = selected.get("dad", doc_replica)

            if doc_replica == "none" or sample_info.get("sample_valid") is False:
                n_skipped += 1
                continue

            replicas = sample_info.get("replicas", {})
            doc_data = replicas.get(doc_replica, {})
            dad_data = replicas.get(dad_replica, {})

            # Combinar dades (DOC de doc_replica, DAD de dad_replica)
            export_data = dict(doc_data)  # Copiar DOC data
            export_data["replica"] = doc_replica
            export_data["replica_dad"] = dad_replica

            # Si DAD ve d'altra rèplica, substituir
            if dad_replica != doc_replica and "df_dad" in dad_data:
                export_data["df_dad"] = dad_data["df_dad"]
                # Propagar dad_export3d_path de la rèplica DAD
                if dad_data.get("dad_export3d_path"):
                    export_data["dad_export3d_path"] = dad_data["dad_export3d_path"]

            # Afegir quantificació si existeix
            if "quantification" in sample_info:
                export_data["quantification"] = sample_info["quantification"]

            # Nom del fitxer (sense rèplica al nom — la rèplica consta al full ID)
            base = _sample_filename_base(sample_name, mode)
            filename = _make_unique_filename(base, ".xlsx", used_filenames)

            filepath = output_path / filename

            # Exportar Excel
            result = write_final_excel(
                str(filepath),
                sample_name,
                export_data,
                calibration_data,
                mode,
                config,
            )

            # Exportar CSV RAW (DOC cru + DAD 101λ)
            if export_raw:
                try:
                    rf = write_csv_raw(raw_dir, sample_name, export_data,
                                       mode, csv_separator, used_raw)
                    results["raw_files"].extend(rf)
                except Exception as e_csv:
                    logger.warning(f"CSV RAW {sample_name}: {e_csv}")

            # Exportar CSV PROCESSED (DOC net + fraccions)
            if export_processed:
                try:
                    pf = write_csv_processed(proc_dir, sample_name, export_data,
                                             mode, config, csv_separator, used_proc)
                    results["processed_files"].extend(pf)
                except Exception as e_csv:
                    logger.warning(f"CSV PROC {sample_name}: {e_csv}")

            results["files"].append({
                "sample": sample_name,
                "path": str(filepath),
                "doc_replica": doc_replica,
                "dad_replica": dad_replica,
                "has_bp": False,
            })
            results["n_exported"] += 1

        except Exception as e:
            results["errors"].append(f"{sample_name}: {str(e)}")
            results["n_errors"] += 1

    if progress_callback:
        progress_callback(100, "Exportació completada")

    results["n_skipped"] = n_skipped
    # Èxit conscient de la cobertura: no és èxit si no s'ha exportat res
    # (abans success=True encara que totes les mostres s'haguessin saltat).
    results["success"] = results["n_errors"] == 0 and results.get("n_exported", 0) > 0
    if results.get("n_exported", 0) == 0 and not results["errors"]:
        results.setdefault("warnings", []).append(
            f"No s'ha exportat cap mostra ({n_skipped} saltades). Revisa la selecció "
            "i els filtres (blancs/controls/mostres invàlides).")
    return results


# =============================================================================
# GENERACIÓ DE RESUM
# =============================================================================

def _build_summary_rows(samples_grouped, mode="COLUMN", config=None,
                        seq_name=None, seq_date=None):
    """Construeix les files del SUMMARY (font ÚNICA per a Excel i CSV).

    Esquema pensat per a ANÀLISI EXTERN (tidy):
    - Columnes d'identitat primer (Seq, Seq_Date, Mode) perquè es puguin
      CONCATENAR seqüències diferents i agrupar per campanya.
    - Conc_ppm SEMPRE numèric o buit (mai text tipus "NO VÀLIDA"), amb una
      columna booleana Valid a part. Així una eina externa parseja la columna
      com a numèrica sense trencar-se.
    """
    config = config or DEFAULT_EXPORT_CONFIG

    def _ident(row):
        """Prepend columnes d'identitat de seqüència."""
        base = {"Seq": seq_name or "", "Seq_Date": seq_date or "", "Mode": mode}
        base.update(row)
        return base

    summary_rows = []
    for sample_name in sorted(samples_grouped.keys()):
        sample_info = samples_grouped[sample_name]

        # Light samples (BLANK/CONTROL) — fila simplificada
        if sample_info.get("analysis_type") == "light":
            selected = sample_info.get("selected", {})
            doc_replica = selected.get("doc", "1")
            replicas = sample_info.get("replicas", {})
            doc_data = replicas.get(doc_replica, {})
            sample_type = sample_info.get("sample_type", "BLANK")
            summary_rows.append(_ident({
                "Sample": sample_name,
                "Type": sample_type,
                "Inj_Index": doc_data.get("injection_index", ""),
                "DOC_Replica": doc_replica,
                "DAD_Replica": None,
                "Valid": True,
                "Conc_ppm": None,
                "Area_total": doc_data.get("area_total"),
                "A_UIB": None,
                "ppm_UIB": None,
                "A_254": None,
                "SNR_Direct": doc_data.get("snr"),
                "SNR_UIB": None,
                "SNR_254": None,
                "R2_DOC": None,
                "R2_DAD": None,
                "Anomalies": "",
                "Max_Severity": "",
                "Warnings": "",
                "HCI": None,
                "HCI_Character": "",
            }))
            continue

        selected = sample_info.get("selected", {})
        quantification = sample_info.get("quantification", {})
        comparison = sample_info.get("comparison", {})

        doc_replica = selected.get("doc", "1")
        dad_replica = selected.get("dad", "1")

        is_invalid = (doc_replica == "none"
                      or sample_info.get("sample_valid") is False
                      or sample_info.get("skip_quantification", False))

        doc_data = sample_info.get("replicas", {}).get(doc_replica, {})
        snr_info = doc_data.get("snr_info", {})
        sample_type = doc_data.get("sample_type", "SAMPLE")

        doc_warnings = comparison.get("doc", {}).get("warnings", []) if comparison else []
        dad_warnings = comparison.get("dad", {}).get("warnings", []) if comparison else []
        all_warnings = doc_warnings + dad_warnings
        anomalies = doc_data.get("anomalies", [])

        dad_data = sample_info.get("replicas", {}).get(dad_replica, {}) if dad_replica != "none" else {}
        dad_areas = (dad_data.get("areas") or {})
        area_254 = dad_areas.get("A254", {}).get("total", 0)
        snr_info_dad = dad_data.get("snr_info_dad") or {}
        snr_254 = snr_info_dad.get("A254", {}).get("snr", 0)

        r2_doc = comparison.get("doc", {}).get("pearson", 0) if comparison else 0
        r2_dad = comparison.get("dad", {}).get("pearson_min", 0) if comparison else 0

        areas_uib = doc_data.get("areas_uib") or {}
        area_uib = areas_uib.get("total", 0)
        ppm_uib = quantification.get("concentration_ppm_uib")

        # Conc_ppm SEMPRE numèric o None (mai text) + Valid a part
        conc = None if is_invalid else quantification.get("concentration_ppm")

        summary_rows.append(_ident({
            "Sample": sample_name,
            "Type": sample_type,
            "Inj_Index": doc_data.get("injection_index", ""),
            "DOC_Replica": "" if doc_replica == "none" else doc_replica,
            "DAD_Replica": "" if dad_replica == "none" else dad_replica,
            "Valid": not is_invalid,
            "Conc_ppm": conc,
            "Area_total": quantification.get("area_total"),
            "A_UIB": area_uib if area_uib else None,
            "ppm_UIB": ppm_uib,
            "A_254": area_254 if area_254 else None,
            "SNR_Direct": snr_info.get("snr_direct"),
            "SNR_UIB": snr_info.get("snr_uib"),
            "SNR_254": snr_254 if snr_254 else None,
            "R2_DOC": round(r2_doc, 4) if r2_doc > 0 else None,
            "R2_DAD": round(r2_dad, 4) if r2_dad > 0 else None,
            "Anomalies": "; ".join(
                (a.get("code", "") + ("[R]" if a.get("repaired") else "")) if isinstance(a, dict)
                else str(a) for a in anomalies
            ) if anomalies else "",
            "Max_Severity": _get_max_severity(anomalies),
            "Warnings": "; ".join(
                (w.get("code", "") if isinstance(w, dict) else str(w)) for w in all_warnings
            ) if all_warnings else "",
            "HCI": quantification.get("hci"),
            "HCI_Character": quantification.get("hci_character", ""),
        }))

    return summary_rows


def generate_summary_excel(
    samples_grouped: dict,
    output_path: str,
    calibration_data: dict = None,
    mode: str = "COLUMN",
    config: dict = None,
    seq_name: str = None,
    seq_date: str = None,
):
    """
    Genera un Excel resum amb totes les mostres.

    Fulls:
        SUMMARY: Una fila per mostra amb concentració, SNR, warnings, BP info
        CALIBRATION: Info de calibració usada

    Args:
        samples_grouped: Dict amb mostres agrupades
        output_path: Camí del fitxer Excel
        calibration_data: Dades de calibració
        mode: "BP" o "COLUMN"
        config: Configuració

    Returns:
        dict amb info d'exportació
    """
    config = config or DEFAULT_EXPORT_CONFIG

    # === FULL SUMMARY (font única de files) ===
    summary_rows = _build_summary_rows(samples_grouped, mode, config,
                                       seq_name=seq_name, seq_date=seq_date)
    df_summary = pd.DataFrame(summary_rows)

    # === FULL CALIBRATION ===
    cal_rows = []
    if calibration_data:
        for key, value in calibration_data.items():
            if not isinstance(value, (list, dict)):
                cal_rows.append((key, value))

    df_cal = pd.DataFrame(cal_rows, columns=["Field", "Value"]) if cal_rows else pd.DataFrame()

    # Escriure Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="SUMMARY", index=False)
        if not df_cal.empty:
            df_cal.to_excel(writer, sheet_name="CALIBRATION", index=False)

    return {
        "success": True,
        "path": output_path,
        "n_samples": len(summary_rows),
    }


# =============================================================================
# PATCH EXCEL PER RECALIBRACIÓ
# =============================================================================

def patch_excel_calibration(excel_path, new_rf, new_intercept, mode="COLUMN",
                             new_rf_uib=None, new_intercept_uib=None):
    """
    Patch quirúrgic d'un Excel individual (per mostra) amb nous RF/intercept.

    Modifica NOMÉS les cel·les que depenen de la calibració:
    - Full ID: Concentration_ppm*, RF_mass_cal_*, Intercept_*
    - Full RESULTS: fila BP_ppm (si existeix)

    No regenera tot l'Excel — patch via openpyxl.

    Args:
        excel_path: Path al fitxer Excel (.xlsx)
        new_rf: Nou RF mass cal per direct
        new_intercept: Nou intercept per direct
        mode: "COLUMN" o "BP"
        new_rf_uib: RF per UIB (None = no tocar)
        new_intercept_uib: Intercept per UIB (None = no tocar)

    Returns:
        dict amb {success, patched_fields, errors}
    """
    import openpyxl

    result = {"success": False, "patched_fields": 0, "errors": []}

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        result["errors"].append(f"Error obrint Excel: {e}")
        return result

    n_patched = 0
    mode_key = mode.lower()
    signal_key = "direct"

    # === FULL ID ===
    if "ID" in wb.sheetnames:
        ws = wb["ID"]
        # El full ID és clau-valor: col A = nom camp, col B = valor
        # Primer llegir àrea i volum per recalcular concentració
        area_total = None
        volume_uL = None
        area_uib = None

        for row in ws.iter_rows(min_col=1, max_col=2):
            key = row[0].value
            if key == "Area_DOC_total":
                area_total = row[1].value
            elif key == "INJ_Volume_uL":
                volume_uL = row[1].value

        if area_total is not None and volume_uL and volume_uL > 0 and new_rf > 0:
            # Recalcular concentracions
            ppm = max(0, float(area_total) - new_intercept) * 1000 / (new_rf * float(volume_uL))

            for row in ws.iter_rows(min_col=1, max_col=2):
                key = row[0].value
                if key == "Concentration_ppm":
                    row[1].value = round(ppm, 3)
                    n_patched += 1
                elif key == "Concentration_ppm_Direct":
                    row[1].value = round(ppm, 3)
                    n_patched += 1
                elif key == "Concentration_ppm_UIB" and new_rf_uib and new_rf_uib > 0:
                    # Cal llegir àrea UIB — buscar Area_UIB_total o recalcular
                    # Per ara, marcar com a pendent
                    pass
                elif key == f"RF_mass_cal_{signal_key}_{mode_key}":
                    row[1].value = new_rf
                    n_patched += 1
                elif key == f"Intercept_{signal_key}_{mode_key}":
                    row[1].value = new_intercept
                    n_patched += 1

    # === FULL ID_BP (si existeix, per COLUMN amb BP vinculada) ===
    if "ID_BP" in wb.sheetnames:
        ws_bp = wb["ID_BP"]
        bp_area = None
        bp_volume = None
        for row in ws_bp.iter_rows(min_col=1, max_col=2):
            key = row[0].value
            if key == "Area_BP_total":
                bp_area = row[1].value
            elif key == "INJ_Volume_uL":
                bp_volume = row[1].value

        # BP usa el seu propi RF (no el COLUMN)
        # No patchejar BP des d'aquí — caldria RF_BP separat

    # === FULL RESULTS (fila BP_ppm) ===
    if "RESULTS" in wb.sheetnames:
        ws_res = wb["RESULTS"]
        for row in ws_res.iter_rows(min_col=1, max_col=3):
            if row[0].value == "BP_ppm":
                # Buscar àrea BP a la fila anterior (BP_total)
                bp_total_val = None
                for r2 in ws_res.iter_rows(min_col=1, max_col=3):
                    if r2[0].value == "BP_total" and r2[2].value not in (None, "-"):
                        bp_total_val = r2[2].value
                        break
                # BP_ppm es calcula amb RF_BP, no amb RF_COLUMN
                # No patchejar aquí si no tenim RF_BP
                break

    try:
        wb.save(excel_path)
        result["success"] = True
        result["patched_fields"] = n_patched
    except Exception as e:
        result["errors"].append(f"Error guardant Excel: {e}")

    return result


# =============================================================================
# EXPORTACIÓ CSV (FAIR — format obert)
# =============================================================================

# Cadència de downsampling DAD RAW (0.04 min = 2.4s)
DAD_RAW_TARGET_DT_MIN = 0.04


def _csv_metadata_header(fields, comment_char="#"):
    """Genera línies de capçalera amb metadades per CSV.

    Args:
        fields: llista de tuples (nom, valor)
        comment_char: caràcter de comentari

    Returns:
        str amb les línies de metadades
    """
    lines = []
    for name, value in fields:
        lines.append(f"{comment_char} {name}: {value}")
    return "\n".join(lines)


def _downsample_2d(t, data_2d, target_dt):
    """Downsample matriu 2D (temps × columnes) per bin-average.

    Args:
        t: array temps (n_points,)
        data_2d: array 2D (n_points, n_cols) o DataFrame
        target_dt: cadència objectiu (min)

    Returns:
        (t_new, data_new) arrays downsampled
    """
    t = np.asarray(t, dtype=float)
    if hasattr(data_2d, 'values'):
        data_2d = data_2d.values
    data_2d = np.asarray(data_2d, dtype=float)

    dt_median = np.median(np.diff(t))
    if dt_median >= target_dt * 0.8:
        return t, data_2d  # ja prou espaiats

    t_min, t_max = t[0], t[-1]
    bins = np.arange(t_min, t_max + target_dt, target_dt)
    n_bins = len(bins) - 1
    if n_bins < 2:
        return t, data_2d

    t_new = np.zeros(n_bins)
    data_new = np.zeros((n_bins, data_2d.shape[1] if data_2d.ndim > 1 else 1))
    if data_2d.ndim == 1:
        data_2d = data_2d.reshape(-1, 1)

    indices = np.digitize(t, bins) - 1
    indices = np.clip(indices, 0, n_bins - 1)

    for b in range(n_bins):
        mask = indices == b
        if mask.any():
            t_new[b] = np.mean(t[mask])
            data_new[b] = np.mean(data_2d[mask], axis=0)
        else:
            t_new[b] = (bins[b] + bins[b + 1]) / 2
            # Interpolar des del punt més proper
            nearest = np.argmin(np.abs(t - t_new[b]))
            data_new[b] = data_2d[nearest]

    return t_new, data_new


def write_csv_raw(
    out_dir: str,
    sample_name: str,
    sample_data: dict,
    mode: str = "COLUMN",
    separator: str = ";",
    used_filenames: set = None,
):
    """
    Exporta dades RAW (sense processar) per una mostra a subcarpeta RAW/.

    Genera:
    - {sample}_HPSEC_{C|B}_DOC_RAW.csv: DOC Direct RAW + UIB RAW
    - {sample}_HPSEC_{C|B}_DAD_RAW.csv: Export3D complet (101λ, downsampled dt=0.04 min)
      Per BP: 1 sola fila a t_max amb totes les λ

    Args:
        out_dir: Directori RAW/ de sortida
        sample_name: Nom de la mostra
        sample_data: Dict amb dades de la rèplica seleccionada
        mode: "BP" o "COLUMN"
        separator: Separador CSV
        used_filenames: Set per evitar col·lisions

    Returns:
        list de fitxers generats
    """
    if used_filenames is None:
        used_filenames = set()

    raw_path = Path(out_dir)
    raw_path.mkdir(parents=True, exist_ok=True)
    files_created = []
    base = _sample_filename_base(sample_name, mode)

    # --- DOC RAW ---
    t_doc = sample_data.get("t_doc")
    if t_doc is not None:
        t_doc = np.asarray(t_doc)
        data = {"time_min": t_doc}

        # DOC Direct RAW (senyal cru, sense shift/baseline/smoothing)
        y_doc_raw = sample_data.get("y_doc_raw")
        if y_doc_raw is not None and len(y_doc_raw) == len(t_doc):
            data["DOC_Direct_RAW_mAU"] = np.asarray(y_doc_raw)

        # DOC UIB RAW
        y_uib_raw = sample_data.get("y_doc_uib_raw")
        if y_uib_raw is not None and len(y_uib_raw) == len(t_doc):
            data["DOC_UIB_RAW_mAU"] = np.asarray(y_uib_raw)

        # Si no tenim RAW, usar el senyal disponible
        if "DOC_Direct_RAW_mAU" not in data and "DOC_UIB_RAW_mAU" not in data:
            y_doc = sample_data.get("y_doc")
            if y_doc is not None and len(y_doc) == len(t_doc):
                data["DOC_Direct_RAW_mAU"] = np.asarray(y_doc)

        if len(data) > 1:  # hi ha dades a part del temps
            meta = [
                ("Sample", sample_name),
                ("Method", mode),
                ("Data_Type", "RAW"),
                ("Signal", "DOC"),
                ("Export_Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
                ("Export_Module", f"hpsec_export v{__version__}"),
            ]
            fname = _make_unique_filename(f"{base}_DOC_RAW", ".csv", used_filenames)
            fpath = raw_path / fname
            header = _csv_metadata_header(meta)
            df = pd.DataFrame(data)
            with open(fpath, 'w', encoding='utf-8', newline='') as f:
                f.write(header + "\n")
                df.to_csv(f, sep=separator, index=False, float_format="%.6g")
            files_created.append(str(fpath))
            logger.info(f"CSV RAW DOC: {fpath}")

    # --- DAD RAW (101λ, downsampled) ---
    dad_export3d_path = sample_data.get("dad_export3d_path")
    if dad_export3d_path and os.path.exists(dad_export3d_path):
        try:
            from hpsec_import import llegir_dad_export3d
            df_dad_full, status = llegir_dad_export3d(dad_export3d_path, wavelengths_to_keep=None)
            if df_dad_full is not None and status.startswith("OK"):
                # Trobar columna temps
                t_col = None
                for candidate in ["time (min)", "Time"]:
                    if candidate in df_dad_full.columns:
                        t_col = candidate
                        break
                if t_col is None:
                    t_col = df_dad_full.columns[0]

                t_dad = df_dad_full[t_col].values
                wl_cols = [c for c in df_dad_full.columns if c != t_col]

                is_bp = mode == "BP" or sample_data.get("is_bp", False)

                if is_bp:
                    # BP: 1 sola fila a t_max (espectre complet)
                    peak_info = sample_data.get("peak_info", {})
                    t_max = peak_info.get("t_max")
                    if t_max is None:
                        # Estimar t_max des del pic DOC
                        if t_doc is not None:
                            y_net = sample_data.get("y_doc_net")
                            if y_net is not None:
                                t_max = t_doc[np.argmax(y_net)]
                    if t_max is not None:
                        idx = np.argmin(np.abs(t_dad - t_max))
                        row_data = {"wavelength_nm": [float(c) for c in wl_cols
                                                       if _is_numeric(c)]}
                        row_vals = [float(df_dad_full[c].iloc[idx]) for c in wl_cols
                                    if _is_numeric(c)]
                        row_data["absorbance_mAU"] = row_vals
                        meta = [
                            ("Sample", sample_name),
                            ("Method", "BP"),
                            ("Data_Type", "RAW"),
                            ("Signal", "DAD_spectrum_at_tmax"),
                            ("t_max_min", f"{t_max:.4f}"),
                            ("N_wavelengths", len(row_data["wavelength_nm"])),
                            ("Export_Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
                            ("Export_Module", f"hpsec_export v{__version__}"),
                        ]
                        fname = _make_unique_filename(f"{base}_DAD_RAW", ".csv", used_filenames)
                        fpath = raw_path / fname
                        header = _csv_metadata_header(meta)
                        df_out = pd.DataFrame(row_data)
                        with open(fpath, 'w', encoding='utf-8', newline='') as f:
                            f.write(header + "\n")
                            df_out.to_csv(f, sep=separator, index=False, float_format="%.6g")
                        files_created.append(str(fpath))
                        logger.info(f"CSV RAW DAD BP spectrum: {fpath}")
                else:
                    # COLUMN: totes les λ, downsample temps a dt=0.04 min
                    wl_numeric = [c for c in wl_cols if _is_numeric(c)]
                    dad_matrix = df_dad_full[wl_numeric].values
                    t_ds, data_ds = _downsample_2d(t_dad, dad_matrix, DAD_RAW_TARGET_DT_MIN)

                    meta = [
                        ("Sample", sample_name),
                        ("Method", mode),
                        ("Data_Type", "RAW"),
                        ("Signal", "DAD_Export3D"),
                        ("N_wavelengths", len(wl_numeric)),
                        ("N_timepoints_original", len(t_dad)),
                        ("N_timepoints_downsampled", len(t_ds)),
                        ("Downsample_dt_min", DAD_RAW_TARGET_DT_MIN),
                        ("Export_Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
                        ("Export_Module", f"hpsec_export v{__version__}"),
                    ]
                    fname = _make_unique_filename(f"{base}_DAD_RAW", ".csv", used_filenames)
                    fpath = raw_path / fname

                    header = _csv_metadata_header(meta)
                    # Construir DataFrame amb temps + λ
                    out_data = {"time_min": t_ds}
                    for j, wl in enumerate(wl_numeric):
                        out_data[wl] = data_ds[:, j]
                    df_out = pd.DataFrame(out_data)

                    with open(fpath, 'w', encoding='utf-8', newline='') as f:
                        f.write(header + "\n")
                        df_out.to_csv(f, sep=separator, index=False, float_format="%.6g")
                    files_created.append(str(fpath))
                    logger.info(f"CSV RAW DAD ({len(wl_numeric)}λ, {len(t_ds)} pts): {fpath}")
        except Exception as e:
            logger.warning(f"CSV RAW DAD {sample_name}: {e}")

    return files_created


def _is_numeric(s):
    """Comprova si un string és numèric (per columnes λ del DAD)."""
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def write_csv_processed(
    out_dir: str,
    sample_name: str,
    sample_data: dict,
    mode: str = "COLUMN",
    config: dict = None,
    separator: str = ";",
    used_filenames: set = None,
):
    """
    Exporta dades PROCESSED per una mostra a subcarpeta PROCESSED/.

    Genera:
    - {sample}_HPSEC_{C|B}_DOC.csv: DOC net (shift+baseline+smoothing+repair) + 6λ DAD
    - {sample}_HPSEC_{C|B}_FRACTIONS.csv: fraccions integrades + ppm

    Args:
        out_dir: Directori PROCESSED/ de sortida
        sample_name: Nom de la mostra
        sample_data: Dict amb dades de la rèplica seleccionada
        mode: "BP" o "COLUMN"
        config: Configuració (opcional)
        separator: Separador CSV
        used_filenames: Set per evitar col·lisions

    Returns:
        list de fitxers generats
    """
    if used_filenames is None:
        used_filenames = set()
    config = config or DEFAULT_EXPORT_CONFIG
    target_wls = config.get("target_wavelengths", [220, 252, 254, 272, 290, 362])

    proc_path = Path(out_dir)
    proc_path.mkdir(parents=True, exist_ok=True)
    files_created = []
    base = _sample_filename_base(sample_name, mode)

    t_doc = sample_data.get("t_doc")
    if t_doc is None:
        return files_created

    t_doc = np.asarray(t_doc)

    # --- DOC processat + DAD 6λ ---
    data = {"time_min": t_doc}

    y_doc_net = sample_data.get("y_doc_net")
    if y_doc_net is not None:
        data["DOC_net_mAU"] = np.asarray(y_doc_net)

    y_uib_net = sample_data.get("y_doc_uib_net")
    if y_uib_net is not None and len(y_uib_net) == len(t_doc):
        data["DOC_UIB_net_mAU"] = np.asarray(y_uib_net)

    y_direct_net = sample_data.get("y_doc_direct_net")
    if y_direct_net is not None and len(y_direct_net) == len(t_doc):
        data["DOC_Direct_net_mAU"] = np.asarray(y_direct_net)

    # DAD 6λ processat
    df_dad = sample_data.get("df_dad")
    if df_dad is not None and not df_dad.empty:
        t_dad = None
        for cand in ["time (min)", "Time"]:
            if cand in df_dad.columns:
                t_dad = df_dad[cand].values
                break

        if t_dad is not None:
            for wl in target_wls:
                dad_vals = _get_dad_column(df_dad, wl)
                if dad_vals is not None:
                    if len(t_dad) != len(t_doc):
                        data[f"A{wl}_mAU"] = np.interp(t_doc, t_dad, dad_vals)
                    else:
                        data[f"A{wl}_mAU"] = dad_vals

    meta = [
        ("Sample", sample_name),
        ("Method", mode),
        ("Data_Type", "PROCESSED"),
        ("Signal", "DOC+DAD"),
        ("Replica_DOC", sample_data.get("replica", "?")),
        ("Replica_DAD", sample_data.get("replica_dad", sample_data.get("replica", "?"))),
        ("Export_Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Export_Module", f"hpsec_export v{__version__}"),
    ]
    quantification = sample_data.get("quantification", {})
    if quantification:
        conc = quantification.get("concentration_ppm")
        if conc is not None:
            meta.append(("Concentration_ppm", round(conc, 3)))

    fname = _make_unique_filename(f"{base}_DOC", ".csv", used_filenames)
    fpath = proc_path / fname
    header = _csv_metadata_header(meta)
    df = pd.DataFrame(data)
    with open(fpath, 'w', encoding='utf-8', newline='') as f:
        f.write(header + "\n")
        df.to_csv(f, sep=separator, index=False, float_format="%.6g")
    files_created.append(str(fpath))
    logger.info(f"CSV PROCESSED DOC: {fpath}")

    # --- Fraccions integrades (només COLUMN, BP no té fraccions) ---
    if mode == "BP":
        return files_created

    fractions = config.get("time_fractions", {})
    fraction_names = list(fractions.keys())

    if y_doc_net is not None:
        rows = []
        y_doc_net_arr = np.asarray(y_doc_net)
        for frac_name in fraction_names:
            t_start, t_end = fractions[frac_name]
            t_range = f"{t_start}-{t_end}"
            row = {"Fraction": frac_name, "Range_min": t_range}

            mask = (t_doc >= t_start) & (t_doc <= t_end)
            row["DOC"] = round(float(trapezoid(y_doc_net_arr[mask], t_doc[mask])), 3) if mask.any() else 0

            # DAD fraccions
            if df_dad is not None and not df_dad.empty and t_dad is not None:
                for wl in target_wls:
                    dad_vals = _get_dad_column(df_dad, wl)
                    if dad_vals is not None:
                        m = (t_dad >= t_start) & (t_dad <= t_end)
                        row[f"A{wl}"] = round(float(trapezoid(dad_vals[m], t_dad[m])), 3) if m.any() else 0
            rows.append(row)

        # Afegir quantificació
        if quantification and not sample_data.get("skip_quantification"):
            for row in rows:
                frac = row["Fraction"]
                frac_q = quantification.get("fractions", {}).get(frac, {})
                if frac_q:
                    row["ppm"] = round(frac_q.get("ppm", 0), 4)

        meta_frac = [
            ("Sample", sample_name),
            ("Method", mode),
            ("Data_Type", "PROCESSED"),
            ("Signal", "FRACTIONS"),
            ("Export_Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        fname_f = _make_unique_filename(f"{base}_FRACTIONS", ".csv", used_filenames)
        fpath_f = proc_path / fname_f
        header_f = _csv_metadata_header(meta_frac)
        df_f = pd.DataFrame(rows)
        with open(fpath_f, 'w', encoding='utf-8', newline='') as f:
            f.write(header_f + "\n")
            df_f.to_csv(f, sep=separator, index=False, float_format="%.6g")
        files_created.append(str(fpath_f))
        logger.info(f"CSV PROCESSED FRACTIONS: {fpath_f}")

    return files_created


def _get_dad_column(df_dad, wl):
    """Obté valors d'una longitud d'ona del DataFrame DAD."""
    for col_name in [str(wl), f"A{wl}", str(int(wl))]:
        if col_name in df_dad.columns:
            return df_dad[col_name].values
    for col in df_dad.columns:
        try:
            if abs(float(col) - wl) < 1:
                return df_dad[col].values
        except (ValueError, TypeError):
            pass
    return None


def generate_summary_csv(
    samples_grouped: dict,
    output_path: str,
    calibration_data: dict = None,
    mode: str = "COLUMN",
    config: dict = None,
    separator: str = ";",
    seq_name: str = None,
    seq_date: str = None,
):
    """
    Genera CSV resum amb totes les mostres (equivalent a SUMMARY.xlsx).

    Args:
        samples_grouped: Dict amb mostres agrupades
        output_path: Camí del fitxer CSV
        calibration_data: Dades de calibració
        mode: "BP" o "COLUMN"
        config: Configuració (opcional)
        separator: Separador CSV (";", ",", o "\t")
    """
    config = config or DEFAULT_EXPORT_CONFIG

    # Metadades
    from hpsec_version import SUITE_VERSION
    meta_fields = [
        ("Suite_Version", f"HPSEC Suite v{SUITE_VERSION}"),
        ("Method", mode),
        ("N_Samples", len(samples_grouped)),
        ("Export_Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Export_Module", f"hpsec_export v{__version__}"),
    ]
    try:
        from hpsec_config import ConfigManager as Config
        cfg_obj = Config()
        meta_fields.append(("Config_Fingerprint", cfg_obj.compute_fingerprint()))
    except Exception:
        pass
    try:
        from hpsec_calibrate import compute_calibration_fingerprint
        meta_fields.append(("Calibration_Fingerprint", compute_calibration_fingerprint()))
    except Exception:
        pass
    meta_fields.append(("Separator", repr(separator)))

    header_text = _csv_metadata_header(meta_fields)

    # Mateixes files que l'Excel (font única _build_summary_rows) → esquema idèntic
    summary_rows = _build_summary_rows(samples_grouped, mode, config,
                                       seq_name=seq_name, seq_date=seq_date)
    df = pd.DataFrame(summary_rows)

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        f.write(header_text + "\n")
        df.to_csv(f, sep=separator, index=False, float_format="%.6g")

    logger.info(f"CSV summary: {output_path}")
    return {"success": True, "path": output_path, "n_samples": len(summary_rows)}


# =============================================================================
# METADATA.JSON (FAIR)
# =============================================================================

def write_metadata_json(
    output_path: str,
    samples_grouped: dict,
    mode: str = "COLUMN",
    calibration_data: dict = None,
    config: dict = None,
    seq_path: str = None,
    export_options: dict = None,
):
    """
    Genera metadata.json amb informació FAIR de traçabilitat.

    Args:
        output_path: Camí del fitxer metadata.json
        samples_grouped: Dict amb mostres agrupades
        mode: "BP" o "COLUMN"
        calibration_data: Dades de calibració
        config: Configuració
        seq_path: Path de la seqüència
        export_options: Opcions d'exportació seleccionades
    """
    from hpsec_version import SUITE_VERSION
    metadata = {
        "suite_version": SUITE_VERSION,
        "export_module": __version__,
        "export_date": datetime.now().isoformat(),
        "encoding": "UTF-8",
        "decimal_separator": ".",
        "method": mode,
        "seq_name": Path(seq_path).name if seq_path else "",
        "n_samples": len(samples_grouped),
    }

    # Llista de mostres
    sample_list = []
    for name, info in samples_grouped.items():
        s = {"name": name}
        s["type"] = info.get("sample_type", info.get("replicas", {}).get(
            info.get("selected", {}).get("doc", "1"), {}
        ).get("sample_type", "SAMPLE"))
        if info.get("quantification"):
            s["concentration_ppm"] = info["quantification"].get("concentration_ppm")
        sample_list.append(s)
    metadata["samples"] = sample_list

    # Fingerprints
    try:
        from hpsec_config import ConfigManager as Config
        cfg_obj = Config()
        metadata["config_fingerprint"] = cfg_obj.compute_fingerprint()
    except Exception:
        pass
    try:
        from hpsec_calibrate import compute_calibration_fingerprint
        metadata["calibration_fingerprint"] = compute_calibration_fingerprint()
    except Exception:
        pass

    # Calibració activa
    if calibration_data:
        cal_info = {}
        for key in ["rf_mass_cal", "intercept", "r2", "n_points", "calibration_date"]:
            if key in calibration_data:
                cal_info[key] = calibration_data[key]
        if cal_info:
            metadata["calibration"] = cal_info

    # Opcions d'exportació
    if export_options:
        metadata["export_options"] = export_options

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False, default=str)

    logger.info(f"metadata.json: {output_path}")


# =============================================================================
# ZIP PACKAGING
# =============================================================================

def create_export_zip(source_dir: str, zip_path: str = None):
    """
    Empaqueta tot el contingut d'exportació en un ZIP.

    Args:
        source_dir: Directori arrel a empaquetar (conté RAW/, PROCESSED/, SUMMARY, metadata)
        zip_path: Camí del ZIP. Si None, usa {source_dir}.zip

    Returns:
        str path del ZIP creat
    """
    source = Path(source_dir)
    if zip_path is None:
        zip_path = str(source) + ".zip"

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fpath in sorted(source.rglob("*")):
            if fpath.is_file():
                arcname = fpath.relative_to(source)
                zf.write(fpath, arcname)

    logger.info(f"ZIP creat: {zip_path} ({os.path.getsize(zip_path) / 1024 / 1024:.1f} MB)")
    return zip_path


# =============================================================================
# EXPORTS
# =============================================================================
__all__ = [
    "write_final_excel",
    "write_light_excel",
    "export_sequence",
    "generate_summary_excel",
    "generate_summary_csv",
    "write_csv_raw",
    "write_csv_processed",
    "write_metadata_json",
    "create_export_zip",
    "patch_excel_calibration",
    "DEFAULT_EXPORT_CONFIG",
]
