#!/usr/bin/env python3
"""
EEM Processing Pipeline - Versió Multi-Mostra per Windows
==========================================================
Pipeline complet per processar centenars d'EEMs del Cary Eclipse.

INSTRUCCIONS:
1. Instal·la Anaconda (https://www.anaconda.com/download)
2. Obre Anaconda Prompt i fes: pip install tensorly openpyxl
3. Executa: python eem_pipeline_multi.py
4. S'obriran finestres per seleccionar les carpetes i opcions
5. Espera i mira la carpeta de resultats!

Estructura de carpetes esperada:
    carpeta_eem/
        FR1643.csv
        FR1644.csv
        ...
    carpeta_uvvis/
        FR1643.xlsx
        FR1644.xlsx
        ...
    Raman.csv  (un sol fitxer)

Sortides:
    resultats/
        resum_indexs.xlsx          <- Taula amb tots els índexs per mostra
        parafac_scores.xlsx        <- Scores PARAFAC per mostra
        parafac_components.png     <- Espectres dels components
        seleccio_components.png    <- Gràfic per triar n_components
        eem_processades/           <- EEMs individuals (opcional)

Autora adaptació: Pipeline per investigació potabilització/DBPs
Basat en millores respecte drEEM (Murphy et al., 2013)
"""

import numpy as np
import csv
import os
import glob
import sys
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.interpolate import interp1d, RectBivariateSpline
from scipy.integrate import trapezoid
import matplotlib
matplotlib.use('Agg')  # Per evitar problemes amb displays
import matplotlib.pyplot as plt
import matplotlib.patheffects as pe
from matplotlib.colors import Normalize
import warnings
warnings.filterwarnings('ignore')


# ============================================================================
# 1. PARSING - Lectura de fitxers del Cary Eclipse
# ============================================================================

def parse_eem_cary(filepath):
    """
    Llegeix un fitxer CSV d'EEM exportat pel Cary Eclipse.
    """
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
    
    # Primera línia: capçaleres amb Ex wavelengths
    header = lines[0].strip().replace('\r', '').split(',')
    ex_wl = []
    ex_cols = []
    for i, h in enumerate(header):
        if 'EX_' in h.upper():
            # Extreure longitud d'ona d'excitació del nom de columna
            parts = h.upper().split('EX_')[1]
            wl = float(parts.split('_')[0])
            ex_wl.append(wl)
            ex_cols.append(i + 1)  # intensitat a la columna següent
    
    ex_wl = np.array(ex_wl)
    
    # Llegir dades d'emissió
    em_wl = []
    intensities = []
    for line in lines[2:]:
        parts = line.strip().replace('\r', '').split(',')
        if not parts[0]:
            break
        try:
            em = float(parts[0])
            em_wl.append(em)
            row = []
            for col_idx in ex_cols:
                if col_idx < len(parts) and parts[col_idx]:
                    row.append(float(parts[col_idx]))
                else:
                    row.append(0.0)
            intensities.append(row)
        except (ValueError, IndexError):
            break
    
    em_wl = np.array(em_wl)
    eem = np.array(intensities)
    
    return ex_wl, em_wl, eem


def parse_raman_cary(filepath):
    """
    Llegeix l'espectre Raman del blanc (aigua MQ).
    """
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
    
    em_wl = []
    intensity = []
    for line in lines[2:]:
        parts = line.strip().replace('\r', '').split(',')
        if len(parts) >= 2 and parts[0] and parts[1]:
            try:
                em_wl.append(float(parts[0]))
                intensity.append(float(parts[1]))
            except ValueError:
                break
    
    return np.array(em_wl), np.array(intensity)


def parse_uvvis(filepath):
    """
    Llegeix l'espectre UV-Vis des d'un xlsx.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    wl = []
    abs_vals = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            try:
                w = float(row[0])
                a = float(row[1])
                wl.append(w)
                abs_vals.append(a)
            except (ValueError, TypeError):
                continue
    
    wl = np.array(wl)
    abs_vals = np.array(abs_vals)
    
    # Ordenar per longitud d'ona creixent
    sort_idx = np.argsort(wl)
    wl = wl[sort_idx]
    abs_vals = abs_vals[sort_idx]
    
    return wl, abs_vals


# ============================================================================
# 2. PREPROCESSAMENT
# ============================================================================

def calculate_raman_area(raman_em, raman_int, low=381, high=426):
    """
    Calcula l'àrea del pic Raman de l'aigua (Ex=350 nm).
    Rang estàndard: Em 381-426 nm.
    """
    mask = (raman_em >= low) & (raman_em <= high)
    if mask.sum() < 3:
        mask = (raman_em >= 371) & (raman_em <= 436)
    area = trapezoid(raman_int[mask], raman_em[mask])
    return area


def remove_scattering_interpolate(ex_wl, em_wl, eem,
                                    rayleigh1_width=15,
                                    rayleigh2_width=15,
                                    raman1_width=12,
                                    raman2_width=12):
    """
    Elimina bandes de scattering per INTERPOLACIÓ (millor que blanking de drEEM).
    
    Rayleigh 1r ordre: Em = Ex
    Rayleigh 2n ordre: Em = 2*Ex  
    Raman 1r ordre: Em ≈ 1/(1/Ex - 3400e-7)
    Raman 2n ordre: Em ≈ 2 * Raman_1r
    """
    eem_clean = eem.copy()
    mask = np.ones_like(eem, dtype=bool)
    
    for j, ex in enumerate(ex_wl):
        raman1_em = 1.0 / (1.0/ex - 3400e-7)
        for i, em in enumerate(em_wl):
            if abs(em - ex) < rayleigh1_width:
                mask[i, j] = False
            if abs(em - 2*ex) < rayleigh2_width:
                mask[i, j] = False
            if abs(em - raman1_em) < raman1_width:
                mask[i, j] = False
            if abs(em - 2*raman1_em) < raman2_width:
                mask[i, j] = False
            if em <= ex:
                mask[i, j] = False
    
    # Interpolar zones eliminades
    for j in range(len(ex_wl)):
        valid = mask[:, j]
        if valid.sum() > 5:
            f_interp = interp1d(em_wl[valid], eem_clean[valid, j],
                               kind='cubic', bounds_error=False, fill_value=0)
            eem_clean[~valid, j] = f_interp(em_wl[~valid])
        else:
            eem_clean[~valid, j] = 0
    
    eem_clean = np.maximum(eem_clean, 0)
    return eem_clean, mask


def correct_ife(ex_wl, em_wl, eem, uvvis_wl, uvvis_abs, pathlength_cm=1.0):
    """
    Correcció d'Inner Filter Effect (Lakowicz).
    F_corr = F_obs * 10^((A_ex + A_em)/2)
    """
    abs_interp = interp1d(uvvis_wl, uvvis_abs, kind='linear',
                          bounds_error=False, fill_value=0)
    
    abs_ex = abs_interp(ex_wl)
    abs_em = abs_interp(em_wl)
    
    correction = np.zeros_like(eem)
    for j in range(len(ex_wl)):
        for i in range(len(em_wl)):
            a_total = (abs_ex[j] + abs_em[i]) * pathlength_cm / 2.0
            correction[i, j] = 10**min(a_total, 1.5)
    
    return eem * correction


def normalize_to_raman_units(eem, raman_area):
    """Normalitza a Raman Units (R.U.)."""
    return eem / raman_area


# ============================================================================
# 3. ÍNDEXS FLUORESCENTS
# ============================================================================

def get_eem_value(ex_wl, em_wl, eem, ex_target, em_target):
    """Valor de l'EEM al punt (ex, em) més proper."""
    j = np.argmin(np.abs(ex_wl - ex_target))
    i = np.argmin(np.abs(em_wl - em_target))
    return eem[i, j]


def get_eem_value_interp(ex_wl, em_wl, eem, ex_target, em_target):
    """Valor interpolat de l'EEM."""
    try:
        spline = RectBivariateSpline(em_wl, ex_wl, eem, kx=1, ky=1)
        return float(spline(em_target, ex_target)[0, 0])
    except:
        return get_eem_value(ex_wl, em_wl, eem, ex_target, em_target)


def calculate_indices(ex_wl, em_wl, eem, uvvis_wl=None, uvvis_abs=None, doc_mg_l=None):
    """
    Calcula tots els índexs fluorescents:
    
    FI:  Fluorescence Index (McKnight et al., 2001) - Em450/Em500 a Ex370
    BIX: Biological Index (Huguet et al., 2009) - Em380/Em430 a Ex310  
    HIX: Humification Index (Ohno, 2002) - àrea Em435-480 / àrea Em300-345 a Ex254
    Pics de Coble: A, C, M, T, B
    SUVA254: si DOC disponible
    """
    idx = {}
    
    # FI
    em450 = get_eem_value_interp(ex_wl, em_wl, eem, 370, 450)
    em500 = get_eem_value_interp(ex_wl, em_wl, eem, 370, 500)
    idx['FI'] = em450 / em500 if em500 != 0 else np.nan
    
    # BIX
    em380 = get_eem_value_interp(ex_wl, em_wl, eem, 310, 380)
    em430 = get_eem_value_interp(ex_wl, em_wl, eem, 310, 430)
    idx['BIX'] = em380 / em430 if em430 != 0 else np.nan
    
    # HIX
    ex254_idx = np.argmin(np.abs(ex_wl - 254))
    mask_high = (em_wl >= 435) & (em_wl <= 480)
    mask_low = (em_wl >= 300) & (em_wl <= 345)
    if mask_high.sum() > 0 and mask_low.sum() > 0:
        area_high = trapezoid(eem[mask_high, ex254_idx], em_wl[mask_high])
        area_low = trapezoid(eem[mask_low, ex254_idx], em_wl[mask_low])
        idx['HIX'] = area_high / (area_high + area_low) if (area_high + area_low) != 0 else np.nan
    else:
        idx['HIX'] = np.nan
    
    # Pics de Coble
    # A: Ex~260 / Em 400-460
    ex_a = np.argmin(np.abs(ex_wl - 260))
    m_a = (em_wl >= 400) & (em_wl <= 460)
    if m_a.sum() > 0:
        idx['Peak_A'] = np.max(eem[m_a, ex_a])
        idx['Peak_A_Em'] = em_wl[m_a][np.argmax(eem[m_a, ex_a])]
    
    # C: Ex~350 / Em 420-480
    ex_c = np.argmin(np.abs(ex_wl - 350))
    m_c = (em_wl >= 420) & (em_wl <= 480)
    if m_c.sum() > 0:
        idx['Peak_C'] = np.max(eem[m_c, ex_c])
        idx['Peak_C_Em'] = em_wl[m_c][np.argmax(eem[m_c, ex_c])]
    
    # M: Ex~312 / Em 380-420
    ex_m = np.argmin(np.abs(ex_wl - 312))
    m_m = (em_wl >= 380) & (em_wl <= 420)
    if m_m.sum() > 0:
        idx['Peak_M'] = np.max(eem[m_m, ex_m])
    
    # T: Ex~275 / Em~340
    idx['Peak_T'] = get_eem_value_interp(ex_wl, em_wl, eem, 275, 340)
    
    # B: Ex~275 / Em~305
    idx['Peak_B'] = get_eem_value_interp(ex_wl, em_wl, eem, 275, 305)
    
    # Ràtios
    if idx.get('Peak_A', 0) > 0 and idx.get('Peak_T', 0) > 0:
        idx['A_over_T'] = idx['Peak_A'] / idx['Peak_T']
    if idx.get('Peak_C', 0) > 0 and idx.get('Peak_A', 0) > 0:
        idx['C_over_A'] = idx['Peak_C'] / idx['Peak_A']
    if idx.get('Peak_C', 0) > 0 and idx.get('Peak_T', 0) > 0:
        idx['C_over_T'] = idx['Peak_C'] / idx['Peak_T']
    
    # Fluorescència total
    idx['Total_Fluorescence'] = trapezoid(trapezoid(eem, em_wl, axis=0), ex_wl)
    
    # UV-Vis related
    if uvvis_wl is not None and uvvis_abs is not None:
        abs_f = interp1d(uvvis_wl, uvvis_abs, kind='linear', bounds_error=False, fill_value=0)
        idx['Abs254'] = float(abs_f(254))
        idx['Abs280'] = float(abs_f(280))
        idx['Abs365'] = float(abs_f(365))
        idx['Abs254_over_Abs365'] = idx['Abs254'] / idx['Abs365'] if idx['Abs365'] > 0.001 else np.nan
        
        if doc_mg_l is not None and doc_mg_l > 0:
            idx['SUVA254'] = idx['Abs254'] / (doc_mg_l * 0.01)  # L/(mg·m)
            idx['DOC_mg_L'] = doc_mg_l
    
    return idx


# ============================================================================
# 4. PARAFAC - ALS amb restriccions
# ============================================================================

def parafac_als(X_3d, n_components, max_iter=500, tol=1e-6,
                n_inits=10, non_negative=True, verbose=True):
    """
    PARAFAC via Alternating Least Squares.
    
    Millores vs drEEM/N-way toolbox:
    - Múltiples inicialitzacions (evita mínims locals)
    - Restricció no-negativitat
    - Convergència monitoritzada
    
    Args:
        X_3d: tensor (n_mostres, n_em, n_ex)
        n_components: nombre de components
        n_inits: inicialitzacions aleatòries
    
    Returns:
        factors: [A(mostres), B(emissió), C(excitació)]
        loss_history: evolució error
        core_consistency: diagnòstic
    """
    n_samples, n_em, n_ex = X_3d.shape
    
    best_loss = np.inf
    best_factors = None
    best_history = None
    
    for init in range(n_inits):
        # Inicialització
        if init == 0:
            # SVD-based
            X_unfold2 = X_3d.reshape(n_em, -1) if n_samples > 0 else X_3d.reshape(n_em, -1)
            try:
                U2, s2, _ = np.linalg.svd(X_unfold2, full_matrices=False)
                B = np.abs(U2[:, :n_components]) * np.sqrt(s2[:n_components])
            except:
                B = np.abs(np.random.randn(n_em, n_components)) + 0.01
            
            X_unfold3 = X_3d.transpose(0, 2, 1).reshape(n_ex, -1)
            try:
                U3, s3, _ = np.linalg.svd(X_unfold3, full_matrices=False)
                C = np.abs(U3[:, :n_components]) * np.sqrt(s3[:n_components])
            except:
                C = np.abs(np.random.randn(n_ex, n_components)) + 0.01
            
            A = np.abs(np.random.randn(n_samples, n_components)) + 0.01
        else:
            A = np.abs(np.random.randn(n_samples, n_components)) + 0.01
            B = np.abs(np.random.randn(n_em, n_components)) + 0.01
            C = np.abs(np.random.randn(n_ex, n_components)) + 0.01
        
        loss_history = []
        
        for iteration in range(max_iter):
            # Update A: mode-1 unfolding X_(1) = A * (B ⊗ C)^T
            #   X_(1) shape: (n_samples, n_em*n_ex)
            #   Khatri-Rao B⊗C shape: (n_em*n_ex, R)
            X_1 = X_3d.reshape(n_samples, -1)
            kr = np.column_stack([np.kron(B[:, r], C[:, r]) for r in range(n_components)])
            gram = kr.T @ kr + 1e-12 * np.eye(n_components)
            A = X_1 @ kr @ np.linalg.inv(gram)
            if non_negative:
                A = np.maximum(A, 0)
            
            # Update B: mode-2 unfolding X_(2) = B * (A ⊗ C)^T
            #   X_(2) shape: (n_em, n_samples*n_ex)
            #   Khatri-Rao A⊗C shape: (n_samples*n_ex, R)
            X_2 = X_3d.transpose(1, 0, 2).reshape(n_em, -1)
            kr = np.column_stack([np.kron(A[:, r], C[:, r]) for r in range(n_components)])
            gram = kr.T @ kr + 1e-12 * np.eye(n_components)
            B = X_2 @ kr @ np.linalg.inv(gram)
            if non_negative:
                B = np.maximum(B, 0)
            
            # Update C: mode-3 unfolding X_(3) = C * (A ⊗ B)^T
            #   X_(3) shape: (n_ex, n_samples*n_em)
            #   Khatri-Rao A⊗B shape: (n_samples*n_em, R)
            X_3 = X_3d.transpose(2, 0, 1).reshape(n_ex, -1)
            kr = np.column_stack([np.kron(A[:, r], B[:, r]) for r in range(n_components)])
            gram = kr.T @ kr + 1e-12 * np.eye(n_components)
            C = X_3 @ kr @ np.linalg.inv(gram)
            if non_negative:
                C = np.maximum(C, 0)
            
            # Normalitzar B i C (absorbir escala a A) per estabilitat numèrica
            for r in range(n_components):
                norm_b = np.linalg.norm(B[:, r])
                norm_c = np.linalg.norm(C[:, r])
                if norm_b > 1e-10 and norm_c > 1e-10:
                    B[:, r] /= norm_b
                    C[:, r] /= norm_c
                    A[:, r] *= norm_b * norm_c
            
            # Calcular loss
            X_hat = np.zeros_like(X_3d)
            for r in range(n_components):
                for s in range(n_samples):
                    X_hat[s] += A[s, r] * np.outer(B[:, r], C[:, r])
            
            loss = np.sum((X_3d - X_hat)**2) / np.sum(X_3d**2)
            loss_history.append(loss)
            
            if iteration > 0 and abs(loss_history[-2] - loss_history[-1]) < tol:
                break
        
        if loss < best_loss:
            best_loss = loss
            best_factors = [A.copy(), B.copy(), C.copy()]
            best_history = loss_history.copy()
    
    explained = (1 - best_loss) * 100
    if verbose:
        print(f"    PARAFAC {n_components} comp: {explained:.2f}% var. explicada, "
              f"{len(best_history)} iter.")
    
    # Core consistency
    cc = estimate_core_consistency(X_3d, best_factors, n_components)
    
    return best_factors, best_history, cc, explained


def estimate_core_consistency(X_3d, factors, n_components):
    """
    Core Consistency Diagnostic (CORCONDIA).
    > 60% = bon model, < 40% = massa components.
    """
    A, B, C = factors
    try:
        A_pinv = np.linalg.pinv(A)
        B_pinv = np.linalg.pinv(B)
        C_pinv = np.linalg.pinv(C)
        
        G = np.tensordot(A_pinv, X_3d, axes=([1], [0]))
        G = np.tensordot(G, B_pinv.T, axes=([1], [0]))
        G = np.tensordot(G, C_pinv.T, axes=([1], [0]))
        
        ideal = np.zeros((n_components, n_components, n_components))
        for r in range(n_components):
            ideal[r, r, r] = G[r, r, r]
        
        cc = 100 * (1 - np.sum((G - ideal)**2) / np.sum(G**2))
        return max(cc, 0)
    except:
        return np.nan


def split_half_validation(X_3d, n_components, n_splits=6, n_inits=5):
    """
    Validació split-half del PARAFAC.
    Divideix el dataset en dues meitats i compara els components obtinguts.
    
    Retorna Tucker's Congruence Coefficient (TCC) per cada component.
    TCC > 0.95 = excel·lent, > 0.85 = acceptable.
    """
    n_samples = X_3d.shape[0]
    if n_samples < 6:
        print("    AVÍS: <6 mostres, split-half no fiable")
        return None
    
    tcc_all = []
    
    for split in range(n_splits):
        # Barrejar i dividir
        perm = np.random.permutation(n_samples)
        half = n_samples // 2
        X_a = X_3d[perm[:half]]
        X_b = X_3d[perm[half:2*half]]
        
        # PARAFAC a cada meitat
        factors_a, _, _, _ = parafac_als(X_a, n_components, n_inits=n_inits, verbose=False)
        factors_b, _, _, _ = parafac_als(X_b, n_components, n_inits=n_inits, verbose=False)
        
        # Tucker's Congruence Coefficient per emissió i excitació
        B_a, C_a = factors_a[1], factors_a[2]
        B_b, C_b = factors_b[1], factors_b[2]
        
        # Emparejar components per màxima similitud
        tccs = []
        for r in range(n_components):
            best_tcc = 0
            for s in range(n_components):
                # TCC per emissió
                num_b = np.dot(B_a[:, r], B_b[:, s])
                den_b = np.sqrt(np.dot(B_a[:, r], B_a[:, r]) * np.dot(B_b[:, s], B_b[:, s]))
                tcc_b = num_b / den_b if den_b > 0 else 0
                
                # TCC per excitació
                num_c = np.dot(C_a[:, r], C_b[:, s])
                den_c = np.sqrt(np.dot(C_a[:, r], C_a[:, r]) * np.dot(C_b[:, s], C_b[:, s]))
                tcc_c = num_c / den_c if den_c > 0 else 0
                
                tcc_mean = (tcc_b + tcc_c) / 2
                if tcc_mean > best_tcc:
                    best_tcc = tcc_mean
            tccs.append(best_tcc)
        
        tcc_all.append(tccs)
    
    tcc_mean = np.mean(tcc_all, axis=0)
    return tcc_mean


# ============================================================================
# 4b. DETECCIÓ D'OUTLIERS I CONTROL DE QUALITAT
# ============================================================================

def detect_outliers_eem(all_eems, sample_names, method='leverage'):
    """
    Detecta outliers en el conjunt d'EEMs.
    
    Mètodes combinats:
    1. Leverage: mostres que tenen una influència desproporcionada 
       en el model (basant-se en la distància al centroide del tensor)
    2. Residuals: mostres que el model no pot reconstruir bé
    3. Intensitat total: mostres amb fluorescència anormalment alta/baixa
    
    Retorna dict amb flags per cada mostra.
    """
    n_samples = len(all_eems)
    X_3d = np.stack(all_eems, axis=0)
    
    flags = {name: [] for name in sample_names}
    
    # --- 1. Intensitat total (IQR) ---
    total_fluor = np.array([eem.sum() for eem in all_eems])
    q1, q3 = np.percentile(total_fluor, [25, 75])
    iqr = q3 - q1
    lower = q1 - 2.0 * iqr  # 2x IQR (menys estricte que 1.5 per a dades ambientals)
    upper = q3 + 2.0 * iqr
    
    for i, name in enumerate(sample_names):
        if total_fluor[i] < lower:
            flags[name].append(f"Fluor. total molt baixa ({total_fluor[i]:.1f}, llindar={lower:.1f})")
        if total_fluor[i] > upper:
            flags[name].append(f"Fluor. total molt alta ({total_fluor[i]:.1f}, llindar={upper:.1f})")
    
    # --- 2. Leverage (distància al centroide) ---
    # Unfoldar el tensor i calcular distàncies de Mahalanobis simplificades
    X_unfold = X_3d.reshape(n_samples, -1)
    centroid = X_unfold.mean(axis=0)
    
    # Distàncies euclidianes al centroide (normalitzades)
    distances = np.array([np.linalg.norm(X_unfold[i] - centroid) for i in range(n_samples)])
    d_mean = distances.mean()
    d_std = distances.std()
    
    for i, name in enumerate(sample_names):
        if distances[i] > d_mean + 3 * d_std:
            flags[name].append(f"Leverage alt (dist={distances[i]:.1f}, llindar={d_mean + 3*d_std:.1f})")
    
    # --- 3. Valors negatius residuals (senyal de problemes instrumentals) ---
    for i, (eem, name) in enumerate(zip(all_eems, sample_names)):
        pct_neg = (eem < -0.001).sum() / eem.size * 100
        if pct_neg > 5:  # >5% de punts negatius
            flags[name].append(f"Molts valors negatius ({pct_neg:.1f}%)")
    
    # --- 4. Saturació del detector ---
    for i, (eem, name) in enumerate(zip(all_eems, sample_names)):
        # En R.U. no tenim un llindar absolut de saturació, però
        # si el màxim és >> que la mitjana dels màxims, pot ser saturació
        max_val = eem.max()
        if i == 0:
            all_maxvals = np.array([e.max() for e in all_eems])
            max_q3 = np.percentile(all_maxvals, 75)
            max_iqr = np.percentile(all_maxvals, 75) - np.percentile(all_maxvals, 25)
        
        if max_val > max_q3 + 3 * max_iqr:
            flags[name].append(f"Possible saturació (max={max_val:.3f})")
    
    # --- 5. Absorbància alta (IFE no fiable si Abs > 1.5) ---
    # Això es comprova durant el preprocessament, aquí és informatiu
    
    # --- 6. Índexs fluorescents fora de rang ---
    # Es comprova a part amb detect_index_outliers()
    
    # Resum
    n_outliers = sum(1 for f in flags.values() if f)
    return flags, distances, total_fluor


def detect_index_outliers(all_indices, sample_names):
    """
    Detecta mostres amb índexs fluorescents fora de rang (IQR).
    """
    keys_to_check = ['FI', 'BIX', 'HIX', 'Peak_A', 'Peak_C', 'Peak_T', 'Total_Fluorescence']
    
    flags = {name: [] for name in sample_names}
    
    for key in keys_to_check:
        values = []
        valid_idx = []
        for i, idx in enumerate(all_indices):
            v = idx.get(key, np.nan)
            if isinstance(v, (int, float)) and not np.isnan(v):
                values.append(v)
                valid_idx.append(i)
        
        if len(values) < 5:
            continue
        
        values = np.array(values)
        q1, q3 = np.percentile(values, [25, 75])
        iqr = q3 - q1
        lower = q1 - 2.0 * iqr
        upper = q3 + 2.0 * iqr
        
        for j, vi in enumerate(valid_idx):
            name = sample_names[vi]
            if values[j] < lower or values[j] > upper:
                flags[name].append(f"{key}={values[j]:.3f} (rang normal: {lower:.3f}-{upper:.3f})")
    
    return flags


def plot_outlier_diagnostics(sample_names, distances, total_fluor,
                             eem_flags, index_flags, output_path=None):
    """
    Gràfic de diagnòstic d'outliers: leverage vs fluorescència total.
    """
    fig, axes = plt.subplots(1, 2, figsize=(16, 7))
    
    # Combinar flags
    all_flagged = set()
    for name in sample_names:
        if eem_flags.get(name, []) or index_flags.get(name, []):
            all_flagged.add(name)
    
    # Colors
    colors = ['red' if n in all_flagged else 'steelblue' for n in sample_names]
    
    # --- Plot 1: Leverage vs Total Fluorescence ---
    ax = axes[0]
    ax.scatter(total_fluor, distances, c=colors, alpha=0.6, edgecolors='k', linewidths=0.5, s=50)
    
    # Marcar outliers amb nom
    for i, name in enumerate(sample_names):
        if name in all_flagged:
            ax.annotate(name, (total_fluor[i], distances[i]),
                       fontsize=7, ha='left', va='bottom', color='red',
                       xytext=(3, 3), textcoords='offset points')
    
    # Llindars
    d_mean = distances.mean()
    d_std = distances.std()
    ax.axhline(d_mean + 3*d_std, color='red', linestyle='--', alpha=0.5, label='Llindar leverage (3σ)')
    
    q1, q3 = np.percentile(total_fluor, [25, 75])
    iqr = q3 - q1
    ax.axvline(q1 - 2*iqr, color='orange', linestyle='--', alpha=0.5, label='Llindar fluor. (2×IQR)')
    ax.axvline(q3 + 2*iqr, color='orange', linestyle='--', alpha=0.5)
    
    ax.set_xlabel('Fluorescència total (R.U.)')
    ax.set_ylabel('Distància al centroide')
    ax.set_title('Diagnòstic d\'outliers: Leverage vs Fluorescència')
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)
    
    # --- Plot 2: Distribució de fluorescència total ---
    ax = axes[1]
    ax.hist(total_fluor, bins=min(30, len(sample_names)//3), color='steelblue', 
            alpha=0.7, edgecolor='black', linewidth=0.5)
    
    for i, name in enumerate(sample_names):
        if name in all_flagged:
            ax.axvline(total_fluor[i], color='red', alpha=0.3, linewidth=1)
    
    ax.axvline(q1 - 2*iqr, color='orange', linestyle='--', label='Llindar 2×IQR')
    ax.axvline(q3 + 2*iqr, color='orange', linestyle='--')
    
    ax.set_xlabel('Fluorescència total (R.U.)')
    ax.set_ylabel('Freqüència')
    ax.set_title('Distribució de fluorescència total')
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    if output_path:
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return fig


def export_quality_report(sample_names, eem_flags, index_flags, output_path):
    """
    Exporta un informe de qualitat (QA/QC) a Excel amb flags per mostra.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control de Qualitat"
    
    # Estils
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Capçalera
    headers = ['Mostra', 'Estat', 'Problemes EEM', 'Problemes Índexs', 'Recomanació']
    for j, h in enumerate(headers):
        cell = ws.cell(row=1, column=j+1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Dades
    n_ok = 0
    n_warn = 0
    n_bad = 0
    
    for i, name in enumerate(sample_names):
        e_flags = eem_flags.get(name, [])
        i_flags = index_flags.get(name, [])
        
        row = i + 2
        ws.cell(row=row, column=1, value=name).border = thin_border
        
        # Estat
        if not e_flags and not i_flags:
            estat = "OK"
            fill = ok_fill
            reco = "Incloure al PARAFAC"
            n_ok += 1
        elif e_flags:
            estat = "REVISAR"
            fill = bad_fill
            reco = "Revisar manualment, considerar excloure"
            n_bad += 1
        else:
            estat = "AVÍS"
            fill = warn_fill
            reco = "Incloure però amb precaució"
            n_warn += 1
        
        cell = ws.cell(row=row, column=2, value=estat)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=3, value='; '.join(e_flags) if e_flags else '-').border = thin_border
        ws.cell(row=row, column=4, value='; '.join(i_flags) if i_flags else '-').border = thin_border
        ws.cell(row=row, column=5, value=reco).border = thin_border
    
    # Resum al final
    row_sum = len(sample_names) + 3
    ws.cell(row=row_sum, column=1, value="RESUM").font = Font(bold=True)
    ws.cell(row=row_sum + 1, column=1, value=f"OK: {n_ok}")
    ws.cell(row=row_sum + 1, column=1).fill = ok_fill
    ws.cell(row=row_sum + 2, column=1, value=f"Avís: {n_warn}")
    ws.cell(row=row_sum + 2, column=1).fill = warn_fill
    ws.cell(row=row_sum + 3, column=1, value=f"Revisar: {n_bad}")
    ws.cell(row=row_sum + 3, column=1).fill = bad_fill
    
    # Ajustar amplades
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 35
    
    wb.save(output_path)
    print(f"  Informe QA/QC guardat a: {output_path}")


# ============================================================================
# 5. VISUALITZACIÓ
# ============================================================================

def plot_eem_contour(ex_wl, em_wl, eem, title="EEM", ax=None,
                     cmap='jet', levels=25, show_coble=True):
    """Gràfic de contorn de l'EEM."""
    if ax is None:
        fig, ax = plt.subplots(1, 1, figsize=(10, 7))
    
    EX, EM = np.meshgrid(ex_wl, em_wl)
    vmax = np.percentile(eem[eem > 0], 99) if (eem > 0).any() else 1
    levels_arr = np.linspace(0, vmax, levels)
    
    cf = ax.contourf(EX, EM, eem, levels=levels_arr, cmap=cmap, extend='max')
    ax.contour(EX, EM, eem, levels=levels_arr, colors='k', linewidths=0.2, alpha=0.3)
    plt.colorbar(cf, ax=ax, label='Intensitat (R.U.)')
    
    if show_coble:
        peaks = {'A': (260, 430), 'C': (350, 450), 'M': (312, 400),
                 'T': (275, 340), 'B': (275, 305)}
        for name, (ex, em) in peaks.items():
            if ex_wl[0] <= ex <= ex_wl[-1] and em_wl[0] <= em <= em_wl[-1]:
                ax.plot(ex, em, 'w*', markersize=12, markeredgecolor='k', markeredgewidth=0.5)
                ax.annotate(name, (ex, em), textcoords="offset points",
                           xytext=(5, 5), fontsize=9, fontweight='bold',
                           color='white',
                           path_effects=[pe.withStroke(linewidth=2, foreground='black')])
    
    ax.set_xlabel('Excitació (nm)')
    ax.set_ylabel('Emissió (nm)')
    ax.set_title(title)
    return ax


def plot_parafac_components(ex_wl, em_wl, factors, component_names=None, output_path=None):
    """Visualitza components PARAFAC: espectres Em, Ex i EEMs reconstruïdes."""
    A, B, C = factors
    n_comp = B.shape[1]
    
    if component_names is None:
        component_names = [f'C{i+1}' for i in range(n_comp)]
    
    colors = plt.cm.Set1(np.linspace(0, 1, max(n_comp, 3)))
    
    # Figura amb espectres + EEMs individuals
    fig = plt.figure(figsize=(6 * n_comp, 10))
    
    # Fila 1: Espectres d'emissió
    ax_em = fig.add_subplot(2, 1, 1)
    for r in range(n_comp):
        b_norm = B[:, r] / B[:, r].max() if B[:, r].max() > 0 else B[:, r]
        ax_em.plot(em_wl, b_norm, color=colors[r], linewidth=2, label=component_names[r])
    ax_em.set_xlabel('Emissió (nm)')
    ax_em.set_ylabel('Intensitat normalitzada')
    ax_em.set_title('Espectres d\'emissió dels components')
    ax_em.legend()
    ax_em.grid(True, alpha=0.3)
    
    # Fila 2: EEM de cada component
    for r in range(n_comp):
        ax = fig.add_subplot(2, n_comp, n_comp + r + 1)
        eem_comp = np.outer(B[:, r], C[:, r])
        EX, EM = np.meshgrid(ex_wl, em_wl)
        vmax = np.percentile(eem_comp, 99) if eem_comp.max() > 0 else 1
        ax.contourf(EX, EM, eem_comp, levels=15, cmap='jet')
        ax.set_title(f'{component_names[r]}')
        ax.set_xlabel('Ex (nm)')
        if r == 0:
            ax.set_ylabel('Em (nm)')
    
    plt.tight_layout()
    if output_path:
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return fig


def plot_component_selection(comp_results, output_path=None):
    """Gràfic per decidir el nombre de components."""
    n_comps = [r['n_components'] for r in comp_results]
    var_exp = [r['var_explained'] for r in comp_results]
    cc = [r['core_consistency'] for r in comp_results]
    
    fig, ax1 = plt.subplots(figsize=(10, 6))
    
    color1 = 'steelblue'
    ax1.plot(n_comps, var_exp, 'o-', color=color1, linewidth=2, markersize=8, label='Var. explicada (%)')
    ax1.set_xlabel('Nombre de components', fontsize=12)
    ax1.set_ylabel('Variància explicada (%)', color=color1, fontsize=12)
    ax1.tick_params(axis='y', labelcolor=color1)
    ax1.set_ylim(min(var_exp) - 5, 102)
    
    ax2 = ax1.twinx()
    color2 = 'darkorange'
    ax2.plot(n_comps, cc, 's--', color=color2, linewidth=2, markersize=8, label='Core Consistency (%)')
    ax2.set_ylabel('Core Consistency (%)', color=color2, fontsize=12)
    ax2.tick_params(axis='y', labelcolor=color2)
    ax2.axhline(60, color='red', linestyle=':', alpha=0.5, label='Llindar CC=60%')
    ax2.set_ylim(-5, 105)
    
    ax1.set_xticks(n_comps)
    ax1.set_title('Selecció del nombre de components PARAFAC', fontsize=14)
    
    # Llegenda combinada
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='center right', fontsize=10)
    
    ax1.grid(True, alpha=0.3)
    plt.tight_layout()
    
    if output_path:
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return fig


def plot_scores_boxplot(A, sample_names, component_names=None, output_path=None):
    """Distribució dels scores PARAFAC."""
    n_comp = A.shape[1]
    if component_names is None:
        component_names = [f'C{i+1}' for i in range(n_comp)]
    
    fig, ax = plt.subplots(figsize=(8, 5))
    bp = ax.boxplot([A[:, r] for r in range(n_comp)], labels=component_names,
                    patch_artist=True)
    colors = plt.cm.Set1(np.linspace(0, 1, max(n_comp, 3)))
    for patch, color in zip(bp['boxes'], colors):
        patch.set_facecolor(color)
        patch.set_alpha(0.6)
    
    ax.set_ylabel('Fmax (R.U.)')
    ax.set_title('Distribució dels scores PARAFAC (Fmax)')
    ax.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    if output_path:
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return fig


# ============================================================================
# 6. EXPORTACIÓ A EXCEL
# ============================================================================

def export_indices_to_excel(all_indices, sample_names, output_path):
    """
    Exporta tots els índexs a un Excel ben formatejat.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Índexs Fluorescents"
    
    # Estils
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill2 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Determinar totes les columnes
    all_keys = ['FI', 'BIX', 'HIX', 'Peak_A', 'Peak_A_Em', 'Peak_C', 'Peak_C_Em',
                'Peak_M', 'Peak_T', 'Peak_B', 'A_over_T', 'C_over_A', 'C_over_T',
                'Total_Fluorescence', 'Abs254', 'Abs280', 'Abs365', 'Abs254_over_Abs365',
                'SUVA254', 'DOC_mg_L']
    
    # Escriure capçalera
    ws.cell(row=1, column=1, value="Mostra").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border
    
    for j, key in enumerate(all_keys):
        cell = ws.cell(row=1, column=j+2, value=key)
        cell.font = header_font
        cell.fill = header_fill if j < 14 else header_fill2
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Escriure dades
    for i, (name, indices) in enumerate(zip(sample_names, all_indices)):
        ws.cell(row=i+2, column=1, value=name).border = thin_border
        for j, key in enumerate(all_keys):
            val = indices.get(key, '')
            cell = ws.cell(row=i+2, column=j+2)
            if isinstance(val, float) and not np.isnan(val):
                cell.value = round(val, 4)
                cell.number_format = '0.0000'
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    # Ajustar amplada de columnes
    for col in range(1, len(all_keys) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['A'].width = 20
    
    # Segona pestanya: interpretació
    ws2 = wb.create_sheet("Interpretació")
    interpretations = [
        ("Índex", "Descripció", "Valors referència"),
        ("FI", "Fluorescence Index (McKnight et al., 2001)", 
         "<1.4 terrestre | 1.4-1.9 mixt | >1.9 microbià/autòcton"),
        ("BIX", "Biological/Freshness Index (Huguet et al., 2009)",
         "<0.7 terrestre | 0.7-1.0 mixt | >1.0 autòcton recent"),
        ("HIX", "Humification Index (Ohno, 2002)",
         "<0.5 poc humificat | 0.5-0.8 moderat | >0.8 molt humificat"),
        ("Peak_A", "Àcids húmics (Ex~260/Em~400-460, Coble 1996)", "R.U."),
        ("Peak_C", "Àcids húmics (Ex~350/Em~420-480, Coble 1996)", "R.U."),
        ("Peak_M", "Húmic marí/microbià (Ex~312/Em~380-420)", "R.U."),
        ("Peak_T", "Triptòfan/proteic (Ex~275/Em~340)", "R.U."),
        ("Peak_B", "Tirosina/proteic (Ex~275/Em~305)", "R.U."),
        ("A/T", "Ràtio húmic/proteic", ">3 predomini húmic"),
        ("C/A", "Ràtio húmics C/A", "Relacionat amb aromaticitat"),
        ("SUVA254", "Specific UV Absorbance at 254nm", 
         "<2 baixa aromaticitat | 2-4 moderada | >4 alta aromaticitat"),
        ("Abs254/Abs365", "Ràtio E2/E3 (Peuravuori & Pihlaja, 1997)",
         "Invers del pes molecular; valors alts = PM baix"),
    ]
    for i, row_data in enumerate(interpretations):
        for j, val in enumerate(row_data):
            cell = ws2.cell(row=i+1, column=j+1, value=val)
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border
    
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 55
    
    wb.save(output_path)
    print(f"  Índexs guardats a: {output_path}")


def export_parafac_scores(A, sample_names, component_names, output_path):
    """Exporta scores PARAFAC (Fmax) a Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PARAFAC Scores (Fmax)"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Capçalera
    ws.cell(row=1, column=1, value="Mostra").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for j, name in enumerate(component_names):
        cell = ws.cell(row=1, column=j+2, value=f"Fmax_{name}")
        cell.font = header_font
        cell.fill = header_fill
    
    # Dades
    for i, sname in enumerate(sample_names):
        ws.cell(row=i+2, column=1, value=sname).border = thin_border
        for j in range(len(component_names)):
            cell = ws.cell(row=i+2, column=j+2, value=round(float(A[i, j]), 4))
            cell.number_format = '0.0000'
            cell.border = thin_border
    
    for col in range(1, len(component_names) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    wb.save(output_path)
    print(f"  Scores PARAFAC guardats a: {output_path}")


# ============================================================================
# 7. PIPELINE PRINCIPAL MULTI-MOSTRA
# ============================================================================

def find_matching_files(eem_folder, uvvis_folder, output_folder=None):
    """
    Troba parells de fitxers EEM+UV-Vis amb el mateix nom.
    Genera un informe detallat de concordances i discrepàncies.
    """
    # Buscar CSVs a la carpeta EEM
    eem_files = glob.glob(os.path.join(eem_folder, '*.csv'))
    eem_dict = {}
    for f in eem_files:
        name = os.path.splitext(os.path.basename(f))[0]
        eem_dict[name] = f
    
    # Buscar corresponents UV-Vis
    uvvis_files = glob.glob(os.path.join(uvvis_folder, '*.xlsx'))
    uvvis_dict = {}
    for f in uvvis_files:
        name = os.path.splitext(os.path.basename(f))[0]
        uvvis_dict[name] = f
    
    # Trobar parells i discrepàncies
    common = sorted(set(eem_dict.keys()) & set(uvvis_dict.keys()))
    eem_only = sorted(set(eem_dict.keys()) - set(uvvis_dict.keys()))
    uvvis_only = sorted(set(uvvis_dict.keys()) - set(eem_dict.keys()))
    
    # Informe detallat
    print(f"\n    ┌─────────────────────────────────────────────┐")
    print(f"    │  INFORME DE CONCORDANÇA DE FITXERS           │")
    print(f"    ├─────────────────────────────────────────────┤")
    print(f"    │  EEM (.csv) trobats:    {len(eem_dict):>5}                │")
    print(f"    │  UV-Vis (.xlsx) trobats: {len(uvvis_dict):>4}                │")
    print(f"    │  Parells concordants:   {len(common):>5}                │")
    print(f"    │  EEM sense UV-Vis:      {len(eem_only):>5}                │")
    print(f"    │  UV-Vis sense EEM:      {len(uvvis_only):>5}                │")
    print(f"    └─────────────────────────────────────────────┘")
    
    if eem_only:
        print(f"\n    ⚠ EEMs SENSE UV-Vis ({len(eem_only)} fitxers):")
        for name in eem_only[:20]:
            print(f"      - {name}")
        if len(eem_only) > 20:
            print(f"      ... i {len(eem_only) - 20} més")
    
    if uvvis_only:
        print(f"\n    ⚠ UV-Vis SENSE EEM ({len(uvvis_only)} fitxers):")
        for name in uvvis_only[:20]:
            print(f"      - {name}")
        if len(uvvis_only) > 20:
            print(f"      ... i {len(uvvis_only) - 20} més")
    
    # Guardar informe a fitxer si tenim output_folder
    if output_folder and (eem_only or uvvis_only):
        os.makedirs(output_folder, exist_ok=True)
        report_path = os.path.join(output_folder, 'informe_concordanca.txt')
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("INFORME DE CONCORDANÇA DE FITXERS\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"EEM (.csv) trobats:     {len(eem_dict)}\n")
            f.write(f"UV-Vis (.xlsx) trobats: {len(uvvis_dict)}\n")
            f.write(f"Parells concordants:    {len(common)}\n\n")
            
            if eem_only:
                f.write(f"EEMs SENSE UV-Vis ({len(eem_only)}):\n")
                for name in eem_only:
                    f.write(f"  {name}\n")
                f.write("\n")
            
            if uvvis_only:
                f.write(f"UV-Vis SENSE EEM ({len(uvvis_only)}):\n")
                for name in uvvis_only:
                    f.write(f"  {name}\n")
        
        print(f"\n    Informe guardat a: {report_path}")
    
    eem_paths = [eem_dict[n] for n in common]
    uvvis_paths = [uvvis_dict[n] for n in common]
    
    return eem_paths, uvvis_paths, common


def run_pipeline(eem_folder, uvvis_folder, raman_file, output_folder,
                 n_parafac=None, max_parafac_test=7,
                 pathlength_cm=1.0, doc_values=None,
                 save_individual_eems=False):
    """
    ====================================================================
    PIPELINE PRINCIPAL - Executa tot automàticament
    ====================================================================
    
    Paràmetres:
        eem_folder:     Carpeta amb els CSV d'EEM
        uvvis_folder:   Carpeta amb els XLSX d'UV-Vis
        raman_file:     Path al fitxer Raman (blanc MQ)
        output_folder:  Carpeta de sortida (es crea si no existeix)
        
        n_parafac:      Nombre de components PARAFAC (None = automàtic)
        max_parafac_test: Màxim components a testar per selecció automàtica
        pathlength_cm:  Pas òptic de la cubeta
        doc_values:     Dict {nom_mostra: DOC_mg_L} o None
        save_individual_eems: Guardar EEM processada de cada mostra com PNG
    """
    
    print("\n" + "=" * 70)
    print("  EEM PROCESSING PIPELINE")
    print("  Millores respecte drEEM: interpolació scattering, IFE, índexs,")
    print("  PARAFAC multi-init amb CORCONDIA i split-half validation")
    print("=" * 70)
    
    # Crear carpeta de sortida
    os.makedirs(output_folder, exist_ok=True)
    if save_individual_eems:
        os.makedirs(os.path.join(output_folder, 'eem_processades'), exist_ok=True)
    
    # --- 1. Trobar fitxers ---
    print(f"\n[1] Buscant fitxers...")
    print(f"    EEM folder:   {eem_folder}")
    print(f"    UV-Vis folder: {uvvis_folder}")
    
    eem_paths, uvvis_paths, sample_names = find_matching_files(eem_folder, uvvis_folder, output_folder)
    n_samples = len(sample_names)
    print(f"    Trobats {n_samples} parells EEM + UV-Vis")
    
    if n_samples == 0:
        print("\n  ERROR: No s'han trobat parells de fitxers!")
        print("  Comprova que els CSVs d'EEM i els XLSX d'UV-Vis tenen el MATEIX nom")
        return None
    
    # --- 2. Raman ---
    print(f"\n[2] Llegint Raman...")
    raman_em, raman_int = parse_raman_cary(raman_file)
    raman_area = calculate_raman_area(raman_em, raman_int)
    print(f"    Àrea Raman: {raman_area:.2f}")
    
    # --- 3. Processar cada mostra ---
    print(f"\n[3] Processant {n_samples} mostres...")
    all_eems = []
    all_indices = []
    common_ex = None
    common_em = None
    errors = []
    
    t_start = time.time()
    
    for i, (eem_f, uv_f, name) in enumerate(zip(eem_paths, uvvis_paths, sample_names)):
        try:
            # Lectura
            ex_wl, em_wl, eem_raw = parse_eem_cary(eem_f)
            uvvis_wl, uvvis_abs = parse_uvvis(uv_f)
            
            if common_ex is None:
                common_ex = ex_wl
                common_em = em_wl
            
            # Preprocessament
            eem_clean, _ = remove_scattering_interpolate(ex_wl, em_wl, eem_raw)
            eem_ife = correct_ife(ex_wl, em_wl, eem_clean, uvvis_wl, uvvis_abs, pathlength_cm)
            eem_ru = normalize_to_raman_units(eem_ife, raman_area)
            
            all_eems.append(eem_ru)
            
            # Índexs
            doc = doc_values.get(name, None) if doc_values else None
            indices = calculate_indices(ex_wl, em_wl, eem_ru, uvvis_wl, uvvis_abs, doc)
            all_indices.append(indices)
            
            # EEM individual (opcional)
            if save_individual_eems:
                fig, ax = plt.subplots(figsize=(8, 6))
                plot_eem_contour(ex_wl, em_wl, eem_ru, title=name, ax=ax)
                fig.savefig(os.path.join(output_folder, 'eem_processades', f'{name}.png'),
                           dpi=100, bbox_inches='tight')
                plt.close(fig)
            
            # Progrés
            if (i + 1) % 10 == 0 or (i + 1) == n_samples:
                elapsed = time.time() - t_start
                rate = (i + 1) / elapsed
                remaining = (n_samples - i - 1) / rate
                print(f"    [{i+1}/{n_samples}] {name} "
                      f"({elapsed:.0f}s, ~{remaining:.0f}s restants)")
        
        except Exception as e:
            errors.append((name, str(e)))
            print(f"    ERROR en {name}: {e}")
    
    if errors:
        print(f"\n  ⚠ {len(errors)} mostres amb errors:")
        for name, err in errors[:10]:
            print(f"    - {name}: {err}")
    
    # Filtrar mostres vàlides
    valid_names = [n for n in sample_names if n not in [e[0] for e in errors]]
    n_valid = len(all_eems)
    print(f"\n    {n_valid} mostres processades correctament")
    
    # --- 4. Exportar índexs ---
    print(f"\n[4] Exportant índexs a Excel...")
    export_indices_to_excel(all_indices, valid_names,
                           os.path.join(output_folder, 'resum_indexs.xlsx'))
    
    # --- 4b. Detecció d'outliers i QA/QC ---
    print(f"\n[4b] Control de qualitat i detecció d'outliers...")
    
    eem_flags, distances, total_fluor = detect_outliers_eem(all_eems, valid_names)
    index_flags = detect_index_outliers(all_indices, valid_names)
    
    # Comptar i reportar
    n_eem_outliers = sum(1 for f in eem_flags.values() if f)
    n_idx_outliers = sum(1 for f in index_flags.values() if f)
    n_any_flag = sum(1 for n in valid_names if eem_flags.get(n, []) or index_flags.get(n, []))
    
    print(f"    Mostres amb problemes d'EEM:    {n_eem_outliers}")
    print(f"    Mostres amb índexs anòmals:     {n_idx_outliers}")
    print(f"    Total mostres flagejades:       {n_any_flag} / {n_valid}")
    
    if n_any_flag > 0:
        print(f"\n    Mostres flagejades:")
        for name in valid_names:
            e_flags = eem_flags.get(name, [])
            i_flags = index_flags.get(name, [])
            if e_flags or i_flags:
                all_f = e_flags + i_flags
                print(f"      {name}: {'; '.join(all_f[:3])}")
    
    # Gràfic d'outliers
    plot_outlier_diagnostics(valid_names, distances, total_fluor,
                            eem_flags, index_flags,
                            os.path.join(output_folder, 'diagnostic_outliers.png'))
    
    # Informe QA/QC a Excel
    export_quality_report(valid_names, eem_flags, index_flags,
                         os.path.join(output_folder, 'informe_qualitat.xlsx'))
    
    # Filtrar outliers per al PARAFAC (excloure els que tenen flags d'EEM greus)
    clean_indices = []
    clean_eems = []
    clean_names = []
    excluded_names = []
    
    for i, name in enumerate(valid_names):
        e_flags = eem_flags.get(name, [])
        # Excloure només els greus (leverage alt, saturació, molts negatius)
        serious = any('Leverage' in f or 'saturació' in f or 'negatius' in f for f in e_flags)
        if serious:
            excluded_names.append(name)
        else:
            clean_eems.append(all_eems[i])
            clean_indices.append(all_indices[i])
            clean_names.append(name)
    
    if excluded_names:
        print(f"\n    ⚠ {len(excluded_names)} mostres excloses del PARAFAC:")
        for name in excluded_names:
            print(f"      - {name}")
        print(f"    ({len(clean_eems)} mostres restants per PARAFAC)")
    
    n_for_parafac = len(clean_eems)
    
    # --- 5. PARAFAC multi-mostra ---
    if n_for_parafac >= 4:
        print(f"\n[5] PARAFAC multi-mostra ({n_for_parafac} mostres)...")
        X_3d = np.stack(clean_eems, axis=0)
        print(f"    Tensor: {X_3d.shape} (mostres × emissió × excitació)")
        
        # Selecció de components
        max_test = min(max_parafac_test, n_valid - 1, 7)
        print(f"\n    Testant 1 a {max_test} components...")
        comp_results = []
        for n in range(1, max_test + 1):
            factors, history, cc, var_exp = parafac_als(X_3d, n, n_inits=10)
            comp_results.append({
                'n_components': n,
                'var_explained': var_exp,
                'core_consistency': cc,
            })
        
        # Gràfic de selecció
        plot_component_selection(comp_results,
                                os.path.join(output_folder, 'seleccio_components.png'))
        
        # Selecció automàtica si no s'especifica
        if n_parafac is None:
            # Regla: últim n on CORCONDIA > 60%
            for r in reversed(comp_results):
                if r['core_consistency'] > 60:
                    n_parafac = r['n_components']
                    break
            if n_parafac is None:
                n_parafac = 2
            print(f"\n    Selecció automàtica: {n_parafac} components "
                  f"(CORCONDIA > 60%)")
        
        # PARAFAC final
        print(f"\n    PARAFAC final amb {n_parafac} components (15 inits)...")
        factors, loss_history, cc, var_exp = parafac_als(
            X_3d, n_parafac, n_inits=15, max_iter=1000)
        
        print(f"    Variància explicada: {var_exp:.2f}%")
        print(f"    Core consistency: {cc:.1f}%")
        
        # Split-half validation
        if n_valid >= 10:
            print(f"\n    Validació split-half...")
            tcc = split_half_validation(X_3d, n_parafac, n_splits=8, n_inits=5)
            if tcc is not None:
                comp_names = [f'C{i+1}' for i in range(n_parafac)]
                for name, val in zip(comp_names, tcc):
                    status = "✓" if val > 0.95 else ("~" if val > 0.85 else "✗")
                    print(f"      {name}: TCC = {val:.3f} {status}")
        
        # Noms de components
        comp_names = [f'C{i+1}' for i in range(n_parafac)]
        
        # Identificar components basant-se en Ex/Em màxims
        A_scores, B_em, C_ex = factors
        print(f"\n    Components identificats:")
        for r in range(n_parafac):
            em_max = common_em[np.argmax(B_em[:, r])]
            ex_max = common_ex[np.argmax(C_ex[:, r])]
            
            # Classificació automàtica
            if em_max > 400 and ex_max < 270:
                tipo = "Tipus húmic A (terrestre)"
            elif em_max > 400 and ex_max > 300:
                tipo = "Tipus húmic C (terrestre)"
            elif 370 < em_max < 420 and 290 < ex_max < 320:
                tipo = "Tipus húmic M (microbià)"
            elif 300 < em_max < 380 and ex_max < 290:
                tipo = "Tipus proteic T (triptòfan)"
            elif em_max < 320 and ex_max < 290:
                tipo = "Tipus proteic B (tirosina)"
            else:
                tipo = "No classificat"
            
            print(f"      C{r+1}: Ex_max={ex_max:.0f} nm, Em_max={em_max:.0f} nm → {tipo}")
        
        # Figures PARAFAC
        plot_parafac_components(common_ex, common_em, factors, comp_names,
                               os.path.join(output_folder, 'parafac_components.png'))
        
        plot_scores_boxplot(A_scores, clean_names, comp_names,
                           os.path.join(output_folder, 'parafac_scores_boxplot.png'))
        
        # Exportar scores
        export_parafac_scores(A_scores, clean_names, comp_names,
                             os.path.join(output_folder, 'parafac_scores.xlsx'))
    
    else:
        print(f"\n[5] PARAFAC: necessites ≥4 mostres netes (tens {n_for_parafac})")
    
    # --- Resum final ---
    elapsed_total = time.time() - t_start
    print(f"\n{'=' * 70}")
    print(f"  COMPLETAT en {elapsed_total:.1f} segons")
    print(f"  {n_valid} mostres processades")
    print(f"  {n_any_flag} mostres amb flags de qualitat")
    if excluded_names:
        print(f"  {len(excluded_names)} mostres excloses del PARAFAC")
    print(f"  Resultats a: {os.path.abspath(output_folder)}")
    print(f"")
    print(f"  FITXERS GENERATS:")
    print(f"    resum_indexs.xlsx        - Índexs fluorescents per mostra")
    print(f"    informe_qualitat.xlsx    - QA/QC amb flags per mostra")
    print(f"    diagnostic_outliers.png  - Gràfic d'outliers")
    if n_for_parafac >= 4:
        print(f"    parafac_scores.xlsx      - Fmax dels components")
        print(f"    parafac_components.png   - Espectres dels components")
        print(f"    seleccio_components.png  - Selecció nombre components")
    if eem_flags or index_flags:
        for name in valid_names:
            if eem_flags.get(name, []) or index_flags.get(name, []):
                break  # hi ha flags
        print(f"    informe_concordanca.txt  - Fitxers sense parella")
    print(f"{'=' * 70}")
    
    return {
        'sample_names': valid_names,
        'indices': all_indices,
        'tensor': np.stack(all_eems, axis=0) if all_eems else None,
        'ex_wl': common_ex,
        'em_wl': common_em,
        'errors': errors
    }


# ============================================================================
# ============================  CONFIGURACIÓ INTERACTIVA  ====================
# ============================================================================
#
# INSTRUCCIONS:
#   1. Fes DOBLE CLIC sobre aquest fitxer (.py)
#      O bé obre Anaconda Prompt i escriu: python eem_pipeline_multi.py
#   2. S'obriran finestres per seleccionar carpetes i fitxers
#   3. Espera que acabi i mira la carpeta de resultats
#
# ============================================================================

def ask_user_interactive():
    """
    Demana a l'usuari les carpetes i opcions mitjançant finestres de Windows.
    Si no es pot obrir la finestra, demana per terminal.
    """
    import tkinter as tk
    from tkinter import filedialog, messagebox
    
    # Crear finestra root (invisible)
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)  # Que aparegui davant
    
    print("\n" + "=" * 60)
    print("  EEM PIPELINE - Selecció de dades")
    print("  (s'obriran finestres per triar carpetes)")
    print("=" * 60)
    
    # 1. Carpeta EEM
    messagebox.showinfo("EEM Pipeline", 
                        "PRIMER: Selecciona la CARPETA amb els fitxers EEM (.csv)")
    eem_folder = filedialog.askdirectory(
        title="Selecciona la carpeta amb els EEM (.csv)")
    if not eem_folder:
        print("\n  Cancel·lat per l'usuari.")
        sys.exit()
    print(f"\n  EEM folder: {eem_folder}")
    
    # 2. Carpeta UV-Vis
    messagebox.showinfo("EEM Pipeline",
                        "SEGON: Selecciona la CARPETA amb els fitxers UV-Vis (.xlsx)")
    uvvis_folder = filedialog.askdirectory(
        title="Selecciona la carpeta amb els UV-Vis (.xlsx)")
    if not uvvis_folder:
        print("\n  Cancel·lat per l'usuari.")
        sys.exit()
    print(f"  UV-Vis folder: {uvvis_folder}")
    
    # 3. Fitxer Raman
    messagebox.showinfo("EEM Pipeline",
                        "TERCER: Selecciona el FITXER Raman del blanc (.csv)")
    raman_file = filedialog.askopenfilename(
        title="Selecciona el fitxer Raman (.csv)",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
    if not raman_file:
        print("\n  Cancel·lat per l'usuari.")
        sys.exit()
    print(f"  Raman: {raman_file}")
    
    # 4. Carpeta de sortida
    messagebox.showinfo("EEM Pipeline",
                        "QUART: Selecciona (o crea) la CARPETA on guardar els RESULTATS")
    output_folder = filedialog.askdirectory(
        title="Selecciona la carpeta de resultats (sortida)")
    if not output_folder:
        # Si no tria, crear una subcarpeta 'resultats' al costat dels EEM
        output_folder = os.path.join(os.path.dirname(eem_folder), 'resultats')
    print(f"  Output: {output_folder}")
    
    # 5. Opcions addicionals
    print("\n  --- Opcions ---")
    
    # Pas òptic
    pathlength = 1.0
    resp = messagebox.askyesno("Pas òptic", 
                                "El pas òptic de la cubeta és 1 cm?\n\n"
                                "Sí = 1 cm (estàndard)\n"
                                "No = et demanaré el valor")
    if not resp:
        # Demanar per terminal
        try:
            val = input("  Pas òptic de la cubeta (cm): ")
            pathlength = float(val)
        except:
            pathlength = 1.0
    print(f"  Pas òptic: {pathlength} cm")
    
    # PARAFAC automàtic?
    auto_parafac = messagebox.askyesno("PARAFAC",
                                        "Vols que el nombre de components PARAFAC\n"
                                        "es seleccioni AUTOMÀTICAMENT?\n\n"
                                        "Sí = automàtic (recomanat)\n"
                                        "No = et demanaré el nombre")
    n_parafac = None
    if not auto_parafac:
        try:
            val = input("  Nombre de components PARAFAC (2-7): ")
            n_parafac = int(val)
        except:
            n_parafac = None
    print(f"  PARAFAC: {'automàtic' if n_parafac is None else f'{n_parafac} components'}")
    
    # Guardar EEMs individuals?
    save_eems = messagebox.askyesno("EEMs individuals",
                                     "Vols guardar una imatge PNG de cada EEM processada?\n\n"
                                     "(Pot trigar una mica més amb centenars de mostres)")
    print(f"  Guardar EEMs individuals: {'Sí' if save_eems else 'No'}")
    
    root.destroy()
    
    return {
        'eem_folder': eem_folder,
        'uvvis_folder': uvvis_folder,
        'raman_file': raman_file,
        'output_folder': output_folder,
        'n_parafac': n_parafac,
        'pathlength': pathlength,
        'save_eems': save_eems
    }


def ask_user_terminal():
    """
    Versió per terminal (sense finestres) per si tkinter no funciona.
    """
    print("\n" + "=" * 60)
    print("  EEM PIPELINE - Configuració")
    print("=" * 60)
    print("\n  Enganxa els paths complets (pots arrossegar carpetes aquí):\n")
    
    eem_folder = input("  Carpeta EEM (.csv): ").strip().strip('"').strip("'")
    uvvis_folder = input("  Carpeta UV-Vis (.xlsx): ").strip().strip('"').strip("'")
    raman_file = input("  Fitxer Raman (.csv): ").strip().strip('"').strip("'")
    output_folder = input("  Carpeta resultats: ").strip().strip('"').strip("'")
    
    if not output_folder:
        output_folder = os.path.join(os.path.dirname(eem_folder), 'resultats')
    
    print(f"\n  Pas òptic cubeta? [1.0]: ", end="")
    val = input().strip()
    pathlength = float(val) if val else 1.0
    
    print(f"  Components PARAFAC? [auto]: ", end="")
    val = input().strip()
    n_parafac = int(val) if val else None
    
    return {
        'eem_folder': eem_folder,
        'uvvis_folder': uvvis_folder,
        'raman_file': raman_file,
        'output_folder': output_folder,
        'n_parafac': n_parafac,
        'pathlength': pathlength,
        'save_eems': True
    }


if __name__ == "__main__":
    
    # Intentar mode interactiu amb finestres, si no, terminal
    try:
        import tkinter as tk
        # Test que tkinter funciona
        root = tk.Tk()
        root.withdraw()
        root.destroy()
        config = ask_user_interactive()
    except Exception as e:
        print(f"\n  (No s'han pogut obrir finestres: {e})")
        print("  Utilitzant mode terminal...\n")
        config = ask_user_terminal()
    
    # Executar pipeline
    print("\n  Començant processament...\n")
    
    results = run_pipeline(
        eem_folder=config['eem_folder'],
        uvvis_folder=config['uvvis_folder'],
        raman_file=config['raman_file'],
        output_folder=config['output_folder'],
        n_parafac=config['n_parafac'],
        pathlength_cm=config['pathlength'],
        doc_values=None,
        save_individual_eems=config['save_eems']
    )
    
    # Mantenir la finestra oberta al final
    print("\n\n  Prem ENTER per tancar...")
    input()
