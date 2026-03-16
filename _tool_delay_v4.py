"""
Delay tool v4 — TOC + DAD 254 SUPERPOSATS amb assignacions per injecció.
=========================================================================
Un sol gràfic amb:
- TOC continu (eix Y esquerre, blau)
- DAD 254 reconstruït (eix Y dret, verd)
- Bandes verticals per cada injecció (colorades per nom)
- Màxims TOC i DAD marcats dins cada banda
- Slider delay mou les bandes TOC

Usage:
    python _tool_delay_v4.py 231
"""
import sys, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button
from scipy.ndimage import uniform_filter1d

from hpsec_config import get_data_folders
from hpsec_consolidate import extract_seq_number, detect_seq_type

FLUSH_TIME_MIN = 3.637


def find_seq_path(arg):
    if os.path.isdir(arg):
        return arg
    for folder in get_data_folders():
        full = os.path.join(folder, arg)
        if os.path.isdir(full):
            return full
        try:
            num = int(arg)
            for d in os.listdir(folder):
                if extract_seq_number(d) == num:
                    return os.path.join(folder, d)
        except ValueError:
            pass
    return None


def read_export3d_254(filepath):
    try:
        with open(filepath, encoding='utf-16') as fh:
            lines = fh.readlines()
        header = lines[0].strip().split(',')
        idx254 = next((i for i, h in enumerate(header) if '254' in h), None)
        if idx254 is None:
            return None, None
        times, vals = [], []
        for line in lines[1:]:
            fields = line.strip().split(',')
            if len(fields) > idx254:
                try:
                    times.append(float(fields[0]))
                    vals.append(float(fields[idx254]))
                except ValueError:
                    pass
        return np.array(times), np.array(vals)
    except:
        return None, None


def main():
    if len(sys.argv) < 2:
        print("Usage: python _tool_delay_v4.py <SEQ>")
        sys.exit(1)

    seq_path = find_seq_path(sys.argv[1])
    if not seq_path:
        print(f"ERROR: '{sys.argv[1]}' no trobat")
        sys.exit(1)

    seq_name = os.path.basename(seq_path)
    method = detect_seq_type(seq_name)
    print(f"Loading {seq_name} ({method})...")

    # MasterFile
    mfs = [f for f in os.listdir(seq_path)
           if 'MasterFile' in f and f.endswith('.xlsx')
           and '~' not in f and 'pre_vol' not in f and 'backup' not in f]
    if not mfs:
        mfs = [f for f in os.listdir(seq_path)
               if 'MasterFile' in f and f.endswith('.xlsx')
               and '~' not in f and 'pre_vol' not in f]
    mf_path = os.path.join(seq_path, mfs[0])

    # TOC
    df_toc = pd.read_excel(mf_path, sheet_name='2-TOC', header=6)
    toc_col = [c for c in df_toc.columns if 'toc' in str(c).lower()][0]
    date_col_toc = [c for c in df_toc.columns if 'started' in str(c).lower()]
    y_toc = pd.to_numeric(df_toc[toc_col], errors='coerce').values
    dt = 4.0 / 60.0
    t_toc = np.arange(len(y_toc)) * dt
    y_clean = np.nan_to_num(y_toc, nan=np.nanmedian(y_toc[~np.isnan(y_toc)]))
    y_smooth = uniform_filter1d(y_clean, size=5)

    t0_toc = None
    if date_col_toc:
        dates = pd.to_datetime(df_toc[date_col_toc[0]], errors='coerce')
        valid = dates.dropna()
        if len(valid) > 0:
            t0_toc = valid.iloc[0]

    # HPLC injections
    df_hplc = pd.read_excel(mf_path, sheet_name='1-HPLC-SEQ')
    name_col = [c for c in df_hplc.columns
                if 'sample' in str(c).lower() and 'name' in str(c).lower()][0]
    hplc_date_col = [c for c in df_hplc.columns if 'acquired' in str(c).lower()][0]

    injs = []
    for _, row in df_hplc.iterrows():
        try:
            name = str(row[name_col]).strip()
            dt_val = pd.to_datetime(row[hplc_date_col])
            t_min = (dt_val - t0_toc).total_seconds() / 60.0 if t0_toc else 0
            is_khp = 'khp' in name.lower()
            is_ctrl = any(x in name.lower() for x in ['mq', 'naoh', 'blanc', 'blnc'])
            injs.append({'name': name, 't_min': t_min,
                         'is_khp': is_khp, 'is_control': is_ctrl})
        except:
            pass

    cadence = np.median(np.diff([inj['t_min'] for inj in injs])) if len(injs) >= 2 else 11.0

    # DAD 254 per injection
    e3d = os.path.join(seq_path, 'Export3d')
    e3d_files = sorted(os.listdir(e3d)) if os.path.isdir(e3d) else []

    name_counter = {}
    dad_data = {}  # inj_index -> (t_rel, y254, t_max_rel, y_max)

    for j, inj in enumerate(injs):
        nk = inj['name'].lower().replace(' ', '').replace('-', '')
        name_counter[nk] = name_counter.get(nk, 0) + 1
        rep = name_counter[nk]

        for f in e3d_files:
            fb = ''.join(c for c in f.lower().replace('.csv', '') if c.isalnum())
            nk_clean = ''.join(c for c in nk if c.isalnum())
            if nk_clean in fb and f'r{rep}' in f.lower():
                t_d, y_d = read_export3d_254(os.path.join(e3d, f))
                if t_d is not None and len(t_d) > 5:
                    pk = np.argmax(y_d)
                    dad_data[j] = {
                        't_rel': t_d, 'y': y_d,
                        't_max_rel': t_d[pk], 'y_max': y_d[pk],
                    }
                break

    print(f"TOC: {len(y_toc)} pts | HPLC: {len(injs)} inj | DAD: {len(dad_data)} segments")

    # Current delay
    current_delay = FLUSH_TIME_MIN
    df_info = pd.read_excel(mf_path, sheet_name='0-INFO', header=None)
    hora_hplc = hora_toc = None
    for _, row in df_info.iterrows():
        key = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        val = row.iloc[1] if len(row) > 1 else None
        if 'hora hplc' in key: hora_hplc = val
        elif 'hora toc' in key: hora_toc = val
        elif 'net delay' in key and 'suite' not in key:
            try:
                v = float(val)
                if not np.isnan(v): current_delay = v
            except: pass
    if current_delay == FLUSH_TIME_MIN and hora_hplc and hora_toc:
        try:
            def to_min(t):
                if hasattr(t, 'hour'):
                    return t.hour * 60 + t.minute + t.second / 60
                return sum(int(p)*m for p, m in zip(str(t).split(':'), [60, 1]))
            current_delay = FLUSH_TIME_MIN - (to_min(hora_hplc) - to_min(hora_toc))
        except: pass

    # === FIGURE ===
    fig, ax = plt.subplots(figsize=(20, 8))
    plt.subplots_adjust(bottom=0.15, top=0.92)

    # TOC (left Y axis)
    ax.plot(t_toc, y_smooth, 'b-', lw=0.5, alpha=0.7, label='TOC')
    ax.set_ylabel('TOC (ppb)', color='blue')
    ax.tick_params(axis='y', labelcolor='blue')

    # DAD 254 (right Y axis)
    ax2 = ax.twinx()
    ax2.set_ylabel('DAD 254 (mAU)', color='green')
    ax2.tick_params(axis='y', labelcolor='green')

    # Plot DAD segments positioned at HPLC time (no delay — direct from HPLC)
    for j, inj in enumerate(injs):
        if j not in dad_data:
            continue
        dd = dad_data[j]
        t_abs = dd['t_rel'] + inj['t_min']  # absolute time from HPLC
        color = '#E74C3C' if inj['is_khp'] else ('#95A5A6' if inj['is_control'] else '#27AE60')
        ax2.plot(t_abs, dd['y'], color=color, lw=0.6, alpha=0.6)
        # DAD peak marker
        ax2.plot(inj['t_min'] + dd['t_max_rel'], dd['y_max'], '^',
                 color=color, ms=5, zorder=6)

    ax.set_xlabel('min (des de inici TOC)')

    # === Injection bands + labels (move with delay) ===
    bands = []
    band_labels = []
    toc_markers = []

    cmap = plt.colormaps['tab20']

    def draw_bands(delay_val):
        for b in bands:
            b.remove()
        bands.clear()
        for lb in band_labels:
            lb.remove()
        band_labels.clear()
        for m in toc_markers:
            m.remove()
        toc_markers.clear()

        y_top_toc = ax.get_ylim()[1]

        for j, inj in enumerate(injs):
            # TOC window: centered at t_hplc + delay + t_retention_estimate
            # For BP, peak is ~1-2 min after injection reaches TOC
            t_center_toc = inj['t_min'] + delay_val + 1.0  # +1 for retention
            hw = cadence / 2

            if inj['is_khp']:
                color = '#E74C3C'
                alpha = 0.10
            elif inj['is_control']:
                color = '#95A5A6'
                alpha = 0.03
            else:
                ci = j % 20
                color = cmap(ci)
                alpha = 0.06

            # Band covering the injection window
            b = ax.axvspan(t_center_toc - hw * 0.4, t_center_toc + hw * 0.6,
                           alpha=alpha, color=color, zorder=0)
            bands.append(b)

            # Label at top
            short = inj['name'][:7]
            lb = ax.annotate(f"{j+1}:{short}",
                             (t_center_toc, y_top_toc * 0.97),
                             fontsize=4, rotation=90, va='top', ha='center',
                             color=color, alpha=0.8)
            band_labels.append(lb)

            # Find TOC max within this window
            mask = (t_toc >= t_center_toc - hw*0.4) & (t_toc <= t_center_toc + hw*0.6)
            if mask.any():
                y_win = y_smooth[mask]
                t_win = t_toc[mask]
                pk_idx = np.argmax(y_win)
                m = ax.plot(t_win[pk_idx], y_win[pk_idx], 'rv', ms=4, zorder=7)
                toc_markers.extend(m)

        # Title with match count
        n_match = 0
        n_samples = 0
        for j, inj in enumerate(injs):
            if inj['is_control']:
                continue
            n_samples += 1
            t_c = inj['t_min'] + delay_val + 1.0
            hw = cadence / 2
            mask = (t_toc >= t_c - hw*0.4) & (t_toc <= t_c + hw*0.6)
            if mask.any():
                y_win = y_smooth[mask]
                bl = np.percentile(y_win, 20)
                pk = np.max(y_win - bl)
                noise = np.std(y_win[y_win-bl < np.percentile(y_win-bl, 30)]) if len(y_win) > 5 else 1
                if noise > 0 and pk/noise > 3:
                    n_match += 1

        pct = n_match/n_samples*100 if n_samples > 0 else 0
        fig.suptitle(
            f"{seq_name} ({method}) — delay={delay_val:.1f} min — "
            f"TOC pics dins finestra: {n_match}/{n_samples} ({pct:.0f}%)\n"
            f"Blau=TOC | Verd=DAD254 | Bandes=assignacio per injecció",
            fontsize=11, fontweight='bold')

    draw_bands(current_delay)

    # Slider
    ax_slider = plt.axes([0.15, 0.04, 0.50, 0.025])
    slider = Slider(ax_slider, 'Delay (min)',
                    current_delay - 30, current_delay + 30,
                    valinit=current_delay, valstep=0.1)

    def on_slider(val):
        draw_bands(val)
        fig.canvas.draw_idle()
    slider.on_changed(on_slider)

    # Auto button
    ax_auto = plt.axes([0.70, 0.04, 0.07, 0.025])
    btn_auto = Button(ax_auto, 'Auto')
    def on_auto(e):
        # Use DAD-TOC matching to find median delay
        delays = []
        for j, inj in enumerate(injs):
            if inj['is_control'] or j not in dad_data:
                continue
            dd = dad_data[j]
            t_dad_abs = inj['t_min'] + dd['t_max_rel']
            # Search TOC peak near this position
            search_start = t_dad_abs
            search_end = t_dad_abs + 15  # max 15 min delay
            mask = (t_toc >= search_start) & (t_toc <= search_end)
            if mask.any():
                y_w = y_smooth[mask]
                t_w = t_toc[mask]
                pk = np.argmax(y_w)
                if y_w[pk] > np.percentile(y_w, 20) + 10:
                    delay_real = t_w[pk] - t_dad_abs
                    delays.append(delay_real)
        if delays:
            # The delay for the slider is: delay_real - 1.0 (we add 1.0 for retention)
            med = np.median(delays)
            slider.set_val(med - 1.0 + injs[0]['t_min'] - injs[0]['t_min'])
            # Actually: slider delay = t_toc_peak - t_hplc - retention
            # t_toc_peak = t_hplc + slider_delay + 1.0
            # slider_delay = t_toc_peak - t_hplc - 1.0 = (t_dad + delay_real) - t_hplc - 1.0
            # = dad_t_max_rel + delay_real - 1.0
            slider.set_val(med)
            print(f"Auto: delay mediana = {med:.2f} min (de {len(delays)} mostres)")
    btn_auto.on_clicked(on_auto)

    # Save
    ax_save = plt.axes([0.78, 0.04, 0.07, 0.025])
    btn_save = Button(ax_save, 'GUARDAR')
    def on_save(e):
        import openpyxl
        val = slider.val
        wb = openpyxl.load_workbook(mf_path)
        ws = wb['0-INFO']
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value and 'net delay' in str(row[0].value).lower() and 'suite' not in str(row[0].value).lower():
                ws.cell(row=row[0].row, column=2, value=round(val, 3))
                wb.save(mf_path)
                print(f"GUARDAT: {val:.3f} min")
                return
        lr = ws.max_row + 1
        ws.cell(row=lr, column=1, value='Net delay (Suite)')
        ws.cell(row=lr, column=2, value=round(val, 3))
        wb.save(mf_path)
        print(f"GUARDAT: {val:.3f} min")
    btn_save.on_clicked(on_save)

    # Reset
    ax_reset = plt.axes([0.86, 0.04, 0.07, 0.025])
    btn_reset = Button(ax_reset, 'Reset')
    btn_reset.on_clicked(lambda e: slider.set_val(current_delay))

    fig.text(0.15, 0.01,
             f"MF delay: {current_delay:.2f} | Cadencia: {cadence:.1f} min | "
             f"DAD: {len(dad_data)}/{len(injs)} inj",
             fontsize=8, color='#666')

    plt.show()


if __name__ == "__main__":
    main()
