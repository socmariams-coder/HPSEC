"""
hpsec_import.py - Mòdul d'importació de dades HPSEC (Fase 1)
=============================================================

FASE 1 del pipeline de 5 fases:
- Crear MasterFile si no existeix (des de rawdata v11/v12)
- Llegir fitxers font (.D, UIB, DAD)
- Aparellar mostres (matching)
- Identificar tipus (KHP, Control, Mostra)
- Aplicar correcció de baseline (preprocessament bàsic)
- Guardar dades RAW i NET (amb baseline restada)

NO fa:
- Validació KHP (Fase 2: CALIBRAR)
- Alineació ni càlcul d'àrees (Fase 3: PROCESSAR)
- Comparació rèpliques (Fase 4: REVISAR)
- Escriptura Excel finals (Fase 5: EXPORTAR)

NOTA: La correcció de baseline s'aplica aquí per permetre que
Calibrate (Fase 2) treballi sobre dades netes per detectar pics KHP.

Usat per HPSEC_Suite.py
"""

__version__ = "1.2.0"
__version_date__ = "2026-01-30"

import os
import re
import glob
import json
import logging
from datetime import datetime
from difflib import SequenceMatcher

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

from hpsec_config import get_config
from hpsec_utils import get_baseline_value
from hpsec_core import detect_timeout, detect_sequence_timeouts, map_timeouts_to_injection, downsample_to_cadence
from hpsec_migrate_master import migrate_single

# Import sistema d'avisos estructurats
from hpsec_warnings import (
    create_anomaly, get_max_warning_level, WarningLevel,
)

# =============================================================================
# CONSTANTS
# =============================================================================
CONFIDENCE_THRESHOLD = 85.0  # Llindar per acceptar match automàticament
DATA_FOLDER_NAME = "data"  # Subcarpeta dins CHECK per JSONs


def _safe_float(val):
    """Converteix un valor a float de forma segura (per uib_sensitivity, etc.).

    Suporta formats com "1-700" (sensibilitat UIB del 0-CHECK) → 700.
    """
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        # Format "X-NNN" (p.ex. "1-700" del 0-CHECK) → agafar últim número
        import re
        m = re.search(r'(\d+)\s*$', str(val))
        if m:
            return float(m.group(1))
        return None


# =============================================================================
# CARPETA DADES (JSONs)
# =============================================================================
def get_data_folder(seq_path, create=True):
    """
    Retorna la carpeta on guardar JSONs i dades processades.

    Ubicació: SEQ_PATH/CHECK/data/

    Args:
        seq_path: Ruta a la carpeta SEQ
        create: Si True, crea la carpeta si no existeix

    Returns:
        Path a la carpeta de dades
    """
    check_folder = os.path.join(seq_path, "CHECK")
    data_folder = os.path.join(check_folder, DATA_FOLDER_NAME)

    if create:
        os.makedirs(data_folder, exist_ok=True)

    return data_folder

# NOTA: Funcions de baseline centralitzades a hpsec_utils.py
# Usar get_baseline_value() i get_baseline_stats() directament


# =============================================================================
# UTILITATS BÀSIQUES
# =============================================================================

def normalize_key(s):
    """
    Normalitza string per matching.

    - Elimina guions (-) i espais
    - MANTÉ underscores (_) perquè són significatius en noms de fitxers
    - Converteix a majúscules

    Exemples:
        "MQ-1" -> "MQ1"
        "MQ_1" -> "MQ_1"  (underscore mantingut)
        "NaOH 0.1mM" -> "NAOH0.1MM"
    """
    result = re.sub(r"[\-\s]+", "", str(s or ""))  # Només elimina guions i espais
    return result.upper()


def generate_agilent_control_name(base_name, occurrence_num):
    """
    Genera nom segons lògica Agilent per controls repetits.

    Agilent/ChemStation assigna noms així quan una mostra es repeteix:
    - 1a aparició: nom base (ex: "MQ")
    - 2a aparició: nom base + "1" (ex: "MQ1")
    - 3a aparició: nom base + "2" (ex: "MQ2")

    Args:
        base_name: Nom base del control (ex: "MQ", "NaOH")
        occurrence_num: Número d'aparició (1, 2, 3, ...)

    Returns:
        str: Nom segons convenció Agilent

    Exemples:
        generate_agilent_control_name("MQ", 1) -> "MQ"
        generate_agilent_control_name("MQ", 2) -> "MQ1"
        generate_agilent_control_name("MQ", 3) -> "MQ2"
    """
    if occurrence_num <= 1:
        return base_name
    else:
        return f"{base_name}{occurrence_num - 1}"


def normalize_rep(rep):
    """Normalitza número de rèplica."""
    if rep is None:
        return ""
    s = str(rep).strip()
    if not s:
        return ""
    try:
        return str(int(s))
    except Exception:
        return s


def is_khp(name):
    """Detecta si és mostra KHP."""
    return "KHP" in str(name).upper()


def extract_khp_conc(filename):
    """
    Extreu la concentració de KHP del nom del fitxer (en ppm C).

    Patrons suportats:
    - KHP2, KHP_2, KHP-2 -> 2.0 ppm
    - KHP2.5, KHP 0.25 -> 2.5 / 0.25 ppm
    - KHP 1 ppm -> 1.0 ppm
    - KHP 500 ppb -> 0.5 ppm
    - KHP 100 (sense unitat, >=100) -> 0.1 ppm (assumeix ppb)

    Exclou noms especials: BUFFER, MQ, FI, MIX, BLANK, NaOH, mM.

    Args:
        filename: Nom del fitxer o mostra

    Returns:
        float: Concentració en ppm C, o 0.0 si no trobada/exclosa
    """
    name = str(filename).strip()
    name_upper = name.upper()

    # Excloure noms especials (no són estàndards de concentració)
    _KHP_EXCLUDE = (r'\bBUFFER\b', r'\bMQ\b', r'\bFI\b', r'\bMIX\b',
                    r'\bBLANC\b', r'\bBLANK\b', r'\bNaOH\b')
    for pat in _KHP_EXCLUDE:
        if re.search(pat, name_upper):
            return 0.0

    # "KHP 20mM" → concentració molar, no ppm C directe
    if "MM" in name_upper:
        return 0.0

    # Regex unificat: "KHP 1", "KHP1", "KHP 1 ppm", "KHP 500 ppb", "KHP 0.25"
    match = re.search(
        r'KHP[_\-\s]*(\d+\.?\d*)\s*(ppm|ppb)?',
        name, re.IGNORECASE
    )
    if not match:
        return 0.0

    raw_str = match.group(1)  # El string numèric original (preserva zeros inicials)
    value = float(raw_str)
    unit = (match.group(2) or "").upper()

    if unit == "PPB":
        value /= 1000.0  # ppb → ppm

    # Convenció zeros inicials sense separador: "KHP01"=0.1, "KHP025"=0.25, "KHP05"=0.5
    # Detectar: pegat a "KHP" (sense espai/guió), comença per "0", sense punt decimal, sense unitat
    # Exclou "KHP 01 ppm" (amb espai+unitat) i "KHP0.25" (amb punt)
    if not unit and '.' not in raw_str and raw_str.startswith('0') and len(raw_str) >= 2:
        # Verificar que el número va pegat a KHP (sense separador)
        match_no_sep = re.search(r'KHP(\d)', name, re.IGNORECASE)
        if match_no_sep:
            value = float('0.' + raw_str[1:])  # "01"->"0.1", "025"->"0.25", "05"->"0.5"
            return value

    # Sense unitat explícita i valor >= 100: assumir ppb
    # Ex: "KHP 250" en context DOC = 250 ppb = 0.25 ppm
    if not unit and value >= 100:
        value /= 1000.0

    return value


def is_control_injection(sample_name, config=None):
    """
    Verifica si una mostra és una injecció de control (MQ, NaOH, etc.).

    Args:
        sample_name: Nom de la mostra
        config: Configuració (si None, es llegeix de get_config())

    Returns:
        True si és una injecció de control
    """
    if config is None:
        config = get_config()

    control_patterns = config.get("control_injections", "patterns", default=[])
    sample_upper = sample_name.upper()

    for pattern in control_patterns:
        if pattern.upper() in sample_upper:
            return True
    return False


def is_blank_injection(sample_name, config=None):
    """
    Verifica si una mostra és un blanc (MQ, H2O, etc.).

    Args:
        sample_name: Nom de la mostra
        config: Configuració (si None, es llegeix de get_config())

    Returns:
        True si és una injecció blanc
    """
    if config is None:
        config = get_config()

    blank_patterns = config.get("blank_injections", "patterns", default=[])
    sample_upper = sample_name.upper()

    for pattern in blank_patterns:
        if pattern.upper() in sample_upper:
            return True
    return False


def classify_reference_standard(sample_name, config=None):
    """
    Classifica una mostra com a patró de referència i retorna el subtipus.

    Subtipus (definits a hpsec_config.json → sample_types):
    - PR_C: Patró de referència de carboni (SUWANNEE, SRNOM, HA/FA) → quantificar
    - PR_I: Patró de referència inorgànic (Br, NO3, CaCO3) → NO quantificar
    - PR_N: Patró de referència de nitrogen → quantificar
    - PR:   Patró genèric (fallback si no encaixa en cap subtipus)

    Args:
        sample_name: Nom de la mostra
        config: Configuració (si None, es llegeix de get_config())

    Returns:
        Subtipus string ("PR_C", "PR_I", "PR_N") o None si no és PR.
    """
    if config is None:
        config = get_config()

    sample_upper = sample_name.upper().strip()
    sample_types = config.get("sample_types", default={})

    # Ordre de prioritat: subtipus específics primer
    pr_type_keys = [
        ("PATRÓ_REF_C", "PR_C"),
        ("PATRÓ_REF_I", "PR_I"),
        ("PATRÓ_REF_N", "PR_N"),
    ]

    for config_key, type_code in pr_type_keys:
        pr_config = sample_types.get(config_key, {})
        patterns = pr_config.get("patterns", [])
        suffixes = pr_config.get("suffixes", [])

        for pattern in patterns:
            if pattern.upper() in sample_upper:
                return type_code

        for suffix in suffixes:
            if sample_upper.endswith(suffix.upper()):
                return type_code

    return None



def obtenir_seq(folder):
    """Extreu ID de seqüència del nom de carpeta."""
    nom = os.path.basename(os.path.normpath(folder))
    m = re.search(r"(\d+[A-Za-z]?)", nom)
    return m.group(1) if m else "000"


def seq_tag(seq, bp):
    """Afegeix _BP si és seqüència BP."""
    if bp and "BP" not in seq.upper():
        return f"{seq}_BP"
    return seq


def split_sample_rep(sample_desc):
    """Separa nom de mostra i rèplica."""
    s = str(sample_desc or "").strip()
    if not s:
        return "", None
    m = re.match(r"^(.*?)(?:[_\-\s]?R(\d+))\s*$", s, flags=re.IGNORECASE)
    if m and m.group(1).strip():
        return m.group(1).strip(), m.group(2)
    return s, None


def clean_sample_name(sample_desc):
    """Neteja nom de mostra."""
    s = str(sample_desc or "").strip()
    return re.sub(r"[^A-Za-z0-9]+", "", s) or "SAMPLE"


def skip_sample_direct(sample_desc):
    """Determina si s'ha d'ignorar una mostra (PRE-HPLC, POST-HPLC, etc.)."""
    d = str(sample_desc or "").strip().upper()
    if d in ("", "0"):
        return True
    if "PRE-HPLC" in d or ("POST" in d and "HPLC" in d):
        return True
    return False


# =============================================================================
# DETECCIÓ MODE (BP vs COLUMN)
# =============================================================================

def is_bp_seq(folder, sample_dad_files=None):
    """
    Detecta si és seqüència BP basant-se en:
    1. Nom de la carpeta (conté _BP)
    2. Durada dels cromatogrames DAD (≤18 min = BP)
    """
    name = os.path.basename(os.path.normpath(folder)).upper()
    if re.search(r"(^|[_\-\s])BP($|[_\-\s])", name):
        return True

    if sample_dad_files:
        for p in sample_dad_files[:3]:
            try:
                df, st = llegir_dad_export3d(p)
                if st.startswith("OK") and not df.empty and "time (min)" in df.columns:
                    tmax = pd.to_numeric(df["time (min)"], errors="coerce").max()
                    if pd.notna(tmax) and float(tmax) <= 18.0:
                        return True
            except Exception:
                continue
    return False


def detect_mode_from_folder(seq_path):
    """
    Detecta el mode de la seqüència (BP o COLUMN) i el tipus de dades (UIB/DIRECT/DUAL).

    Returns:
        dict amb method ("BP"/"COLUMN") i data_mode ("UIB"/"DIRECT"/"DUAL")
    """
    result = {
        "method": "COLUMN",
        "data_mode": "UIB",
    }

    # Detectar BP pel nom
    folder_name = os.path.basename(seq_path).upper()
    if "_BP" in folder_name or folder_name.endswith("BP"):
        result["method"] = "BP"

    # Detectar tipus de dades
    path_csv = os.path.join(seq_path, "CSV")
    has_uib = False
    has_direct = False

    if os.path.isdir(path_csv):
        uib_files = glob.glob(os.path.join(path_csv, "*UIB1B*"))
        has_uib = len(uib_files) > 0

    # Buscar master file per Direct
    master_path, master_format = trobar_excel_mestre(seq_path)
    if master_path and master_format == "NEW":
        master_data = llegir_masterfile_nou(master_path)
        if master_data.get("toc") is not None:
            has_direct = True

    if has_uib and has_direct:
        result["data_mode"] = "DUAL"
    elif has_direct:
        result["data_mode"] = "DIRECT"
    else:
        result["data_mode"] = "UIB"

    return result


# =============================================================================
# LECTURA FITXERS MASTER
# =============================================================================

def detect_master_format(filepath):
    """
    Detecta el format del fitxer mestre.

    Returns:
        "NEW" si és nou MasterFile (0-INFO, 1-HPLC-SEQ, 2-TOC, 4-TOC_CALC)
        "OLD" si és format antic (0-CHECK, 2-TOC, 4-SEQ_DATA)
        None si no és reconegut
    """
    try:
        with pd.ExcelFile(filepath, engine="openpyxl") as xl:
            sheets = set(xl.sheet_names)

            # Format nou: té 0-INFO i 4-TOC_CALC (o 1-HPLC-SEQ)
            if "0-INFO" in sheets and ("4-TOC_CALC" in sheets or "1-HPLC-SEQ" in sheets):
                return "NEW"

            # Format antic: té 0-CHECK i 4-SEQ_DATA
            if any(s.lower() == "0-check" for s in sheets) and "4-SEQ_DATA" in sheets:
                return "OLD"

            # Compatibilitat: si té 2-TOC i 4-SEQ_DATA, és antic
            if "2-TOC" in sheets and "4-SEQ_DATA" in sheets:
                return "OLD"

            return None
    except Exception:
        return None


def trobar_excel_mestre(folder_seq, prefer_new=True):
    """
    Troba Excel mestre (compatible amb format antic i nou).

    Args:
        folder_seq: Carpeta de la seqüència
        prefer_new: Si True, prioritza MasterFile nou sobre format antic

    Returns:
        Tuple (path, format) on format és "NEW" o "OLD", o (None, None) si no es troba
    """
    candidats = []
    for ext in ("*.xlsx", "*.XLSX", "*.xlsm", "*.XLSM"):
        candidats.extend(glob.glob(os.path.join(folder_seq, ext)))
    candidats = [p for p in candidats if not os.path.basename(p).startswith("~$")]

    # Separar per tipus
    new_format = []
    old_format = []

    for p in sorted(candidats):
        fmt = detect_master_format(p)
        if fmt == "NEW":
            new_format.append(p)
        elif fmt == "OLD":
            old_format.append(p)

    # Prioritzar format NEW (MasterFile)
    # NO retornar format OLD (v11/v12) - cal migrar-los primer
    if new_format:
        # Prioritzar fitxers amb "MasterFile" al nom
        for p in new_format:
            if "MasterFile" in os.path.basename(p):
                return p, "NEW"
        return new_format[0], "NEW"

    # Si només hi ha format OLD (v11/v12), retornar None per disparar migrate
    # El migrate crearà el MasterFile a partir del v11/v12
    return None, None


def read_master_date(seq_folder):
    """Llegeix data del fitxer mestre (compatible amb format antic i nou)."""
    xls = glob.glob(os.path.join(seq_folder, "*.xlsx"))
    for f in xls:
        if "~$" in os.path.basename(f):
            continue
        try:
            with pd.ExcelFile(f, engine="openpyxl") as xf:

                # Primer intentar format nou (0-INFO)
                if "0-INFO" in xf.sheet_names:
                    df = pd.read_excel(xf, sheet_name="0-INFO", header=None, engine="openpyxl")
                    for i, row in df.iterrows():
                        if str(row.iloc[0]).strip().lower() == "date":
                            val = row.iloc[1]
                            if pd.notna(val):
                                return str(val)
                    continue

                # Fallback format antic (0-CHECK)
                sheet = None
                for s in xf.sheet_names:
                    if str(s).strip().lower() == "0-check":
                        sheet = s
                        break
                if sheet is None:
                    sheet = 0
                df = pd.read_excel(xf, sheet_name=sheet, header=None, engine="openpyxl")
                val = df.iloc[1, 1]
                if pd.notna(val):
                    return str(val)
        except Exception:
            pass
    return ""


def llegir_masterfile_nou(filepath):
    """
    Llegeix el nou format MasterFile.

    Returns:
        dict amb:
            - info: dict amb metadata de 0-INFO
            - hplc_seq: DataFrame de 1-HPLC-SEQ
            - toc: DataFrame de 2-TOC
            - toc_calc: DataFrame de 4-TOC_CALC
            - dad_khp: DataFrame de 3-DAD_KHP (si existeix)
    """
    result = {
        "info": {},
        "hplc_seq": None,
        "toc": None,
        "toc_calc": None,
        "dad_khp": None,
        "filepath": filepath,
    }

    try:
        with pd.ExcelFile(filepath, engine="openpyxl") as xl:

            # 0-INFO
            if "0-INFO" in xl.sheet_names:
                df_info = pd.read_excel(xl, sheet_name="0-INFO", header=None, engine="openpyxl")
                for _, row in df_info.iterrows():
                    key = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    val = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None
                    if key:
                        result["info"][key] = val

                # Sensibilitat UIB a B5 (fila 4, columna 1 en 0-indexed)
                if len(df_info) > 4 and len(df_info.columns) > 1:
                    uib_sens = df_info.iloc[4, 1]
                    if pd.notna(uib_sens):
                        result["info"]["uib_sensitivity"] = uib_sens

            # 1-HPLC-SEQ (o 1-HPLC-SEQ_RAW per format antic v11)
            if "1-HPLC-SEQ" in xl.sheet_names:
                result["hplc_seq"] = pd.read_excel(xl, sheet_name="1-HPLC-SEQ", engine="openpyxl")
            elif "1-HPLC-SEQ_RAW" in xl.sheet_names:
                result["hplc_seq"] = pd.read_excel(xl, sheet_name="1-HPLC-SEQ_RAW", engine="openpyxl")

            # 2-TOC
            if "2-TOC" in xl.sheet_names:
                result["toc"] = pd.read_excel(xl, sheet_name="2-TOC", header=6, engine="openpyxl")

            # 4-TOC_CALC (o 4-SEQ_DATA per format antic v11)
            if "4-TOC_CALC" in xl.sheet_names:
                result["toc_calc"] = pd.read_excel(xl, sheet_name="4-TOC_CALC", engine="openpyxl")
            elif "4-SEQ_DATA" in xl.sheet_names:
                result["toc_calc"] = pd.read_excel(xl, sheet_name="4-SEQ_DATA", engine="openpyxl")

            # 3-DAD_KHP (opcional)
            if "3-DAD_KHP" in xl.sheet_names:
                result["dad_khp"] = pd.read_excel(xl, sheet_name="3-DAD_KHP", header=1, engine="openpyxl")

    except PermissionError:
        result["error"] = (
            f"No es pot llegir el MasterFile: el fitxer està obert a Excel o sense permisos. "
            f"Tancar '{os.path.basename(filepath)}' i tornar a importar."
        )
    except Exception as e:
        result["error"] = str(e)

    return result


def llegir_master_direct(mestre):
    """
    Llegeix format antic de master (2-TOC & 4-SEQ_DATA).

    Returns:
        (df_toc, df_seq) o (None, None) si error
    """
    try:
        with pd.ExcelFile(mestre, engine="openpyxl") as xl:

            df_toc = None
            df_seq = None

            # Llegir 2-TOC
            if "2-TOC" in xl.sheet_names:
                df_toc = pd.read_excel(xl, sheet_name="2-TOC", header=6, engine="openpyxl")

            # Llegir 4-SEQ_DATA
            if "4-SEQ_DATA" in xl.sheet_names:
                df_seq = pd.read_excel(xl, sheet_name="4-SEQ_DATA", engine="openpyxl")

            return df_toc, df_seq
    except PermissionError:
        raise PermissionError(
            f"No es pot llegir el MasterFile: el fitxer està obert a Excel o sense permisos. "
            f"Tancar '{os.path.basename(mestre)}' i tornar a importar."
        )
    except Exception:
        return None, None


def extract_doc_from_masterfile(toc_df, row_start, row_end, t_start=None, detect_timeouts=True,
                                  max_duration_min=None):
    """
    Extreu segment DOC del nou MasterFile.

    Args:
        toc_df: DataFrame de 2-TOC
        row_start: Fila inicial (1-indexed, relatiu a fila 8 del TOC original)
        row_end: Fila final
        t_start: Timestamp d'inici (opcional, per calcular temps relatiu)
        detect_timeouts: Si True, també detecta i retorna info de timeouts
        max_duration_min: Durada màxima en minuts (per truncar última mostra)
                         Si None, no trunca. Típic: 75-80 min per COLUMN

    Returns:
        Si detect_timeouts=False: DataFrame amb columns "time (min)" i "DOC"
        Si detect_timeouts=True: tuple (DataFrame, timeout_info dict)
    """
    empty_timeout = {
        "n_timeouts": 0, "n_major_timeouts": 0, "timeouts": [],
        "dt_median_sec": 0, "dt_max_sec": 0, "zone_summary": {},
        "severity": "OK", "warning_message": ""
    }

    if toc_df is None or toc_df.empty:
        if detect_timeouts:
            return pd.DataFrame(), empty_timeout
        return pd.DataFrame()

    # Trobar columnes de temps i senyal
    time_col = None
    sig_col = None

    for col in toc_df.columns:
        col_str = str(col).lower()
        if 'date' in col_str and 'start' in col_str:
            time_col = col
        elif 'toc' in col_str and 'ppb' in col_str:
            sig_col = col
        elif 'tc' in col_str and 'ppb' in col_str and sig_col is None:
            sig_col = col

    # Fallback a posició si no es troben
    if time_col is None and len(toc_df.columns) > 3:
        time_col = toc_df.columns[3]
    if sig_col is None and len(toc_df.columns) > 5:
        sig_col = toc_df.columns[5]

    if time_col is None or sig_col is None:
        if detect_timeouts:
            return pd.DataFrame(), empty_timeout
        return pd.DataFrame()

    # Extreure segment (ajustar índexs)
    # row_start/row_end són 1-indexed des de fila 8 del Excel
    start_idx = max(int(row_start) - 8, 0)
    end_idx = max(int(row_end) - 8, start_idx)

    seg = toc_df.iloc[start_idx:end_idx + 1].copy()

    # Calcular temps relatiu
    seg_time = pd.to_datetime(seg[time_col], errors="coerce")
    if t_start is not None:
        t0 = pd.to_datetime(t_start, errors="coerce")
    else:
        t0 = seg_time.iloc[0] if len(seg_time) > 0 else None

    if t0 is not None:
        t_min = (seg_time - t0).dt.total_seconds() / 60.0
    else:
        t_min = pd.Series(range(len(seg))) * 0.1  # Fallback

    # --- Detectar i corregir timestamps arrodonits al minut ---
    # Alguns exports Sievers perden la resolució de segons al 2-TOC.
    # Es detecta per la proporció de timestamps duplicats (cadència real = 4s, ~15 pts/min).
    _toc_minute_precision = False
    if len(t_min) > 30:
        n_unique = t_min.nunique()
        ratio_unique = n_unique / len(t_min)
        # Normal: ~100% únics. Arrodonit al minut: ~6-7% únics (1/15)
        if ratio_unique < 0.15:
            # Reconstruir cadència uniforme dins de cada minut
            t_min_corrected = np.empty(len(t_min), dtype=np.float64)
            pos = 0
            for t_val in sorted(t_min.unique()):
                mask = t_min.values == t_val
                n_pts = mask.sum()
                offsets = np.arange(n_pts) * (1.0 / n_pts)
                t_min_corrected[pos:pos + n_pts] = t_val + offsets
                pos += n_pts
            t_min = pd.Series(t_min_corrected, index=t_min.index)
            _toc_minute_precision = True
            import logging
            logging.getLogger(__name__).info(
                "TOC timestamps arrodonits al minut detectats: %d punts, "
                "%d unics -> cadencia reconstruida (%.0f pts/min)",
                len(t_min), n_unique, len(t_min) / n_unique
            )

    df_doc = pd.DataFrame({
        "time (min)": t_min.values,
        "DOC": pd.to_numeric(seg[sig_col], errors="coerce").values
    })
    df_doc = df_doc.dropna(subset=["time (min)", "DOC"])

    # Truncar si supera la durada màxima (útil per última mostra)
    if max_duration_min is not None and len(df_doc) > 0:
        max_time = df_doc["time (min)"].max()
        if max_time > max_duration_min:
            df_doc = df_doc[df_doc["time (min)"] <= max_duration_min].copy()

    # Detectar timeouts si s'ha sol·licitat
    if detect_timeouts:
        timeout_info = detect_timeout(df_doc["time (min)"].values)
        if _toc_minute_precision:
            timeout_info["toc_minute_precision"] = True
        return df_doc, timeout_info

    return df_doc



# =============================================================================
# LECTURA FITXERS UIB (DOC)
# =============================================================================

def llegir_doc_uib(path):
    """
    Llegeix fitxer DOC format UIB (CSV tab-separated, UTF-16 o UTF-8).

    Returns:
        (DataFrame, status): DataFrame amb columns ['time (min)', 'DOC'], status string
    """
    for enc in ["utf-16", "utf-8"]:
        try:
            df = pd.read_csv(path, sep="\t", encoding=enc, header=None, engine="python")
            df = df.iloc[:, [0, 1]]
            df.columns = ["time (min)", "DOC"]
            return df, f"OK{' (UTF-8)' if enc == 'utf-8' else ''}"
        except Exception:
            continue
    return pd.DataFrame(), "Error"


def netejar_nom_uib(nom_fitxer):
    """
    Neteja nom de fitxer UIB i extreu mostra/rèplica.

    Returns:
        (mostra, rep): Tuple amb nom de mostra i número de rèplica
    """
    stem = os.path.splitext(nom_fitxer)[0]
    stem = re.sub(r"UIB1B\d*", "", stem, flags=re.IGNORECASE)
    tokens = [t for t in re.split(r"[_\-\s]+", stem) if t]

    rep = None
    if tokens:
        last = tokens[-1]
        if last.isdigit():
            rep = last
            tokens = tokens[:-1]
        else:
            m = re.fullmatch(r"R(\d+)", last, flags=re.IGNORECASE)
            if m:
                rep = m.group(1)
                tokens = tokens[:-1]

    if rep is None:
        rep = "1"

    mostra = "_".join(tokens).strip("_")
    if not mostra:
        mostra = os.path.splitext(nom_fitxer)[0]
    return mostra, rep


# =============================================================================
# LECTURA FITXERS DAD
# =============================================================================

def llegir_dad_export3d(path, wavelengths_to_keep=None):
    """
    Llegeix fitxer DAD Export3D (CSV comma-separated).

    Args:
        path: Camí al fitxer CSV
        wavelengths_to_keep: Llista de wavelengths (int) a conservar. Si None, retorna tot.

    Returns:
        (DataFrame, status): DataFrame amb temps i wavelengths, status string
    """
    # Construir filtre usecols per evitar parsejar columnes innecessàries
    wl_set = set(int(w) for w in wavelengths_to_keep) if wavelengths_to_keep else None

    def _col_filter(col_name):
        """Retorna True per columnes a conservar (temps + wavelengths seleccionades)."""
        try:
            v = float(str(col_name).strip())
            return int(v) in wl_set
        except (ValueError, TypeError):
            return True  # Conservar columnes no numèriques (temps)

    for enc in ["utf-16", "utf-8"]:
        try:
            usecols = _col_filter if wl_set else None
            df = pd.read_csv(path, sep=",", encoding=enc, engine="python",
                             usecols=usecols)
            if df.shape[1] == 0:
                return pd.DataFrame(), "Buit"
            cols = list(df.columns)
            cols[0] = "time (min)"
            out_cols = [cols[0]]
            for c in cols[1:]:
                sc = str(c).strip()
                try:
                    v = float(sc)
                    out_cols.append(str(int(v)) if v.is_integer() else str(v))
                except Exception:
                    out_cols.append(sc)
            df.columns = out_cols
            return df, f"OK{' (UTF-8)' if enc == 'utf-8' else ''}"
        except Exception:
            continue
    return pd.DataFrame(), "Error"


def llegir_dad_1a(path, wavelength="254"):
    """
    Llegeix fitxer DAD1A (format Agilent: UTF-16, tab-separated, sense capçalera).

    Args:
        path: Camí al fitxer DAD1A
        wavelength: Nom de la columna de wavelength (default "254")

    Returns:
        (DataFrame, status): DataFrame amb columns ['time (min)', wavelength], status string
    """
    for enc in ["utf-16", "utf-16-le", "utf-8", "latin1"]:
        try:
            df = pd.read_csv(path, encoding=enc, sep="\t", header=None,
                           names=["time (min)", wavelength])
            df["time (min)"] = pd.to_numeric(df["time (min)"], errors="coerce")
            df[wavelength] = pd.to_numeric(df[wavelength], errors="coerce")
            df = df.dropna()

            if len(df) < 10:
                continue

            return df, f"OK (DAD1A, {enc})"
        except Exception:
            continue
    return pd.DataFrame(), "Error"


def llegir_dad_amb_fallback(path_export3d, path_dad1a=None, wavelength="254"):
    """
    Llegeix DAD amb fallback: primer Export3D, després DAD1A.

    Returns:
        (t, y, source): Arrays de temps i senyal, i string indicant la font
    """
    # Primer intentar Export3D
    if path_export3d and os.path.exists(path_export3d):
        df, status = llegir_dad_export3d(path_export3d)
        if status.startswith("OK") and not df.empty:
            if wavelength in df.columns:
                t = pd.to_numeric(df["time (min)"], errors="coerce").to_numpy()
                y = pd.to_numeric(df[wavelength], errors="coerce").to_numpy()
                valid = np.isfinite(t) & np.isfinite(y)
                if np.sum(valid) > 10:
                    return t[valid], y[valid], "Export3D"

    # Fallback a DAD1A
    if path_dad1a and os.path.exists(path_dad1a):
        df, status = llegir_dad_1a(path_dad1a, wavelength)
        if status.startswith("OK") and not df.empty:
            t = df["time (min)"].to_numpy()
            y = df[wavelength].to_numpy()
            if len(t) > 10:
                return t, y, "DAD1A"

    return None, None, "NOT_FOUND"


def list_dad_files(folder_export3d, folder_csv=None):
    """
    Llista tots els fitxers DAD (matriu 3D completa).

    NOMÉS busca a Export3d - els fitxers CSV són UIB (DOC), no DAD.

    Returns:
        Llista de paths a fitxers DAD
    """
    dad_files = []

    # Només Export3d conté fitxers DAD (matriu 3D)
    if folder_export3d and os.path.isdir(folder_export3d):
        for ext in ("*.csv", "*.CSV"):
            dad_files.extend(glob.glob(os.path.join(folder_export3d, ext)))

    # Excloure fitxers UIB (contenen "UIB1B") per si de cas
    dad_files = [f for f in dad_files if "UIB1B" not in os.path.basename(f).upper()]

    return sorted(set(dad_files))


# =============================================================================
# MATCHING MOSTRES (UIB ↔ HPLC-SEQ)
# =============================================================================

def _add_unique(lst, item):
    """Append item to list if not already present."""
    if item not in lst:
        lst.append(item)


def get_valid_samples_from_hplc_seq(master_data):
    """
    Extreu la llista de mostres vàlides de 1-HPLC-SEQ.

    Args:
        master_data: Dict retornat per llegir_masterfile_nou()

    Returns:
        list de noms de mostres (normalitzats) que pertanyen a la SEQ
    """
    df_hplc = master_data.get("hplc_seq")
    if df_hplc is None or df_hplc.empty:
        return []

    valid_samples = []

    # Buscar columna "Sample Name"
    sample_col = None
    for col in df_hplc.columns:
        if 'sample' in str(col).lower() and 'name' in str(col).lower():
            sample_col = col
            break

    if sample_col is None:
        return []

    # Extreure noms únics i normalitzar
    for val in df_hplc[sample_col].dropna().unique():
        name = str(val).strip()
        if name:
            _add_unique(valid_samples, name)
            _add_unique(valid_samples, normalize_key(name))

    return valid_samples


def match_sample_confidence(sample_name, valid_samples):
    """
    Cerca la millor coincidència d'una mostra amb la llista vàlida.

    Args:
        sample_name: Nom de la mostra (del fitxer UIB)
        valid_samples: Llista (o set) de mostres vàlides

    Returns:
        dict amb matched, best_match, confidence (0-100), match_type
    """
    # Accept both set and list for backward compat
    if isinstance(valid_samples, set):
        valid_samples = sorted(valid_samples)

    result = {
        "matched": False,
        "best_match": None,
        "confidence": 0.0,
        "match_type": "NOT_FOUND"
    }

    if not valid_samples:
        return {
            "matched": True,
            "best_match": sample_name,
            "confidence": 100.0,
            "match_type": "NO_VALIDATION"
        }

    sample_norm = normalize_key(sample_name)
    sample_upper = sample_name.upper()

    # 1. Coincidència exacta (100%)
    if sample_name in valid_samples:
        return {
            "matched": True,
            "best_match": sample_name,
            "confidence": 100.0,
            "match_type": "EXACT"
        }

    # 2. Coincidència normalitzada (95%)
    for vs in valid_samples:
        if normalize_key(vs) == sample_norm:
            return {
                "matched": True,
                "best_match": vs,
                "confidence": 95.0,
                "match_type": "NORMALIZED"
            }

    # 3. Case-insensitive (90%)
    for vs in valid_samples:
        if vs.upper() == sample_upper:
            return {
                "matched": True,
                "best_match": vs,
                "confidence": 90.0,
                "match_type": "CASE_INSENSITIVE"
            }

    # 4. Variants (85%)
    variants = [
        sample_name.replace("_", " "),
        sample_name.replace(" ", "_"),
        sample_name.replace("-", "_"),
        sample_name.replace("_", "-"),
        sample_name.replace("_", ""),
        sample_name.replace(" ", ""),
    ]
    for v in variants:
        v_norm = normalize_key(v)
        v_upper = v.upper()
        for vs in valid_samples:
            if vs == v or normalize_key(vs) == v_norm or vs.upper() == v_upper:
                return {
                    "matched": True,
                    "best_match": vs,
                    "confidence": 85.0,
                    "match_type": "VARIANT"
                }

    # 5. Fuzzy matching
    best_ratio = 0.0
    best_match = None

    for vs in valid_samples:
        ratio1 = SequenceMatcher(None, sample_upper, vs.upper()).ratio()
        ratio2 = SequenceMatcher(None, sample_norm, normalize_key(vs)).ratio()
        ratio = max(ratio1, ratio2)

        if ratio > best_ratio:
            best_ratio = ratio
            best_match = vs

    # Només suggerir si >95% i no és diferència numèrica
    sample_base = re.sub(r'\d+$', '', sample_name)

    if best_ratio >= 0.95:
        best_base = re.sub(r'\d+$', '', best_match) if best_match else ""
        if sample_base == best_base:
            return result

        confidence = best_ratio * 100
        return {
            "matched": True,
            "best_match": best_match,
            "confidence": confidence,
            "match_type": "FUZZY"
        }

    return result


def is_sample_in_seq(sample_name, valid_samples):
    """Verifica si una mostra pertany a la seqüència."""
    match_info = match_sample_confidence(sample_name, valid_samples)
    return match_info["matched"] and match_info["confidence"] >= CONFIDENCE_THRESHOLD


# =============================================================================
# MATCHING INTEL·LIGENT D'ORFES
# =============================================================================

def extract_sample_from_filename(filename):
    """
    Extreu el nom de mostra i rèplica d'un nom de fitxer.

    Exemples:
        "MOSTRA_A_R1.csv" → ("MOSTRA_A", 1)
        "KHP2_R2_UIB1B.csv" → ("KHP2", 2)
        "MOSTRA-B-F_1.csv" → ("MOSTRA-B-F", 1)
    """
    stem = os.path.splitext(os.path.basename(filename))[0]

    # Eliminar sufixos comuns (UIB1B, DAD1A, etc.)
    stem_clean = re.sub(r"_?UIB1B\d*", "", stem, flags=re.IGNORECASE)
    stem_clean = re.sub(r"_?DAD1A\d*", "", stem_clean, flags=re.IGNORECASE)
    # Netejar underscores/guions trailing (ex: MQ1__ → MQ1)
    stem_clean = stem_clean.rstrip("_-")

    # Buscar patró _R# o _#
    match_r = re.search(r"[_\-]R(\d+)$", stem_clean, flags=re.IGNORECASE)
    match_us = re.search(r"_(\d+)$", stem_clean)

    if match_r:
        sample = stem_clean[:match_r.start()]
        rep = int(match_r.group(1))
    elif match_us:
        sample = stem_clean[:match_us.start()]
        rep = int(match_us.group(1))
    else:
        sample = stem_clean
        rep = 1

    return sample, rep


def compute_orphan_suggestions(samples, orphan_files, file_type="dad"):
    """
    Calcula suggeriments de matching entre mostres sense dades i fitxers orfes.

    Args:
        samples: Dict de mostres {nom: {type, replicas: {rep: {dad, uib, ...}}}}
        orphan_files: Llista de paths de fitxers orfes
        file_type: "dad" o "uib"

    Returns:
        Dict de suggeriments:
        {
            (sample_name, replica): {
                "suggested_file": path,
                "suggested_filename": nom_fitxer,
                "confidence": 0-100,
                "match_type": "FUZZY" | "NORMALIZED" | "EXACT",
                "auto_assign": True si confidence > 85
            }
        }
    """
    suggestions = {}

    if not orphan_files:
        return suggestions

    # Construir índex d'orfes: {(sample_norm, rep): [files]}
    orphan_index = {}
    for fpath in orphan_files:
        sample, rep = extract_sample_from_filename(fpath)
        sample_norm = normalize_key(sample)
        key = (sample_norm, rep)
        if key not in orphan_index:
            orphan_index[key] = []
        orphan_index[key].append({
            "path": fpath,
            "filename": os.path.basename(fpath),
            "sample_extracted": sample,
        })

    # Per cada mostra/rèplica sense dades del tipus especificat
    for sample_name, sample_info in samples.items():
        for rep_str, rep_data in sample_info.get("replicas", {}).items():
            rep_num = int(rep_str) if rep_str.isdigit() else 1

            # Comprovar si falta el tipus de dades
            has_data = False
            if file_type == "dad":
                has_data = rep_data.get("dad") is not None
            elif file_type == "uib":
                has_data = rep_data.get("uib") is not None and rep_data["uib"].get("t") is not None

            if has_data:
                continue  # Ja té dades, no cal suggeriment

            sample_norm = normalize_key(sample_name)

            # Buscar el millor match entre els orfes
            sample_norm = normalize_key(sample_name)
            best_match = None
            best_confidence = 0
            best_match_type = "NONE"

            for (orphan_sample_norm, orphan_rep), orphan_list in orphan_index.items():
                # Primer, comprovar si la rèplica coincideix
                rep_matches = (orphan_rep == rep_num)
                orphan_sample_orig = orphan_list[0]["sample_extracted"].upper()

                # Calcular similitud del nom amb diverses estratègies
                if sample_norm == orphan_sample_norm:
                    # Match exacte (normalitzat)
                    confidence = 100 if rep_matches else 80
                    match_type = "EXACT"
                elif sample_norm.endswith(orphan_sample_norm) or sample_name.upper().endswith(orphan_sample_orig):
                    # La mostra ACABA amb el nom de l'orfe (ex: "3S101H HA" acaba amb "HA")
                    confidence = 95 if rep_matches else 75
                    match_type = "SUFFIX"
                elif orphan_sample_norm in sample_norm or orphan_sample_orig in sample_name.upper():
                    # L'orfe està CONTINGUT dins la mostra
                    confidence = 85 if rep_matches else 65
                    match_type = "CONTAINS"
                else:
                    # Fuzzy match
                    ratio1 = SequenceMatcher(None, sample_norm, orphan_sample_norm).ratio()
                    ratio2 = SequenceMatcher(None, sample_name.upper(), orphan_sample_orig).ratio()
                    ratio = max(ratio1, ratio2)

                    # Penalitzar si la rèplica no coincideix
                    if not rep_matches:
                        ratio *= 0.7

                    confidence = ratio * 100
                    match_type = "FUZZY"


                if confidence > best_confidence:
                    best_confidence = confidence
                    best_match = orphan_list[0]  # Primer fitxer de la llista
                    best_match_type = match_type

            # Guardar suggeriment si la confiança és suficient
            if best_match and best_confidence >= 60:
                suggestions[(sample_name, rep_num)] = {
                    "suggested_file": best_match["path"],
                    "suggested_filename": best_match["filename"],
                    "confidence": round(best_confidence, 1),
                    "match_type": best_match_type,
                    "auto_assign": best_confidence >= 85,
                }

    return suggestions


def apply_orphan_suggestions(samples, suggestions, orphan_files, file_type="dad"):
    """
    Aplica els suggeriments d'alta confiança automàticament.

    Args:
        samples: Dict de mostres (es modifica in-place)
        suggestions: Dict de suggeriments de compute_orphan_suggestions
        orphan_files: Llista d'orfes (es modifica per treure els assignats)
        file_type: "dad" o "uib"

    Returns:
        Tuple (applied_count, remaining_orphans)
    """
    applied = []

    for (sample_name, rep_num), suggestion in suggestions.items():
        if not suggestion.get("auto_assign"):
            continue

        # Trobar la mostra i rèplica
        if sample_name not in samples:
            continue

        rep_str = str(rep_num)
        if rep_str not in samples[sample_name].get("replicas", {}):
            continue

        rep_data = samples[sample_name]["replicas"][rep_str]

        # Marcar com a assignació suggerida (no carregar dades encara)
        if file_type == "dad":
            rep_data["dad_suggestion"] = {
                "file": suggestion["suggested_file"],
                "filename": suggestion["suggested_filename"],
                "confidence": suggestion["confidence"],
                "match_type": suggestion["match_type"],
                "status": "SUGGESTED",  # Per revisar a la GUI
            }
        elif file_type == "uib":
            rep_data["uib_suggestion"] = {
                "file": suggestion["suggested_file"],
                "filename": suggestion["suggested_filename"],
                "confidence": suggestion["confidence"],
                "match_type": suggestion["match_type"],
                "status": "SUGGESTED",
            }

        applied.append(suggestion["suggested_file"])

    # Treure els fitxers aplicats de la llista d'orfes
    remaining = [f for f in orphan_files if f not in applied]

    return len(applied), remaining


# =============================================================================
# MATCHING DAD FILES
# =============================================================================

def detect_dad_rep_style(dad_files):
    """
    Detecta l'estil de numeració de rèpliques als fitxers DAD.

    Returns:
        "R" si usa _R1, _R2
        "UNDERSCORE" si usa _1, _2
        "" si no es detecta
    """
    r_count = 0
    us_count = 0

    for f in dad_files:
        stem = os.path.splitext(os.path.basename(f))[0]
        if re.search(r'_R\d+$', stem, re.IGNORECASE):
            r_count += 1
        elif re.search(r'_\d+$', stem):
            us_count += 1

    if r_count > us_count:
        return "R"
    elif us_count > r_count:
        return "UNDERSCORE"
    return ""


def dad_sample_rep_from_path(path):
    """
    Extreu nom de mostra i rèplica d'un path DAD.

    Returns:
        (sample, rep, style): Nom mostra, número rèplica, estil detectat
    """
    stem = os.path.splitext(os.path.basename(path))[0]

    # Estil R
    match_r = re.search(r'^(.+?)_R(\d+)$', stem, re.IGNORECASE)
    if match_r:
        return match_r.group(1), match_r.group(2), "R"

    # Estil underscore
    match_us = re.search(r'^(.+?)_(\d+)$', stem)
    if match_us:
        return match_us.group(1), match_us.group(2), "UNDERSCORE"

    return stem, "1", "NONE"


def target_keys_from_desc(sample_desc):
    """Genera claus de matching a partir d'una descripció de mostra."""
    keys = []
    s = str(sample_desc).strip()
    if s:
        keys.append(s)
        keys.append(normalize_key(s))
        keys.append(s.upper())
        keys.append(s.replace("_", ""))
        keys.append(s.replace("-", "_"))
    return keys


def choose_best_candidate(target_key, rep, dad_files, used_files=None):
    """
    Selecciona el millor fitxer DAD per una mostra.

    Returns:
        (path, status) o (None, status)
    """
    used_files = used_files or set()
    target_norm = normalize_key(target_key)
    rep_str = str(rep) if rep else "1"

    candidates = []
    for f in dad_files:
        if f in used_files:
            continue
        sample, file_rep, style = dad_sample_rep_from_path(f)
        if normalize_key(sample) == target_norm and file_rep == rep_str:
            candidates.append((f, 100))  # Match exacte
        elif normalize_key(sample) == target_norm:
            candidates.append((f, 80))   # Match mostra però no rèplica

    if candidates:
        candidates.sort(key=lambda x: -x[1])
        return candidates[0][0], "OK"

    return None, "NOT_FOUND"


# =============================================================================
# MATCHING MASTER ↔ UIB
# =============================================================================

def find_master_row_for_sample(df_seq, mostra_uib, rep_uib):
    """
    Troba la fila del mestre que correspon a una mostra UIB.

    Args:
        df_seq: DataFrame del sheet 1-HPLC-SEQ del mestre
        mostra_uib: Nom de la mostra extret del fitxer UIB
        rep_uib: Rèplica extreta del fitxer UIB

    Returns:
        Series amb la fila trobada, o None si no es troba
    """
    mostra_key = normalize_key(mostra_uib)
    rep_norm = normalize_rep(rep_uib)

    for _, row in df_seq.iterrows():
        sample_desc = str(row.get("SAMPLE_DESC", "")).strip()
        if skip_sample_direct(sample_desc):
            continue

        base_name, rep_explicit = split_sample_rep(sample_desc)
        base_key = normalize_key(base_name)

        # Matching per nom
        if base_key == mostra_key or mostra_key in base_key or base_key in mostra_key:
            # Si tenim rèplica explícita, verificar que coincideix
            if rep_explicit and rep_norm:
                if normalize_rep(rep_explicit) == rep_norm:
                    return row
            else:
                # Sense rèplica explícita, acceptar si el nom coincideix
                return row

    return None


def build_master_sample_index(df_seq):
    """
    Construeix un índex de mostres del mestre per facilitar el matching.

    Args:
        df_seq: DataFrame del sheet 1-HPLC-SEQ del mestre

    Returns:
        Dict amb clau normalitzada -> llista de (idx, row, base_name, rep)
    """
    index = {}
    rep_counter = {}

    for idx, row in df_seq.iterrows():
        sample_desc = str(row.get("SAMPLE_DESC", "")).strip()
        if skip_sample_direct(sample_desc):
            continue

        base_name, rep_explicit = split_sample_rep(sample_desc)
        base_key = normalize_key(base_name)

        if rep_explicit is None:
            rep_counter[base_key] = rep_counter.get(base_key, 0) + 1
            rep = str(rep_counter[base_key])
        else:
            rep = str(rep_explicit)

        if base_key not in index:
            index[base_key] = []
        index[base_key].append({
            "idx": idx,
            "row": row,
            "base_name": base_name,
            "rep": rep,
            "sample_desc": sample_desc
        })

    return index


def match_uib_to_master(mostra_uib, rep_uib, master_index):
    """
    Busca la fila del mestre que correspon a una mostra UIB usant l'índex.

    Args:
        mostra_uib: Nom de la mostra extret del fitxer UIB
        rep_uib: Rèplica extreta del fitxer UIB
        master_index: Índex construït amb build_master_sample_index()

    Returns:
        Dict amb info de la fila trobada, o None si no es troba
    """
    mostra_key = normalize_key(mostra_uib)
    rep_norm = normalize_rep(rep_uib)

    # Buscar coincidència exacta
    if mostra_key in master_index:
        for entry in master_index[mostra_key]:
            if entry["rep"] == rep_norm:
                return entry
        # Si no hi ha coincidència de rèplica, retornar la primera
        if master_index[mostra_key]:
            return master_index[mostra_key][0]

    # Buscar coincidència parcial
    for key, entries in master_index.items():
        if mostra_key in key or key in mostra_key:
            for entry in entries:
                if entry["rep"] == rep_norm:
                    return entry
            if entries:
                return entries[0]

    return None


# =============================================================================
# VERIFICACIÓ FITXERS (QA/QC)
# =============================================================================

def detect_replica_anomalies(filenames):
    """
    Detecta anomalies en la numeració de rèpliques (ex: R12 en lloc de R2).

    Returns:
        List of dicts amb anomalies detectades
    """
    anomalies = []
    samples = {}

    for fname in filenames:
        stem = os.path.splitext(fname)[0]
        stem_clean = re.sub(r"_?UIB1B\d*", "", stem, flags=re.IGNORECASE)

        match_r = re.search(r"[_\-]R(\d+)$", stem_clean, flags=re.IGNORECASE)
        match_us = re.search(r"_(\d+)$", stem_clean)

        if match_r:
            base = stem_clean[:match_r.start()]
            rep = int(match_r.group(1))
            style = "R"
        elif match_us:
            base = stem_clean[:match_us.start()]
            rep = int(match_us.group(1))
            style = "UNDERSCORE"
        else:
            base = stem_clean
            rep = 1
            style = "NONE"

        base_key = re.sub(r"[^A-Za-z0-9]+", "", base).upper()
        if base_key not in samples:
            samples[base_key] = []
        samples[base_key].append({
            "filename": fname,
            "base": base,
            "rep": rep,
            "style": style
        })

    # Detectar anomalies
    for base_key, reps in samples.items():
        if len(reps) < 2:
            continue

        rep_nums = sorted([r["rep"] for r in reps])

        for i, rep_num in enumerate(rep_nums):
            expected = i + 1
            if rep_num != expected and rep_num > 3:
                for r in reps:
                    if r["rep"] == rep_num:
                        proposed_rep = expected
                        original = r["filename"]

                        if r["style"] == "R":
                            proposed = re.sub(
                                r"([_\-])R\d+(\.[^.]+)$",
                                rf"\1R{proposed_rep}\2",
                                original,
                                flags=re.IGNORECASE
                            )
                        elif r["style"] == "UNDERSCORE":
                            proposed = re.sub(
                                r"_\d+(\.[^.]+)$",
                                rf"_{proposed_rep}\1",
                                original
                            )
                        else:
                            proposed = original

                        if proposed != original:
                            anomalies.append({
                                "original": original,
                                "proposed": proposed,
                                "reason": f"R{rep_num} sembla error tipogràfic (esperat R{proposed_rep})",
                                "confidence": "HIGH" if rep_num > 9 else "MEDIUM",
                                "apply": False
                            })

    return anomalies


def check_sequence_files(seq_path, used_uib_files=None, used_dad_files=None):
    """
    Verifica els fitxers d'una seqüència i detecta discrepàncies.

    Returns:
        Dict amb info de verificació (orfes, anomalies, etc.)
    """
    used_uib_files = used_uib_files or set()
    used_dad_files = used_dad_files or set()

    path_csv = os.path.join(seq_path, "CSV")
    path_3d = os.path.join(seq_path, "Export3d")
    if not os.path.isdir(path_3d):
        path_3d = os.path.join(seq_path, "Export3D")

    result = {
        "seq_path": seq_path,
        "seq_name": os.path.basename(seq_path),
        "uib": {"found": [], "used": [], "orphan": [], "count_found": 0, "count_used": 0, "count_orphan": 0},
        "dad": {"found": [], "used": [], "orphan": [], "count_found": 0, "count_used": 0, "count_orphan": 0},
        "anomalies": [],
        "proposed_renames": [],
        "has_issues": False,
        "low_confidence_matches": [],
        "seq_orphan_files": [],
    }

    # Fitxers UIB
    if os.path.isdir(path_csv):
        uib_files = []
        for ext in ("*.csv", "*.CSV", "*.txt", "*.TXT"):
            uib_files.extend(glob.glob(os.path.join(path_csv, f"*UIB1B*{ext}")))
        uib_files = sorted(set(uib_files))

        result["uib"]["found"] = [os.path.basename(f) for f in uib_files]
        result["uib"]["count_found"] = len(uib_files)

        for f in uib_files:
            if f in used_uib_files:
                result["uib"]["used"].append(os.path.basename(f))
            else:
                result["uib"]["orphan"].append(os.path.basename(f))

        result["uib"]["count_used"] = len(result["uib"]["used"])
        result["uib"]["count_orphan"] = len(result["uib"]["orphan"])

        uib_anomalies = detect_replica_anomalies(result["uib"]["found"])
        result["proposed_renames"].extend(uib_anomalies)

    # Fitxers DAD
    if os.path.isdir(path_3d):
        dad_files = []
        for ext in ("*.csv", "*.CSV"):
            dad_files.extend(glob.glob(os.path.join(path_3d, ext)))
        dad_files = [f for f in dad_files if "UIB1B" not in os.path.basename(f).upper()]
        dad_files = sorted(set(dad_files))

        result["dad"]["found"] = [os.path.basename(f) for f in dad_files]
        result["dad"]["count_found"] = len(dad_files)

        for f in dad_files:
            if f in used_dad_files:
                result["dad"]["used"].append(os.path.basename(f))
            else:
                result["dad"]["orphan"].append(os.path.basename(f))

        result["dad"]["count_used"] = len(result["dad"]["used"])
        result["dad"]["count_orphan"] = len(result["dad"]["orphan"])

        dad_anomalies = detect_replica_anomalies(result["dad"]["found"])
        result["proposed_renames"].extend(dad_anomalies)

    # Determinar si hi ha problemes
    result["has_issues"] = (
        result["uib"]["count_orphan"] > 0 or
        result["dad"]["count_orphan"] > 0 or
        len(result["proposed_renames"]) > 0
    )

    return result


# =============================================================================
# FUNCIÓ PRINCIPAL: import_sequence() v2
# =============================================================================

def parse_injections_from_masterfile(master_data, config=None):
    """
    Parseja les injeccions del MasterFile (fulla 1-HPLC-SEQ).

    Gestiona controls repetits (NaOH, MQ) assignant-los _1, _2, etc.
    per ordre d'aparició.

    Returns:
        list of dict: Llista d'injeccions amb info normalitzada
        list of str: Warnings (números no correlatius, etc.)
    """
    config = config or get_config()
    injections = []
    warnings = []

    # Obtenir DataFrame de HPLC-SEQ (pot ser 'hplc_seq' o 'seq')
    df_seq = master_data.get("hplc_seq")
    if df_seq is None:
        df_seq = master_data.get("seq")
    if df_seq is None or (hasattr(df_seq, 'empty') and df_seq.empty):
        return [], ["ERROR: No s'ha trobat fulla 1-HPLC-SEQ al MasterFile"], 0

    # ==========================================================================
    # IDENTIFICAR COLUMNES PER NOM (primer match)
    # El MasterFile ha d'estar ben format: columnes A-F amb dades.
    # Si falta alguna columna → error clar indicant quina fulla revisar.
    # ==========================================================================
    errors = []

    sample_col = None
    inj_col = None
    line_col = None
    volume_col = None
    sample_rep_col = None

    for col in df_seq.columns:
        col_lower = str(col).lower().strip()
        if col_lower == "sample_rep":
            sample_rep_col = col
        elif ("sample" in col_lower and "name" in col_lower) and sample_col is None:
            sample_col = col
        elif (col_lower in ("inj#", "inj")) and inj_col is None:
            inj_col = col
        elif (col_lower in ("line#", "line")) and line_col is None:
            line_col = col
        elif ("volume" in col_lower or "vol" in col_lower) and volume_col is None:
            volume_col = col

    # Prioritat volum: 1) Capçalera explícita (ja detectada amunt)
    #                  2) 0-INFO (fiable)
    #                  3) Columna N índex 13 (heurístic, NOMÉS si no BP)

    # 0-INFO B4 (Inj_Volume) — font fiable
    info_volume = None
    info_data = master_data.get("info", {})
    for key, val in info_data.items():
        if "volume" in str(key).lower() or "vol" in str(key).lower():
            try:
                info_volume = float(val)
            except (ValueError, TypeError):
                pass
            break

    # Columna N (índex 13) per volum — prioritat sobre 0-INFO
    # Heurístic: columna sense capçalera a posició 13 amb valors numèrics 50-1000.
    # S'aplica a TOTS els modes (COLUMN i BP). Els volums per injecció són
    # la font de veritat (per-injecció > global 0-INFO).
    if volume_col is None:
        col_list = list(df_seq.columns)
        if len(col_list) > 13:
            potential_vol_col = col_list[13]
            try:
                sample_vals = df_seq[potential_vol_col].dropna().head(5)
                if len(sample_vals) > 0:
                    numeric_vals = pd.to_numeric(sample_vals, errors='coerce').dropna()
                    if len(numeric_vals) > 0 and all(50 <= v <= 1000 for v in numeric_vals):
                        volume_col = potential_vol_col
                        logger.info("Volume heuristic: columna index-13 '%s' detectada amb valors %s",
                                    potential_vol_col,
                                    [float(v) for v in numeric_vals.head(3)])
                        # Validació creuada: si 0-INFO té volum i difereix de col N
                        if info_volume is not None:
                            col_n_vals = set(int(v) for v in numeric_vals)
                            if len(col_n_vals) == 1 and int(info_volume) == list(col_n_vals)[0]:
                                pass  # Coincideixen, tot OK
                            elif len(col_n_vals) >= 1 and int(info_volume) not in col_n_vals:
                                warnings.append(
                                    f"VOLUM DISCREPANT: 0-INFO diu {int(info_volume)}µL "
                                    f"però columna N (index-13) té {sorted(col_n_vals)}µL. "
                                    f"S'utilitzen els volums per injecció de la columna N.")
            except Exception as e:
                logger.debug("Volume column index-13 heuristic failed: %s", e)

    # --- VALIDACIÓ: columnes obligatòries han de tenir dades ---
    # Detectar si les dades estan desplaçades (columnes .1 amb dades, originals buides)
    def _detect_offset_hint():
        for col in df_seq.columns:
            if ".1" in str(col) and df_seq[col].notna().sum() > 0:
                return (" Les dades semblen estar a columnes desplaçades (G-L en lloc de A-F)."
                        " Obrir el MasterFile, moure les dades a les columnes A-F i prémer Re-importar.")
        return ""

    if line_col is None:
        hint = _detect_offset_hint()
        return [], [f"ERROR fulla 1-HPLC-SEQ: no s'ha trobat columna 'Line#'.{hint} Revisar el MasterFile."], 0

    n_data_rows = df_seq[line_col].notna().sum()
    if n_data_rows == 0:
        hint = _detect_offset_hint()
        return [], [f"ERROR fulla 1-HPLC-SEQ: columna 'Line#' (A) buida.{hint} Revisar el MasterFile."], 0

    if sample_col is None:
        return [], ["ERROR fulla 1-HPLC-SEQ: no s'ha trobat columna 'Sample Name'. Revisar el MasterFile."], 0

    sample_data_rows = df_seq[sample_col].notna().sum()
    if sample_data_rows == 0:
        hint = _detect_offset_hint()
        return [], [f"ERROR fulla 1-HPLC-SEQ: columna 'Sample Name' (D) buida.{hint} Revisar el MasterFile."], 0

    # Comptador per controls repetits
    control_counts = {}  # base_name -> total count
    control_sets = {}    # base_name -> current set number
    control_last_inj = {} # base_name -> last inj_num (per detectar nou set)
    control_patterns = ["naoh", "mq", "blank", "buffer", "wash", "h2o", "milli", "blk", "blanc"]

    # Pre-calcular tots els noms de mostra per detectar duplicats exactes
    all_sample_names = [
        str(row.get(sample_col, "")).strip()
        for _, row in df_seq.iterrows()
        if str(row.get(sample_col, "")).strip() not in ["", "nan"]
    ]
    # Set de noms únics per detectar col·lisions amb generate_agilent_control_name
    # Ex: MQ (3a aparició) → "MQ1", però "MQ1" ja existeix com a mostra diferent
    all_sample_names_set = set(all_sample_names)

    # Identificar columnes crítiques per validació
    rt_col = None
    area_col = None
    for col in df_seq.columns:
        col_lower = str(col).lower().strip()
        if col_lower == "rt" or "retention" in col_lower:
            rt_col = col
        elif col_lower == "area":
            area_col = col

    # Columnes addicionals: Location, Method, Date
    location_col = None
    method_col = None
    date_col = None
    for col in df_seq.columns:
        col_lower = str(col).lower().strip()
        if col_lower == "location" and location_col is None:
            location_col = col
        elif ("method" in col_lower) and method_col is None:
            method_col = col
        elif ("date" in col_lower or "acquired" in col_lower) and date_col is None:
            date_col = col

    critical_cols = {
        'Line#': line_col,           # A
        'Inj#': inj_col,             # B
        'Location': location_col,     # C
        'Sample Name': sample_col,    # D
        'Method': method_col,         # E
        'Date': date_col,             # F
        'Volume': volume_col,         # N
    }
    # Filtrar None
    critical_cols = {k: v for k, v in critical_cols.items() if v is not None}

    # Comptar total de files amb dades (per validació posterior)
    total_rows_with_line = 0
    skipped_rows = []
    incomplete_rows = []  # Files amb cel·les crítiques buides

    # Processar cada fila
    prev_line = 0
    for idx, row in df_seq.iterrows():
        # Detectar si la fila té número de línia (indica que és una injecció real)
        line_val = row.get(line_col) if line_col else None
        has_line_num = line_val is not None and str(line_val).strip() not in ["", "nan"]
        if has_line_num:
            total_rows_with_line += 1

            # Validar cel·les crítiques buides
            missing_cols = []
            for col_name, col_ref in critical_cols.items():
                if col_ref is not None:
                    val = row.get(col_ref)
                    if val is None or str(val).strip() in ["", "nan", "NaN"]:
                        missing_cols.append(col_name)
            if missing_cols:
                try:
                    line_num_val = int(line_val)
                    incomplete_rows.append((line_num_val, missing_cols))
                except (ValueError, TypeError):
                    pass

        sample_name = str(row.get(sample_col, "")).strip()
        if not sample_name or sample_name.lower() in ["nan", ""]:
            # Si té número de línia però no té nom, és un error al MasterFile
            if has_line_num:
                try:
                    line_num_val = int(line_val)
                    skipped_rows.append(line_num_val)
                    warnings.append(f"⚠️ Línia {line_num_val}: falta nom de mostra (columna buida)")
                except (ValueError, TypeError):
                    pass
            continue

        # Obtenir número de línia/injecció
        line_num = row.get(line_col) if line_col else idx + 1
        inj_num = row.get(inj_col) if inj_col else 1

        try:
            line_num = int(line_num)
        except (ValueError, TypeError):
            line_num = idx + 1

        try:
            inj_num = int(inj_num)
        except (ValueError, TypeError):
            inj_num = 1

        # Check correlativitat
        if prev_line > 0 and line_num != prev_line + 1:
            warnings.append(f"Injeccions no correlatives: {prev_line} -> {line_num}")
        prev_line = line_num

        sample_lower = sample_name.lower()

        # Inicialitzar variables que s'usen després del if/else
        is_repeated_control = False
        original_name_count = 1

        # =====================================================================
        # ASSIGNAR NOM ÚNIC: Prioritzar Sample_Rep del MasterFile (migrate)
        # =====================================================================
        if sample_rep_col is not None:
            sample_rep_val = row.get(sample_rep_col)
            if sample_rep_val and str(sample_rep_val).strip() not in ["", "nan"]:
                # Sample_Rep format: "NOM_R1" o "NOM1_R2" → extreure NOM sense _R{num}
                sample_rep = str(sample_rep_val).strip()
                # Extreure part del nom (sense _R{num} al final)
                match = re.match(r'^(.+)_R(\d+)$', sample_rep)
                if match:
                    unique_name = match.group(1)
                    effective_inj_num = int(match.group(2))
                else:
                    unique_name = sample_rep
                    effective_inj_num = inj_num

                # Determinar tipus
                if is_khp(sample_name):
                    sample_type = "KHP"
                elif (pr_type := classify_reference_standard(sample_name, config)):
                    sample_type = pr_type
                elif is_blank_injection(sample_name, config):
                    sample_type = "BLANK"
                elif is_control_injection(sample_name, config):
                    sample_type = "CONTROL"
                elif "test" in sample_lower:
                    sample_type = "TEST"
                else:
                    sample_type = "SAMPLE"
            else:
                # Sample_Rep buit - fallback a nom original
                unique_name = sample_name
                effective_inj_num = inj_num
                if is_khp(sample_name):
                    sample_type = "KHP"
                elif (pr_type := classify_reference_standard(sample_name, config)):
                    sample_type = pr_type
                elif is_blank_injection(sample_name, config):
                    sample_type = "BLANK"
                elif is_control_injection(sample_name, config):
                    sample_type = "CONTROL"
                elif "test" in sample_lower:
                    sample_type = "TEST"
                else:
                    sample_type = "SAMPLE"
        else:
            # =====================================================================
            # FALLBACK: MasterFile antic sense Sample_Rep - generar nom al vol
            # =====================================================================
            is_repeated_control = any(p in sample_lower for p in control_patterns)

            if is_repeated_control:
                # Només eliminar espais i guions alts, mantenir xifres i underscores
                base_name = re.sub(r'[\s\-]+', '', sample_name).strip()
                if not base_name:
                    base_name = sample_name

                if base_name not in control_sets:
                    control_sets[base_name] = 1
                    control_last_inj[base_name] = 0

                if inj_num <= control_last_inj.get(base_name, 0):
                    control_sets[base_name] += 1

                control_last_inj[base_name] = inj_num
                current_set = control_sets[base_name]

                if base_name not in control_counts:
                    control_counts[base_name] = 0
                control_counts[base_name] += 1

                original_name_count = all_sample_names.count(sample_name)
                if original_name_count > 1:
                    unique_name = generate_agilent_control_name(base_name, current_set)
                    # Evitar col·lisió amb noms de mostra reals al MasterFile
                    # Ex: MQ (3a aparició) → "MQ1", però "MQ1" ja existeix com a mostra diferent
                    while unique_name in all_sample_names_set and unique_name != base_name:
                        current_set += 1
                        control_sets[base_name] = current_set
                        unique_name = generate_agilent_control_name(base_name, current_set)
                else:
                    unique_name = sample_name
                # Diferenciar BLANK de CONTROL
                if is_blank_injection(sample_name, config):
                    sample_type = "BLANK"
                else:
                    sample_type = "CONTROL"
                effective_inj_num = inj_num
            else:
                unique_name = sample_name
                effective_inj_num = inj_num
                if is_khp(sample_name):
                    sample_type = "KHP"
                elif (pr_type := classify_reference_standard(sample_name, config)):
                    sample_type = pr_type
                elif is_blank_injection(sample_name, config):
                    sample_type = "BLANK"
                elif is_control_injection(sample_name, config):
                    sample_type = "CONTROL"
                elif "test" in sample_lower:
                    sample_type = "TEST"
                else:
                    sample_type = "SAMPLE"

        # Obtenir volum d'injecció — NO suposar cap valor per defecte
        # Prioritat: 1) Columna explícita/heurística, 2) 0-INFO global
        # Si cap font té dades → inj_volume = None (warning emès al final)
        inj_volume = None
        if volume_col is not None:
            try:
                vol_val = row.get(volume_col)
                if vol_val is not None and str(vol_val).strip() not in ["", "nan"]:
                    inj_volume = float(vol_val)
            except (ValueError, TypeError):
                pass
        # Fallback: volum de 0-INFO (comú per tota la seqüència)
        if inj_volume is None and info_volume is not None:
            inj_volume = info_volume

        # Info de set per controls (per matching amb fitxers MQ1_R1, MQ2_R1, etc.)
        # NOMÉS usar set/rep si hi ha duplicats exactes - si els noms ja són únics, fer match directe
        if is_repeated_control and original_name_count > 1:
            control_set_num = control_sets.get(base_name)
            control_rep_in_set = inj_num
        else:
            control_set_num = None
            control_rep_in_set = None


        # Extreure location/method/date usant els noms de columna detectats
        inj_location = None
        if location_col is not None:
            val = row.get(location_col)
            if val is not None and str(val).strip() not in ("", "nan"):
                inj_location = str(val).strip()

        inj_method = None
        if method_col is not None:
            val = row.get(method_col)
            if val is not None and str(val).strip() not in ("", "nan"):
                inj_method = str(val).strip()

        inj_date = ""
        if date_col is not None:
            val = row.get(date_col)
            if val is not None and str(val).strip() not in ("", "nan"):
                inj_date = str(val).strip()

        injections.append({
            "line_num": line_num,
            "inj_num": effective_inj_num,  # Replica efectiva
            "inj_num_original": inj_num,   # Replica original del MasterFile
            "sample_name_original": sample_name,
            "sample_name": unique_name,
            "sample_type": sample_type,
            "inj_volume": inj_volume,  # Volum d'injecció en µL (pot ser None)
            "inj_location": inj_location,
            "inj_method": inj_method,
            "inj_date": inj_date,
            "control_set": control_set_num,  # Número de set per controls (MQ1, MQ2, ...)
            "control_rep": control_rep_in_set,  # Rèplica dins del set (R1, R2, ...)
            "row_data": row.to_dict(),
        })

    # Validar: volums d'injecció
    inj_without_volume = [inj for inj in injections if inj.get("inj_volume") is None]
    if inj_without_volume:
        n_missing = len(inj_without_volume)
        n_total = len(injections)
        if n_missing == n_total:
            warnings.insert(0,
                f"VOLUM DESCONEGUT: Cap injecció té volum assignat. "
                f"No s'ha trobat capçalera 'Volume', ni 0-INFO, ni columna index-13 vàlida. "
                f"Caldrà assignar volums manualment o afegir-los al MasterFile.")
        else:
            lines_missing = [str(inj.get("line_num", "?")) for inj in inj_without_volume[:5]]
            lines_str = ", ".join(lines_missing)
            if n_missing > 5:
                lines_str += f"... (+{n_missing - 5})"
            warnings.insert(0,
                f"VOLUM PARCIAL: {n_missing}/{n_total} injeccions sense volum (lines: {lines_str}). "
                f"Revisar MasterFile.")

    # Validar: comparar total de files amb Line# vs injeccions processades
    if total_rows_with_line > len(injections):
        missing = total_rows_with_line - len(injections)
        warnings.insert(0, f"ATENCIÓ: {missing} injecció(ns) no processada(es) - revisar MasterFile (files: {skipped_rows})")

    # Validar: files amb cel·les crítiques buides
    if incomplete_rows:
        # Agrupar per columnes que falten
        by_missing = {}
        for line_num, cols in incomplete_rows:
            key = tuple(sorted(cols))
            if key not in by_missing:
                by_missing[key] = []
            by_missing[key].append(line_num)

        for cols, lines in by_missing.items():
            lines_str = ", ".join(str(l) for l in lines[:5])
            if len(lines) > 5:
                lines_str += f"... (+{len(lines)-5})"
            warnings.append(
                f"⚠️ FILES INCOMPLETES: Línies {lines_str} - falta: {', '.join(cols)}"
            )

    # Afegir errors de validació MasterFile al principi dels warnings
    if errors:
        warnings = errors + warnings

    # Retornar també el total de línies del MasterFile per validació posterior
    return injections, warnings, total_rows_with_line


def _toc_calc_has_minute_precision(toc_calc_df):
    """Detecta si el 4-TOC_CALC té timestamps arrodonits al minut (necessita regeneració)."""
    if toc_calc_df is None or toc_calc_df.empty:
        return False
    col = "Temps_Relatiu (min)"
    if col not in toc_calc_df.columns:
        return False
    assigned = toc_calc_df[toc_calc_df["Inj_Index"] > 0]
    if len(assigned) <= 30:
        return False
    t = assigned[col].dropna()
    return len(t) > 0 and t.nunique() / len(t) < 0.15


def _save_toc_calc_to_masterfile(master_path, toc_calc_df):
    """
    Guarda el 4-TOC_CALC calculat al MasterFile.

    Si el full ja existeix, el sobreescriu. Si no, el crea al final.

    Returns:
        True si s'ha guardat correctament, False si ha fallat.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(str(master_path))

        # Eliminar full existent si existeix
        if '4-TOC_CALC' in wb.sheetnames:
            del wb['4-TOC_CALC']

        ws = wb.create_sheet('4-TOC_CALC')

        # Escriure capçaleres
        headers = list(toc_calc_df.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Escriure dades
        for row_idx, (_, row) in enumerate(toc_calc_df.iterrows(), 2):
            for col_idx, header in enumerate(headers, 1):
                val = row[header]
                # Convertir numpy types a Python natiu
                if hasattr(val, 'item'):
                    val = val.item()
                elif pd.isna(val):
                    val = None
                ws.cell(row=row_idx, column=col_idx, value=val)

        wb.save(str(master_path))
        wb.close()
        logger.info(f"4-TOC_CALC guardat al MasterFile ({len(toc_calc_df)} files)")
        return True

    except Exception as e:
        logger.warning(f"No s'ha pogut guardar 4-TOC_CALC al MasterFile: {e}")
        return False


def compute_toc_calc(master_data, toc_df):
    """
    Calcula 4-TOC_CALC in-memory quan no existeix al MasterFile (ex: plantilla).

    Replica la lògica de _create_masterfile de hpsec_migrate_master.py:
    - Extreu timestamps TOC de 2-TOC (columna D, fila 8+)
    - Extreu timestamps HPLC de 1-HPLC-SEQ (Acquired Date)
    - Calcula net_delay des de 0-INFO (hora HPLC - hora TOC)
    - Assigna cada punt TOC a la injecció corresponent

    Returns:
        pd.DataFrame amb columnes [TOC_Row, Sample, Temps_Relatiu (min), Inj_Index]
        o None si no es pot calcular.
    """
    FLUSH_TIME_MIN = 3.637

    # 1. Obtenir HPLC-SEQ amb timestamps
    df_seq = master_data.get("hplc_seq")
    if df_seq is None:
        df_seq = master_data.get("seq")
    if df_seq is None or (hasattr(df_seq, 'empty') and df_seq.empty):
        logger.debug("compute_toc_calc: No hi ha 1-HPLC-SEQ")
        return None

    # Trobar columnes rellevants
    date_col = sample_col = sample_rep_col = None
    for col in df_seq.columns:
        col_lower = str(col).lower().strip()
        if 'acquired date' in col_lower or ('injection' in col_lower and 'date' in col_lower):
            date_col = col
        elif ('sample' in col_lower and 'name' in col_lower) and sample_col is None:
            sample_col = col
        elif col_lower == 'sample_rep':
            sample_rep_col = col

    if date_col is None or sample_col is None:
        logger.debug(f"compute_toc_calc: Falten columnes: date_col={date_col}, sample_col={sample_col}")
        return None

    # Parsejar timestamps HPLC
    df_hplc = df_seq[df_seq[sample_col].notna()].copy()
    try:
        df_hplc[date_col] = pd.to_datetime(df_hplc[date_col])
    except Exception as e:
        logger.debug(f"compute_toc_calc: Error parsejant dates HPLC: {e}")
        return None

    df_hplc = df_hplc.sort_values(date_col).reset_index(drop=True)
    hplc_times = df_hplc[date_col].values

    # Obtenir noms de mostra (preferir Sample_Rep si existeix)
    if sample_rep_col and sample_rep_col in df_hplc.columns:
        hplc_samples = df_hplc[sample_rep_col].values
    else:
        # Crear Sample_Rep: nom_Rx per a cada rèplica
        counts = {}
        sample_reps = []
        for name in df_hplc[sample_col].values:
            name_str = str(name).strip()
            counts[name_str] = counts.get(name_str, 0) + 1
            sample_reps.append(f"{name_str}_R{counts[name_str]}")
        hplc_samples = sample_reps

    if len(hplc_times) == 0:
        logger.debug("compute_toc_calc: No hi ha timestamps HPLC")
        return None

    # 2. Extreure timestamps TOC de 2-TOC (columna D, fila 8+)
    toc_timestamps = []
    TOC_DATA_START_ROW = 8  # Les dades TOC comencen a la fila 8 (1-indexed)
    if toc_df is not None and toc_df.shape[1] > 3:
        # toc_df ja té header=6, així que la fila 0 del df correspon a la fila 7 de l'Excel
        # Les dades comencen a la fila 8 de l'Excel = fila 1 del df (index 0-based)
        for i in range(len(toc_df)):
            # El TOC_Row a l'Excel és la fila real del full (i + header_offset + 1)
            # Com que header=6, fila 0 del df = fila 8 de l'Excel
            excel_row = i + TOC_DATA_START_ROW
            # Columna D (index 3) del df original conté el timestamp
            # Però amb header=6, la columna pot tenir diferent index
            # Intentar trobar la columna de timestamp per nom o posició
            val = None
            # Primer buscar per posició (columna D = index 3)
            col_list = list(toc_df.columns)
            if len(col_list) > 3:
                val = toc_df.iloc[i, 3]

            if pd.notna(val):
                try:
                    toc_timestamps.append((excel_row, pd.to_datetime(val)))
                except Exception:
                    pass

    if not toc_timestamps:
        logger.debug("compute_toc_calc: No s'han trobat timestamps TOC")
        return None

    # 3. Calcular net_delay des de 0-INFO
    info = master_data.get("info", {})
    hora_hplc_clock = None
    hora_toc_clock = None

    for key, val in info.items():
        key_lower = str(key).lower()
        if 'hora hplc' in key_lower or 'hora_hplc' in key_lower:
            hora_hplc_clock = val
        elif 'hora toc' in key_lower or 'hora_toc' in key_lower:
            hora_toc_clock = val

    net_delay_min = FLUSH_TIME_MIN
    if hora_hplc_clock and hora_toc_clock:
        try:
            def to_minutes(t):
                if hasattr(t, 'hour'):
                    return t.hour * 60 + t.minute + t.second / 60
                parts = str(t).split(':')
                return int(parts[0]) * 60 + int(parts[1])

            hplc_min = to_minutes(hora_hplc_clock)
            toc_min = to_minutes(hora_toc_clock)
            desfase_min = hplc_min - toc_min
            net_delay_min = FLUSH_TIME_MIN - desfase_min
        except Exception:
            net_delay_min = FLUSH_TIME_MIN

    # 4. Calcular mapping TOC → injecció
    # Pre-margin: el pic DOC s'eixampla per dispersió al reactor TOC, la pujada
    # pot començar ABANS de l'hora d'injecció HPLC. Si una fila TOC cau dins
    # [inj_start_{N+1} - pre_margin, inj_start_{N+1}], assignar a injecció N+1.
    pre_margin_min = 1.5  # minuts — dispersió reactor TOC
    try:
        from hpsec_config import get_config
        pre_margin_min = get_config().get("sequence", "toc_pre_margin_min", default=1.5)
    except Exception:
        pass
    pre_margin_ns = pre_margin_min * 60 * 1e9  # nanoseconds

    rows = []
    for toc_row, toc_time in toc_timestamps:
        hora_hplc = toc_time - pd.Timedelta(minutes=net_delay_min)
        hora_hplc_ns = hora_hplc.value
        inj_index = int((hplc_times <= hora_hplc).sum())

        # Check pre-margin: si cau just abans de la injecció següent, reassignar
        if inj_index < len(hplc_times):
            next_start_ns = pd.Timestamp(hplc_times[inj_index]).value
            if (next_start_ns - hora_hplc_ns) <= pre_margin_ns:
                inj_index += 1

        if 0 < inj_index <= len(hplc_samples):
            sample = hplc_samples[inj_index - 1]
            inj_start = pd.Timestamp(hplc_times[inj_index - 1])
            temps_rel = (hora_hplc - inj_start).total_seconds() / 60.0
        else:
            sample = ''
            temps_rel = None

        rows.append({
            "TOC_Row": toc_row,
            "Sample": sample,
            "Temps_Relatiu (min)": round(temps_rel, 3) if temps_rel is not None else None,
            "Inj_Index": inj_index,
        })

    if not rows:
        return None

    toc_calc_df = pd.DataFrame(rows)

    # --- Corregir timestamps arrodonits al minut ---
    # Si el 2-TOC té timestamps sense segons, els Temps_Relatiu queden escalonats.
    # Reconstruir cadència uniforme dins de cada injecció.
    assigned = toc_calc_df[toc_calc_df["Inj_Index"] > 0]
    if len(assigned) > 30:
        t_rel = assigned["Temps_Relatiu (min)"]
        n_unique = t_rel.nunique()
        ratio = n_unique / len(t_rel)
        if ratio < 0.15:
            # Reconstruir per cada injecció
            for inj_idx in toc_calc_df["Inj_Index"].unique():
                if inj_idx <= 0:
                    continue
                mask = toc_calc_df["Inj_Index"] == inj_idx
                group = toc_calc_df.loc[mask, "Temps_Relatiu (min)"]
                if group.isna().all():
                    continue
                t_vals = sorted(group.dropna().unique())
                if len(t_vals) < 2:
                    continue
                corrected = np.empty(mask.sum(), dtype=np.float64)
                pos = 0
                for t_val in t_vals:
                    sub_mask = group.values == t_val
                    n_pts = sub_mask.sum()
                    offsets = np.arange(n_pts) * (1.0 / n_pts)
                    corrected[pos:pos + n_pts] = t_val + offsets
                    pos += n_pts
                # Handle NaN entries at the end
                if pos < len(corrected):
                    corrected[pos:] = np.nan
                toc_calc_df.loc[mask, "Temps_Relatiu (min)"] = np.round(corrected, 3)
            logger.info(
                "compute_toc_calc: timestamps arrodonits al minut detectats, "
                "cadencia reconstruida"
            )

    logger.debug(f"compute_toc_calc: Calculat 4-TOC_CALC amb {len(toc_calc_df)} files")
    return toc_calc_df


def _strip_all(s):
    """Normalitza eliminant TOTS els separadors (espais, guions, underscores) i majúscules."""
    return re.sub(r"[\-\s_]+", "", str(s or "")).upper()


def _match_khp_dad_from_masterfile(sample_name, original_name, inj_num, master_khp_data):
    """
    Busca dades DAD 254nm per un KHP al full 3-DAD_KHP del MasterFile.

    Strategy:
    1. Exact match: busca columna que contingui el nom exacte + _R{inj_num}
    2. Base match: busca totes les columnes que continguin el nom base del KHP
       (sense concentració repetida) i agafa la N-èsima per inj_num.

    Format 3-DAD_KHP: columnes amb nom KHP seguides de time/value columns.
    Row 0 conté "time (min)" / "value (mAU)" com a headers.

    Returns:
        (dad_dict, source_str) o (None, None) si no trobat
    """
    def _try_load_col(col):
        """Intenta carregar time+value des d'una columna nomenada."""
        col_idx = master_khp_data.columns.get_loc(col)
        if col_idx + 1 >= len(master_khp_data.columns):
            return None
        time_col = col
        value_col = master_khp_data.columns[col_idx + 1]
        time_data = master_khp_data[time_col].iloc[1:] if len(master_khp_data) > 1 else master_khp_data[time_col]
        value_data = master_khp_data[value_col].iloc[1:] if len(master_khp_data) > 1 else master_khp_data[value_col]
        df_khp = pd.DataFrame({
            "time (min)": pd.to_numeric(time_data, errors="coerce"),
            "254": pd.to_numeric(value_data, errors="coerce"),
        }).dropna().reset_index(drop=True)
        if not df_khp.empty and len(df_khp) > 5:
            return {"path": "MasterFile:3-DAD_KHP", "df": df_khp, "source": "masterfile"}
        return None

    # --- Strategy 1: exact key match (incloent _R{inj_num}) ---
    khp_keys = [
        f"{sample_name}_{inj_num}_R{inj_num}",
        f"{original_name}_{inj_num}_R{inj_num}",
        f"{original_name}_R{inj_num}",
        f"{sample_name}_R{inj_num}",
    ]

    for khp_key in khp_keys:
        key_norm = normalize_key(khp_key)
        key_stripped = _strip_all(khp_key)

        for col in master_khp_data.columns:
            if str(col).startswith("Unnamed"):
                continue
            col_norm = normalize_key(str(col))
            col_stripped = _strip_all(str(col))

            if (col_norm == key_norm or key_norm in col_norm
                    or col_stripped == key_stripped or key_stripped in col_stripped):
                result = _try_load_col(col)
                if result:
                    return result, "masterfile"
                break

    # --- Strategy 2: base name match (sense _R{N}), agafar la inj_num-èsima ---
    # Construir nom base: "KHP100PPB" o "KHP 100 ppb" → stripped = "KHP100PPB"
    base_names = list(dict.fromkeys([_strip_all(sample_name), _strip_all(original_name)]))

    # Trobar totes les columnes que continguin el nom base
    matching_cols = []
    for col in master_khp_data.columns:
        if str(col).startswith("Unnamed"):
            continue
        col_stripped = _strip_all(str(col))
        for base in base_names:
            if base in col_stripped:
                matching_cols.append(col)
                break

    # Agafar la inj_num-èsima columna (1-based)
    idx = int(inj_num) - 1 if str(inj_num).isdigit() else 0
    if 0 <= idx < len(matching_cols):
        result = _try_load_col(matching_cols[idx])
        if result:
            logger.debug("3-DAD_KHP: base match %s[%d] → col '%s'",
                         base_names[0], idx, matching_cols[idx])
            return result, "masterfile"

    # --- Strategy 3: generic "KHP" prefix match ---
    # Per SEQs amb noms genèrics al MasterFile (KHP_R1, KHP 1_R2) que no contenen
    # la concentració, buscar totes les columnes que comencin per "KHP" i assignar
    # per inj_num. Això funciona per SEQs amb un sol tipus de KHP (la majoria).
    all_khp_cols = []
    for col in master_khp_data.columns:
        if str(col).startswith("Unnamed"):
            continue
        col_stripped = _strip_all(str(col))
        if col_stripped.startswith("KHP"):
            all_khp_cols.append(col)

    if all_khp_cols and 0 <= idx < len(all_khp_cols):
        result = _try_load_col(all_khp_cols[idx])
        if result:
            logger.debug("3-DAD_KHP: generic KHP[%d] → col '%s'", idx, all_khp_cols[idx])
            return result, "masterfile"

    return None, None


def _validate_toc_assignments(result):
    """
    Valida assignacions TOC: detecta solapaments entre injeccions i injeccions sense DOC.

    Afegeix warnings al result quan:
    - Dues injeccions comparteixen el mateix rang de files TOC
    - Una injecció no té dades DOC Direct assignades
    """
    # Recopilar (sample, rep, row_start, row_end) de totes les rèpliques
    assignments = []
    no_doc_injections = []

    for sample_name, sample_data in result.get("samples", {}).items():
        for rep_key, rep_data in sample_data.get("replicas", {}).items():
            direct = rep_data.get("direct")
            if direct and direct.get("row_start") and direct.get("row_end"):
                assignments.append({
                    "sample": sample_name,
                    "rep": rep_key,
                    "row_start": direct["row_start"],
                    "row_end": direct["row_end"],
                    "line_num": rep_data.get("injection_info", {}).get("line_num", 0),
                })
            else:
                line_num = rep_data.get("injection_info", {}).get("line_num", "?")
                no_doc_injections.append(f"{sample_name}_R{rep_key} (línia {line_num})")

    # Detectar injeccions sense DOC Direct
    if no_doc_injections:
        n = len(no_doc_injections)
        if n <= 5:
            names = ", ".join(no_doc_injections)
        else:
            names = ", ".join(no_doc_injections[:5]) + f"... (+{n-5} més)"
        result["warnings"].append(
            f"⚠️ {n} injeccions sense DOC Direct (no assignades al 4-TOC_CALC): {names}"
        )

    # Detectar solapaments
    if len(assignments) < 2:
        return

    # Ordenar per row_start
    assignments.sort(key=lambda a: a["row_start"])

    overlaps = []
    for i in range(len(assignments) - 1):
        a = assignments[i]
        b = assignments[i + 1]
        # Solapament si b comença abans que a acabi
        if b["row_start"] <= a["row_end"]:
            overlap_rows = a["row_end"] - b["row_start"] + 1
            # Solapament petit (1-2 files) pot ser normal (fronteres)
            if overlap_rows > 5:
                overlaps.append(
                    f"{a['sample']}_R{a['rep']} [{a['row_start']}-{a['row_end']}] ↔ "
                    f"{b['sample']}_R{b['rep']} [{b['row_start']}-{b['row_end']}] "
                    f"({overlap_rows} files solapades)"
                )

    # Detectar rangs idèntics (dues injeccions amb exactament el mateix rang)
    from collections import Counter
    range_counts = Counter((a["row_start"], a["row_end"]) for a in assignments)
    duplicate_ranges = {rng: cnt for rng, cnt in range_counts.items() if cnt > 1}

    if duplicate_ranges:
        for (rs, re_), cnt in duplicate_ranges.items():
            affected = [
                f"{a['sample']}_R{a['rep']}"
                for a in assignments
                if a["row_start"] == rs and a["row_end"] == re_
            ]
            if len(affected) <= 5:
                names = ", ".join(affected)
            else:
                names = ", ".join(affected[:5]) + f"... (+{len(affected)-5} més)"
            result["warnings"].append(
                f"⚠️ CRÍTIC: {cnt} injeccions comparteixen el mateix rang TOC "
                f"[{rs}-{re_}]: {names}. "
                f"Probable error d'assignació — les dades DOC seran idèntiques!"
            )

    if overlaps and not duplicate_ranges:
        n = len(overlaps)
        if n <= 3:
            detail = "; ".join(overlaps)
        else:
            detail = "; ".join(overlaps[:3]) + f"... (+{n-3} més)"
        result["warnings"].append(
            f"⚠️ {n} solapaments TOC detectats: {detail}"
        )


def find_data_for_injection(injection, seq_path, uib_files, dad_files, dad_csv_files,
                            master_khp_data, used_files, config=None,
                            toc_df=None, toc_calc_df=None,
                            valid_sample_names=None,
                            seq_timeouts=None, method=None):
    """
    Busca dades per una injecció des de múltiples fonts.

    Prioritat:
    - DOC Direct: MasterFile 2-TOC via 4-TOC_CALC (SEMPRE)
    - DOC UIB: CSV UIB (si disponible)
    - DAD 254 (KHP): Export3d > CSV DAD > MasterFile 3-DAD_KHP

    Args:
        valid_sample_names: Set de noms de mostra normalitzats vàlids (per evitar
                           fallback matching quan el nom del fitxer és una mostra vàlida)

    Returns:
        dict amb dades trobades
    """
    config = config or get_config()
    result = {
        "has_data": False,
        "direct": None,  # DOC Direct del MasterFile
        "uib": None,     # DOC UIB del CSV
        "dad": None,
        "dad_source": None,  # "export3d", "csv", "masterfile"
    }

    sample_name = injection["sample_name"]
    inj_num = injection["inj_num"]
    line_num = injection["line_num"]
    sample_type = injection["sample_type"]
    original_name = injection["sample_name_original"]
    inj_num_original = injection.get("inj_num_original", inj_num)
    control_set = injection.get("control_set")  # Número de set per controls (MQ1, MQ2, ...)
    control_rep = injection.get("control_rep")  # Rèplica dins del set (R1, R2, ...)

    # 0. Extreure DOC Direct del MasterFile (SEMPRE si disponible)
    if toc_df is not None and toc_calc_df is not None and not toc_calc_df.empty:
        row_start = None
        row_end = None

        # Detectar format (nou vs antic)
        is_new_format = "Sample" in toc_calc_df.columns and "TOC_Row" in toc_calc_df.columns
        is_old_format = "SAMPLE_ID" in toc_calc_df.columns and "Row initial" in toc_calc_df.columns

        if is_new_format:
            # Format nou: 4-TOC_CALC amb Sample, TOC_Row, Inj_Index
            # Prioritzar cerca per line_num (Inj_Index) que és més fiable per controls
            mask = None
            match_method = None

            # 1. Primer intentar per Inj_Index (line_num) - més fiable
            if "Inj_Index" in toc_calc_df.columns:
                mask = (toc_calc_df["Inj_Index"] == line_num)
                if not mask.any():
                    mask = None
                else:
                    match_method = "inj_index"

            # 2. Si no, intentar per nom exacte amb rèplica
            if mask is None:
                sample_key_r = f"{original_name}_R{inj_num_original}"
                mask = (toc_calc_df["Sample"] == sample_key_r)
                if not mask.any():
                    # 3. Fallback: nom parcial (RESTRINGIT)
                    # Treure sufixos _1, _2 del original_name per matching més flexible
                    original_base = re.sub(r'[_\-]?\d+$', '', original_name).strip()
                    if original_base and len(original_base) >= 3:
                        mask = (toc_calc_df["Sample"].astype(str).str.contains(
                            original_base, case=False, na=False))
                        match_method = "fallback_partial"
                    else:
                        mask = None
                else:
                    match_method = "exact_name"

            sample_rows = toc_calc_df[mask] if mask is not None else pd.DataFrame()
            if not sample_rows.empty:
                toc_rows = sample_rows["TOC_Row"].dropna()
                if len(toc_rows) > 0:
                    candidate_start = int(toc_rows.min())
                    candidate_end = int(toc_rows.max())
                    row_span = candidate_end - candidate_start

                    # Validació: si el rang és excessiu (>2500 files ≈ 2x normal),
                    # el fallback ha assignat files d'altres injeccions → descartar
                    MAX_REASONABLE_ROWS = 2500  # ~1200 normal, 2x marge
                    if match_method == "fallback_partial" and row_span > MAX_REASONABLE_ROWS:
                        logger.warning(
                            f"TOC fallback massa ampli per {sample_name} (line {line_num}): "
                            f"rows {candidate_start}-{candidate_end} ({row_span} files). Descartant."
                        )
                        # No assignar — millor no tenir DOC que tenir-lo erroni
                    else:
                        row_start = candidate_start
                        row_end = candidate_end

        elif is_old_format:
            # Format antic: 4-SEQ_DATA amb SAMPLE_ID, Row initial, Row Final
            # SAMPLE_ID pot ser numero (1, 2, 3...) o text (Pre, Post, MQ)
            mask = (toc_calc_df["SAMPLE_ID"].astype(str) == str(original_name))
            if not mask.any():
                # Provar amb Line# si existeix
                if "Line#" in toc_calc_df.columns:
                    mask = (toc_calc_df["Line#"] == line_num)

            sample_rows = toc_calc_df[mask]
            if not sample_rows.empty:
                # Usar primera fila que coincideixi
                row_data = sample_rows.iloc[0]
                try:
                    row_start = int(row_data.get("Row initial", 0))
                    row_end = int(row_data.get("Row Final", 0))
                except (ValueError, TypeError):
                    row_start = None
                    row_end = None

        if row_start is not None and row_end is not None and row_start > 0:
            # Extreure DOC Direct (sense detecció timeout local — ve de seq_timeouts)
            max_dur = config.get("max_duration_min", 80.0)
            df_doc = extract_doc_from_masterfile(
                toc_df, row_start, row_end, detect_timeouts=False,
                max_duration_min=max_dur
            )

            if not df_doc.empty:
                t_direct = df_doc["time (min)"].values
                y_direct = df_doc["DOC"].values

                # Timeout info: font única = detecció global
                is_bp_mode = (method or "").upper() == "BP"
                timeout_info = map_timeouts_to_injection(
                    seq_timeouts or [], row_start, row_end,
                    t_rel=t_direct, is_bp=is_bp_mode
                )

                result["direct"] = {
                    "path": "MasterFile:2-TOC",
                    "t": t_direct,
                    "y": y_direct,
                    "row_start": row_start,
                    "row_end": row_end,
                    "timeout_info": timeout_info,
                    "timeout_at_boundary": timeout_info["timeout_at_boundary"],
                    "boundary_timeout": timeout_info["boundary_timeout"],
                }
                result["has_data"] = True

    # 1. Buscar UIB a CSV
    for uib_path in uib_files:
        if uib_path in used_files.get("uib", set()):
            continue

        filename = os.path.basename(uib_path)
        file_sample, file_rep = netejar_nom_uib(filename)

        # Match per nom original o nom únic
        file_sample_norm = normalize_key(file_sample)
        original_norm = normalize_key(original_name)
        unique_norm = normalize_key(sample_name)

        # També comparar amb el nom base sense sufix _1, _2, etc.
        # (El MasterFile pot tenir "NaOH 0.1mM_1" però el fitxer "NAOH 0.1MM")
        original_base_norm = re.sub(r'\d+$', '', original_norm)  # Treure números del final

        # Match directe
        name_match = (file_sample_norm == original_norm or
                      file_sample_norm == unique_norm or
                      file_sample_norm == original_base_norm)

        # Match per controls amb lògica Agilent:
        # - Set 1: unique_name="MQ", fitxers MQ_1.CSV, MQ_2.CSV
        # - Set 2: unique_name="MQ1", fitxers MQ1_1.CSV, MQ1_2.CSV
        # Fallback: Si no hi ha match directe, extreure base+número del fitxer
        # IMPORTANT: No usar fallback si el nom del fitxer és una mostra vàlida
        # (ex: no fer match MQ1 file amb MQ injection si MQ1 existeix com a mostra)
        file_control_num = None
        if not name_match:
            # Verificar si el nom del fitxer és una mostra vàlida (no usar fallback)
            file_is_valid_sample = (valid_sample_names and file_sample_norm in valid_sample_names)

            if not file_is_valid_sample:
                # Extreure base i número del nom del fitxer (fallback per dades antigues)
                match = re.match(r'^(.+?)(\d+)$', file_sample_norm)
                if match:
                    file_sample_base = match.group(1)
                    file_control_num = int(match.group(2))
                    # Comparar base amb original
                    if file_sample_base == original_norm:
                        name_match = True

        if name_match:
            # Verificar rèplica
            try:
                file_rep_int = int(file_rep) if file_rep else 1
            except ValueError:
                file_rep_int = 1

            # Lògica de matching rèplica:
            # - Si tenim control_set i file_control_num (fallback): comparar ambdós
            # - Altrament: comparar rèplica del fitxer amb inj_num
            if control_set is not None and file_control_num is not None:
                # Control: fitxer MQ1_R2 → set 1, rep 2
                rep_match = (file_control_num == control_set and file_rep_int == control_rep)
            elif file_control_num is not None:
                # Fallback: comparar número del nom amb inj_num
                rep_match = (file_control_num == inj_num)
            else:
                # Mostra normal: comparar rèplica del fitxer amb inj_num
                rep_match = (file_rep_int == inj_num)

            if rep_match:
                # Llegir UIB
                df_uib, status = llegir_doc_uib(uib_path)
                if status.startswith("OK"):
                    used_files.setdefault("uib", set()).add(uib_path)

                    t_uib = df_uib["time (min)"].values if not df_uib.empty else None
                    y_uib = df_uib["DOC"].values if not df_uib.empty else None

                    # Downsample UIB a la cadència DOC Direct (~4s/punt)
                    # UIB CSV té dt≈0.005 min (12x més dens que Direct dt≈0.067).
                    # Sense downsample, el SG i derivades es comporten diferent.
                    if t_uib is not None and len(t_uib) > 10:
                        t_uib, y_uib = downsample_to_cadence(t_uib, y_uib)

                    # Propagar timeout_info de Direct a UIB (mateix detector TOC)
                    uib_timeout = (result.get("direct") or {}).get("timeout_info", {})
                    result["uib"] = {
                        "path": uib_path,
                        "df": df_uib,  # df original (per referència)
                        "t": t_uib,
                        "y": y_uib,
                        "timeout_info": uib_timeout,
                        "timeout_at_boundary": uib_timeout.get("timeout_at_boundary", False),
                    }
                    result["has_data"] = True
                break

    # 2. Buscar DAD a Export3d (prioritat 1)
    for dad_path in dad_files:
        if dad_path in used_files.get("dad", set()):
            continue

        filename = os.path.basename(dad_path)
        file_sample, file_rep, _ = dad_sample_rep_from_path(dad_path)

        file_sample_norm = normalize_key(file_sample)
        original_norm = normalize_key(original_name)
        unique_norm = normalize_key(sample_name)

        # També comparar amb el nom base sense sufix numèric
        original_base_norm = re.sub(r'\d+$', '', original_norm)

        # Match directe
        name_match = (file_sample_norm == original_norm or
                      file_sample_norm == unique_norm or
                      file_sample_norm == original_base_norm)

        # Match per controls amb lògica Agilent (fallback per dades antigues)
        # IMPORTANT: No usar fallback si el nom del fitxer és una mostra vàlida
        file_control_num = None
        file_sample_base = None
        if not name_match:
            # Verificar si el nom del fitxer és una mostra vàlida (no usar fallback)
            file_is_valid_sample = (valid_sample_names and file_sample_norm in valid_sample_names)

            if not file_is_valid_sample:
                # Extreure base i número del nom del fitxer (MQ1 → base=MQ, num=1)
                match = re.match(r'^(.+?)(\d+)$', file_sample_norm)
                if match:
                    file_sample_base = match.group(1)
                    file_control_num = int(match.group(2))
                    # Comparar base amb original
                    if file_sample_base == original_norm or file_sample_base == original_base_norm:
                        name_match = True

        if name_match:
            try:
                file_rep_int = int(file_rep) if file_rep else 1
            except ValueError:
                file_rep_int = 1

            # Lògica de matching rèplica:
            # - Si tenim control_set i file_control_num (fallback): comparar ambdós
            # - Altrament: comparar rèplica del fitxer amb inj_num
            if control_set is not None and file_control_num is not None:
                # Control: fitxer MQ1_R2 → set 1, rep 2
                rep_match = (file_control_num == control_set and file_rep_int == control_rep)
            elif file_control_num is not None:
                # Fallback: comparar número del nom amb inj_num
                rep_match = (file_control_num == inj_num)
            else:
                # Mostra normal: comparar rèplica del fitxer amb inj_num
                rep_match = (file_rep_int == inj_num)

            if rep_match:
                wl_keep = config.get("wavelengths", "selected") if config else None
                df_dad, status = llegir_dad_export3d(dad_path, wavelengths_to_keep=wl_keep)
                if status.startswith("OK"):
                    used_files.setdefault("dad", set()).add(dad_path)
                    result["dad"] = {
                        "path": dad_path,
                        "df": df_dad,
                    }
                    result["dad_source"] = "export3d"
                    result["has_data"] = True
                break

    # 2b. Fallback: buscar DAD a CSV (DAD1A format) si no hem trobat Export3d
    if result["dad"] is None and dad_csv_files:
        for dad_path in dad_csv_files:
            if dad_path in used_files.get("dad", set()):
                continue

            filename = os.path.basename(dad_path)
            # Extreure nom mostra del fitxer CSV (format típic: MOSTRA_DAD1A.csv)
            file_sample = filename.replace("_DAD1A", "").replace("_DAD", "")
            file_sample = os.path.splitext(file_sample)[0]

            # Intentar extreure rèplica del nom
            match = re.match(r'^(.+?)_R(\d+)$', file_sample, re.IGNORECASE)
            if match:
                file_sample_base = match.group(1)
                file_rep = int(match.group(2))
            else:
                file_sample_base = file_sample
                file_rep = 1

            file_sample_norm = normalize_key(file_sample_base)
            original_norm = normalize_key(original_name)

            if file_sample_norm == original_norm and file_rep == inj_num:
                df_dad, status = llegir_dad_1a(dad_path)
                if status.startswith("OK") and not df_dad.empty:
                    used_files.setdefault("dad", set()).add(dad_path)
                    result["dad"] = {
                        "path": dad_path,
                        "df": df_dad,
                    }
                    result["dad_source"] = "csv"
                    result["has_data"] = True
                    break

    # 3. Si és KHP i no tenim DAD, buscar a MasterFile 3-DAD_KHP
    if sample_type == "KHP" and result["dad"] is None and master_khp_data is not None:
        dad_result, dad_src = _match_khp_dad_from_masterfile(
            sample_name, original_name, inj_num, master_khp_data
        )
        if dad_result is not None:
            result["dad"] = dad_result
            result["dad_source"] = dad_src
            result["has_data"] = True
            logger.debug("3-DAD_KHP: SUCCESS! %d files per %s (inj %s)",
                         len(dad_result["df"]), sample_name, inj_num)
        else:
            logger.debug("3-DAD_KHP: NO MATCH per %s (inj %s)", sample_name, inj_num)

    return result


# =============================================================================
# GENERACIÓ D'AVISOS ESTRUCTURATS PER IMPORTACIÓ
# =============================================================================

def _generate_import_warnings(result: dict) -> list:
    """
    Genera avisos estructurats a partir del resultat d'importació.

    Tots els avisos usen create_anomaly() (font única: ANOMALY_CATALOG).

    Args:
        result: Dict del resultat de import_sequence()

    Returns:
        Llista d'avisos estructurats
    """
    warnings = []

    # 1. Errors crítics (BLOCKER)
    for error in result.get("errors", []):
        if "no data" in error.lower() or "buida" in error.lower() or "empty" in error.lower():
            warnings.append(create_anomaly("IMP_NO_DATA"))
        elif "uib" in error.lower() and ("missing" in error.lower() or "falt" in error.lower()):
            warnings.append(create_anomaly("IMP_MISSING_UIB"))
        elif "dad" in error.lower() and ("missing" in error.lower() or "falt" in error.lower()):
            warnings.append(create_anomaly("IMP_MISSING_DAD"))
        else:
            anomaly = create_anomaly(
                "IMP_NO_DATA",
                details={"message": error},
            )
            anomaly["message"] = error
            warnings.append(anomaly)

    # 2. Fitxers orfes
    orphan_uib = result.get("orphan_files", {}).get("uib", [])
    orphan_dad = result.get("orphan_files", {}).get("dad", [])
    n_orphan = len(orphan_uib) + len(orphan_dad)

    if n_orphan > 0:
        all_names = [os.path.basename(f) for f in orphan_uib + orphan_dad]
        anomaly = create_anomaly(
            "IMP_ORPHAN_FILES",
            details={"n": n_orphan, "files": ", ".join(all_names),
                      "uib": orphan_uib, "dad": orphan_dad},
        )
        # Incloure noms al message perquè la barra d'avisos els mostri
        anomaly["message"] = f"Fitxers sense assignar: {', '.join(all_names)}"
        warnings.append(anomaly)

    # 3. Warnings del manifest (strings) — extreure info accionable
    for w_str in result.get("warnings", []):
        w_lower = w_str.lower() if isinstance(w_str, str) else ""
        if "incompleta" in w_lower or "duplicat" in w_lower or "duplicada" in w_lower:
            anomaly = create_anomaly("IMP_INCOMPLETE")
            # Usar el text original que ja és descriptiu
            anomaly["message"] = w_str.lstrip("\u26a0\ufe0f ").strip()
            warnings.append(anomaly)

    # 4. Cromatogrames truncats (DOC massa curt)
    # COLUMN: mínim 30 min = 450 punts (dt=4s)
    # BP: mínim 4 min = 60 punts (dt=4s)
    is_bp = result.get("method", "").upper() == "BP"
    min_doc_points = 60 if is_bp else 450
    samples = result.get("samples", {})
    if isinstance(samples, dict):
        for sample_name, sample_data in samples.items():
            if not isinstance(sample_data, dict):
                continue
            replicas = sample_data.get("replicas", {})
            if not isinstance(replicas, dict):
                continue
            for rep_key, rep in replicas.items():
                if not isinstance(rep, dict):
                    continue
                direct = rep.get("direct", {})
                if isinstance(direct, dict):
                    n_pts = direct.get("n_points", 0)
                    if 0 < n_pts < min_doc_points:
                        min_min = 4 if is_bp else 30
                        anomaly = create_anomaly(
                            "IMP_SHORT_CHROMATOGRAM",
                            sample=sample_name,
                            details={
                                "n_points": n_pts,
                                "min_points": min_doc_points,
                                "minutes": round(n_pts * 4 / 60, 1),
                                "min_minutes": min_min,
                                "replica": rep_key,
                            },
                        )
                        anomaly["message"] = (
                            f"{sample_name} R{rep_key}: cromatograma truncat "
                            f"({round(n_pts * 4 / 60, 1)} min, mínim {min_min} min)"
                        )
                        warnings.append(anomaly)

    # 5. TOC amb timestamps arrodonits al minut (cadència reconstruïda)
    if isinstance(samples, dict):
        for sample_data in samples.values():
            if not isinstance(sample_data, dict):
                continue
            replicas = sample_data.get("replicas", {})
            if not isinstance(replicas, dict):
                continue
            for rep in replicas.values():
                if not isinstance(rep, dict):
                    continue
                direct = rep.get("direct", {})
                if isinstance(direct, dict):
                    ti = direct.get("timeout_info", {})
                    if isinstance(ti, dict) and ti.get("toc_minute_precision"):
                        warnings.append(create_anomaly("IMP_TOC_MINUTE_PRECISION"))
                        # Només un avís per seqüència (afecta totes les mostres)
                        return warnings

    return warnings


def import_sequence(seq_path, config=None, progress_callback=None):
    """
    FASE 1: Importar dades RAW d'una seqüència (v2).

    FONT DE VERITAT: MasterFile fulla 1-HPLC-SEQ

    1. Llegeix MasterFile per obtenir llista completa d'injeccions
    2. Verifica correlativitat (warning si falten números)
    3. Per cada injecció, busca dades a CSV/Export3d/MasterFile
    4. Cross-check fitxers orfes

    Args:
        seq_path: Ruta a la carpeta de la seqüència
        config: Configuració (opcional)
        progress_callback: Funció callback(pct, msg) per progrés

    Returns:
        dict amb:
        - success: bool
        - seq_path: path original
        - seq_name: nom de la seqüència
        - method: "BP" o "COLUMN"
        - data_mode: "UIB", "DIRECT", o "DUAL"
        - injections: llista d'injeccions del MasterFile
        - samples: {nom: {type, replicas: {rep: {uib, dad}}}}
        - khp_samples, control_samples: llistes
        - orphan_files: fitxers no assignats
        - errors, warnings: llistes
    """
    config = config or get_config()

    result = {
        "success": False,
        "seq_path": seq_path,
        "seq_name": os.path.basename(seq_path),
        "method": "COLUMN",
        "data_mode": "UIB",
        "date": "",
        "master_data": None,
        "master_file": None,
        "master_format": None,
        "injections": [],
        "samples": {},
        "khp_samples": [],
        "control_samples": [],
        "valid_samples": [],
        "orphan_files": {"uib": [], "dad": []},
        "file_check": None,
        "errors": [],
        "warnings": [],
    }

    def report_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    try:
        report_progress(5, "Detectant mode...")

        # 1. Detectar mode (BP/COLUMN)
        mode_info = detect_mode_from_folder(seq_path)
        result["method"] = mode_info["method"]
        result["data_mode"] = mode_info["data_mode"]

        report_progress(10, "Llegint MasterFile...")

        # 2. Llegir MasterFile (FONT DE VERITAT)
        master_path, master_format = trobar_excel_mestre(seq_path)
        result["master_format"] = master_format
        result["master_file"] = master_path

        if not master_path:
            report_progress(12, "MasterFile no trobat, intentant crear...")
            migrate_result = migrate_single(seq_path)

            if migrate_result.get("status") == "ok":
                master_path = migrate_result.get("file")
                master_format = "NEW"
                result["master_format"] = master_format
                result["master_file"] = master_path
                result["warnings"].append(
                    f"MasterFile creat des de rawdata ({migrate_result.get('rows', 0)} files TOC)"
                )
            elif migrate_result.get("status") == "exists":
                master_path = migrate_result.get("file")
                master_format = "NEW"
                result["master_format"] = master_format
                result["master_file"] = master_path

        if not master_path:
            result["errors"].append("No s'ha pogut obtenir MasterFile - font de veritat necessària")
            return result

        # Llegir MasterFile
        if master_format == "NEW":
            result["master_data"] = llegir_masterfile_nou(master_path)
        else:
            df_toc, df_seq = llegir_master_direct(master_path)
            result["master_data"] = {
                "toc": df_toc,
                "seq": df_seq,
                "filepath": master_path,
            }

        # Propagar method a master_data (perquè parse_injections pugui accedir-hi)
        result["master_data"]["method"] = result["method"]

        # Comprovar si llegir_masterfile_nou ha fallat (ex: fitxer obert a Excel)
        if result["master_data"].get("error"):
            result["errors"].append(result["master_data"]["error"])
            return result

        result["date"] = read_master_date(seq_path)

        # Extreure sensibilitat UIB si disponible (de 0-INFO B5)
        master_info = result["master_data"].get("info", {})
        uib_sensitivity = master_info.get("uib_sensitivity")
        if uib_sensitivity is not None:
            # Guard: pot ser string des d'Excel
            try:
                uib_sensitivity = float(uib_sensitivity)
            except (ValueError, TypeError):
                uib_sensitivity = None
            if uib_sensitivity is not None:
                result["uib_sensitivity"] = uib_sensitivity

        report_progress(20, "Parsejant injeccions del MasterFile...")

        # 3. Parsejar injeccions del MasterFile (FONT DE VERITAT)
        injections, parse_warnings, master_line_count = parse_injections_from_masterfile(result["master_data"], config)
        result["injections"] = injections
        result["master_line_count"] = master_line_count  # Total línies al MasterFile (Line#)
        result["warnings"].extend(parse_warnings)

        if not injections:
            # Usar el missatge específic del parsing si existeix (ex: "ERROR fulla 1-HPLC-SEQ...")
            specific_errors = [w for w in parse_warnings if w.upper().startswith("ERROR")]
            if specific_errors:
                result["errors"].extend(specific_errors)
            else:
                result["errors"].append("No s'han trobat injeccions al MasterFile")
            return result

        # Crear llista de mostres vàlides
        for inj in injections:
            _add_unique(result["valid_samples"], inj["sample_name"])
            _add_unique(result["valid_samples"], inj["sample_name_original"])

        report_progress(30, "Llistant fitxers disponibles...")

        # 4. Llistar fitxers disponibles
        path_csv = os.path.join(seq_path, "CSV")
        uib_files = []
        dad_csv_files = []
        if os.path.isdir(path_csv):
            for ext in ("*.csv", "*.CSV", "*.txt", "*.TXT"):
                uib_files.extend(glob.glob(os.path.join(path_csv, f"*UIB1B*{ext}")))
                dad_csv_files.extend(glob.glob(os.path.join(path_csv, f"*DAD*{ext}")))
            uib_files = sorted(set(uib_files))
            dad_csv_files = sorted(set(dad_csv_files))

        path_3d = os.path.join(seq_path, "Export3d")
        if not os.path.isdir(path_3d):
            path_3d = os.path.join(seq_path, "Export3D")
        dad_files = list_dad_files(path_3d) if os.path.isdir(path_3d) else []

        # Llegir dades addicionals del MasterFile
        master_khp_data = None
        toc_df = None
        toc_calc_df = None
        has_3dad_khp_sheet = False
        try:
            with pd.ExcelFile(master_path) as xl:
                if "3-DAD_KHP" in xl.sheet_names:
                    master_khp_data = pd.read_excel(master_path, sheet_name="3-DAD_KHP")
                    has_3dad_khp_sheet = True
                if "2-TOC" in xl.sheet_names:
                    toc_df = pd.read_excel(master_path, sheet_name="2-TOC", header=6)
                if "4-TOC_CALC" in xl.sheet_names:
                    toc_calc_df = pd.read_excel(master_path, sheet_name="4-TOC_CALC")
                elif "4-SEQ_DATA" in xl.sheet_names:
                    toc_calc_df = pd.read_excel(master_path, sheet_name="4-SEQ_DATA")
        except Exception as e:
            logger.warning(f"Error llegint fulls addicionals del MasterFile: {e}")

        # Detectar 4-TOC_CALC existent amb timestamps arrodonits al minut
        _needs_regen = _toc_calc_has_minute_precision(toc_calc_df)
        if _needs_regen:
            logger.info("4-TOC_CALC existent amb timestamps arrodonits, regenerant...")

        # Si 4-TOC_CALC no existeix, és buit, o té timestamps degradats → calcular
        if _needs_regen or toc_calc_df is None or (hasattr(toc_calc_df, 'empty') and toc_calc_df.empty):
            if toc_df is not None:
                result["warnings"].append(
                    "4-TOC_CALC no trobat al MasterFile, calculant automàticament..."
                )
                toc_calc_df = compute_toc_calc(result["master_data"], toc_df)
                if toc_calc_df is not None and not toc_calc_df.empty:
                    # Actualitzar master_data amb el toc_calc calculat
                    result["master_data"]["toc_calc"] = toc_calc_df
                    # Guardar 4-TOC_CALC al MasterFile per a futures importacions
                    saved = _save_toc_calc_to_masterfile(master_path, toc_calc_df)
                    if not saved:
                        result["warnings"].append(
                            "⚠️ 4-TOC_CALC calculat però no s'ha pogut escriure al MasterFile "
                            "(fitxer obert a Excel?). Es recalcularà a la propera importació."
                        )
                else:
                    result["warnings"].append(
                        "⚠️ No s'ha pogut calcular 4-TOC_CALC. DOC Direct no disponible."
                    )
            else:
                result["warnings"].append(
                    "⚠️ 4-TOC_CALC i 2-TOC no disponibles. DOC Direct no disponible."
                )

        # Warning KHP DAD: només si NO hi ha cap de les 3 fonts (3-DAD_KHP, CSV DAD, Export3D)
        has_export3d = os.path.isdir(path_3d) and len(dad_files) > 0
        has_dad_csv = len(dad_csv_files) > 0
        has_3dad_khp_data = has_3dad_khp_sheet and master_khp_data is not None and not master_khp_data.empty
        has_khp = any(inj.get("sample_type") == "KHP" for inj in injections)
        if has_khp and not has_export3d and not has_dad_csv and not has_3dad_khp_data:
            result["warnings"].append(
                "KHP sense dades DAD: no s'ha trobat Export3D, CSV DAD, ni full 3-DAD_KHP amb dades."
            )

        report_progress(40, "Processant injeccions...")

        # 4b. Detecció global de timeouts al flux TOC complet
        seq_timeouts = detect_sequence_timeouts(toc_df) if toc_df is not None else []
        if seq_timeouts:
            logger.info("Detectats %d timeouts al flux TOC complet", len(seq_timeouts))
        result["seq_timeouts"] = seq_timeouts

        # 5. Processar cada injecció
        used_files = {"uib": set(), "dad": set()}
        total_inj = len(injections)

        # Crear set de noms de mostra normalitzats per evitar fallback matching incorrecte
        # Això evita que fitxer "MQ1" faci match amb injecció "MQ" quan "MQ1" és mostra vàlida
        valid_sample_names = {normalize_key(inj["sample_name"]) for inj in injections}

        for i, inj in enumerate(injections):
            pct = 40 + int(45 * (i + 1) / max(total_inj, 1))
            report_progress(pct, f"Processant {inj['sample_name']}...")

            sample_name = inj["sample_name"]
            sample_type = inj["sample_type"]
            inj_num = inj["inj_num"]

            # Buscar dades per aquesta injecció
            data = find_data_for_injection(
                inj, seq_path, uib_files, dad_files, dad_csv_files,
                master_khp_data, used_files, config,
                toc_df=toc_df, toc_calc_df=toc_calc_df,
                valid_sample_names=valid_sample_names,
                seq_timeouts=seq_timeouts,
                method=result.get("method")
            )

            # Classificar mostra
            if sample_type == "KHP":
                if sample_name not in result["khp_samples"]:
                    result["khp_samples"].append(sample_name)
            elif sample_type == "CONTROL":
                if sample_name not in result["control_samples"]:
                    result["control_samples"].append(sample_name)

            # Crear entrada a samples
            if sample_name not in result["samples"]:
                result["samples"][sample_name] = {
                    "type": sample_type,
                    "original_name": inj["sample_name_original"],
                    "replicas": {}
                }

            # Processar dades DOC Direct si existeixen
            direct_data = data.get("direct")
            if direct_data and direct_data.get("t") is not None:
                t_direct = direct_data["t"]
                y_direct = direct_data["y"]
                y_direct_net = None
                baseline_direct = None

                if len(t_direct) > 10:
                    mode = "BP" if result["method"] == "BP" else "COLUMN"
                    baseline_direct = get_baseline_value(t_direct, y_direct, mode=mode)
                    y_direct_net = y_direct - baseline_direct

                direct_data["y_net"] = y_direct_net
                direct_data["baseline"] = baseline_direct

            # Processar dades UIB si existeixen
            uib_data = data.get("uib")
            if uib_data and uib_data.get("t") is not None:
                t_uib = uib_data["t"]
                y_uib = uib_data["y"]
                y_uib_net = None
                baseline_uib = None

                if len(t_uib) > 10:
                    mode = "BP" if result["method"] == "BP" else "COLUMN"
                    baseline_uib = get_baseline_value(t_uib, y_uib, mode=mode)
                    y_uib_net = y_uib - baseline_uib

                uib_data["y_net"] = y_uib_net
                uib_data["baseline"] = baseline_uib

            # Guardar rèplica
            rep_key = str(inj_num)
            # Detectar sobreescriptura (duplicats al MasterFile)
            if rep_key in result["samples"][sample_name]["replicas"]:
                result["warnings"].append(
                    f"⚠️ ATENCIÓ: '{sample_name}' rèplica {rep_key} duplicada (línia {inj['line_num']}) - revisar MasterFile"
                )
            result["samples"][sample_name]["replicas"][rep_key] = {
                "direct": direct_data,  # DOC Direct del MasterFile
                "uib": uib_data,        # DOC UIB del CSV (si disponible)
                "dad": data.get("dad"),
                "dad_source": data.get("dad_source"),
                "has_data": data.get("has_data", False),
                "injection_info": {
                    "line_num": inj["line_num"],
                    "inj_num": inj_num,
                    "inj_volume": inj.get("inj_volume"),
                    "inj_location": inj.get("inj_location"),
                    "inj_method": inj.get("inj_method"),
                    "inj_date": inj.get("inj_date", ""),
                },
            }

        # 5b. Validar assignacions TOC: detectar solapaments i injeccions sense DOC
        _validate_toc_assignments(result)

        report_progress(90, "Verificant fitxers orfes...")

        # 6. Identificar fitxers orfes (no assignats a cap mostra)
        orphan_uib = [f for f in uib_files if f not in used_files["uib"]]
        orphan_dad = [f for f in dad_files if f not in used_files["dad"]]


        # 7. Matching intel·ligent: suggerir assignacions per orfes
        report_progress(92, "Calculant suggeriments de matching...")

        # Guardar tots els orfes ABANS d'aplicar suggeriments (per comptar punts)
        all_orphan_uib = orphan_uib.copy()
        all_orphan_dad = orphan_dad.copy()

        dad_suggestions = compute_orphan_suggestions(
            result["samples"], orphan_dad, file_type="dad"
        )
        uib_suggestions = compute_orphan_suggestions(
            result["samples"], orphan_uib, file_type="uib"
        )

        # Aplicar suggeriments d'alta confiança (>= 85%)
        dad_applied, orphan_dad = apply_orphan_suggestions(
            result["samples"], dad_suggestions, orphan_dad, file_type="dad"
        )
        uib_applied, orphan_uib = apply_orphan_suggestions(
            result["samples"], uib_suggestions, orphan_uib, file_type="uib"
        )

        # Guardar orfes restants (sense suggeriments aplicats)
        result["orphan_files"] = {
            "uib": orphan_uib,
            "dad": orphan_dad,
        }

        # Guardar TOTS els orfes originals (per comptar punts)
        result["all_orphan_files"] = {
            "uib": all_orphan_uib,
            "dad": all_orphan_dad,
        }

        result["orphan_suggestions"] = {
            "dad": {f"{k[0]}_R{k[1]}": v for k, v in dad_suggestions.items()},
            "uib": {f"{k[0]}_R{k[1]}": v for k, v in uib_suggestions.items()},
        }

        if dad_applied:
            result["warnings"].append(
                f"Suggerits {dad_applied} fitxers DAD per revisar (matching automàtic)"
            )
        if uib_applied:
            result["warnings"].append(
                f"Suggerits {uib_applied} fitxers UIB per revisar (matching automàtic)"
            )

        if orphan_uib:
            names = [os.path.basename(f) for f in orphan_uib]
            result["warnings"].append(
                f"UIB orfes: {', '.join(names)} → Assignar a Importar"
            )
        if orphan_dad:
            names = [os.path.basename(f) for f in orphan_dad]
            result["warnings"].append(
                f"DAD orfes: {', '.join(names)} → Assignar a Importar"
            )

        report_progress(95, "Finalitzant...")

        # 7. Estadístiques
        samples_with_data = sum(
            1 for s in result["samples"].values()
            for r in s["replicas"].values()
            if r.get("has_data")
        )
        samples_without_data = sum(
            1 for s in result["samples"].values()
            for r in s["replicas"].values()
            if not r.get("has_data")
        )

        # Comptador de rèpliques reals (pot ser < total_injections si hi ha sobreescriptures)
        total_replicas_actual = sum(
            len(s["replicas"]) for s in result["samples"].values()
        )

        # Comptadors per tipus de senyal (rèpliques amb cada tipus)
        doc_direct_count = sum(
            1 for s in result["samples"].values()
            for r in s["replicas"].values()
            if r.get("direct") is not None
        )
        uib_count = sum(
            1 for s in result["samples"].values()
            for r in s["replicas"].values()
            if r.get("uib") is not None
        )
        dad_count = sum(
            1 for s in result["samples"].values()
            for r in s["replicas"].values()
            if r.get("dad") is not None
        )

        result["stats"] = {
            "master_line_count": result.get("master_line_count", len(injections)),  # Line# al MasterFile
            "total_injections": len(injections),  # Injeccions parsejades del MasterFile
            "total_replicas_imported": total_replicas_actual,  # Rèpliques reals (després de dedup)
            "total_samples": len(result["samples"]),
            "samples_with_data": samples_with_data,
            "samples_without_data": samples_without_data,
            "khp_count": len(result["khp_samples"]),
            "control_count": len(result["control_samples"]),
            "uib_files_used": len(used_files["uib"]),
            "dad_files_used": len(used_files["dad"]),
            "orphan_uib": len(orphan_uib),
            "orphan_dad": len(orphan_dad),
            "doc_direct_count": doc_direct_count,
            "uib_count": uib_count,
            "dad_count": dad_count,
        }

        # Warning clar si s'han perdut injeccions per sobreescriptura
        if total_replicas_actual < len(injections):
            lost = len(injections) - total_replicas_actual
            result["warnings"].insert(0,
                f"⚠️ IMPORTACIÓ INCOMPLETA: {total_replicas_actual}/{len(injections)} "
                f"injeccions importades. {lost} perdudes per Inj# duplicat al MasterFile. "
                f"Cal corregir Sample_Rep al MasterFile i reimportar."
            )

        # Validar que els KHP tenen dades DOC (necessari per calibrar)
        khp_without_doc = []
        for khp_name in result["khp_samples"]:
            sample = result["samples"].get(khp_name, {})
            replicas = sample.get("replicas", {})
            has_any_doc = False
            for rep_data in replicas.values():
                direct = rep_data.get("direct") if rep_data else None
                uib = rep_data.get("uib") if rep_data else None
                direct_t = direct.get("t") if direct else None
                uib_t = uib.get("t") if uib else None
                if (direct_t is not None and len(direct_t) > 0) or \
                   (uib_t is not None and len(uib_t) > 0):
                    has_any_doc = True
                    break
            if not has_any_doc:
                khp_without_doc.append(khp_name)

        if khp_without_doc:
            result["warnings"].append(
                f"⚠️ KHP sense dades DOC: {', '.join(khp_without_doc)} - no es podrà calibrar (falten fitxers TOC?)"
            )
            result["khp_without_doc"] = khp_without_doc

        # Detectar si les PRIMERES mostres no tenen dades (DOC sobreescrit)
        # Ordenar per line_num per veure l'ordre d'injecció
        injection_order = []
        for sample_name, sample in result["samples"].items():
            for rep_key, rep_data in sample.get("replicas", {}).items():
                inj = rep_data.get("injection", {})
                line_num = inj.get("line_num", 999)
                has_data = rep_data.get("has_data", False)
                injection_order.append((line_num, sample_name, rep_key, has_data))

        injection_order.sort(key=lambda x: x[0])

        # Comptar primeres injeccions sense dades consecutives
        first_without_data = []
        found_with_data = False
        for line_num, sample_name, rep_key, has_data in injection_order:
            if has_data:
                found_with_data = True
                break
            else:
                first_without_data.append(f"{sample_name}_R{rep_key}")

        # Si hi ha primeres sense dades PERÒ després n'hi ha amb dades = sobreescrit
        if first_without_data and found_with_data:
            n_lost = len(first_without_data)
            samples_lost = list(dict.fromkeys([x.rsplit('_R', 1)[0] for x in first_without_data]))
            result["warnings"].append(
                f"⚠️ DOC SOBREESCRIT: Les primeres {n_lost} injeccions ({', '.join(samples_lost[:5])}"
                f"{'...' if len(samples_lost) > 5 else ''}) no tenen dades. "
                f"Tip: Esborra registre TOC abans de llençar SEQ o revisa durada."
            )
            result["doc_overwritten"] = {
                "n_lost": n_lost,
                "samples_lost": samples_lost,
                "first_with_data": injection_order[n_lost][1] if n_lost < len(injection_order) else None
            }

        # Detectar i TALLAR mostres amb dades ANÒMALES
        # 1. Calcular durada mitjana de les injeccions normals
        is_bp = result.get("method", "").upper() == "BP"
        min_points = 50 if is_bp else 200
        min_duration = 5.0 if is_bp else 30.0
        max_expected = 15.0 if is_bp else 85.0  # Límit inicial

        all_durations = []
        for sample_name, sample in result["samples"].items():
            for rep_key, rep_data in sample.get("replicas", {}).items():
                direct = rep_data.get("direct", {})
                if not direct:
                    continue
                t_array = direct.get("t")
                if t_array is not None and len(t_array) > min_points:
                    t_max = float(t_array[-1])
                    # Només incloure duracions "normals" per calcular mitjana
                    if min_duration < t_max < max_expected * 2:
                        all_durations.append(t_max)

        # Calcular durada referència (mediana de les normals)
        if all_durations:
            all_durations.sort()
            median_duration = all_durations[len(all_durations) // 2]
            # Permetre 5% extra per variabilitat
            trim_threshold = median_duration * 1.05
        else:
            trim_threshold = max_expected

        truncated_samples = []
        postrun_samples = []
        trimmed_samples = []

        for sample_name, sample in result["samples"].items():
            for rep_key, rep_data in sample.get("replicas", {}).items():
                direct = rep_data.get("direct", {})
                if not direct:
                    continue

                t_array = direct.get("t")
                y_array = direct.get("y")
                y_net = direct.get("y_net")

                if t_array is None or len(t_array) == 0:
                    continue

                import numpy as np
                t_array = np.asarray(t_array)
                n_points = len(t_array)
                t_max = float(t_array[-1])
                sample_id = f"{sample_name}_R{rep_key}"

                # Detectar dades truncades (massa poques)
                if n_points > 0 and (n_points < min_points or t_max < min_duration):
                    truncated_samples.append(f"{sample_id} ({n_points}pts, {t_max:.1f}min)")

                # Detectar i TALLAR postrun
                if t_max > trim_threshold * 1.5:  # >50% més llarg que normal
                    original_duration = t_max
                    original_points = n_points

                    # Trobar índex on tallar
                    trim_idx = np.searchsorted(t_array, trim_threshold)
                    if trim_idx < len(t_array) - 10:  # Deixar mínim 10 punts marge
                        # Tallar arrays
                        direct["t"] = t_array[:trim_idx].tolist()
                        if y_array is not None:
                            direct["y"] = np.asarray(y_array)[:trim_idx].tolist()
                        if y_net is not None:
                            direct["y_net"] = np.asarray(y_net)[:trim_idx].tolist()

                        # Actualitzar metadata
                        direct["t_max_original"] = original_duration
                        direct["n_points_original"] = original_points
                        direct["trimmed_at"] = trim_threshold
                        direct["postrun_removed"] = True

                        trimmed_samples.append(
                            f"{sample_id} ({original_duration:.0f}→{trim_threshold:.0f}min)"
                        )
                    else:
                        # No es pot tallar prou, només warning
                        postrun_samples.append(
                            f"{sample_id} ({t_max:.0f}min = {t_max/60:.1f}h)"
                        )

        if truncated_samples:
            result["warnings"].append(
                f"⚠️ DADES TRUNCADES: {', '.join(truncated_samples[:3])}"
                f"{'...' if len(truncated_samples) > 3 else ''} - poques dades, run interromput?"
            )
            result["truncated_samples"] = truncated_samples

        if trimmed_samples:
            # No warning - el tall és automàtic i transparent
            # Només guardem metadata per traçabilitat
            result["trimmed_samples"] = trimmed_samples
            result["trim_threshold"] = trim_threshold

        if postrun_samples:
            result["warnings"].append(
                f"⚠️ POSTRUN NO TALLAT: {', '.join(postrun_samples[:3])}"
                f"{'...' if len(postrun_samples) > 3 else ''} - revisar manualment"
            )
            result["postrun_samples"] = postrun_samples

        result["success"] = True
        report_progress(100, "Importació completada")

    except Exception as e:
        import traceback
        result["errors"].append(str(e))
        result["errors"].append(traceback.format_exc())

    # Alliberar 2-TOC DataFrame (les dades DOC ja estan extretes a rep_data["direct"]["t"/"y"])
    if result.get("master_data") and result["master_data"].get("toc") is not None:
        result["master_data"]["toc"] = None

    # Generar avisos estructurats (nou sistema)
    result["warnings_structured"] = _generate_import_warnings(result)
    result["warning_level"] = get_max_warning_level(result["warnings_structured"])

    return result


# =============================================================================
# IMPORT PACK (SIBLINGS)
# =============================================================================

def import_sequence_pack(seq_paths, config=None, progress_callback=None):
    """
    Importa múltiples carpetes siblings com un pack unificat.

    Quan l'equip s'atura i es reinicia, es creen carpetes com:
    282_SEQ, 282B_SEQ, 282C_SEQ

    Aquesta funció les importa totes i fusiona els resultats.

    Args:
        seq_paths: Llista de paths (siblings) o path únic
        config: Configuració (opcional)
        progress_callback: Callback(pct, msg)

    Returns:
        dict unificat amb totes les mostres dels siblings
    """
    # Si és un sol path, convertir a llista
    if isinstance(seq_paths, str):
        seq_paths = [seq_paths]

    if not seq_paths:
        return {"success": False, "errors": ["Cap path proporcionat"]}

    # Si només hi ha un path, importar directament
    if len(seq_paths) == 1:
        return import_sequence(seq_paths[0], config, progress_callback)

    def report_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    # Ordenar paths (282_SEQ abans de 282B_SEQ)
    seq_paths = sorted(seq_paths)
    primary_path = seq_paths[0]  # El principal és el primer (sense lletra)

    report_progress(0, f"Importació pack: {len(seq_paths)} siblings")

    # Importar cada sibling i guardar el seu manifest individual
    imported_results = []
    for i, path in enumerate(seq_paths):
        pct_start = int(100 * i / len(seq_paths))
        pct_end = int(100 * (i + 1) / len(seq_paths))

        def sub_progress(pct, msg):
            real_pct = pct_start + int((pct_end - pct_start) * pct / 100)
            report_progress(real_pct, f"[{os.path.basename(path)}] {msg}")

        result = import_sequence(path, config, sub_progress)
        imported_results.append(result)

        # Guardar manifest individual de cada sibling (per poder re-carregar independentment)
        if result.get("success"):
            save_import_manifest(result)

    # Fusionar resultats
    report_progress(95, "Fusionant resultats...")
    merged = _merge_import_results(imported_results, primary_path)

    report_progress(100, f"Pack importat: {len(seq_paths)} siblings fusionats")

    return merged


def _merge_import_results(results, primary_path):
    """
    Fusiona múltiples resultats d'importació en un de sol.

    Args:
        results: Llista de dicts retornats per import_sequence()
        primary_path: Path principal (on es guarden els resultats)

    Returns:
        dict fusionat
    """
    if not results:
        return {"success": False, "errors": ["Cap resultat per fusionar"]}

    if len(results) == 1:
        return results[0]

    # Usar el primer com a base
    primary = results[0]

    merged = {
        "success": True,
        "seq_path": primary_path,
        "seq_name": os.path.basename(primary_path),
        "method": primary.get("method", "COLUMN"),
        "data_mode": primary.get("data_mode", "DUAL"),
        "date": primary.get("date", ""),
        "is_pack": True,
        "pack_sources": [r.get("seq_path") for r in results],
        "pack_count": len(results),
        "master_data": primary.get("master_data"),
        "master_file": primary.get("master_file"),
        "master_format": primary.get("master_format"),
        "injections": [],
        "samples": {},
        "khp_samples": [],
        "control_samples": [],
        "valid_samples": [],
        "orphan_files": {"uib": [], "dad": []},
        "errors": [],
        "warnings": [],
    }

    # Afegir info de pack als warnings
    sibling_names = [os.path.basename(r.get("seq_path", "")) for r in results]
    merged["warnings"].append(
        f"📦 PACK FUSIONAT: {', '.join(sibling_names)}"
    )

    # Fusionar dades de cada resultat
    seen_samples = {}  # Per detectar duplicats

    for idx, result in enumerate(results):
        source_name = os.path.basename(result.get("seq_path", f"sibling_{idx}"))

        # Fusionar errors/warnings
        for err in result.get("errors", []):
            merged["errors"].append(f"[{source_name}] {err}")
        for warn in result.get("warnings", []):
            # No duplicar el warning de pack
            if "PACK FUSIONAT" not in warn:
                merged["warnings"].append(f"[{source_name}] {warn}")

        # Fusionar injeccions
        for inj in result.get("injections", []):
            inj_copy = inj.copy()
            inj_copy["source_seq"] = source_name
            merged["injections"].append(inj_copy)

        # Fusionar samples
        for sample_name, sample_data in result.get("samples", {}).items():
            if sample_name not in merged["samples"]:
                merged["samples"][sample_name] = {
                    "type": sample_data.get("type", "SAMPLE"),
                    "original_name": sample_data.get("original_name", sample_name),
                    "replicas": {},
                    "sources": []
                }

            # Afegir font
            merged["samples"][sample_name]["sources"].append(source_name)

            # Fusionar rèpliques (evitar sobreescriure)
            for rep_key, rep_data in sample_data.get("replicas", {}).items():
                # En un pack, cada sibling és una mesura independent (reinici equip).
                # NO deduplicar entre siblings — sempre afegir amb clau única.
                unique_key = rep_key
                if unique_key in merged["samples"][sample_name]["replicas"]:
                    unique_key = f"{rep_key}_{source_name}"

                rep_copy = rep_data.copy() if rep_data else {}
                rep_copy["source_seq"] = source_name
                merged["samples"][sample_name]["replicas"][unique_key] = rep_copy

        # Fusionar llistes KHP/Control
        for khp in result.get("khp_samples", []):
            if khp not in merged["khp_samples"]:
                merged["khp_samples"].append(khp)

        for ctrl in result.get("control_samples", []):
            if ctrl not in merged["control_samples"]:
                merged["control_samples"].append(ctrl)

        # Fusionar valid_samples
        for vs in result.get("valid_samples", []):
            _add_unique(merged["valid_samples"], vs)

        # Fusionar orphan_files
        merged["orphan_files"]["uib"].extend(result.get("orphan_files", {}).get("uib", []))
        merged["orphan_files"]["dad"].extend(result.get("orphan_files", {}).get("dad", []))

    # Renumerar rèpliques seqüencialment per sample (evitar claus com "2_297B_SEQ")
    # Ordre: primer les del sibling primari, després secundaris, dins cada un per rep_key original
    source_order = {name: i for i, name in enumerate(sibling_names)}
    for sample_name, sample_data in merged["samples"].items():
        old_replicas = sample_data.get("replicas", {})
        if len(old_replicas) <= 2:
            # 2 o menys rèpliques: no renumerar (comportament idèntic a SEQ normal)
            continue
        sorted_items = sorted(
            old_replicas.items(),
            key=lambda kv: (source_order.get(kv[1].get("source_seq", ""), 99), kv[0])
        )
        new_replicas = {}
        for new_idx, (old_key, rep_data) in enumerate(sorted_items, start=1):
            rep_data["original_rep_num"] = old_key
            new_replicas[str(new_idx)] = rep_data
        sample_data["replicas"] = new_replicas

    # Estadístiques fusionades
    merged["stats"] = {
        "total_injections": len(merged["injections"]),
        "total_samples": len(merged["samples"]),
        "khp_count": len(merged["khp_samples"]),
        "control_count": len(merged["control_samples"]),
        "pack_sources": len(results),
    }

    # Check success
    merged["success"] = not any("error" in str(e).lower() for e in merged["errors"])

    return merged


def _replicas_identical(rep1, rep2):
    """Comprova si dues rèpliques tenen les mateixes dades (evitar duplicats)."""
    if not rep1 or not rep2:
        return False

    # Comparar per injection_info
    inj1 = rep1.get("injection_info", {})
    inj2 = rep2.get("injection_info", {})

    if inj1.get("line_num") == inj2.get("line_num") and \
       inj1.get("inj_num") == inj2.get("inj_num"):
        return True

    # Comparar per dades DOC
    direct1 = rep1.get("direct", {})
    direct2 = rep2.get("direct", {})

    t1 = direct1.get("t")
    t2 = direct2.get("t")

    if t1 is not None and t2 is not None:
        if len(t1) == len(t2) and len(t1) > 0:
            # Comparar primer i últim punt
            try:
                if abs(t1[0] - t2[0]) < 0.001 and abs(t1[-1] - t2[-1]) < 0.001:
                    return True
            except (IndexError, TypeError):
                pass

    return False


def _resolve_sibling_path(source_seq, primary_seq_path, pack_sources, config=None):
    """Retorna el path de la carpeta SEQ d'on ve la rèplica (sibling o principal).

    Args:
        source_seq: Nom del sibling (ex: "297B_SEQ") o "" si principal
        primary_seq_path: Path de la SEQ principal
        pack_sources: Llista de paths de tots els siblings
        config: Configuració (per data_folders)

    Returns:
        Path a la carpeta SEQ corresponent (o primary_seq_path si no es troba)
    """
    if not source_seq or source_seq == os.path.basename(primary_seq_path):
        return primary_seq_path

    # Buscar als pack_sources
    for src_path in (pack_sources or []):
        if os.path.basename(src_path) == source_seq:
            if os.path.isdir(src_path):
                return src_path

    # Buscar a les data_folders
    try:
        from hpsec_config import get_data_folders
        for df_path in get_data_folders(config):
            candidate = os.path.join(df_path, source_seq)
            if os.path.isdir(candidate):
                return candidate
    except Exception:
        pass

    return primary_seq_path


# =============================================================================
# GENERACIÓ DE MANIFEST JSON
# =============================================================================

def generate_import_manifest(imported_data, include_injection_details=True):
    """
    Genera un manifest JSON amb metadata de la importació.

    Útil per:
    - Traçabilitat: saber d'on venen les dades
    - Verificació: comprovar què s'ha detectat
    - Debugging: diagnosticar problemes d'importació

    Args:
        imported_data: Dict retornat per import_sequence()
        include_injection_details: Si True, inclou detall per cada injecció

    Returns:
        dict JSON-serializable amb tota la metadata
    """
    from hpsec_version import SUITE_VERSION
    manifest = {
        "suite_version": SUITE_VERSION,
        "generated_at": datetime.now().isoformat(),
        "import_module": __version__,

        # Info seqüència (path relatiu: només nom SEQ, es reconstrueix amb data_folder del config)
        "sequence": {
            "name": imported_data.get("seq_name", ""),
            "path": os.path.basename(imported_data.get("seq_path", "")),
            "date": str(imported_data.get("date", "")),
            "method": imported_data.get("method", ""),  # COLUMN o BP
            "data_mode": imported_data.get("data_mode", ""),  # DUAL, DIRECT, UIB
            "uib_sensitivity": imported_data.get("uib_sensitivity"),  # Sensibilitat UIB (700, 1000, etc.)
        },

        # MasterFile (path relatiu dins SEQ)
        "master_file": {
            "path": os.path.basename(imported_data.get("master_file", "")),
            "format": imported_data.get("master_format", ""),  # NEW o OLD
            "filename": os.path.basename(imported_data.get("master_file", "")) if imported_data.get("master_file") else "",
        },

        # Pack/siblings info (si aplica)
        "is_pack": imported_data.get("is_pack", False),
        "pack_sources": imported_data.get("pack_sources", []),

        # Estadístiques globals
        "stats": imported_data.get("stats", {}),

        # Timeouts globals del flux TOC (detectats a nivell de seqüència)
        "seq_timeouts": imported_data.get("seq_timeouts", []),

        # Avisos i errors
        "warnings": imported_data.get("warnings", []),
        "warnings_structured": imported_data.get("warnings_structured", []),
        "errors": imported_data.get("errors", []),

        # Fitxers orfes
        "orphan_files": {
            "uib": [os.path.basename(f) for f in imported_data.get("orphan_files", {}).get("uib", [])],
            "dad": [os.path.basename(f) for f in imported_data.get("orphan_files", {}).get("dad", [])],
        },

        # Suggeriments de matching (orfes → mostres)
        "orphan_suggestions": imported_data.get("orphan_suggestions", {"dad": {}, "uib": {}}),

        # Estat de revisió d'avisos
        "orphan_warning_dismissed": imported_data.get("orphan_warning_dismissed", False),
        "warnings_confirmed": imported_data.get("warnings_confirmed", False),
    }

    # Detall per mostra
    samples_detail = []
    for sample_name, sample_info in imported_data.get("samples", {}).items():
        sample_entry = {
            "name": sample_name,
            "original_name": sample_info.get("original_name", sample_name),  # Nom original del MasterFile
            "type": sample_info.get("type", "SAMPLE"),
            "replicas": [],
        }

        for rep_num, rep_data in sample_info.get("replicas", {}).items():
            replica_entry = {
                "replica": rep_num,
                "has_data": rep_data.get("has_data", False),
            }

            # DOC Direct
            direct = rep_data.get("direct") or {}
            if direct.get("t") is not None:
                t_arr = direct["t"]
                _ti = direct.get("timeout_info", {})
                replica_entry["direct"] = {
                    "source": "MasterFile:2-TOC",
                    "row_start": direct.get("row_start"),
                    "row_end": direct.get("row_end"),
                    "n_points": len(t_arr),
                    "t_min": float(min(t_arr)),
                    "t_max": float(max(t_arr)),
                    "baseline": direct.get("baseline"),
                    "has_timeout": _ti.get("n_timeouts", 0) > 0,
                    "timeout_severity": _ti.get("severity", "OK"),
                    "timeout_at_boundary": direct.get("timeout_at_boundary", False),
                }
                if direct.get("boundary_timeout"):
                    replica_entry["direct"]["boundary_timeout"] = direct["boundary_timeout"]
                if _ti.get("n_timeouts", 0) > 0:
                    replica_entry["direct"]["timeout_ranges"] = _ti.get("timeouts", [])
                if _ti.get("toc_minute_precision"):
                    replica_entry["direct"]["toc_minute_precision"] = True
            elif direct.get("row_start") is not None or direct.get("n_points"):
                # Preservar metadades encara que no hi hagi dades reals
                replica_entry["direct"] = {
                    "source": "MasterFile:2-TOC",
                    "row_start": direct.get("row_start"),
                    "row_end": direct.get("row_end"),
                    "n_points": direct.get("n_points", 0),
                    "baseline": direct.get("baseline"),
                    "has_timeout": False,
                }

            # DOC UIB
            uib = rep_data.get("uib") or {}
            if uib.get("t") is not None:
                t_arr = uib["t"]
                _uib_ti = uib.get("timeout_info", {})
                replica_entry["uib"] = {
                    "source": "CSV",
                    "file": os.path.basename(uib.get("path", "") or uib.get("file", "")),
                    "n_points": len(t_arr),
                    "t_min": float(min(t_arr)),
                    "t_max": float(max(t_arr)),
                    "baseline": uib.get("baseline"),
                    "has_timeout": _uib_ti.get("n_timeouts", 0) > 0,
                    "timeout_severity": _uib_ti.get("severity", "OK"),
                    "timeout_at_boundary": uib.get("timeout_at_boundary", False),
                }
                if _uib_ti.get("n_timeouts", 0) > 0:
                    replica_entry["uib"]["timeout_ranges"] = _uib_ti.get("timeouts", [])
                # Afegir info d'assignació manual si existeix
                if uib.get("manual_assignment"):
                    replica_entry["uib"]["manual_assignment"] = True
                    replica_entry["uib"]["manual_file"] = uib.get("manual_file", "")
            elif uib.get("file") or uib.get("n_points") or uib.get("manual_file"):
                # Preservar metadades encara que no hi hagi dades reals
                replica_entry["uib"] = {
                    "source": "CSV",
                    "file": uib.get("file", ""),
                    "n_points": uib.get("n_points", 0),
                    "baseline": uib.get("baseline"),
                }
                if uib.get("manual_assignment"):
                    replica_entry["uib"]["manual_assignment"] = True
                    replica_entry["uib"]["manual_file"] = uib.get("manual_file", "")

            # DAD
            dad = rep_data.get("dad") or {}
            df = dad.get("df")
            if df is not None:
                t_col = df.columns[0]
                replica_entry["dad"] = {
                    "source": rep_data.get("dad_source", "unknown"),
                    "file": os.path.basename(dad.get("path", "") or dad.get("file", "")),
                    "n_points": len(df),
                    "n_wavelengths": len(df.columns) - 1,  # -1 per columna temps
                    "t_min": float(pd.to_numeric(df[t_col], errors='coerce').min()),
                    "t_max": float(pd.to_numeric(df[t_col], errors='coerce').max()),
                    "wavelengths_range": f"{df.columns[1]}-{df.columns[-1]}" if len(df.columns) > 1 else "none",
                }
                # Afegir info d'assignació manual si existeix
                if dad.get("manual_assignment"):
                    replica_entry["dad"]["manual_assignment"] = True
                    replica_entry["dad"]["manual_file"] = dad.get("manual_file") or dad.get("file", "")
            elif dad.get("file") or dad.get("manual_file") or dad.get("n_points"):
                # Preservar metadades encara que no hi hagi DataFrame
                replica_entry["dad"] = {
                    "source": rep_data.get("dad_source", "unknown"),
                    "file": dad.get("file", "") or dad.get("path", ""),
                    "n_points": dad.get("n_points", 0),
                }
                if dad.get("manual_assignment"):
                    replica_entry["dad"]["manual_assignment"] = True
                    replica_entry["dad"]["manual_file"] = dad.get("manual_file") or dad.get("file", "")

            # Suggeriments de matching (si n'hi ha)
            dad_suggestion = rep_data.get("dad_suggestion")
            if dad_suggestion:
                replica_entry["dad_suggestion"] = {
                    "file": dad_suggestion.get("filename", ""),
                    "confidence": dad_suggestion.get("confidence", 0),
                    "match_type": dad_suggestion.get("match_type", ""),
                    "status": dad_suggestion.get("status", "SUGGESTED"),
                }

            uib_suggestion = rep_data.get("uib_suggestion")
            if uib_suggestion:
                replica_entry["uib_suggestion"] = {
                    "file": uib_suggestion.get("filename", ""),
                    "confidence": uib_suggestion.get("confidence", 0),
                    "match_type": uib_suggestion.get("match_type", ""),
                    "status": uib_suggestion.get("status", "SUGGESTED"),
                }

            # Info injecció original (si disponible)
            inj_info = rep_data.get("injection_info")
            if inj_info and include_injection_details:
                replica_entry["injection"] = {
                    "line_num": inj_info.get("line_num"),
                    "inj_num": inj_info.get("inj_num"),
                    "inj_volume": inj_info.get("inj_volume"),
                    "location": inj_info.get("inj_location"),
                    "acq_date": inj_info.get("inj_date", ""),
                    "method": inj_info.get("inj_method"),
                }

            # Traçabilitat sibling (packs)
            if rep_data.get("source_seq"):
                replica_entry["source_seq"] = rep_data["source_seq"]
            if rep_data.get("original_rep_num"):
                replica_entry["original_rep_num"] = rep_data["original_rep_num"]

            sample_entry["replicas"].append(replica_entry)

        samples_detail.append(sample_entry)

    manifest["samples"] = samples_detail

    # Resum per tipus
    manifest["summary"] = {
        "total_samples": len([s for s in samples_detail if s["type"] == "SAMPLE"]),
        "total_khp": len([s for s in samples_detail if s["type"] == "KHP"]),
        "total_pr": len([s for s in samples_detail if s["type"].startswith("PR")]),
        "total_blank": len([s for s in samples_detail if s["type"] == "BLANK"]),
        "total_control": len([s for s in samples_detail if s["type"] == "CONTROL"]),
        "total_replicas": sum(len(s["replicas"]) for s in samples_detail),
        "replicas_with_direct": sum(
            1 for s in samples_detail
            for r in s["replicas"]
            if "direct" in r
        ),
        "replicas_with_uib": sum(
            1 for s in samples_detail
            for r in s["replicas"]
            if "uib" in r
        ),
        "replicas_with_dad": sum(
            1 for s in samples_detail
            for r in s["replicas"]
            if "dad" in r
        ),
    }

    return manifest


def save_import_manifest(imported_data, output_path=None):
    """
    Genera i guarda el manifest JSON a un fitxer.

    Args:
        imported_data: Dict retornat per import_sequence()
        output_path: Ruta de sortida (default: SEQ_PATH/CHECK/data/import_manifest.json)

    Returns:
        Path del fitxer generat
    """
    manifest = generate_import_manifest(imported_data)

    if output_path is None:
        seq_path = imported_data.get("seq_path", ".")
        data_folder = get_data_folder(seq_path, create=True)
        output_path = os.path.join(data_folder, "import_manifest.json")

    # Assegurar que la carpeta existeix
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    return output_path


def load_manifest(seq_path):
    """
    Carrega el manifest JSON si existeix.

    Busca primer a la nova ubicació (CHECK/data/) i després a l'antiga (arrel SEQ).

    Args:
        seq_path: Ruta a la carpeta SEQ

    Returns:
        dict amb manifest o None si no existeix
    """
    # Nova ubicació: CHECK/data/import_manifest.json
    data_folder = get_data_folder(seq_path, create=False)
    new_path = os.path.join(data_folder, "import_manifest.json")

    # Antiga ubicació: SEQ/import_manifest.json (compatibilitat)
    old_path = os.path.join(seq_path, "import_manifest.json")

    # Prioritzar nova ubicació
    for manifest_path in [new_path, old_path]:
        if os.path.exists(manifest_path):
            try:
                with open(manifest_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                continue
    return None


# =============================================================================
# HELPERS COMPARTITS: càrrega DOC Direct / UIB
# =============================================================================

def _load_doc_direct(toc_df, row_start, row_end, mode, config, seq_timeouts=None):
    """Carrega DOC Direct des del MasterFile amb baseline i y_net.

    Timeout info derivada de seq_timeouts (detecció global, font única).

    Args:
        toc_df: DataFrame 2-TOC
        row_start, row_end: rang de files (TOC_Row, Excel 1-indexed)
        mode: "COLUMN" o "BP"
        config: dict configuració
        seq_timeouts: llista de timeouts de seqüència (detect_sequence_timeouts)

    Returns:
        dict amb {t, y, y_net, baseline, timeout_info, ...} o None si falla
    """
    if toc_df is None or row_start is None or row_end is None:
        return None
    try:
        max_dur = config.get("max_duration_min", 80.0)
        df_doc = extract_doc_from_masterfile(
            toc_df, row_start, row_end,
            detect_timeouts=False,
            max_duration_min=max_dur
        )
        if df_doc is None or df_doc.empty:
            return None
        t_direct = df_doc["time (min)"].values
        y_direct = df_doc["DOC"].values
        baseline = get_baseline_value(t_direct, y_direct, mode=mode)
        y_net = np.array(y_direct) - baseline

        # Timeout info: font única = detecció global → mapejat per injecció
        is_bp = mode.upper() == "BP"
        timeout_info = map_timeouts_to_injection(
            seq_timeouts or [], row_start, row_end,
            t_rel=t_direct, is_bp=is_bp
        )

        return {
            "path": "MasterFile:2-TOC",
            "t": t_direct,
            "y": y_direct,
            "row_start": row_start,
            "row_end": row_end,
            "n_points": len(t_direct),
            "timeout_info": timeout_info,
            "y_net": y_net,
            "baseline": baseline,
            "timeout_at_boundary": timeout_info["timeout_at_boundary"],
            "boundary_timeout": timeout_info["boundary_timeout"],
        }
    except Exception as e:
        logger.warning("_load_doc_direct: %s", e)
        return None


def _load_uib_csv(uib_file, seq_path, mode):
    """Carrega UIB CSV amb downsample i baseline.

    Args:
        uib_file: nom del fitxer UIB
        seq_path: path de la seqüència
        mode: "COLUMN" o "BP"

    Returns:
        dict amb {t, y, y_net, baseline, df, ...} o None si falla
    """
    if not uib_file:
        return None

    # Buscar fitxer a CSV/ o arrel
    uib_path = os.path.join(seq_path, "CSV", uib_file)
    if not os.path.exists(uib_path):
        for subdir in ["", "csv"]:
            test_path = os.path.join(seq_path, subdir, uib_file) if subdir else os.path.join(seq_path, uib_file)
            if os.path.exists(test_path):
                uib_path = test_path
                break

    if not os.path.exists(uib_path):
        return None

    try:
        df_uib, status = llegir_doc_uib(uib_path)
        if df_uib.empty or "OK" not in status:
            return None

        t_uib = df_uib["time (min)"].values
        y_uib = df_uib["DOC"].values

        # Downsample a cadència DOC Direct
        if len(t_uib) > 10:
            t_uib, y_uib = downsample_to_cadence(t_uib, y_uib)

        baseline = get_baseline_value(t_uib, y_uib, mode=mode)
        y_net = np.array(y_uib) - baseline

        return {
            "path": uib_path,
            "df": df_uib,
            "t": t_uib,
            "y": y_uib,
            "file": uib_file,
            "n_points": len(t_uib),
            "y_net": y_net,
            "baseline": baseline,
        }
    except Exception as e:
        logger.warning("_load_uib_csv: %s — %s", uib_file, e)
        return None


def import_from_manifest(seq_path, manifest=None, config=None, progress_callback=None,
                         load_data=True):
    """
    Reimporta dades guiat pel manifest (molt més ràpid).

    En lloc de detectar/matching, llegeix directament d'on indica el manifest:
    - DOC Direct: files row_start-row_end del MasterFile
    - UIB: fitxer CSV indicat
    - DAD: segons source (export3d o dad1a)

    Args:
        seq_path: Ruta a la carpeta SEQ
        manifest: Manifest carregat (si None, el carrega)
        config: Configuració
        progress_callback: Funció callback(pct, msg)
        load_data: Si True (defecte), llegeix MasterFile i fitxers CSV per obtenir
                   cromatogrames. Si False, només carrega metadades del manifest
                   (molt més ràpid, per mostrar la taula d'importació sense les
                   dades crues). Les dades es poden completar després amb
                   ensure_data_loaded().

    Returns:
        dict equivalent a import_sequence()
    """
    config = config or get_config()

    def report_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    # Carregar manifest si no proporcionat
    if manifest is None:
        manifest = load_manifest(seq_path)
        if manifest is None:
            # Fallback a import normal
            report_progress(0, "Manifest no trobat, important normalment...")
            return import_sequence(seq_path, config, progress_callback)

    report_progress(5, "Llegint manifest...")

    # Verificar si el manifest és antic (no té informació DOC Direct)
    manifest_samples = manifest.get("samples", [])
    has_direct_info = any(
        rep.get("direct") is not None
        for sample in manifest_samples
        for rep in sample.get("replicas", [])
    )

    if not has_direct_info and manifest_samples:
        # Manifest antic sense informació DOC Direct - reimportar completament
        report_progress(0, "Manifest antic detectat, reimportant...")
        logger.info("Manifest antic sense DOC Direct, fent reimportació completa")
        return import_sequence(seq_path, config, progress_callback)

    # Inicialitzar resultat
    seq_info = manifest.get("sequence", {})
    mf_info = manifest.get("master_file", {})

    result = {
        "success": False,
        "seq_path": seq_path,
        "seq_name": seq_info.get("name", os.path.basename(seq_path)),
        "method": seq_info.get("method", "COLUMN"),
        "data_mode": seq_info.get("data_mode", "DUAL"),
        "uib_sensitivity": _safe_float(seq_info.get("uib_sensitivity")),  # Restaurar sensibilitat UIB
        "date": seq_info.get("date", ""),
        "master_file": mf_info.get("path", ""),
        "master_format": mf_info.get("format", "NEW"),
        "master_data": None,
        "injections": [],
        "samples": {},
        "khp_samples": [],
        "control_samples": [],
        "valid_samples": [],
        "orphan_files": {"uib": [], "dad": []},
        "errors": [],
        "warnings": [],
        "from_manifest": True,
        "manifest_date": manifest.get("generated_at", ""),
        "orphan_warning_dismissed": manifest.get("orphan_warning_dismissed", False),
        "warnings_confirmed": manifest.get("warnings_confirmed", False),
    }

    # Verificar que MasterFile existeix (pot ser relatiu al manifest)
    master_path = mf_info.get("path", "")
    if master_path and not os.path.isabs(master_path):
        # Path relatiu → reconstruir absolut dins seq_path
        master_path = os.path.join(seq_path, master_path)
    if not master_path or not os.path.exists(master_path):
        # Intentar trobar-lo
        master_path_new, _ = trobar_excel_mestre(seq_path)
        if master_path_new:
            master_path = master_path_new
            result["master_file"] = master_path
        elif load_data:
            result["errors"].append("MasterFile no trobat")
            return result

    # Llegir MasterFile per obtenir dades DOC Direct + DAD KHP
    # Quan load_data=False, saltem la lectura (molt més ràpid per mostrar la taula)
    toc_df = None
    master_khp_data = None

    if load_data:
        report_progress(10, "Llegint MasterFile...")
        toc_calc_df = None
        try:
            with pd.ExcelFile(master_path, engine="openpyxl") as xl:
                if "2-TOC" in xl.sheet_names:
                    toc_df = pd.read_excel(xl, sheet_name="2-TOC", header=6, engine="openpyxl")
                # FIX F2.2: Llegir 3-DAD_KHP per KHP samples (abans no es llegia des de manifest)
                if "3-DAD_KHP" in xl.sheet_names:
                    master_khp_data = pd.read_excel(xl, sheet_name="3-DAD_KHP", engine="openpyxl")
                    logger.debug("3-DAD_KHP sheet loaded (%d rows)", len(master_khp_data))
                # Comprovar si 4-TOC_CALC existeix i té dades
                if "4-TOC_CALC" in xl.sheet_names:
                    toc_calc_df = pd.read_excel(xl, sheet_name="4-TOC_CALC", engine="openpyxl")
        except PermissionError:
            result["errors"].append(
                f"No es pot llegir el MasterFile: el fitxer està obert a Excel o sense permisos. "
                f"Tancar '{os.path.basename(master_path)}' i tornar a importar."
            )
            return result
        except Exception as e:
            result["errors"].append(f"Error llegint MasterFile: {e}")
            return result

        # Detectar 4-TOC_CALC amb timestamps arrodonits
        _needs_regen = _toc_calc_has_minute_precision(toc_calc_df)

        # Si 4-TOC_CALC és buit o degradat, calcular i guardar al MasterFile
        if (_needs_regen or toc_calc_df is None or toc_calc_df.empty) and toc_df is not None:
            master_data = llegir_masterfile_nou(master_path)
            computed = compute_toc_calc(master_data, toc_df)
            if computed is not None and not computed.empty:
                saved = _save_toc_calc_to_masterfile(master_path, computed)
                if saved:
                    logger.info("4-TOC_CALC generat i guardat al MasterFile des de manifest import")
                else:
                    result["warnings"].append(
                        "⚠️ 4-TOC_CALC calculat però no s'ha pogut escriure al MasterFile."
                    )
            else:
                result["warnings"].append(
                    "⚠️ No s'ha pogut calcular 4-TOC_CALC. DOC Direct no disponible."
                )
        elif (toc_calc_df is None or toc_calc_df.empty) and toc_df is None:
            result["warnings"].append(
                "⚠️ 4-TOC_CALC i 2-TOC no disponibles. DOC Direct no disponible."
            )
    else:
        report_progress(10, "Carregant manifest...")

    report_progress(20, "Processant mostres del manifest...")

    # Pack/siblings: carregar MasterFiles addicionals
    sibling_toc_dfs = {}  # {source_seq_name: toc_df}
    if manifest.get("is_pack") and manifest.get("pack_sources") and load_data:
        from hpsec_config import get_data_folders
        data_folders = get_data_folders(config)
        for sibling_path in manifest["pack_sources"]:
            sibling_name = os.path.basename(sibling_path)
            if sibling_name == os.path.basename(seq_path):
                continue  # Principal ja carregat
            # Trobar la carpeta del sibling
            for df_path in data_folders:
                candidate = os.path.join(df_path, sibling_name)
                if os.path.isdir(candidate):
                    sib_mf, _ = trobar_excel_mestre(candidate)
                    if sib_mf:
                        try:
                            with pd.ExcelFile(sib_mf, engine="openpyxl") as xl:
                                if "2-TOC" in xl.sheet_names:
                                    sib_toc = pd.read_excel(xl, sheet_name="2-TOC", header=6, engine="openpyxl")
                                    sibling_toc_dfs[sibling_name] = sib_toc
                                    logger.info("Pack: carregat 2-TOC de %s (%d files)", sibling_name, len(sib_toc))
                        except Exception as e:
                            logger.warning("Pack: error llegint MasterFile de %s: %s", sibling_name, e)
                    break

    # Detecció global de timeouts al flux TOC complet
    seq_timeouts = detect_sequence_timeouts(toc_df) if toc_df is not None else []

    # Processar cada mostra del manifest
    manifest_samples = manifest.get("samples", [])
    total = len(manifest_samples)

    for i, sample_info in enumerate(manifest_samples):
        pct = 20 + int((i / total) * 70) if total > 0 else 90
        report_progress(pct, f"Llegint {sample_info.get('name', '?')}...")

        sample_name = sample_info.get("name", "")
        sample_type = sample_info.get("type", "SAMPLE")

        if sample_name not in result["samples"]:
            result["samples"][sample_name] = {
                "type": sample_type,
                "original_name": sample_info.get("original_name", sample_name),  # Nom original del MasterFile
                "replicas": {},
            }

        # Classificar
        if sample_type == "KHP" and sample_name not in result["khp_samples"]:
            result["khp_samples"].append(sample_name)
        elif sample_type == "CONTROL" and sample_name not in result["control_samples"]:
            result["control_samples"].append(sample_name)

        _add_unique(result["valid_samples"], sample_name)

        # Processar cada rèplica
        for rep_info in sample_info.get("replicas", []):
            rep_num = str(rep_info.get("replica", "1"))

            rep_data = {
                "direct": None,
                "uib": None,
                "dad": None,
                "dad_source": None,
                "has_data": False,
                "injection_info": rep_info.get("injection"),
            }

            # Traçabilitat sibling
            if rep_info.get("source_seq"):
                rep_data["source_seq"] = rep_info["source_seq"]
            if rep_info.get("original_rep_num"):
                rep_data["original_rep_num"] = rep_info["original_rep_num"]

            # === DOC Direct ===
            direct_info = rep_info.get("direct")
            if direct_info:
                row_start = direct_info.get("row_start")
                row_end = direct_info.get("row_end")

                # Preservar metadades del manifest encara que no es puguin llegir les dades
                rep_data["direct"] = {
                    "path": f"MasterFile:2-TOC",
                    "df": None,
                    "t": None,
                    "y": None,
                    "row_start": row_start,
                    "row_end": row_end,
                    "n_points": direct_info.get("n_points", 0),
                    "timeout_info": {},
                    "y_net": None,
                    "baseline": None,
                }

                # Diagnòstic DOC Direct (només quan load_data=True, sinó és esperat que no hi hagi dades)
                if load_data:
                    if toc_df is None:
                        result["warnings"].append(f"⚠️ {sample_name} R{rep_num}: No s'ha pogut llegir 2-TOC del MasterFile")
                    elif row_start is None or row_end is None:
                        result["warnings"].append(f"⚠️ {sample_name} R{rep_num}: Fila DOC no definida al manifest")

                # Seleccionar toc_df correcte (sibling o principal)
                _rep_source = rep_info.get("source_seq", "")
                _rep_toc = sibling_toc_dfs.get(_rep_source, toc_df) if _rep_source else toc_df

                # Intentar llegir les dades reals si tenim MasterFile
                mode = "BP" if result["method"] == "BP" else "COLUMN"
                doc_result = _load_doc_direct(_rep_toc, row_start, row_end, mode, config,
                                              seq_timeouts=seq_timeouts)
                if doc_result:
                    rep_data["direct"] = doc_result
                    rep_data["has_data"] = True

            # === DOC UIB ===
            uib_info = rep_info.get("uib")
            if uib_info:
                # Prioritzar manual_file si existeix (assignació manual de l'usuari)
                uib_file = uib_info.get("manual_file") or uib_info.get("file", "")

                # Preservar metadades del manifest encara que no es puguin llegir les dades
                rep_data["uib"] = {
                    "path": uib_file,
                    "df": None,
                    "t": None,
                    "y": None,
                    "file": uib_file,
                    "n_points": uib_info.get("n_points", 0),
                    "y_net": None,
                    "baseline": None,
                    "manual_assignment": uib_info.get("manual_assignment", False),
                }

                if uib_file and load_data:
                    mode = "BP" if result["method"] == "BP" else "COLUMN"
                    # Sibling: buscar CSV a la carpeta del sibling
                    _uib_search_path = _resolve_sibling_path(
                        _rep_source, seq_path, manifest.get("pack_sources", []), config)
                    uib_result = _load_uib_csv(uib_file, _uib_search_path, mode)
                    if uib_result:
                        uib_result["manual_assignment"] = uib_info.get("manual_assignment", False)
                        # Propagar timeout_info de Direct a UIB (mateix detector TOC)
                        direct_ti = (rep_data.get("direct") or {}).get("timeout_info", {})
                        uib_result["timeout_info"] = direct_ti
                        uib_result["timeout_at_boundary"] = direct_ti.get("timeout_at_boundary", False)
                        rep_data["uib"] = uib_result
                        rep_data["has_data"] = True

            # === DAD ===
            # Sibling: buscar fitxers a la carpeta del sibling
            _dad_search_path = _resolve_sibling_path(
                _rep_source, seq_path, manifest.get("pack_sources", []), config)
            dad_info = rep_info.get("dad")
            dad_source = None
            if dad_info:
                dad_source = dad_info.get("source", "export3d")
                rep_data["dad_source"] = dad_source

                # Quan load_data=False, preservar metadades sense carregar fitxers
                if not load_data:
                    # Guardar info DAD del manifest per la taula
                    dad_file = dad_info.get("manual_file") or dad_info.get("file", "")
                    dad_n_points = dad_info.get("n_points", 0)
                    if dad_file or dad_n_points:
                        rep_data["dad"] = {
                            "df": None,
                            "path": dad_file,
                            "file": dad_file,
                            "n_points": dad_n_points,
                        }
                        rep_data["dad_source"] = dad_source
                        rep_data["has_data"] = True

                # FIX F2.2: Si DAD ve de 3-DAD_KHP (source="masterfile"), llegir del MasterFile
                elif dad_source == "masterfile" and master_khp_data is not None:
                    original_name = sample_info.get("original_name", sample_name)
                    inj_num = rep_info.get("injection", {}).get("inj_num", rep_num) if rep_info.get("injection") else rep_num
                    dad_result, dad_src = _match_khp_dad_from_masterfile(
                        sample_name, original_name, inj_num, master_khp_data
                    )
                    if dad_result is not None:
                        rep_data["dad"] = dad_result
                        rep_data["dad_source"] = dad_src
                        rep_data["has_data"] = True
                        logger.debug("3-DAD_KHP manifest: loaded %d rows for %s", len(dad_result["df"]), sample_name)

                # Prioritzar manual_file si existeix (assignació manual de l'usuari)
                # (només quan load_data=True — quan False, ja s'han guardat metadades a dalt)
                manual_dad_file = dad_info.get("manual_file") if load_data else None
                dad_file_from_manifest = dad_info.get("file", "") if load_data else ""

                # Si ja carregat des de masterfile (3-DAD_KHP) o metadades, no cal buscar més
                dad_loaded = rep_data.get("dad") is not None

                # 1. Intentar carregar des de manual_file
                wl_keep = config.get("wavelengths", "selected") if config else None

                if manual_dad_file and not dad_loaded:
                    # Buscar el fitxer manual
                    dad_dirs = ["Export3d", "Export3D", "CSV", "csv", ""]
                    for subdir in dad_dirs:
                        test_path = os.path.join(seq_path, subdir, manual_dad_file) if subdir else os.path.join(seq_path, manual_dad_file)
                        if os.path.exists(test_path):
                            try:
                                df_dad, status = llegir_dad_export3d(test_path, wavelengths_to_keep=wl_keep)
                                if df_dad is not None and status.startswith("OK"):
                                    rep_data["dad"] = {
                                        "df": df_dad,
                                        "path": test_path,
                                        "file": manual_dad_file,
                                        "manual_assignment": True,
                                    }
                                    rep_data["dad_source"] = "export3d"
                                    rep_data["has_data"] = True
                                    dad_loaded = True
                                    break
                            except Exception as e:
                                logger.debug("DAD manual export3d load failed for %s: %s", test_path, e)
                    if not dad_loaded:
                        # Intentar amb dad1a
                        for subdir in dad_dirs:
                            test_path = os.path.join(seq_path, subdir, manual_dad_file) if subdir else os.path.join(seq_path, manual_dad_file)
                            if os.path.exists(test_path):
                                try:
                                    df_dad, status = llegir_dad_1a(test_path)
                                    if df_dad is not None and status.startswith("OK"):
                                        rep_data["dad"] = {
                                            "df": df_dad,
                                            "path": test_path,
                                            "file": manual_dad_file,
                                            "manual_assignment": True,
                                        }
                                        rep_data["dad_source"] = "dad1a"
                                        rep_data["has_data"] = True
                                        dad_loaded = True
                                        break
                                except Exception as e:
                                    logger.debug("DAD manual dad1a load failed for %s: %s", test_path, e)

                # 2. Intentar carregar des de file guardat al manifest
                if dad_file_from_manifest and not dad_loaded:
                    dad_dirs = ["Export3d", "Export3D", "CSV", "csv", ""]
                    for subdir in dad_dirs:
                        test_path = os.path.join(seq_path, subdir, dad_file_from_manifest) if subdir else os.path.join(seq_path, dad_file_from_manifest)
                        if os.path.exists(test_path):
                            try:
                                df_dad, status = llegir_dad_export3d(test_path, wavelengths_to_keep=wl_keep)
                                if df_dad is not None and status.startswith("OK"):
                                    rep_data["dad"] = {"df": df_dad, "path": test_path, "file": dad_file_from_manifest}
                                    rep_data["dad_source"] = "export3d"
                                    rep_data["has_data"] = True
                                    dad_loaded = True
                                    break
                            except Exception as e:
                                logger.debug("DAD manifest export3d load failed for %s: %s", test_path, e)

                # 3. Fallback: buscar per nom de mostra (comportament original, només amb load_data)
                if load_data and not dad_loaded and dad_source == "export3d":
                    path_3d = os.path.join(seq_path, "Export3d")
                    if not os.path.isdir(path_3d):
                        path_3d = os.path.join(seq_path, "Export3D")

                    if os.path.isdir(path_3d):
                        # Buscar fitxer per aquesta mostra
                        dad_files = list_dad_files(path_3d)
                        for df_path in dad_files:
                            # Match per nom mostra
                            fname = os.path.basename(df_path).upper()
                            sname = sample_name.upper().replace(" ", "").replace("_", "")
                            if sname in fname.replace(" ", "").replace("_", ""):
                                try:
                                    df_dad, status = llegir_dad_export3d(df_path, wavelengths_to_keep=wl_keep)
                                    if df_dad is not None and status.startswith("OK"):
                                        rep_data["dad"] = {"df": df_dad, "path": df_path}
                                        rep_data["dad_source"] = "export3d"
                                        rep_data["has_data"] = True
                                        break
                                except Exception as e:
                                    logger.debug("DAD fallback export3d load failed for %s: %s", df_path, e)

            # FIX: KHP sense DAD al manifest (seqs antigues pre-FIX F2.2)
            # Si és KHP i encara no s'ha carregat DAD, intentar 3-DAD_KHP del MasterFile
            if sample_type == "KHP" and rep_data.get("dad") is None and master_khp_data is not None:
                rep_data["dad"], rep_data["dad_source"] = _match_khp_dad_from_masterfile(
                    sample_name, sample_info.get("original_name", sample_name),
                    rep_info.get("injection", {}).get("inj_num", rep_num) if rep_info.get("injection") else rep_num,
                    master_khp_data
                )
                if rep_data["dad"] is not None:
                    rep_data["has_data"] = True
                    logger.info("3-DAD_KHP fallback: loaded %d rows for %s (manifest sense DAD)",
                                len(rep_data["dad"]["df"]), sample_name)
                else:
                    logger.warning("KHP %s rep %s: cap font DAD disponible (ni manifest ni 3-DAD_KHP)", sample_name, rep_num)

            result["samples"][sample_name]["replicas"][rep_num] = rep_data

    # Stats
    samples_with_data = sum(
        1 for s in result["samples"].values()
        for r in s.get("replicas", {}).values()
        if r.get("has_data")
    )

    # Recuperar stats complets del manifest (doc_direct_count, uib_count, dad_count, etc.)
    # i sobreescriure amb valors actuals calculats
    saved_stats = manifest.get("stats", {})
    saved_stats.update({
        "total_samples": len(result["samples"]),
        "samples_with_data": samples_with_data,
        "from_manifest": True,
    })
    result["stats"] = saved_stats

    # Recuperar informació d'orfes i suggeriments del manifest
    orphan_info = manifest.get("orphan_files", {})

    # Reconstruir paths complets per fitxers orfes
    def find_full_path(filename, search_dirs):
        """Busca el path complet d'un fitxer."""
        for subdir in search_dirs:
            test_path = os.path.join(seq_path, subdir, filename) if subdir else os.path.join(seq_path, filename)
            if os.path.exists(test_path):
                return test_path
        return None

    # Directoris on buscar cada tipus
    dad_dirs = ["Export3d", "Export3D", "CSV", "csv", ""]
    uib_dirs = ["CSV", "csv", ""]

    # Reconstruir paths per orfes UIB
    orphan_uib_full = []
    for fname in orphan_info.get("uib", []):
        full_path = find_full_path(fname, uib_dirs)
        if full_path:
            orphan_uib_full.append(full_path)

    # Reconstruir paths per orfes DAD
    orphan_dad_full = []
    for fname in orphan_info.get("dad", []):
        full_path = find_full_path(fname, dad_dirs)
        if full_path:
            orphan_dad_full.append(full_path)

    result["orphan_files"] = {
        "uib": orphan_uib_full,
        "dad": orphan_dad_full,
    }

    # Llista completa d'orfes (incloent suggeriments)
    all_orphan_uib = list(orphan_uib_full)
    all_orphan_dad = list(orphan_dad_full)

    # Afegir suggeriments des del manifest a cada rèplica
    for sample_info in manifest_samples:
        sample_name = sample_info.get("name", "")
        if sample_name not in result["samples"]:
            continue

        for rep_info in sample_info.get("replicas", []):
            rep_num = str(rep_info.get("replica", "1"))
            rep_data = result["samples"][sample_name]["replicas"].get(rep_num)
            if not rep_data:
                continue

            # Recuperar suggeriments DAD
            dad_suggestion = rep_info.get("dad_suggestion")
            if dad_suggestion:
                suggested_file = dad_suggestion.get("file", "")
                rep_data["dad_suggestion"] = {
                    "file": suggested_file,
                    "filename": suggested_file,
                    "confidence": dad_suggestion.get("confidence", 0),
                    "match_type": dad_suggestion.get("match_type", "SUGGESTED"),
                    "status": dad_suggestion.get("status", "SUGGESTED"),
                }
                # Afegir a llista completa d'orfes si no hi és
                if suggested_file:
                    full_path = find_full_path(suggested_file, dad_dirs)
                    if full_path and full_path not in all_orphan_dad:
                        all_orphan_dad.append(full_path)

            # Recuperar suggeriments UIB
            uib_suggestion = rep_info.get("uib_suggestion")
            if uib_suggestion:
                suggested_file = uib_suggestion.get("file", "")
                rep_data["uib_suggestion"] = {
                    "file": suggested_file,
                    "filename": suggested_file,
                    "confidence": uib_suggestion.get("confidence", 0),
                    "match_type": uib_suggestion.get("match_type", "SUGGESTED"),
                    "status": uib_suggestion.get("status", "SUGGESTED"),
                }
                # Afegir a llista completa d'orfes si no hi és
                if suggested_file:
                    full_path = find_full_path(suggested_file, uib_dirs)
                    if full_path and full_path not in all_orphan_uib:
                        all_orphan_uib.append(full_path)

    result["all_orphan_files"] = {
        "uib": all_orphan_uib,
        "dad": all_orphan_dad,
    }

    result["success"] = True
    result["data_deferred"] = not load_data  # True quan les dades crues no s'han carregat

    if load_data:
        report_progress(100, "Importació des de manifest completada")
    else:
        report_progress(100, "Manifest carregat (dades diferides)")

    return result


def ensure_data_loaded(imported_data, config=None, progress_callback=None):
    """
    Completa les dades crues d'un resultat d'importació amb data_deferred=True.

    Quan import_from_manifest() es crida amb load_data=False, les metadades es
    carreguen del manifest però els cromatogrames (DOC Direct, UIB, DAD) no es
    llegeixen del disc. Aquesta funció fa la lectura diferida quan les dades
    realment es necessiten (p.ex. abans d'analitzar).

    Modifica imported_data in-place i retorna el mateix dict.

    Args:
        imported_data: dict retornat per import_from_manifest(load_data=False)
        config: Configuració (opcional)
        progress_callback: Funció callback(pct, msg)

    Returns:
        imported_data actualitzat amb dades crues carregades
    """
    if not imported_data or not imported_data.get("data_deferred"):
        return imported_data  # Ja té les dades carregades

    config = config or get_config()
    wl_keep = config.get("wavelengths", "selected") if config else None

    def report_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    report_progress(5, "Carregant dades des del MasterFile...")

    seq_path = imported_data.get("seq_path", "")
    method = imported_data.get("method", "COLUMN")
    mode = "BP" if method == "BP" else "COLUMN"

    # Llegir MasterFile
    master_path = imported_data.get("master_file", "")
    if master_path and not os.path.isabs(master_path):
        master_path = os.path.join(seq_path, master_path)
    if not master_path or not os.path.exists(master_path):
        master_path_new, _ = trobar_excel_mestre(seq_path)
        if master_path_new:
            master_path = master_path_new
            imported_data["master_file"] = master_path
        else:
            logger.error("ensure_data_loaded: MasterFile no trobat per %s", seq_path)
            return imported_data

    toc_df = None
    master_khp_data = None
    toc_calc_df = None
    try:
        with pd.ExcelFile(master_path, engine="openpyxl") as xl:
            if "2-TOC" in xl.sheet_names:
                toc_df = pd.read_excel(xl, sheet_name="2-TOC", header=6, engine="openpyxl")
            if "3-DAD_KHP" in xl.sheet_names:
                master_khp_data = pd.read_excel(xl, sheet_name="3-DAD_KHP", engine="openpyxl")
            if "4-TOC_CALC" in xl.sheet_names:
                toc_calc_df = pd.read_excel(xl, sheet_name="4-TOC_CALC", engine="openpyxl")
    except Exception as e:
        logger.error("ensure_data_loaded: Error llegint MasterFile: %s", e)
        return imported_data

    # Detectar 4-TOC_CALC amb timestamps arrodonits
    _needs_regen = _toc_calc_has_minute_precision(toc_calc_df)

    # Si 4-TOC_CALC és buit o degradat, calcular i guardar al MasterFile
    if (_needs_regen or toc_calc_df is None or toc_calc_df.empty) and toc_df is not None:
        master_data_temp = llegir_masterfile_nou(master_path)
        computed = compute_toc_calc(master_data_temp, toc_df)
        if computed is not None and not computed.empty:
            saved = _save_toc_calc_to_masterfile(master_path, computed)
            if saved:
                logger.info("ensure_data_loaded: 4-TOC_CALC generat i guardat al MasterFile")
            else:
                logger.warning("ensure_data_loaded: 4-TOC_CALC calculat però no escrit")
        else:
            logger.warning("ensure_data_loaded: No s'ha pogut calcular 4-TOC_CALC")
    elif (toc_calc_df is None or toc_calc_df.empty) and toc_df is None:
        logger.warning("ensure_data_loaded: 4-TOC_CALC i 2-TOC no disponibles")

    report_progress(20, "Completant dades de mostres...")

    # Detecció global de timeouts al flux TOC complet
    seq_timeouts = detect_sequence_timeouts(toc_df) if toc_df is not None else []

    samples = imported_data.get("samples", {})
    total = len(samples)

    for i, (sample_name, sample_data) in enumerate(samples.items()):
        pct = 20 + int((i / total) * 70) if total > 0 else 90
        report_progress(pct, f"Carregant {sample_name}...")

        sample_type = sample_data.get("type", "SAMPLE")

        for rep_num, rep_data in sample_data.get("replicas", {}).items():
            # === DOC Direct ===
            direct = rep_data.get("direct")
            if direct and direct.get("t") is None and direct.get("row_start") is not None:
                doc_result = _load_doc_direct(
                    toc_df, direct["row_start"], direct["row_end"], mode, config,
                    seq_timeouts=seq_timeouts
                )
                if doc_result:
                    rep_data["direct"] = doc_result
                    rep_data["has_data"] = True

            # === DOC UIB ===
            uib = rep_data.get("uib")
            if uib and uib.get("t") is None:
                uib_file = uib.get("file", "")
                if uib_file:
                    uib_result = _load_uib_csv(uib_file, seq_path, mode)
                    if uib_result:
                        uib_result["manual_assignment"] = uib.get("manual_assignment", False)
                        # Propagar timeout_info de Direct a UIB (mateix detector TOC)
                        direct_ti = (rep_data.get("direct") or {}).get("timeout_info", {})
                        uib_result["timeout_info"] = direct_ti
                        uib_result["timeout_at_boundary"] = direct_ti.get("timeout_at_boundary", False)
                        rep_data["uib"] = uib_result
                        rep_data["has_data"] = True

            # === DAD ===
            dad = rep_data.get("dad")
            if dad and dad.get("df") is None:
                dad_source = rep_data.get("dad_source", "export3d")
                dad_file = dad.get("file", "")

                # Intentar des de 3-DAD_KHP per KHP
                if dad_source == "masterfile" and master_khp_data is not None:
                    original_name = sample_data.get("original_name", sample_name)
                    inj_info = rep_data.get("injection_info", {})
                    inj_num = inj_info.get("inj_num", rep_num) if inj_info else rep_num
                    dad_result, dad_src = _match_khp_dad_from_masterfile(
                        sample_name, original_name, inj_num, master_khp_data
                    )
                    if dad_result is not None:
                        rep_data["dad"] = dad_result
                        rep_data["dad_source"] = dad_src
                        rep_data["has_data"] = True
                        continue

                # Intentar des de fitxer DAD
                if dad_file:
                    dad_dirs = ["Export3d", "Export3D", "CSV", "csv", ""]
                    for subdir in dad_dirs:
                        test_path = os.path.join(seq_path, subdir, dad_file) if subdir else os.path.join(seq_path, dad_file)
                        if os.path.exists(test_path):
                            try:
                                df_dad, status = llegir_dad_export3d(test_path, wavelengths_to_keep=wl_keep)
                                if df_dad is not None and status.startswith("OK"):
                                    rep_data["dad"] = {
                                        "df": df_dad,
                                        "path": test_path,
                                        "file": dad_file,
                                        "manual_assignment": dad.get("manual_assignment", False),
                                    }
                                    rep_data["dad_source"] = "export3d"
                                    rep_data["has_data"] = True
                                    break
                            except Exception:
                                pass
                            try:
                                df_dad, status = llegir_dad_1a(test_path)
                                if df_dad is not None and status.startswith("OK"):
                                    rep_data["dad"] = {
                                        "df": df_dad,
                                        "path": test_path,
                                        "file": dad_file,
                                        "manual_assignment": dad.get("manual_assignment", False),
                                    }
                                    rep_data["dad_source"] = "dad1a"
                                    rep_data["has_data"] = True
                                    break
                            except Exception:
                                pass

            # KHP fallback DAD
            if sample_type == "KHP" and rep_data.get("dad") is None and master_khp_data is not None:
                original_name = sample_data.get("original_name", sample_name)
                inj_info = rep_data.get("injection_info", {})
                inj_num = inj_info.get("inj_num", rep_num) if inj_info else rep_num
                dad_result, dad_src = _match_khp_dad_from_masterfile(
                    sample_name, original_name, inj_num, master_khp_data
                )
                if dad_result is not None:
                    rep_data["dad"] = dad_result
                    rep_data["dad_source"] = dad_src
                    rep_data["has_data"] = True

    # Alliberar 2-TOC DataFrame (les dades DOC ja estan extretes a rep_data["direct"]["t"/"y"])
    if imported_data.get("master_data") and imported_data["master_data"].get("toc") is not None:
        imported_data["master_data"]["toc"] = None

    imported_data["data_deferred"] = False
    report_progress(100, "Dades carregades")
    return imported_data


# =============================================================================
# FUNCIONS AUXILIARS PER COMPATIBILITAT
# =============================================================================

# Re-exportar funcions que altres mòduls poden necessitar
__all__ = [
    # Carpeta dades
    "get_data_folder",
    "DATA_FOLDER_NAME",
    # Utilitats
    "normalize_key",
    "normalize_rep",
    "is_khp",
    "is_blank_injection",
    "is_control_injection",
    "obtenir_seq",
    "seq_tag",
    "split_sample_rep",
    "clean_sample_name",
    "skip_sample_direct",
    # Detecció mode
    "is_bp_seq",
    "detect_mode_from_folder",
    # Baseline (usar get_baseline_value de hpsec_utils directament)
    # Lectura/creació master
    "migrate_single",
    "detect_master_format",
    "trobar_excel_mestre",
    "read_master_date",
    "llegir_masterfile_nou",
    "llegir_master_direct",
    # Lectura UIB
    "llegir_doc_uib",
    "netejar_nom_uib",
    # Lectura DAD
    "llegir_dad_export3d",
    "llegir_dad_1a",
    "llegir_dad_amb_fallback",
    "list_dad_files",
    # Matching
    "get_valid_samples_from_hplc_seq",
    "match_sample_confidence",
    "is_sample_in_seq",
    "CONFIDENCE_THRESHOLD",
    # Matching intel·ligent d'orfes
    "extract_sample_from_filename",
    "compute_orphan_suggestions",
    "apply_orphan_suggestions",
    # DAD matching
    "detect_dad_rep_style",
    "dad_sample_rep_from_path",
    "target_keys_from_desc",
    "choose_best_candidate",
    # Verificació
    "detect_replica_anomalies",
    "check_sequence_files",
    # Principal
    "import_sequence",
    # Manifest
    "generate_import_manifest",
    "save_import_manifest",
    "load_manifest",
    "import_from_manifest",
    "ensure_data_loaded",
]
