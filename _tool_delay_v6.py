"""
Delay tool v6 — TOC es desplaça, DAD fix com a referencia.
============================================================
- DAD 254 continu (verd) = referencia fixa al temps HPLC
- TOC continu (blau) = es desplaça amb el slider
- Quan el delay es correcte, els pics blaus coincideixen amb els verds
- Timeouts TOC marcats
- Fletxes teclat per ajust fi

Usage:
    python _tool_delay_v6.py 295_SEQ_BP
"""
import sys, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button
from scipy.ndimage import uniform_filter1d

from hpsec_config import get_data_folders
from hpsec_consolidate import extract_seq_number, detect_seq_type


def find_seq_path(arg):
    if os.path.isdir(arg):
        return arg
    for folder in get_data_folders():
        full = os.path.join(folder, arg)
        if os.path.isdir(full):
            return full
        try:
            num = int(arg)
            for d in os.listdir(folder):
                if extract_seq_number(d) == num:
                    return os.path.join(folder, d)
        except ValueError:
            pass
    return None


def read_export3d_254(filepath):
    try:
        with open(filepath, encoding='utf-16') as fh:
            lines = fh.readlines()
        header = lines[0].strip().split(',')
        idx254 = next((i for i, h in enumerate(header) if '254' in h), None)
        if idx254 is None: return None, None
        times, vals = [], []
        for line in lines[1:]:
            fields = line.strip().split(',')
            if len(fields) > idx254:
                try:
                    times.append(float(fields[0]))
                    vals.append(float(fields[idx254]))
                except ValueError: pass
        return np.array(times), np.array(vals)
    except:
        return None, None


def main():
    if len(sys.argv) < 2:
        print("Usage: python _tool_delay_v6.py <SEQ>")
        sys.exit(1)

    seq_path = find_seq_path(sys.argv[1])
    if not seq_path:
        print(f"ERROR: '{sys.argv[1]}' no trobat"); sys.exit(1)

    seq_name = os.path.basename(seq_path)
    method = detect_seq_type(seq_name)
    print(f"Loading {seq_name} ({method})...")

    mfs = [f for f in os.listdir(seq_path)
           if 'MasterFile' in f and f.endswith('.xlsx')
           and '~' not in f and 'pre_vol' not in f and 'backup' not in f]
    if not mfs:
        mfs = [f for f in os.listdir(seq_path)
               if 'MasterFile' in f and f.endswith('.xlsx')
               and '~' not in f and 'pre_vol' not in f]
    mf_path = os.path.join(seq_path, mfs[0])

    # === TOC ===
    df_toc = pd.read_excel(mf_path, sheet_name='2-TOC', header=6)
    toc_col = [c for c in df_toc.columns if 'toc' in str(c).lower()][0]
    date_col_toc = [c for c in df_toc.columns if 'started' in str(c).lower()]
    y_toc_raw = pd.to_numeric(df_toc[toc_col], errors='coerce').values
    dt = 4.0 / 60.0
    t_toc_orig = np.arange(len(y_toc_raw)) * dt
    y_toc = np.nan_to_num(y_toc_raw, nan=np.nanmedian(y_toc_raw[~np.isnan(y_toc_raw)]))
    y_toc_smooth = uniform_filter1d(y_toc, size=5)

    t0_toc = None
    if date_col_toc:
        dates = pd.to_datetime(df_toc[date_col_toc[0]], errors='coerce')
        valid = dates.dropna()
        if len(valid) > 0: t0_toc = valid.iloc[0]

    # === Detect TOC timeouts (gaps in signal) ===
    timeout_t = []
    if date_col_toc:
        ts = pd.to_datetime(df_toc[date_col_toc[0]], errors='coerce')
        ts_diff = ts.diff().dt.total_seconds().values
        for i, d in enumerate(ts_diff):
            if not np.isnan(d) and d > 30:
                timeout_t.append(t_toc_orig[i])

    print(f"TOC: {len(y_toc)} pts, {len(timeout_t)} timeouts")

    # === HPLC ===
    df_hplc = pd.read_excel(mf_path, sheet_name='1-HPLC-SEQ')
    name_col = [c for c in df_hplc.columns
                if 'sample' in str(c).lower() and 'name' in str(c).lower()][0]
    hplc_date_col = [c for c in df_hplc.columns if 'acquired' in str(c).lower()][0]

    injs = []
    for _, row in df_hplc.iterrows():
        try:
            name = str(row[name_col]).strip()
            if pd.isna(name) or name == 'nan': continue
            dt_val = pd.to_datetime(row[hplc_date_col])
            t_min = (dt_val - t0_toc).total_seconds() / 60.0 if t0_toc else 0
            injs.append({'name': name, 't_min': t_min,
                         'is_khp': 'khp' in name.lower(),
                         'is_control': any(x in name.lower()
                                           for x in ['mq', 'naoh', 'blanc', 'blnc'])})
        except: pass

    cadence = np.median(np.diff([i['t_min'] for i in injs])) if len(injs) >= 2 else 11.0

    # === DAD 254 continu ===
    e3d = os.path.join(seq_path, 'Export3d')
    e3d_files = sorted(os.listdir(e3d)) if os.path.isdir(e3d) else []

    name_counter = {}
    dad_t_all, dad_y_all = [], []

    for j, inj in enumerate(injs):
        nk = ''.join(c for c in inj['name'].lower() if c.isalnum())
        name_counter[nk] = name_counter.get(nk, 0) + 1
        rep = name_counter[nk]

        matched = None
        for f in e3d_files:
            fb = f.lower().replace('.csv', '')
            fb_alnum = ''.join(c for c in fb if c.isalnum())
            if nk in fb_alnum and (f'_r{rep}' in fb or f'_{rep}' in fb):
                if 'uib' not in fb:
                    matched = f; break

        if matched is None: continue
        t_rel, y_254 = read_export3d_254(os.path.join(e3d, matched))
        if t_rel is None or len(t_rel) < 5: continue

        # Absolute time = HPLC injection time + relative time in run
        t_abs = t_rel + inj['t_min']
        dad_t_all.append(t_abs)
        dad_y_all.append(y_254)

    has_dad = len(dad_t_all) > 0
    if has_dad:
        t_dad = np.concatenate(dad_t_all)
        y_dad = np.concatenate(dad_y_all)
        order = np.argsort(t_dad)
        t_dad, y_dad = t_dad[order], y_dad[order]
        print(f"DAD 254: {len(t_dad)} pts, {len(dad_t_all)} injeccions")
    else:
        print("DAD 254: no disponible")

    # === Current delay ===
    current_delay = 3.637
    df_info = pd.read_excel(mf_path, sheet_name='0-INFO', header=None)
    hora_hplc = hora_toc = None
    for _, row in df_info.iterrows():
        key = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        val = row.iloc[1] if len(row) > 1 else None
        if 'hora hplc' in key: hora_hplc = val
        elif 'hora toc' in key: hora_toc = val
        elif 'net delay' in key and 'suite' not in key:
            try:
                v = float(val)
                if not np.isnan(v): current_delay = v
            except: pass
    if hora_hplc and hora_toc:
        try:
            def to_min(t):
                if hasattr(t, 'hour'):
                    return t.hour * 60 + t.minute + t.second / 60
                return sum(int(p)*m for p, m in zip(str(t).split(':'), [60, 1]))
            calc = 3.637 - (to_min(hora_hplc) - to_min(hora_toc))
            if current_delay == 3.637: current_delay = calc
        except: pass

    # === FIGURE ===
    fig, ax = plt.subplots(figsize=(22, 7))
    plt.subplots_adjust(bottom=0.15, top=0.90, left=0.06, right=0.92)

    # DAD 254 FIX — eix dret (referencia, NO es mou)
    ax2 = ax.twinx()
    if has_dad:
        ax2.plot(t_dad, y_dad, 'g-', lw=0.5, alpha=0.6)
    ax2.set_ylabel('DAD 254 (mAU)', color='green', fontsize=10)
    ax2.tick_params(axis='y', labelcolor='green')

    ax.set_ylabel('TOC (ppb)', color='blue', fontsize=10)
    ax.tick_params(axis='y', labelcolor='blue')
    ax.set_xlabel('min (temps absolut des de inici TOC)')

    # Injection labels on DAD
    for j, inj in enumerate(injs):
        color = '#E74C3C' if inj['is_khp'] else ('#95A5A6' if inj['is_control'] else '#27AE60')
        ax2.axvline(inj['t_min'], color=color, ls=':', lw=0.3, alpha=0.2)

    # TOC line (MOVABLE) + timeout markers
    toc_line = [None]
    timeout_lines = []

    def draw_toc(delay_val):
        # Save zoom
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        ylim2 = ax2.get_ylim()
        zoomed = xlim != (0, 1) and ylim != (0, 1)

        # Remove old TOC line
        if toc_line[0] is not None:
            toc_line[0].remove()
        for tl in timeout_lines:
            tl.remove()
        timeout_lines.clear()

        # TOC shifted: subtract delay to align with DAD timeline
        # TOC absolute time = t_toc_orig
        # DAD absolute time = t_hplc + t_rel (direct HPLC clock)
        # If delay is correct: t_toc_pic = t_dad_pic + delay
        # So to align: t_toc_shifted = t_toc_orig - delay
        t_shifted = t_toc_orig - delay_val
        toc_line[0], = ax.plot(t_shifted, y_toc_smooth, 'b-', lw=0.5, alpha=0.7)

        # Timeouts shifted
        for tp in timeout_t:
            tp_shifted = tp - delay_val
            tl = ax.axvspan(tp_shifted - 0.3, tp_shifted + 1.0,
                            alpha=0.12, color='red', zorder=0)
            timeout_lines.append(tl)

        # Restore zoom
        if zoomed:
            ax.set_xlim(xlim)
            ax.set_ylim(ylim)
            ax2.set_ylim(ylim2)

        fig.suptitle(
            f"{seq_name} ({method}) — delay={delay_val:.1f} min\n"
            f"Verd=DAD254 (fix) | Blau=TOC (mou amb slider) | "
            f"Vermell=timeout | {len(timeout_t)} TO detectats",
            fontsize=10, fontweight='bold')

    draw_toc(current_delay)

    # === SLIDER ===
    ax_slider = plt.axes([0.15, 0.04, 0.50, 0.025])
    slider = Slider(ax_slider, 'Delay', current_delay - 20, current_delay + 20,
                    valinit=current_delay, valstep=0.1)
    def on_slider(val):
        draw_toc(val)
        fig.canvas.draw_idle()
    slider.on_changed(on_slider)

    # Keyboard: fletxes per ajust fi
    def on_key(event):
        if event.key == 'left':
            slider.set_val(max(slider.val - 0.1, slider.valmin))
        elif event.key == 'right':
            slider.set_val(min(slider.val + 0.1, slider.valmax))
        elif event.key == 'shift+left':
            slider.set_val(max(slider.val - 1.0, slider.valmin))
        elif event.key == 'shift+right':
            slider.set_val(min(slider.val + 1.0, slider.valmax))
    fig.canvas.mpl_connect('key_press_event', on_key)

    # === BUTTONS ===
    ax_auto = plt.axes([0.70, 0.04, 0.07, 0.025])
    btn_auto = Button(ax_auto, 'Auto')
    def on_auto(e):
        if not has_dad: return
        # Cross-correlation
        t_common = np.arange(0, min(t_toc_orig[-1], t_dad[-1]+20), dt)
        best_corr, best_d = -1, current_delay
        for d_try in np.arange(max(0, current_delay-15), current_delay+15, 0.2):
            y_toc_i = np.interp(t_common, t_toc_orig - d_try, y_toc_smooth,
                                left=np.nan, right=np.nan)
            y_dad_i = np.interp(t_common, t_dad, y_dad, left=np.nan, right=np.nan)
            valid = ~np.isnan(y_toc_i) & ~np.isnan(y_dad_i)
            if valid.sum() > 50:
                corr = np.corrcoef(y_toc_i[valid], y_dad_i[valid])[0, 1]
                if corr > best_corr: best_corr, best_d = corr, d_try
        slider.set_val(best_d)
        print(f"Auto: delay={best_d:.1f} (corr={best_corr:.3f})")
    btn_auto.on_clicked(on_auto)

    ax_save = plt.axes([0.78, 0.04, 0.07, 0.025])
    btn_save = Button(ax_save, 'GUARDAR')
    def on_save(e):
        import openpyxl
        val = slider.val
        wb = openpyxl.load_workbook(mf_path)
        ws = wb['0-INFO']
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value and 'net delay' in str(row[0].value).lower() and 'suite' not in str(row[0].value).lower():
                ws.cell(row=row[0].row, column=2, value=round(val, 3))
                wb.save(mf_path); print(f"GUARDAT: {val:.3f}"); return
        lr = ws.max_row + 1
        ws.cell(row=lr, column=1, value='Net delay (Suite)')
        ws.cell(row=lr, column=2, value=round(val, 3))
        wb.save(mf_path); print(f"GUARDAT: {val:.3f}")
    btn_save.on_clicked(on_save)

    ax_reset = plt.axes([0.86, 0.04, 0.07, 0.025])
    btn_reset = Button(ax_reset, 'Reset')
    btn_reset.on_clicked(lambda e: slider.set_val(current_delay))

    fig.text(0.06, 0.01,
             f"MF delay: {current_delay:.2f} | Cadencia: {cadence:.1f}min | "
             f"DAD: {len(dad_t_all)}/{len(injs)} inj | "
             f"Fletxes ←→ ±0.1min, Shift+←→ ±1min",
             fontsize=8, color='#666')

    print(f"DAD verd = referencia fixa")
    print(f"TOC blau = es mou amb slider")
    print(f"Quan coincideixen = delay correcte")
    print(f"Fletxes ←→ per ajust fi")
    plt.show()


if __name__ == "__main__":
    main()
