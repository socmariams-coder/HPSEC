"""
Interactive delay adjustment tool.
===================================
Opens a matplotlib window with:
- Full TOC signal
- Injection assignment markers (move with delay slider)
- Real-time update when slider changes
- Save button to write delay to MasterFile

Usage:
    python _tool_delay_adjust.py 231          # by SEQ number
    python _tool_delay_adjust.py 231_SEQ_BP   # by name
    python _tool_delay_adjust.py C:\path\to\SEQ  # by full path
"""
import sys, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')  # interactive backend
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button
from pathlib import Path

from hpsec_config import get_data_folders
from hpsec_consolidate import extract_seq_number, detect_seq_type

FLUSH_TIME_MIN = 3.637


def find_seq_path(arg):
    """Resolve SEQ path from number, name, or full path."""
    if os.path.isdir(arg):
        return arg

    folders = get_data_folders()
    for folder in folders:
        # Try exact name
        full = os.path.join(folder, arg)
        if os.path.isdir(full):
            return full
        # Try by number
        try:
            num = int(arg)
            for d in os.listdir(folder):
                if extract_seq_number(d) == num:
                    return os.path.join(folder, d)
        except ValueError:
            pass

    return None


def load_data(seq_path):
    """Load TOC signal, HPLC times, and current delay."""
    seq_name = os.path.basename(seq_path)
    method = detect_seq_type(seq_name)

    # Find MasterFile
    mfs = [f for f in os.listdir(seq_path)
           if 'MasterFile' in f and f.endswith('.xlsx')
           and '~' not in f and 'pre_vol' not in f and 'backup' not in f]
    if not mfs:
        mfs = [f for f in os.listdir(seq_path)
               if 'MasterFile' in f and f.endswith('.xlsx')
               and '~' not in f and 'pre_vol' not in f]
    if not mfs:
        print(f"ERROR: No MasterFile a {seq_path}")
        sys.exit(1)

    mf_path = os.path.join(seq_path, mfs[0])

    # 2-TOC
    df_toc_raw = pd.read_excel(mf_path, sheet_name='2-TOC', header=6)
    toc_col = None
    date_col = None
    for c in df_toc_raw.columns:
        cs = str(c).lower()
        if toc_col is None and 'toc' in cs:
            toc_col = c
        if date_col is None and 'started' in cs:
            date_col = c

    if toc_col is None:
        print("ERROR: No TOC column")
        sys.exit(1)

    y_toc = pd.to_numeric(df_toc_raw[toc_col], errors='coerce').values
    t_toc = np.arange(len(y_toc)) * (4.0 / 60.0)

    t0_toc = None
    if date_col:
        dates = pd.to_datetime(df_toc_raw[date_col], errors='coerce')
        valid = dates.dropna()
        if len(valid) > 0:
            t0_toc = valid.iloc[0]

    # 1-HPLC-SEQ
    df_hplc = pd.read_excel(mf_path, sheet_name='1-HPLC-SEQ')
    name_col = None
    hplc_date_col = None
    for c in df_hplc.columns:
        if name_col is None and 'sample' in str(c).lower() and 'name' in str(c).lower():
            name_col = c
        if hplc_date_col is None and 'acquired' in str(c).lower():
            hplc_date_col = c

    injs = []
    if name_col and hplc_date_col and t0_toc:
        for _, row in df_hplc.iterrows():
            try:
                name = str(row[name_col]).strip()
                dt = pd.to_datetime(row[hplc_date_col])
                t_min = (dt - t0_toc).total_seconds() / 60.0

                is_khp = 'khp' in name.lower()
                is_mq = any(x in name.lower() for x in ['mq', 'blanc', 'blnc'])
                is_naoh = 'naoh' in name.lower()

                injs.append({
                    'name': name, 't_min': t_min,
                    'is_khp': is_khp, 'is_mq': is_mq, 'is_naoh': is_naoh,
                    'is_control': is_mq or is_naoh,
                })
            except Exception:
                pass

    # Current delay from 0-INFO (calculated from clocks)
    df_info = pd.read_excel(mf_path, sheet_name='0-INFO', header=None)
    hora_hplc = None
    hora_toc = None
    delay_explicit = None

    for _, row in df_info.iterrows():
        key = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        val = row.iloc[1] if len(row) > 1 else None
        if 'hora hplc' in key:
            hora_hplc = val
        elif 'hora toc' in key:
            hora_toc = val
        elif 'net delay' in key and 'suite' not in key:
            try:
                v = float(val)
                if not np.isnan(v):
                    delay_explicit = v
            except (ValueError, TypeError):
                pass

    current_delay = delay_explicit
    if current_delay is None and hora_hplc and hora_toc:
        try:
            def to_min(t):
                if hasattr(t, 'hour'):
                    return t.hour * 60 + t.minute + t.second / 60
                parts = str(t).split(':')
                return int(parts[0]) * 60 + int(parts[1])
            desfase = to_min(hora_hplc) - to_min(hora_toc)
            current_delay = FLUSH_TIME_MIN - desfase
        except Exception:
            current_delay = FLUSH_TIME_MIN

    if current_delay is None:
        current_delay = FLUSH_TIME_MIN

    # Cadence
    if len(injs) >= 2:
        cadence = np.median(np.diff([inj['t_min'] for inj in injs]))
    else:
        cadence = 11.0 if method == 'BP' else 78.0

    return {
        'seq_name': seq_name,
        'method': method,
        'mf_path': mf_path,
        't_toc': t_toc,
        'y_toc': y_toc,
        'injs': injs,
        'current_delay': current_delay,
        'cadence': cadence,
    }


def count_matches(t_toc, y_toc, injs, delay, cadence):
    """Count how many injection windows have a clear peak."""
    from scipy.ndimage import uniform_filter1d
    dt = t_toc[1] - t_toc[0] if len(t_toc) > 1 else 4.0/60.0
    y_smooth = uniform_filter1d(
        np.nan_to_num(y_toc, nan=np.nanmedian(y_toc[~np.isnan(y_toc)])),
        size=max(3, int(0.3/dt)))

    n_ok = 0
    n_samples = 0
    for inj in injs:
        if inj['is_control']:
            continue
        n_samples += 1
        t_center = inj['t_min'] + delay
        hw = cadence / 2
        mask = (t_toc >= t_center - hw * 0.3) & (t_toc <= t_center + hw * 0.7)
        y_w = y_smooth[mask]
        if len(y_w) < 5:
            continue
        bl = np.percentile(y_w, 20)
        pk = np.max(y_w - bl)
        noise_vals = y_w - bl
        noise_low = noise_vals[noise_vals < np.percentile(noise_vals, 30)]
        noise = np.std(noise_low) if len(noise_low) > 3 else 1.0
        if noise > 0 and pk / noise > 5:
            n_ok += 1

    return n_ok, n_samples


def main():
    if len(sys.argv) < 2:
        print("Usage: python _tool_delay_adjust.py <SEQ_number_or_name>")
        sys.exit(1)

    seq_path = find_seq_path(sys.argv[1])
    if not seq_path:
        print(f"ERROR: No s'ha trobat la SEQ '{sys.argv[1]}'")
        sys.exit(1)

    print(f"Loading {os.path.basename(seq_path)}...")
    data = load_data(seq_path)

    t = data['t_toc']
    y = data['y_toc']
    injs = data['injs']
    cadence = data['cadence']
    method = data['method']
    current_delay = data['current_delay']

    # Smooth for display
    from scipy.ndimage import uniform_filter1d
    dt = t[1] - t[0] if len(t) > 1 else 4.0/60.0
    y_smooth = uniform_filter1d(
        np.nan_to_num(y, nan=np.nanmedian(y[~np.isnan(y)])),
        size=max(3, int(0.3/dt)))

    # --- BUILD FIGURE ---
    fig, ax = plt.subplots(figsize=(18, 7))
    plt.subplots_adjust(bottom=0.22, top=0.92)

    # TOC signal
    ax.plot(t, y, 'b-', lw=0.2, alpha=0.3)
    ax.plot(t, y_smooth, 'b-', lw=0.5, alpha=0.8)
    ax.set_xlabel('min (des de inici TOC)')
    ax.set_ylabel('ppb')

    # Store line references for updating
    vlines = []
    labels = []

    def draw_assignments(delay_val):
        """Draw/update injection assignment markers."""
        # Remove old markers
        for vl in vlines:
            vl.remove()
        vlines.clear()
        for lb in labels:
            lb.remove()
        labels.clear()

        y_top = ax.get_ylim()[1]

        for j, inj in enumerate(injs):
            t_exp = inj['t_min'] + delay_val

            if inj['is_khp']:
                color = '#E74C3C'
                lw = 1.5
            elif inj['is_control']:
                color = '#95A5A6'
                lw = 0.5
            else:
                color = '#27AE60'
                lw = 0.8

            vl = ax.axvline(t_exp, color=color, ls='--', lw=lw, alpha=0.6)
            vlines.append(vl)

            # Label
            short = inj['name'][:7]
            lb = ax.annotate(f"{j+1}:{short}", (t_exp, y_top * 0.98),
                            fontsize=4, rotation=90, va='top', ha='center',
                            color=color, alpha=0.7)
            labels.append(lb)

        # Count matches
        n_ok, n_total = count_matches(t, y, injs, delay_val, cadence)
        pct = n_ok / n_total * 100 if n_total > 0 else 0
        ax.set_title(
            f"{data['seq_name']} ({method}) — delay={delay_val:.2f} min — "
            f"Mostres amb pic: {n_ok}/{n_total} ({pct:.0f}%)",
            fontsize=12, fontweight='bold')

    # Initial draw
    draw_assignments(current_delay)

    # --- SLIDER ---
    ax_slider = plt.axes([0.15, 0.08, 0.55, 0.03])
    # Range: current_delay +/- 30 min
    slider_min = current_delay - 30
    slider_max = current_delay + 30
    slider = Slider(ax_slider, 'Delay (min)', slider_min, slider_max,
                    valinit=current_delay, valstep=0.1)

    def on_slider_change(val):
        draw_assignments(val)
        fig.canvas.draw_idle()

    slider.on_changed(on_slider_change)

    # --- BUTTONS ---
    # Auto-detect button
    ax_auto = plt.axes([0.75, 0.08, 0.08, 0.03])
    btn_auto = Button(ax_auto, 'Auto-detect')

    def on_auto(event):
        """Find optimal delay automatically."""
        from scipy.ndimage import uniform_filter1d as uf1d
        from scipy.signal import find_peaks as sp_peaks

        y_s = uf1d(np.nan_to_num(y, nan=np.nanmedian(y[~np.isnan(y)])), size=5)
        bl = np.percentile(y_s, 15)

        if method == 'BP':
            md = int(8.0 / dt)
        else:
            md = int(40.0 / dt)

        peaks, _ = sp_peaks(y_s, distance=md, prominence=max(5, (y_s.max()-bl)*0.02),
                            height=bl+10)

        if len(peaks) == 0:
            return

        # Find baseline end -> first peak
        bl_std = np.std(y_s[:int(5.0/dt)])
        thresh = bl + max(5*bl_std, 10)
        win = int(0.5/dt)
        bl_end_idx = int(5.0/dt)
        for i in range(bl_end_idx, len(y_s)-win):
            if np.mean(y_s[i:i+win]) > thresh:
                bl_end_idx = i
                break

        # First peak after baseline
        first_pk = None
        for pk in peaks:
            if pk >= bl_end_idx:
                first_pk = pk
                break

        if first_pk is None:
            first_pk = peaks[0]

        hplc_times = [inj['t_min'] for inj in injs]
        if hplc_times:
            est_delay = t[first_pk] - hplc_times[0]
            slider.set_val(est_delay)

    btn_auto.on_clicked(on_auto)

    # Save button
    ax_save = plt.axes([0.85, 0.08, 0.08, 0.03])
    btn_save = Button(ax_save, 'GUARDAR')
    btn_save.color = '#27AE60'
    btn_save.hovercolor = '#219A52'

    def on_save(event):
        """Save current delay to MasterFile 0-INFO."""
        delay_val = slider.val
        mf_path = data['mf_path']

        import openpyxl
        wb = openpyxl.load_workbook(mf_path)
        ws = wb['0-INFO']

        # Find the Net delay row
        saved = False
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value and 'net delay' in str(cell.value).lower() and 'suite' not in str(cell.value).lower():
                # Write value to column B
                ws.cell(row=cell.row, column=2, value=round(delay_val, 3))
                saved = True
                break

        if not saved:
            # Add new row
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1, value='Net delay (Suite)')
            ws.cell(row=last_row, column=2, value=round(delay_val, 3))

        wb.save(mf_path)
        print(f"GUARDAT: delay={delay_val:.3f} min a {os.path.basename(mf_path)}")

        # Update title
        n_ok, n_total = count_matches(t, y, injs, delay_val, cadence)
        pct = n_ok / n_total * 100 if n_total > 0 else 0
        ax.set_title(
            f"{data['seq_name']} ({method}) — delay={delay_val:.2f} min — "
            f"Mostres amb pic: {n_ok}/{n_total} ({pct:.0f}%) — GUARDAT!",
            fontsize=12, fontweight='bold', color='green')
        fig.canvas.draw_idle()

    btn_save.on_clicked(on_save)

    # --- RESET button ---
    ax_reset = plt.axes([0.75, 0.04, 0.08, 0.03])
    btn_reset = Button(ax_reset, 'Reset')

    def on_reset(event):
        slider.set_val(current_delay)

    btn_reset.on_clicked(on_reset)

    # Info text
    fig.text(0.15, 0.02,
             f"Cadencia: {cadence:.1f} min | "
             f"Injeccions: {len(injs)} | "
             f"Delay MasterFile: {current_delay:.2f} min",
             fontsize=9, color='#666')

    print(f"Obert. Delay actual: {current_delay:.2f} min")
    print(f"Usa el slider per ajustar. 'Auto-detect' per estimar. 'GUARDAR' per escriure al MasterFile.")
    plt.show()


if __name__ == "__main__":
    main()
