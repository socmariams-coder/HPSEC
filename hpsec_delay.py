# -*- coding: utf-8 -*-
"""
hpsec_delay.py — Gestio del Net delay HPLC↔TOC als MasterFiles.

Funcions per:
- Llegir el delay actual d'un MasterFile
- Actualitzar el delay i recalcular 4-TOC_CALC
- Estimar l'impacte d'un canvi de delay (quantes injeccions canvien)

El "Net delay" es el parametre clau per assignar files TOC a cada injeccio HPLC.
Es llegeix del MasterFile (full 0-INFO, cel·la B12).

Formula: hora_HPLC = hora_TOC - Net_delay
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)


def read_current_delay(mf_path):
    """
    Llegeix el Net delay actual d'un MasterFile.

    Prioritat: "Net delay (Suite)" (corregit) > B12 (original) > càlcul implícit
    des de hores HPLC/TOC + flush time.

    Args:
        mf_path: Path al MasterFile (.xlsx)

    Returns:
        float o None: Net delay en minuts, o None si no es pot llegir
    """
    import openpyxl

    FLUSH_TIME_MIN = 3.637

    try:
        wb = openpyxl.load_workbook(str(mf_path), read_only=True, data_only=True)
        if '0-INFO' not in wb.sheetnames:
            wb.close()
            return None
        ws = wb['0-INFO']

        # Buscar "Net delay (Suite)" — valor corregit per la Suite (prioritari)
        suite_delay = None
        info = {}
        for row in range(1, ws.max_row + 1):
            key = ws.cell(row=row, column=1).value
            val = ws.cell(row=row, column=2).value
            if key:
                key_lower = str(key).lower()
                if 'net delay (suite)' in key_lower:
                    if val is not None:
                        suite_delay = float(val)
                info[key_lower] = val

        # Fallback: B12 (delay original)
        original_delay = ws['B12'].value
        wb.close()

        if suite_delay is not None:
            return suite_delay
        if original_delay is not None:
            return float(original_delay)

        # Últim recurs: calcular delay implícit des de hores HPLC/TOC + flush time
        # (mateixa lògica que compute_toc_calc a hpsec_import.py)
        hora_hplc_clock = None
        hora_toc_clock = None
        flush_time = FLUSH_TIME_MIN
        for key_lower, val in info.items():
            if 'hora hplc' in key_lower or 'hora_hplc' in key_lower:
                hora_hplc_clock = val
            elif 'hora toc' in key_lower or 'hora_toc' in key_lower:
                hora_toc_clock = val
            elif 'flush time' in key_lower or 'flush_time' in key_lower:
                if val is not None:
                    try:
                        flush_time = float(val)
                    except (ValueError, TypeError):
                        pass

        if hora_hplc_clock and hora_toc_clock:
            try:
                def _to_minutes(t):
                    if hasattr(t, 'hour'):
                        return t.hour * 60 + t.minute + t.second / 60
                    parts = str(t).split(':')
                    return int(parts[0]) * 60 + int(parts[1])

                hplc_min = _to_minutes(hora_hplc_clock)
                toc_min = _to_minutes(hora_toc_clock)
                desfase_min = hplc_min - toc_min
                implicit_delay = flush_time - desfase_min
                logger.debug(f"Delay implícit calculat: flush={flush_time:.3f} - "
                             f"desfase={desfase_min:.3f} = {implicit_delay:.3f} min")
                return implicit_delay
            except Exception as e:
                logger.debug(f"No s'ha pogut calcular delay implícit: {e}")

        return None
    except Exception as e:
        logger.warning(f"Error llegint delay de {mf_path}: {e}")
        return None


def _read_hplc_timestamps(wb):
    """Llegeix timestamps HPLC del full 1-HPLC-SEQ."""
    ws = wb['1-HPLC-SEQ']

    # Buscar columnes
    hplc_header = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        hplc_header.append(str(val).lower() if val else '')

    date_col_idx = None
    sample_col_idx = None
    sample_rep_col_idx = None

    for i, h in enumerate(hplc_header):
        if 'acquired date' in h or ('injection' in h and 'date' in h):
            date_col_idx = i + 1
        elif h == 'sample_rep':
            sample_rep_col_idx = i + 1
        elif 'sample' in h and 'name' in h:
            sample_col_idx = i + 1

    if date_col_idx is None:
        return [], []

    hplc_times = []
    hplc_samples = []
    for row in range(2, ws.max_row + 1):
        dt_val = ws.cell(row=row, column=date_col_idx).value
        if dt_val is None:
            continue
        if isinstance(dt_val, str):
            try:
                dt_val = pd.to_datetime(dt_val)
            except (ValueError, TypeError):
                continue
        elif not hasattr(dt_val, 'hour'):
            continue

        hplc_times.append(pd.Timestamp(dt_val))

        if sample_rep_col_idx:
            sample = ws.cell(row=row, column=sample_rep_col_idx).value or ''
        elif sample_col_idx:
            sample = ws.cell(row=row, column=sample_col_idx).value or ''
        else:
            sample = f'INJ_{row - 1}'
        hplc_samples.append(str(sample))

    return hplc_times, hplc_samples


def _read_toc_timestamps(wb):
    """Llegeix timestamps TOC del full 2-TOC."""
    ws = wb['2-TOC']
    TOC_START = 8

    time_col_toc = None
    for c in range(1, min(ws.max_column + 1, 10)):
        header = ws.cell(row=7, column=c).value
        if header and 'date' in str(header).lower() and 'start' in str(header).lower():
            time_col_toc = c
            break
    if time_col_toc is None:
        time_col_toc = 4  # fallback: columna D

    toc_rows = []
    for row in range(TOC_START, ws.max_row + 1):
        time_val = ws.cell(row=row, column=time_col_toc).value
        if time_val is None:
            continue
        if isinstance(time_val, str):
            try:
                time_val = pd.to_datetime(time_val)
            except (ValueError, TypeError):
                continue
        elif not hasattr(time_val, 'hour'):
            continue
        toc_rows.append((row, pd.Timestamp(time_val)))

    return toc_rows


def _config_pre_margin():
    """Marge de pre-injecció (dispersió del reactor TOC), FONT ÚNICA des de config.
    Així l'eina de delay i la importació usen el mateix valor (abans 1,5 fix
    a l'eina vs config a l'import → assignacions divergents)."""
    try:
        from hpsec_config import get_config
        return float(get_config().get("sequence", "toc_pre_margin_min", default=1.5))
    except Exception:
        return 1.5


def _assign_toc_rows(hplc_times, toc_rows, net_delay_min, pre_margin_min=None):
    """
    Assigna files TOC a injeccions HPLC amb un delay donat.

    Returns:
        list of dict: [{toc_row, inj_index, sample_idx, temps_rel_min}, ...]
    """
    if not hplc_times or not toc_rows:
        return []

    if pre_margin_min is None:
        pre_margin_min = _config_pre_margin()
    hplc_times_ns = np.array([t.value for t in hplc_times])
    pre_margin_ns = pre_margin_min * 60 * 1e9

    assignments = []
    for toc_row, toc_time in toc_rows:
        hora_hplc = toc_time - pd.Timedelta(minutes=net_delay_min)
        hora_hplc_ns = hora_hplc.value

        inj_index = int((hplc_times_ns <= hora_hplc_ns).sum())

        # Check pre-margin
        if inj_index < len(hplc_times_ns):
            if (hplc_times_ns[inj_index] - hora_hplc_ns) <= pre_margin_ns:
                inj_index += 1

        if 0 < inj_index <= len(hplc_times):
            inj_start = hplc_times[inj_index - 1]
            temps_rel = (hora_hplc - inj_start).total_seconds() / 60.0
        else:
            temps_rel = None

        assignments.append({
            'toc_row': toc_row,
            'inj_index': inj_index,
            'temps_rel_min': temps_rel,
        })

    return assignments


def estimate_delay_impact(mf_path, old_delay, new_delay, pre_margin_min=None):
    """
    Estima l'impacte d'un canvi de delay sense modificar el fitxer.

    Args:
        mf_path: Path al MasterFile
        old_delay: Delay actual (min)
        new_delay: Delay nou (min)
        pre_margin_min: Marge de pre-injeccio (min)

    Returns:
        dict: {
            n_total: total files TOC,
            n_changed: files que canvien d'injeccio,
            n_injections: total injeccions HPLC,
            changed_injections: set d'indexos d'injeccio afectats,
            old_assignments: assignacions amb delay antic,
            new_assignments: assignacions amb delay nou,
        }
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(mf_path), read_only=True, data_only=True)

    try:
        hplc_times, hplc_samples = _read_hplc_timestamps(wb)
        toc_rows = _read_toc_timestamps(wb)
    finally:
        wb.close()

    if not hplc_times or not toc_rows:
        return {
            'n_total': len(toc_rows),
            'n_changed': 0,
            'n_injections': len(hplc_times),
            'changed_injections': set(),
        }

    old_assign = _assign_toc_rows(hplc_times, toc_rows, old_delay, pre_margin_min)
    new_assign = _assign_toc_rows(hplc_times, toc_rows, new_delay, pre_margin_min)

    n_changed = 0
    changed_inj = set()
    for old_a, new_a in zip(old_assign, new_assign):
        if old_a['inj_index'] != new_a['inj_index']:
            n_changed += 1
            changed_inj.add(old_a['inj_index'])
            changed_inj.add(new_a['inj_index'])

    return {
        'n_total': len(toc_rows),
        'n_changed': n_changed,
        'n_injections': len(hplc_times),
        'changed_injections': changed_inj,
        'old_assignments': old_assign,
        'new_assignments': new_assign,
    }


def update_masterfile_delay(mf_path, net_delay_min, pre_margin_min=None,
                            backup=True):
    """
    Actualitza el Net delay i recalcula 4-TOC_CALC al MasterFile.

    Args:
        mf_path: Path al MasterFile (.xlsx)
        net_delay_min: Nou delay en minuts
        pre_margin_min: Marge dispersio reactor TOC (min)
        backup: Si True, crea backup abans de modificar

    Returns:
        dict: {success, n_assigned, n_total, old_delay, new_delay, backup_path}
    """
    import openpyxl

    mf_path = Path(mf_path)
    result = {
        'success': False,
        'n_assigned': 0,
        'n_total': 0,
        'old_delay': None,
        'new_delay': net_delay_min,
        'backup_path': None,
    }

    try:
        # Backup
        if backup:
            backup_name = mf_path.stem + f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            backup_path = mf_path.parent / backup_name
            shutil.copy2(str(mf_path), str(backup_path))
            result['backup_path'] = str(backup_path)
            logger.info(f"Backup creat: {backup_path.name}")

        wb = openpyxl.load_workbook(str(mf_path))

        # Llegir delay antic (B12 original)
        ws_info = wb['0-INFO']
        original_b12 = ws_info['B12'].value
        result['old_delay'] = original_b12

        # NO sobreescriure B12 — afegir fila "Net delay (Suite)" per audit trail
        # Buscar si ja existeix una fila "Net delay (Suite)"
        suite_row = None
        last_row = ws_info.max_row
        for row in range(1, last_row + 1):
            key = ws_info.cell(row=row, column=1).value
            if key and 'net delay (suite)' in str(key).lower():
                suite_row = row
                break

        if suite_row is None:
            # Afegir noves files al final
            suite_row = last_row + 1
            ws_info.cell(row=suite_row, column=1, value='Net delay (Suite)')

        ws_info.cell(row=suite_row, column=2, value=round(net_delay_min, 4))

        # Llegir timestamps
        hplc_times, hplc_samples = _read_hplc_timestamps(wb)
        toc_rows = _read_toc_timestamps(wb)

        if not hplc_times or not toc_rows:
            logger.warning("No s'han trobat timestamps HPLC o TOC")
            wb.close()
            return result

        # Recalcular 4-TOC_CALC
        if '4-TOC_CALC' in wb.sheetnames:
            del wb['4-TOC_CALC']
        ws_calc = wb.create_sheet('4-TOC_CALC')

        ws_calc['A1'] = 'TOC_Row'
        ws_calc['B1'] = 'Sample'
        ws_calc['C1'] = 'Temps_Relatiu (min)'
        ws_calc['D1'] = 'Inj_Index'

        assignments = _assign_toc_rows(hplc_times, toc_rows, net_delay_min, pre_margin_min)
        result['n_total'] = len(assignments)

        for i, a in enumerate(assignments):
            excel_row = i + 2
            ws_calc.cell(row=excel_row, column=1, value=a['toc_row'])

            inj_idx = a['inj_index']
            if 0 < inj_idx <= len(hplc_samples):
                ws_calc.cell(row=excel_row, column=2, value=hplc_samples[inj_idx - 1])
                if a['temps_rel_min'] is not None:
                    ws_calc.cell(row=excel_row, column=3, value=round(a['temps_rel_min'], 3))
                ws_calc.cell(row=excel_row, column=4, value=inj_idx)
                result['n_assigned'] += 1
            else:
                ws_calc.cell(row=excel_row, column=2, value='')
                ws_calc.cell(row=excel_row, column=4, value=0)

        wb.save(str(mf_path))
        wb.close()

        result['success'] = True
        logger.info(f"MasterFile actualitzat: delay={net_delay_min:.3f}, "
                     f"{result['n_assigned']}/{result['n_total']} files assignades")

    except Exception as e:
        logger.error(f"Error actualitzant MasterFile: {e}")
        result['error'] = str(e)

    return result
