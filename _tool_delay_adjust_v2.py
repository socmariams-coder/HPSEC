"""
Interactive delay adjustment tool v2 — with DAD 254 overlay.
=============================================================
- Full TOC signal with detected peak maxima (triangles)
- Reconstructed DAD 254 continuous signal (from Export3D)
- Peak maxima for both signals marked
- Slider to adjust delay
- Delay = difference between TOC peak and DAD peak positions

Usage:
    python _tool_delay_adjust_v2.py 231
    python _tool_delay_adjust_v2.py 274_SEQ
"""
import sys, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button
from pathlib import Path
from scipy.signal import find_peaks
from scipy.ndimage import uniform_filter1d

from hpsec_config import get_data_folders
from hpsec_consolidate import extract_seq_number, detect_seq_type

FLUSH_TIME_MIN = 3.637


def find_seq_path(arg):
    if os.path.isdir(arg):
        return arg
    folders = get_data_folders()
    for folder in folders:
        full = os.path.join(folder, arg)
        if os.path.isdir(full):
            return full
        try:
            num = int(arg)
            for d in os.listdir(folder):
                if extract_seq_number(d) == num:
                    return os.path.join(folder, d)
        except ValueError:
            pass
    return None


def load_toc(mf_path):
    """Load TOC continuous signal."""
    df = pd.read_excel(mf_path, sheet_name='2-TOC', header=6)
    toc_col = None
    for c in df.columns:
        if 'toc' in str(c).lower():
            toc_col = c
            break
    y = pd.to_numeric(df[toc_col], errors='coerce').values if toc_col else np.array([])
    t = np.arange(len(y)) * (4.0 / 60.0)
    return t, y


def load_hplc_times(mf_path, t0_toc):
    """Load HPLC injection info."""
    df = pd.read_excel(mf_path, sheet_name='1-HPLC-SEQ')
    name_col = None
    date_col = None
    for c in df.columns:
        if name_col is None and 'sample' in str(c).lower() and 'name' in str(c).lower():
            name_col = c
        if date_col is None and 'acquired' in str(c).lower():
            date_col = c

    injs = []
    if name_col and date_col and t0_toc:
        for _, row in df.iterrows():
            try:
                name = str(row[name_col]).strip()
                dt = pd.to_datetime(row[date_col])
                t_min = (dt - t0_toc).total_seconds() / 60.0
                injs.append({
                    'name': name, 't_min': t_min,
                    'is_khp': 'khp' in name.lower(),
                    'is_control': any(x in name.lower() for x in ['mq', 'naoh', 'blanc', 'blnc']),
                })
            except:
                pass
    return injs


def load_dad254_continuous(seq_path, injs):
    """Reconstruct continuous DAD 254 signal from Export3D files.

    Position each injection's DAD signal at its HPLC timestamp.
    Returns (t_dad, y_dad) arrays.
    """
    e3d = os.path.join(seq_path, 'Export3d')
    if not os.path.isdir(e3d):
        return None, None

    files = sorted(os.listdir(e3d))
    if not files:
        return None, None

    # Map injection names to Export3D files
    # Pattern: {name}_R{n}.csv
    segments = []

    for j, inj in enumerate(injs):
        # Find matching Export3D file(s)
        name_clean = inj['name'].replace(' ', '_').replace('-', '_')
        name_lower = name_clean.lower()
        rep_num = 1

        # Count how many of this name we've seen
        same_name_count = sum(1 for k in range(j)
                              if injs[k]['name'] == inj['name'])
        rep_num = same_name_count + 1

        # Try different file patterns
        candidates = []
        for f in files:
            f_lower = f.lower().replace('.csv', '')
            # Match by name prefix + _R{n}
            for name_try in [name_lower, name_lower.replace('_', ''),
                             inj['name'].lower().replace(' ', '')]:
                if f_lower.startswith(name_try) or name_try in f_lower:
                    candidates.append(f)
                    break

        # Filter to the right replica
        matched = None
        for f in candidates:
            if f'_R{rep_num}.' in f or f'_R{rep_num}_' in f:
                matched = f
                break
        if matched is None and candidates:
            # Try just the first match
            if rep_num <= len(candidates):
                matched = candidates[rep_num - 1]

        if matched is None:
            continue

        # Read the Export3D file
        fpath = os.path.join(e3d, matched)
        try:
            # Export3D is UTF-16 with comma-separated, first column is time
            df = pd.read_csv(fpath, encoding='utf-16', sep=None, engine='python',
                             header=0)

            # If single column, it's comma-separated within
            if df.shape[1] == 1:
                # Parse the header to get wavelength columns
                header = df.columns[0]
                wavelengths = header.split(',')
                # Re-read properly
                from io import StringIO
                with open(fpath, encoding='utf-16') as fh:
                    content = fh.read()
                df = pd.read_csv(StringIO(content), header=0)

            # Find time and 254nm columns
            cols = list(df.columns)
            # First column should be time
            if len(cols) < 2:
                continue

            # Parse column names as wavelengths
            # Header format: ",200.00000,202.00000,...,254.00000,..."
            t_col_idx = 0
            col254_idx = None
            for ci, c in enumerate(cols):
                try:
                    wl = float(str(c).strip())
                    if abs(wl - 254.0) < 1.0:
                        col254_idx = ci
                        break
                except ValueError:
                    pass

            if col254_idx is None:
                continue

            # Get data
            time_data = []
            val_data = []
            for _, row in df.iterrows():
                row_str = str(row.iloc[0]).split(',')
                if len(row_str) > col254_idx:
                    try:
                        t_val = float(row_str[0])
                        y_val = float(row_str[col254_idx])
                        time_data.append(t_val)
                        val_data.append(y_val)
                    except ValueError:
                        pass

            if time_data:
                # Position at HPLC injection time (absolute)
                t_offset = inj['t_min']
                t_abs = np.array(time_data) + t_offset
                y_abs = np.array(val_data)
                segments.append((t_abs, y_abs, inj['name'], j))

        except Exception as e:
            pass

    if not segments:
        return None, None

    # Concatenate all segments
    all_t = np.concatenate([s[0] for s in segments])
    all_y = np.concatenate([s[1] for s in segments])

    # Sort by time
    order = np.argsort(all_t)
    return all_t[order], all_y[order], segments


def get_current_delay(mf_path):
    """Read or calculate current delay."""
    df_info = pd.read_excel(mf_path, sheet_name='0-INFO', header=None)
    hora_hplc = None
    hora_toc = None
    delay_explicit = None

    for _, row in df_info.iterrows():
        key = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        val = row.iloc[1] if len(row) > 1 else None
        if 'hora hplc' in key:
            hora_hplc = val
        elif 'hora toc' in key:
            hora_toc = val
        elif 'net delay' in key and 'suite' not in key:
            try:
                v = float(val)
                if not np.isnan(v):
                    delay_explicit = v
            except (ValueError, TypeError):
                pass

    if delay_explicit is not None:
        return delay_explicit

    if hora_hplc and hora_toc:
        try:
            def to_min(t):
                if hasattr(t, 'hour'):
                    return t.hour * 60 + t.minute + t.second / 60
                parts = str(t).split(':')
                return int(parts[0]) * 60 + int(parts[1])
            desfase = to_min(hora_hplc) - to_min(hora_toc)
            return FLUSH_TIME_MIN - desfase
        except:
            pass

    return FLUSH_TIME_MIN


def main():
    if len(sys.argv) < 2:
        print("Usage: python _tool_delay_adjust_v2.py <SEQ>")
        sys.exit(1)

    seq_path = find_seq_path(sys.argv[1])
    if not seq_path:
        print(f"ERROR: No s'ha trobat '{sys.argv[1]}'")
        sys.exit(1)

    seq_name = os.path.basename(seq_path)
    method = detect_seq_type(seq_name)
    print(f"Loading {seq_name} ({method})...")

    # Find MasterFile
    mfs = [f for f in os.listdir(seq_path)
           if 'MasterFile' in f and f.endswith('.xlsx')
           and '~' not in f and 'pre_vol' not in f and 'backup' not in f]
    if not mfs:
        mfs = [f for f in os.listdir(seq_path)
               if 'MasterFile' in f and f.endswith('.xlsx')
               and '~' not in f and 'pre_vol' not in f]
    mf_path = os.path.join(seq_path, mfs[0])

    # Load data
    t_toc, y_toc = load_toc(mf_path)

    # Get TOC t0 for HPLC time alignment
    df_toc_raw = pd.read_excel(mf_path, sheet_name='2-TOC', header=6)
    date_col = [c for c in df_toc_raw.columns if 'started' in str(c).lower()]
    t0_toc = None
    if date_col:
        dates = pd.to_datetime(df_toc_raw[date_col[0]], errors='coerce')
        valid = dates.dropna()
        if len(valid) > 0:
            t0_toc = valid.iloc[0]

    injs = load_hplc_times(mf_path, t0_toc)
    current_delay = get_current_delay(mf_path)

    cadence = np.median(np.diff([inj['t_min'] for inj in injs])) if len(injs) >= 2 else (11.0 if method == 'BP' else 78.0)

    # Smooth TOC
    dt = t_toc[1] - t_toc[0] if len(t_toc) > 1 else 4.0/60.0
    y_clean = np.nan_to_num(y_toc, nan=np.nanmedian(y_toc[~np.isnan(y_toc)]))
    y_smooth = uniform_filter1d(y_clean, size=max(3, int(0.3/dt)))

    # Detect TOC peaks
    baseline = np.percentile(y_smooth, 15)
    if method == 'BP':
        pk_dist = int(6.0 / dt)
    else:
        pk_dist = int(40.0 / dt)
    toc_peaks, _ = find_peaks(y_smooth, distance=pk_dist,
                               prominence=max(5, (y_smooth.max()-baseline)*0.02),
                               height=baseline + 10)

    print(f"TOC: {len(y_toc)} pts, {len(toc_peaks)} pics detectats")

    # Load DAD 254
    dad_result = load_dad254_continuous(seq_path, injs)
    has_dad = dad_result is not None and dad_result[0] is not None
    if has_dad:
        t_dad, y_dad, dad_segments = dad_result
        print(f"DAD 254: {len(t_dad)} pts, {len(dad_segments)} segments")
    else:
        print("DAD 254: no disponible")

    # === FIGURE ===
    fig, axes = plt.subplots(2 if has_dad else 1, 1,
                              figsize=(18, 9 if has_dad else 6),
                              sharex=True)
    plt.subplots_adjust(bottom=0.18, top=0.93, hspace=0.1)

    if not has_dad:
        axes = [axes]
    ax_toc = axes[0]
    ax_dad = axes[1] if has_dad else None

    # --- TOC plot ---
    ax_toc.plot(t_toc, y_clean, 'b-', lw=0.15, alpha=0.3)
    ax_toc.plot(t_toc, y_smooth, 'b-', lw=0.5, alpha=0.8)

    # TOC peak maxima (fixed, don't move with slider)
    ax_toc.plot(t_toc[toc_peaks], y_smooth[toc_peaks], 'rv', ms=5, alpha=0.6,
                label=f'{len(toc_peaks)} pics TOC', zorder=5)
    # Number peaks
    for i, pk in enumerate(toc_peaks):
        ax_toc.annotate(f'{i+1}', (t_toc[pk], y_smooth[pk]+5),
                        fontsize=5, ha='center', color='red', alpha=0.7)

    ax_toc.set_ylabel('TOC (ppb)')
    ax_toc.legend(fontsize=7, loc='upper right')

    # --- DAD plot ---
    if has_dad and ax_dad is not None:
        for t_seg, y_seg, name, j in dad_segments:
            inj = injs[j]
            color = '#E74C3C' if inj['is_khp'] else ('#95A5A6' if inj['is_control'] else '#27AE60')
            ax_dad.plot(t_seg, y_seg, color=color, lw=0.5, alpha=0.6)

        # DAD peak maxima per segment
        for t_seg, y_seg, name, j in dad_segments:
            if len(y_seg) > 5:
                pk_idx = np.argmax(y_seg)
                if y_seg[pk_idx] > 0.1:  # minimum height
                    ax_dad.plot(t_seg[pk_idx], y_seg[pk_idx], '^', color='green',
                               ms=5, alpha=0.7, zorder=5)

        ax_dad.set_ylabel('DAD 254 (mAU)')
        ax_dad.set_xlabel('min (des de inici TOC)')

    if not has_dad:
        ax_toc.set_xlabel('min (des de inici TOC)')

    # --- Injection markers (move with delay) ---
    vlines_toc = []
    vlines_dad = []
    marker_labels = []

    def draw_markers(delay_val):
        for vl in vlines_toc:
            vl.remove()
        vlines_toc.clear()
        for vl in vlines_dad:
            vl.remove()
        vlines_dad.clear()
        for lb in marker_labels:
            lb.remove()
        marker_labels.clear()

        y_top = ax_toc.get_ylim()[1]

        for j, inj in enumerate(injs):
            # TOC expected position = HPLC time + delay
            t_toc_exp = inj['t_min'] + delay_val

            if inj['is_khp']:
                color = '#E74C3C'
                lw = 1.2
            elif inj['is_control']:
                color = '#95A5A6'
                lw = 0.4
            else:
                color = '#2E86AB'
                lw = 0.6

            vl = ax_toc.axvline(t_toc_exp, color=color, ls=':', lw=lw, alpha=0.5)
            vlines_toc.append(vl)

            # Label
            short = inj['name'][:6]
            lb = ax_toc.annotate(f"{j+1}:{short}", (t_toc_exp, y_top * 0.97),
                                fontsize=4, rotation=90, va='top', ha='center',
                                color=color, alpha=0.6)
            marker_labels.append(lb)

            # DAD markers at HPLC time (no delay — DAD is direct)
            if has_dad and ax_dad:
                vl2 = ax_dad.axvline(inj['t_min'], color=color, ls=':', lw=lw, alpha=0.3)
                vlines_dad.append(vl2)

        # Count matches
        n_ok = 0
        n_samples = 0
        for inj in injs:
            if inj['is_control']:
                continue
            n_samples += 1
            t_exp = inj['t_min'] + delay_val
            dists = np.abs(t_toc[toc_peaks] - t_exp) if len(toc_peaks) > 0 else np.array([999])
            if len(dists) > 0 and np.min(dists) < cadence * 0.3:
                n_ok += 1

        pct = n_ok / n_samples * 100 if n_samples > 0 else 0
        fig.suptitle(
            f"{seq_name} ({method}) — delay={delay_val:.2f} min — "
            f"Pics alineats: {n_ok}/{n_samples} ({pct:.0f}%)",
            fontsize=12, fontweight='bold')

    draw_markers(current_delay)

    # --- SLIDER ---
    ax_slider = plt.axes([0.15, 0.06, 0.50, 0.03])
    slider = Slider(ax_slider, 'Delay (min)',
                    current_delay - 30, current_delay + 30,
                    valinit=current_delay, valstep=0.1)

    def on_slider(val):
        draw_markers(val)
        fig.canvas.draw_idle()
    slider.on_changed(on_slider)

    # --- AUTO-DETECT ---
    ax_auto = plt.axes([0.70, 0.06, 0.08, 0.03])
    btn_auto = Button(ax_auto, 'Auto-detect')

    def on_auto(event):
        # Find first TOC peak after baseline
        bl_std = np.std(y_smooth[:int(5.0/dt)])
        thresh = baseline + max(5*bl_std, 10)
        win = int(0.5/dt)
        bl_end = int(5.0/dt)
        for i in range(bl_end, len(y_smooth)-win):
            if np.mean(y_smooth[i:i+win]) > thresh:
                bl_end = i
                break
        first_pk = None
        for pk in toc_peaks:
            if pk >= bl_end:
                first_pk = pk
                break
        if first_pk is None and len(toc_peaks) > 0:
            first_pk = toc_peaks[0]
        if first_pk is not None and injs:
            est = t_toc[first_pk] - injs[0]['t_min']
            slider.set_val(est)
    btn_auto.on_clicked(on_auto)

    # --- SAVE ---
    ax_save = plt.axes([0.80, 0.06, 0.08, 0.03])
    btn_save = Button(ax_save, 'GUARDAR')

    def on_save(event):
        import openpyxl
        delay_val = slider.val
        wb = openpyxl.load_workbook(mf_path)
        ws = wb['0-INFO']
        saved = False
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value and 'net delay' in str(cell.value).lower() and 'suite' not in str(cell.value).lower():
                ws.cell(row=cell.row, column=2, value=round(delay_val, 3))
                saved = True
                break
        if not saved:
            lr = ws.max_row + 1
            ws.cell(row=lr, column=1, value='Net delay (Suite)')
            ws.cell(row=lr, column=2, value=round(delay_val, 3))
        wb.save(mf_path)
        print(f"GUARDAT: delay={delay_val:.3f} min")
        fig.suptitle(fig._suptitle.get_text() + " — GUARDAT!",
                     fontsize=12, fontweight='bold', color='green')
        fig.canvas.draw_idle()
    btn_save.on_clicked(on_save)

    # --- RESET ---
    ax_reset = plt.axes([0.70, 0.02, 0.08, 0.03])
    btn_reset = Button(ax_reset, 'Reset')
    def on_reset(event):
        slider.set_val(current_delay)
    btn_reset.on_clicked(on_reset)

    # Info
    fig.text(0.15, 0.01,
             f"Cadencia: {cadence:.1f} min | Inj: {len(injs)} | "
             f"Delay MF: {current_delay:.2f} min | "
             f"Pics TOC: {len(toc_peaks)}"
             + (f" | DAD: {len(dad_segments)} segments" if has_dad else ""),
             fontsize=8, color='#666')

    print(f"Delay actual: {current_delay:.2f} min")
    print(f"Slider per ajustar | Auto-detect | GUARDAR al MasterFile")
    plt.show()


if __name__ == "__main__":
    main()
