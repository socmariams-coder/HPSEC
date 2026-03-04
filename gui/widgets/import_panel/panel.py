"""
HPSEC Suite - Import Panel v3.0
================================

Panel per a la fase 1: Importació de seqüències.
- Columnes separades per punts i fitxers
- Colors segons tipus de match (EXACT/FUZZY/MANUAL/NONE)
- Dropdown per assignació manual d'orfes
- Verificació obligatòria per FUZZY
"""

import os
import json
from pathlib import Path

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QComboBox, QApplication
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont, QColor, QBrush, QCursor

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))

from hpsec_import import (
    import_sequence, load_manifest, import_from_manifest,
    generate_import_manifest, save_import_manifest,
    llegir_doc_uib, llegir_dad_export3d, llegir_dad_1a,
    get_baseline_value
)
from hpsec_core import downsample_to_cadence
import logging
import numpy as np

logger = logging.getLogger(__name__)

# Importar components del paquet
from .delegates import ComboBoxDelegate, FileAssignmentDelegate
from .worker import ImportWorker
from .dialogs import OrphanFilesDialog, ChromatogramPreviewDialog

# Importar estils compartits
from gui.widgets.styles import (
    PANEL_MARGINS, PANEL_SPACING,
    COLOR_SUCCESS, COLOR_WARNING, COLOR_ERROR,
    apply_panel_layout
)

CONFIG_PATH = Path(__file__).parent.parent.parent.parent / "hpsec_config.json"

# Colors per tipus de match
MATCH_COLORS = {
    "EXACT": QColor("#D5F5E3"),    # Verd clar
    "FUZZY": QColor("#FCF3CF"),    # Groc/taronja clar
    "MANUAL": QColor("#D6EAF8"),   # Blau clar
    "NONE": QColor("#FADBD8"),     # Vermell/rosa clar
    "NORMAL": QColor("#FFFFFF"),   # Blanc
}


class _SortableImportItem(QTableWidgetItem):
    """Item que ordena per UserRole si existeix, sinó per text."""
    def __lt__(self, other):
        my_data = self.data(Qt.UserRole)
        other_data = other.data(Qt.UserRole) if other else None
        if my_data is not None and other_data is not None:
            try:
                return float(my_data) < float(other_data)
            except (TypeError, ValueError):
                pass
        return self.text() < (other.text() if other else "")


def load_sample_types_config():
    """Carrega configuració de tipus de mostra."""
    default_types = {
        "MOSTRA": {"label": "MOSTRA", "color": "#2E86AB", "patterns": []},
        "PATRÓ_CAL": {"label": "PATRÓ_CAL", "color": "#2A9D8F", "patterns": ["KHP"]},
        "PATRÓ_REF": {"label": "PATRÓ_REF", "color": "#9B59B6", "patterns": ["REF", "QC"]},
        "CONTROL": {"label": "CONTROL", "color": "#F6AE2D", "patterns": ["NaOH", "CONTROL"]},
        "BLANC": {"label": "BLANC", "color": "#888888", "patterns": ["MQ", "BLANK", "BLK"]},
    }
    try:
        if CONFIG_PATH.exists():
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config.get("sample_types", default_types)
    except:
        pass
    return default_types


def detect_sample_type(sample_name, original_type, config):
    """Detecta el tipus de mostra basant-se en patrons del config."""
    name_upper = sample_name.upper()

    for type_key, type_info in config.items():
        patterns = type_info.get("patterns", [])
        for pattern in patterns:
            if pattern.upper() in name_upper:
                return type_key

    translations = {
        "SAMPLE": "MOSTRA",
        "KHP": "PATRÓ_CAL",
        "CONTROL": "CONTROL",
        "BLANK": "BLANC",
    }
    return translations.get(original_type.upper(), "MOSTRA")


class ImportPanel(QWidget):
    """Panel d'importació de seqüències."""

    import_completed = Signal(dict)
    warnings_dismissed = Signal()  # Senyal quan s'han descartat els avisos

    # Columnes base (s'ajusten segons mode a _setup_table_columns)
    COL_INJ = 0
    COL_MOSTRA = 1
    COL_TIPUS = 2
    COL_REP = 3
    COL_INJ_VOL = 4  # Volum d'injecció (µL)
    COL_DIRECT_PTS = 5
    COL_DIRECT_FILE = 6
    # Columnes dinàmiques (s'ajusten segons mode de dades)
    # Per DUAL: UIB=7,8, DAD=9,10, SEM_DOC=11, SEM_UIB=12, SEM_DAD=13
    # Per DIRECT: DAD=7,8, SEM_DOC=9, SEM_DAD=10
    COL_UIB_PTS_ACTUAL = 7
    COL_UIB_FILE_ACTUAL = 8
    COL_DAD_PTS_ACTUAL = 9
    COL_DAD_FILE_ACTUAL = 10
    # Semàfors al final (s'ajusten a _setup_table_columns)
    COL_SEM_DOC = 11
    COL_SEM_UIB = 12  # None en mode DIRECT
    COL_SEM_DAD = 13

    # Tipus de mostra que requereixen assignació obligatòria de fitxers
    TYPES_REQUIRE_ASSIGNMENT = {"MOSTRA", "PATRÓ_CAL", "PATRÓ_REF"}
    # Tipus de mostra que permeten assignació opcional
    TYPES_OPTIONAL_ASSIGNMENT = {"CONTROL", "BLANC"}

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.seq_path = None
        self.existing_manifest = None
        self.imported_data = None
        self.worker = None
        self.sample_types_config = load_sample_types_config()
        self._sample_data = []
        self._orphan_files = {"uib": [], "dad": []}
        self._match_types = {}  # (row, col) -> match_type
        self._unverified_fuzzy = set()  # Set of rows needing verification
        self._manual_assignments = {}  # (sample_name, replica) -> {col: filename}
        self._data_mode = "DUAL"  # DUAL, DIRECT, UIB
        self._import_warnings = []  # Warnings d'importació
        self._loaded_from_manifest = False  # Si s'ha carregat des de manifest existent
        self._orphan_warning_dismissed = False  # Si l'usuari ha marcat l'avís d'orfes com revisat
        self._warnings_confirmed = False  # Si l'usuari ha confirmat els warnings (FUZZY, etc.)
        self._warnings_confirmed_by = None  # G05: Qui ha confirmat (traçabilitat)

        self._setup_ui()

    def reset(self):
        """Reinicia el panel al seu estat inicial."""
        self.seq_path = None
        self.existing_manifest = None
        self.imported_data = None
        self.worker = None
        self._sample_data = []
        self._orphan_files = {"uib": [], "dad": []}
        self._match_types = {}
        self._unverified_fuzzy = set()
        self._manual_assignments = {}
        self._data_mode = "DUAL"
        self._import_warnings = []
        self._loaded_from_manifest = False
        self._orphan_warning_dismissed = False
        self._warnings_confirmed = False
        self._warnings_confirmed_by = None

        # Reset UI elements
        self.info_frame.setVisible(False)
        self.table_help.setVisible(False)
        self.samples_table.setRowCount(0)
        self.samples_table.setVisible(False)

        # Mostrar placeholder
        if hasattr(self, 'placeholder'):
            self.placeholder.setVisible(True)

    def _setup_ui(self):
        """Configura la interfície del panel."""
        layout = QVBoxLayout(self)
        apply_panel_layout(layout)

        # === INFO BARRA (resum injeccions) ===
        self.info_frame = QFrame()
        self.info_frame.setVisible(False)
        self.info_frame.setFixedHeight(28)
        info_layout = QHBoxLayout(self.info_frame)
        info_layout.setContentsMargins(0, 2, 0, 2)
        info_layout.setSpacing(16)

        self.total_label = QLabel()
        self.total_label.setStyleSheet("font-weight: bold; color: #2E86AB;")
        info_layout.addWidget(self.total_label)

        # Botó per obrir carpeta SEQ
        self.btn_open_folder = QPushButton("📁 Obrir carpeta")
        self.btn_open_folder.setToolTip("Obrir carpeta de la seqüència (per revisar MasterFile)")
        self.btn_open_folder.setStyleSheet(
            "QPushButton { border: 1px solid #ccc; border-radius: 3px; padding: 2px 8px; "
            "font-size: 11px; color: #555; background: #f8f8f8; }"
            "QPushButton:hover { background: #e8e8e8; color: #333; }"
        )
        self.btn_open_folder.setFixedHeight(22)
        self.btn_open_folder.clicked.connect(self._open_seq_folder)
        info_layout.addWidget(self.btn_open_folder)

        info_layout.addStretch()

        # Comptador de fitxers UIB i DAD
        self.files_label = QLabel()
        self.files_label.setStyleSheet("color: #666;")
        info_layout.addWidget(self.files_label)

        layout.addWidget(self.info_frame)

        # === TAULA DE MOSTRES ===
        # Nota d'ajuda
        self.table_help = QLabel("💡 Doble-clic a una fila per veure la gràfica")
        self.table_help.setStyleSheet("color: #666; font-size: 11px; padding: 2px;")
        self.table_help.setVisible(False)
        layout.addWidget(self.table_help)

        self.samples_table = QTableWidget()
        self.samples_table.setToolTip("Doble-clic per veure gràfica DOC + DAD 254nm")

        # Amagar numeració automàtica de files
        self.samples_table.verticalHeader().setVisible(False)

        # NO usar colors alternats (interfereixen amb colors de match)
        self.samples_table.setAlternatingRowColors(False)

        # Permetre ordenar
        self.samples_table.setSortingEnabled(True)

        self.samples_table.cellDoubleClicked.connect(self._on_cell_double_clicked)
        self.samples_table.cellChanged.connect(self._on_cell_changed)
        self.samples_table.setEditTriggers(
            QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed
        )
        self.samples_table.setVisible(False)

        layout.addWidget(self.samples_table, 1)

        # === PLACEHOLDER ===
        self.placeholder = QLabel("Preparant importació...")
        self.placeholder.setAlignment(Qt.AlignCenter)
        self.placeholder.setStyleSheet("color: #888; font-size: 14px;")
        layout.addWidget(self.placeholder, 1)

    def _get_file_options(self, row, col):
        """Retorna opcions pel dropdown de fitxers."""
        options = ["(cap)"]

        # Determinar si és UIB o DAD segons la columna
        if self._data_mode == "DIRECT":
            if col == self.COL_DAD_FILE_ACTUAL:
                orphans = self._orphan_files.get("dad", [])
            else:
                return options
        else:
            if col == self.COL_UIB_FILE_ACTUAL:
                orphans = self._orphan_files.get("uib", [])
            elif col == self.COL_DAD_FILE_ACTUAL:
                orphans = self._orphan_files.get("dad", [])
            else:
                return options

        # Afegir valor actual si existeix
        current = self.samples_table.item(row, col)
        if current and current.text() and current.text() != "-" and current.text() != "(cap)":
            options.append(current.text())

        # Afegir orfes
        for f in orphans:
            fname = Path(f).name if "/" in f or "\\" in f else f
            if fname not in options:
                options.append(fname)

        return options

    def set_sequence_path(self, path):
        self.seq_path = path
        self.main_window.seq_path = path
        self._check_manifest()

    def load_from_dashboard(self, seq_path):
        """Carrega una seqüència des del Dashboard - auto-carrega si hi ha manifest."""
        self.set_sequence_path(seq_path)

        # Si main_window ja té imported_data (pre-carregat per _preload_completed_stages),
        # reutilitzar-lo directament sense reimportar (evita doble lectura MasterFile)
        if self.main_window.imported_data and self.main_window.imported_data.get("success"):
            logger.info("Reutilitzant imported_data pre-carregat (skip MasterFile re-read)")
            self._display_preloaded_data(self.main_window.imported_data)
            return

        # Si hi ha manifest existent, carregar automàticament (async amb progrés)
        if self.existing_manifest:
            self._auto_load_from_manifest()

    def _display_preloaded_data(self, result):
        """Mostra dades ja carregades sense re-importar ni re-guardar manifest."""
        self.placeholder.setVisible(False)
        self.imported_data = result
        self._loaded_from_manifest = True
        self._orphan_warning_dismissed = result.get("orphan_warning_dismissed", False)
        self._warnings_confirmed = result.get("warnings_confirmed", False)
        self._warnings_confirmed_by = result.get("warnings_confirmed_by", None)

        self._show_results(result)
        self.main_window.enable_tab(1)
        self.main_window.set_status("Importació carregada", 3000)

        self.import_completed.emit({
            'success': True,
            'warnings': result.get('warnings', []),
            'orphan_files': result.get('orphan_files', {}),
            'warnings_confirmed': self._warnings_confirmed,
            'orphan_warning_dismissed': self._orphan_warning_dismissed,
        })

    def _go_to_dashboard(self):
        """Torna a la llista de seqüències."""
        self.main_window.show_dashboard()

    def _check_manifest(self):
        self.existing_manifest = load_manifest(self.seq_path)
        self.samples_table.setVisible(False)
        self.table_help.setVisible(False)
        self.info_frame.setVisible(False)
        self.placeholder.setVisible(True)

    def _auto_load_from_manifest(self):
        """Carrega automàticament des del manifest existent (sense llegir MasterFile)."""
        self._loaded_from_manifest = True
        self.main_window.show_progress(0)
        self.placeholder.setText("Carregant manifest...")
        self.placeholder.setVisible(True)

        # load_data=False: només metadades, NO llegeix MasterFile ni fitxers CSV/DAD
        # Les dades crues es carregaran quan es necessitin (ensure_data_loaded)
        self.worker = ImportWorker(
            self.seq_path,
            use_manifest=True,
            manifest=self.existing_manifest,
            load_data=False
        )
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_import_finished)
        self.worker.error.connect(self._on_import_error)
        self.worker.start()

    def _run_import(self, force_reimport=False):
        """Executa importació. Si force_reimport=True, reimporta tot."""
        if not self.seq_path:
            return

        self.main_window.show_progress(0)

        # Si ja hi ha manifest i no forcem reimportació, usar-lo
        use_manifest = self.existing_manifest and not force_reimport
        self._loaded_from_manifest = use_manifest

        self.worker = ImportWorker(
            self.seq_path,
            use_manifest=use_manifest,
            manifest=self.existing_manifest if use_manifest else None
        )
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_import_finished)
        self.worker.error.connect(self._on_import_error)
        self.worker.start()

    def _on_progress(self, pct, msg):
        self.main_window.show_progress(pct)
        self.main_window.set_status(msg)
        # Actualitzar placeholder amb el pas actual
        if self.placeholder.isVisible():
            self.placeholder.setText(msg)

    def _on_import_finished(self, result):
        self.main_window.show_progress(-1)
        self.placeholder.setVisible(False)

        if not result.get("success"):
            errors = result.get("errors", ["Error desconegut"])
            error_msg = "\n\n".join(errors)
            QMessageBox.critical(self, "Error d'Importació", error_msg)
            # Guardar JSON amb error per persistència entre sessions
            try:
                seq_path = result.get("seq_path") or self.seq_path
                if seq_path:
                    import json, os
                    from datetime import datetime
                    data_folder = os.path.join(seq_path, "CHECK", "data")
                    os.makedirs(data_folder, exist_ok=True)
                    from hpsec_version import SUITE_VERSION
                    error_json = {
                        "suite_version": SUITE_VERSION,
                        "generated_at": datetime.now().isoformat(),
                        "success": False,
                        "errors": errors,
                        "warnings": [],
                    }
                    with open(os.path.join(data_folder, "import_manifest.json"), "w",
                              encoding="utf-8") as f:
                        json.dump(error_json, f, indent=2, ensure_ascii=False)
            except Exception as e:
                logger.warning(f"No s'ha pogut guardar JSON d'error: {e}")
            self.import_completed.emit({'success': False, 'errors': errors})
            return

        # Verificar si cal preguntar la sensibilitat UIB
        self._check_uib_sensitivity(result)

        self.imported_data = result
        self.main_window.imported_data = result
        # Restaurar estat "revisat" d'orfes i warnings si es van marcar anteriorment
        self._orphan_warning_dismissed = result.get("orphan_warning_dismissed", False)
        self._warnings_confirmed = result.get("warnings_confirmed", False)
        self._warnings_confirmed_by = result.get("warnings_confirmed_by", None)  # G05: traçabilitat

        self._show_results(result)

        # Nota: Els avisos es gestionen des del wizard header

        try:
            manifest_path = save_import_manifest(result)
            self.main_window.mark_manifest_saved()
            # Verificar que DAD s'ha guardat al manifest
            n_dad = sum(
                1 for s in result.get("samples", {}).values()
                for r in s.get("replicas", {}).values()
                if r.get("dad") is not None
            )
            logger.info(f"Manifest guardat: {manifest_path} (DAD: {n_dad} rèpliques)")
        except Exception as e:
            logger.warning(f"No s'ha pogut guardar manifest: {e}")

        # Auto-generar PDF d'importació
        try:
            from generate_import_report import generate_import_report
            seq_path = result.get("seq_path") or self.main_window.seq_path
            if seq_path:
                pdf = generate_import_report(seq_path)
                if pdf:
                    logger.info(f"Report importació: {pdf}")
        except ImportError:
            pass
        except Exception as e:
            logger.warning(f"No s'ha pogut generar report d'importació: {e}")

        self.main_window.enable_tab(1)
        self.main_window.set_status("Importació completada", 5000)

        # Emetre senyal per al wizard
        self.import_completed.emit({
            'success': True,
            'warnings': result.get('warnings', []),
            'warnings_structured': result.get('warnings_structured', []),
            'orphan_files': result.get('orphan_files', {}),
            'warnings_confirmed': self._warnings_confirmed,
            'orphan_warning_dismissed': self._orphan_warning_dismissed,
        })

    # Nota: _show_warnings_bar eliminada - avisos es gestionen des del wizard header

    def _check_uib_sensitivity(self, result):
        """Verifica si cal preguntar la sensibilitat UIB i actualitza el MasterFile."""
        # Només per mode DUAL o UIB
        data_mode = result.get("data_mode", "")
        if data_mode not in ["DUAL", "UIB"]:
            return

        # Verificar si la sensibilitat UIB ja ve del backend (MasterFile 0-INFO B5)
        uib_sens = result.get("uib_sensitivity")
        if uib_sens is not None and str(uib_sens).strip() not in ("", "None"):
            logger.debug(f"Sensibilitat UIB del MasterFile: {uib_sens}")
            return

        # Si el manifest existent ja té la sensibilitat (preguntada anteriorment)
        if self.existing_manifest:
            manifest_sens = (self.existing_manifest.get("sequence") or {}).get("uib_sensitivity")
            if manifest_sens is not None and str(manifest_sens).strip() not in ("", "None"):
                result["uib_sensitivity"] = float(manifest_sens)
                logger.debug(f"Sensibilitat UIB del manifest: {manifest_sens}")
                return

        # No s'ha trobat — preguntar a l'usuari
        from PySide6.QtWidgets import QInputDialog
        text, ok = QInputDialog.getText(
            self,
            "Sensibilitat UIB",
            "No s'ha trobat la sensibilitat UIB al MasterFile (0-INFO B5).\n"
            "Indica-la (ex: 700, 1000) o deixa buit si no aplica:",
            text=""
        )

        if ok and text.strip():
            try:
                sens_value = float(text.strip())
            except ValueError:
                logger.warning(f"Valor UIB sensitivity no numèric: {text}")
                return

            result["uib_sensitivity"] = sens_value

            # Escriure al MasterFile per no tornar a preguntar
            master_file = result.get("master_file")
            if master_file and os.path.exists(master_file):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(master_file)
                    if "0-INFO" in wb.sheetnames:
                        ws = wb["0-INFO"]
                        ws["B5"] = sens_value
                        wb.save(master_file)
                        logger.info(f"Sensibilitat UIB escrita al MasterFile: {sens_value}")
                except Exception as e:
                    logger.warning(f"No s'ha pogut actualitzar MasterFile: {e}")

    def _on_import_error(self, error_msg):
        self.main_window.show_progress(-1)
        QMessageBox.critical(self, "Error", f"Error durant la importació:\n{error_msg}")

    # =========================================================================
    # MÈTODES AUXILIARS PER _show_results (descomposició)
    # =========================================================================

    def _init_results_state(self):
        """Inicialitza l'estat per mostrar resultats."""
        self.placeholder.setVisible(False)
        self._match_types = {}
        self._unverified_fuzzy = set()
        self._manual_assignments = {}
        if not self._loaded_from_manifest:
            self._orphan_warning_dismissed = False
            self._warnings_confirmed = False

    def _process_orphan_files(self, manifest, samples, result):
        """Processa i filtra fitxers orfes."""
        # Orfes del manifest (noms) per mostrar - només els no suggerits
        self._orphan_files = {
            "uib": manifest.get("orphan_files", {}).get("uib", []),
            "dad": manifest.get("orphan_files", {}).get("dad", []),
        }
        # TOTS els orfes amb paths complets (per comptar punts, inclou suggerits)
        self._orphan_files_full = {
            "uib": result.get("all_orphan_files", result.get("orphan_files", {})).get("uib", []),
            "dad": result.get("all_orphan_files", result.get("orphan_files", {})).get("dad", []),
        }

        # Si carregat des de manifest, filtrar orfes que ja estan assignats a mostres
        if self._loaded_from_manifest:
            assigned_uib = set()
            assigned_dad = set()
            for sample in samples:
                for rep in sample.get("replicas", []):
                    uib_info = rep.get("uib", {})
                    dad_info = rep.get("dad", {})
                    if uib_info and uib_info.get("file"):
                        assigned_uib.add(Path(uib_info["file"]).name)
                    if dad_info and dad_info.get("file") and not dad_info.get("file", "").startswith("["):
                        assigned_dad.add(Path(dad_info["file"]).name)
            # Treure fitxers assignats de la llista d'orfes
            self._orphan_files["uib"] = [f for f in self._orphan_files["uib"] if Path(f).name not in assigned_uib]
            self._orphan_files["dad"] = [f for f in self._orphan_files["dad"] if Path(f).name not in assigned_dad]

    def _build_injection_list(self, samples):
        """Construeix llista plana d'injeccions ordenada per line_num."""
        all_injections = []
        for sample in samples:
            original_type = sample.get("type", "SAMPLE")
            sample_type = detect_sample_type(
                sample["name"], original_type, self.sample_types_config
            )
            original_name = sample.get("original_name", sample["name"])

            for rep in sample.get("replicas", []):
                inj_info = rep.get("injection_info") or rep.get("injection") or {}
                line_num = inj_info.get("line_num")
                if line_num is None:
                    d = rep.get("direct", {})
                    line_num = d.get("row_start") if d else 999999
                if line_num is None:
                    line_num = 999999

                all_injections.append({
                    "sample_name": sample["name"],
                    "original_name": original_name,
                    "sample_type": sample_type,
                    "rep": rep,
                    "line_num": line_num,
                })

        all_injections.sort(key=lambda x: (x["line_num"], x["sample_name"]))
        return all_injections

    def _update_info_bar(self, result, all_injections):
        """Actualitza la barra d'informació amb resum de la seqüència."""
        total_injections = len(all_injections)
        stats = result.get("stats", {})
        master_line_count = stats.get("master_line_count", result.get("master_line_count", total_injections))
        # Rèpliques reals importades (pot ser < total_injections si Inj# duplicats)
        replicas_imported = stats.get("total_replicas_imported", total_injections)

        # Recollir volums d'injecció
        volumes = []
        for inj in all_injections:
            rep = inj.get("rep", {})
            inj_info = rep.get("injection_info") or rep.get("injection") or {}
            vol = inj_info.get("inj_volume")
            if vol is not None:
                volumes.append(vol)

        # Construir resum
        method = result.get("method", "COLUMN")
        info_parts = []

        # Injeccions: sempre mostrar importades vs total MasterFile
        if master_line_count > replicas_imported:
            info_parts.append(f"⚠️ Importades: {replicas_imported}/{master_line_count}")
            has_warning = True
        elif master_line_count > total_injections:
            info_parts.append(f"⚠️ Importades: {total_injections}/{master_line_count}")
            has_warning = True
        else:
            info_parts.append(f"Importades: {total_injections}/{master_line_count}")
            has_warning = False

        info_parts.append(method)
        info_parts.append(self._data_mode)

        # Volum d'injecció
        if volumes:
            vol_min, vol_max = min(volumes), max(volumes)
            if vol_min == vol_max:
                info_parts.append(f"{int(vol_min)}µL")
            else:
                info_parts.append(f"{int(vol_min)}-{int(vol_max)}µL")

        # Sensibilitat UIB
        if self._data_mode in ["DUAL", "UIB"]:
            uib_sens = result.get("uib_sensitivity")
            if uib_sens is not None:
                try:
                    sens_val = int(float(uib_sens))
                    info_parts.append(f"UIB:{sens_val}ppb")
                except (ValueError, TypeError):
                    info_parts.append(f"UIB:{uib_sens}")

        self.total_label.setText(" · ".join(info_parts))
        self.total_label.setStyleSheet(
            "font-weight: bold; color: #E74C3C;" if has_warning else "font-weight: bold; color: #2E86AB;"
        )

        # Comptar fitxers per tipus de senyal
        n_total = stats.get("total_replicas_imported", total_injections)
        doc_direct = stats.get("doc_direct_count", 0)
        uib_count = stats.get("uib_count", 0)
        dad_count = stats.get("dad_count", 0)

        files_parts = []
        files_parts.append(f"DOC: {doc_direct}/{n_total}")
        if self._data_mode in ["DUAL", "UIB"]:
            files_parts.append(f"UIB: {uib_count}/{n_total}")
        files_parts.append(f"DAD: {dad_count}/{n_total}")
        self.files_label.setText(" · ".join(files_parts))
        self.info_frame.setVisible(True)

        # Warning destacat si injeccions incompletes (rèpliques perdudes o línies faltants)
        if master_line_count > replicas_imported:
            missing = master_line_count - replicas_imported
            self._import_warnings.append(
                f"INCOMPLETA: {replicas_imported}/{master_line_count} injeccions. "
                f"Falten {missing} (possibles Inj# duplicats al MasterFile). "
                f"Obrir carpeta i corregir Sample_Rep."
            )

    def _open_seq_folder(self):
        """Obre la carpeta de la seqüència a l'explorador de fitxers."""
        import subprocess
        seq_path = self.seq_path or (
            self.main_window.imported_data.get("seq_path")
            if self.main_window.imported_data else None
        )

        # Si no tenim path absolut, reconstruir des de config + nom SEQ
        if not seq_path or not os.path.isdir(seq_path):
            seq_name = os.path.basename(seq_path or "")
            if not seq_name and self.imported_data:
                seq_name = self.imported_data.get('seq_name', '')
            if seq_name:
                try:
                    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    paths = config.get("paths", {})
                    folders = paths.get("data_folders", [])
                    if not folders:
                        single = paths.get("data_folder", "")
                        if single:
                            folders = [single]
                    for data_folder in folders:
                        alt_path = os.path.join(data_folder, seq_name)
                        if os.path.isdir(alt_path):
                            seq_path = alt_path
                            break
                except Exception:
                    pass

        if seq_path and os.path.isdir(seq_path):
            norm_path = os.path.normpath(seq_path)
            subprocess.Popen(f'explorer "{norm_path}"')
        else:
            QMessageBox.information(self, "Info",
                f"No s'ha trobat la carpeta de la seqüència.\n\nPath: {seq_path}")

    def _populate_row_basic(self, row, injection_num, inj):
        """Omple les columnes bàsiques d'una fila (Inj, Mostra, Tipus, Rep, Vol, Direct)."""
        sample_name = inj["sample_name"]
        original_name = inj.get("original_name", sample_name)
        sample_type = inj["sample_type"]
        rep = inj["rep"]

        # Inj (SortableImportItem per ordenació numèrica correcta)
        inj_item = _SortableImportItem(str(injection_num))
        inj_item.setData(Qt.UserRole, injection_num)
        inj_item.setTextAlignment(Qt.AlignCenter)
        inj_item.setFlags(inj_item.flags() & ~Qt.ItemIsEditable)
        self.samples_table.setItem(row, self.COL_INJ, inj_item)

        # Mostra
        if original_name != sample_name:
            name_item = QTableWidgetItem(original_name)
            name_item.setToolTip(f"Nom únic: {sample_name}\nNom MasterFile: {original_name}")
            name_item.setForeground(QBrush(QColor("#2E86AB")))
        else:
            name_item = QTableWidgetItem(sample_name)
        name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
        name_item.setData(Qt.UserRole, sample_name)
        self.samples_table.setItem(row, self.COL_MOSTRA, name_item)

        # Tipus
        type_item = QTableWidgetItem(sample_type)
        type_item.setTextAlignment(Qt.AlignCenter)
        self.samples_table.setItem(row, self.COL_TIPUS, type_item)

        # Rep
        rep_item = QTableWidgetItem(str(rep.get("replica", "?")))
        rep_item.setTextAlignment(Qt.AlignCenter)
        rep_item.setFlags(rep_item.flags() & ~Qt.ItemIsEditable)
        self.samples_table.setItem(row, self.COL_REP, rep_item)

        # Volum
        inj_info = rep.get("injection_info") or rep.get("injection") or {}
        inj_vol = inj_info.get("inj_volume")
        vol_text = f"{int(inj_vol)}" if inj_vol else "-"
        vol_item = QTableWidgetItem(vol_text)
        vol_item.setTextAlignment(Qt.AlignCenter)
        vol_item.setFlags(vol_item.flags() & ~Qt.ItemIsEditable)
        self.samples_table.setItem(row, self.COL_INJ_VOL, vol_item)

        # Direct
        d = rep.get("direct", {})
        direct_pts = d.get("n_points", 0) if d else 0
        row_start = d.get("row_start", "") if d else ""
        row_end = d.get("row_end", "") if d else ""
        direct_file = f"{row_start}-{row_end}" if row_start and row_end else "-"

        self._add_simple_cell(row, self.COL_DIRECT_PTS, str(direct_pts) if direct_pts else "-")
        self._add_simple_cell(row, self.COL_DIRECT_FILE, direct_file)

        return direct_pts

    def _populate_row_uib(self, row, rep, sample_name, sample_type, requires_assignment, optional_can_assign):
        """Omple les columnes UIB d'una fila. Retorna (review_signals, missing_signals, needs_review)."""
        review_signals = []
        missing_signals = []
        needs_review = False

        u = rep.get("uib", {})
        uib_pts = u.get("n_points", 0) if u else 0
        uib_file = u.get("file", "") if u else ""
        uib_suggestion = rep.get("uib_suggestion")

        if uib_file:
            uib_file = Path(uib_file).name

        self._add_simple_cell(row, self.COL_UIB_PTS_ACTUAL, str(uib_pts) if uib_pts else "-")

        if uib_suggestion and (requires_assignment or optional_can_assign):
            suggested_file = uib_suggestion.get("file", "")
            confidence = uib_suggestion.get("confidence", 0)
            suggestion_status = uib_suggestion.get("status", "SUGGESTED")
            replica_num = rep.get("replica", 1)
            display_name = f"{sample_name}_R{replica_num}"

            # Mostrar cel·la de suggeriment si no està confirmat, independentment de la font
            if suggestion_status == "CONFIRMED" or self._warnings_confirmed:
                self._add_simple_cell(row, self.COL_UIB_FILE_ACTUAL, display_name)
            else:
                self._add_suggestion_cell(row, self.COL_UIB_FILE_ACTUAL, suggested_file, confidence, display_name)
                if requires_assignment:
                    review_signals.append(f"UIB {int(confidence)}%")
                needs_review = True

            n_points = self._count_file_points(suggested_file, "uib")
            if n_points > 0:
                self.samples_table.item(row, self.COL_UIB_PTS_ACTUAL).setText(str(n_points))

        elif uib_file:
            # Fitxer ja assignat (des de manifest o durant importació)
            replica_num = rep.get("replica", 1)
            display_name = f"{sample_name}_R{replica_num}"
            self._add_simple_cell(row, self.COL_UIB_FILE_ACTUAL, display_name)
            if not uib_pts:
                n_points = self._count_file_points(uib_file, "uib")
                if n_points > 0:
                    self.samples_table.item(row, self.COL_UIB_PTS_ACTUAL).setText(str(n_points))

        elif not uib_pts and self._orphan_files.get("uib") and requires_assignment:
            self._add_file_cell(row, self.COL_UIB_FILE_ACTUAL, "-", editable=True)
            if not self._warnings_confirmed:
                missing_signals.append("UIB")
                needs_review = True

        elif optional_can_assign and self._orphan_files.get("uib"):
            display_val = uib_file if uib_file else "-"
            self._add_file_cell(row, self.COL_UIB_FILE_ACTUAL, display_val, editable=True)

        else:
            self._add_simple_cell(row, self.COL_UIB_FILE_ACTUAL, uib_file if uib_file else "-")

        return review_signals, missing_signals, needs_review, uib_pts

    def _populate_row_dad(self, row, rep, sample_name, sample_type, requires_assignment, optional_can_assign):
        """Omple les columnes DAD d'una fila. Retorna (review_signals, missing_signals, needs_review)."""
        review_signals = []
        missing_signals = []
        needs_review = False

        dad = rep.get("dad", {})
        dad_pts = dad.get("n_points", 0) if dad else 0
        dad_suggestion = rep.get("dad_suggestion")

        # Obtenir el fitxer DAD
        dad_file = ""
        if dad:
            dad_file = dad.get("file", "")
            if not dad_file and dad_pts > 0:
                source = dad.get("source", "")
                if source == "masterfile":
                    dad_file = "[MasterFile]"
                elif source in ["export3d", "csv"]:
                    dad_file = f"[{source}]"
            elif dad_file:
                dad_file = Path(dad_file).name if "/" in dad_file or "\\" in dad_file else dad_file

        self._add_simple_cell(row, self.COL_DAD_PTS_ACTUAL, str(dad_pts) if dad_pts else "-")

        if dad_suggestion and (requires_assignment or optional_can_assign):
            suggested_file = dad_suggestion.get("file", "")
            confidence = dad_suggestion.get("confidence", 0)
            suggestion_status = dad_suggestion.get("status", "SUGGESTED")
            replica_num = rep.get("replica", 1)
            display_name = f"{sample_name}_R{replica_num}"

            # Mostrar cel·la de suggeriment si no està confirmat, independentment de la font
            if suggestion_status == "CONFIRMED" or self._warnings_confirmed:
                self._add_simple_cell(row, self.COL_DAD_FILE_ACTUAL, display_name)
            else:
                self._add_suggestion_cell(row, self.COL_DAD_FILE_ACTUAL, suggested_file, confidence, display_name)
                if requires_assignment:
                    review_signals.append(f"DAD {int(confidence)}%")
                needs_review = True

            n_points = self._count_file_points(suggested_file, "dad")
            if n_points > 0:
                self.samples_table.item(row, self.COL_DAD_PTS_ACTUAL).setText(str(n_points))
                dad_pts = n_points

        elif dad_file and not dad_file.startswith("["):
            replica_num = rep.get("replica", 1)
            display_name = f"{sample_name}_R{replica_num}"
            self._add_simple_cell(row, self.COL_DAD_FILE_ACTUAL, display_name)
            if not dad_pts:
                n_points = self._count_file_points(dad_file, "dad")
                if n_points > 0:
                    self.samples_table.item(row, self.COL_DAD_PTS_ACTUAL).setText(str(n_points))

        elif not dad_pts and not dad_file and self._orphan_files.get("dad") and requires_assignment:
            self._add_file_cell(row, self.COL_DAD_FILE_ACTUAL, "-", editable=True)
            if not self._warnings_confirmed:
                missing_signals.append("DAD")
                needs_review = True

        elif optional_can_assign and self._orphan_files.get("dad"):
            display_val = dad_file if dad_file else "-"
            self._add_file_cell(row, self.COL_DAD_FILE_ACTUAL, display_val, editable=True)

        else:
            self._add_simple_cell(row, self.COL_DAD_FILE_ACTUAL, dad_file if dad_file else "-")

        return review_signals, missing_signals, needs_review, dad_pts

    # Símbols semàfor: forma diferent per cada estat (accessibilitat daltonisme)
    _SEM_SHAPES = {
        "#27AE60": "\u2714",  # ✔ tick verd = OK
        "#F39C12": "\u25B2",  # ▲ triangle groc = revisar
        "#E74C3C": "\u2716",  # ✖ creu vermella = falta
        "#BDC3C7": "\u25CB",  # ○ cercle buit gris = sense dades
    }

    # Prioritat per ordenació (menor = més atenció)
    _SEM_PRIORITY = {
        "#E74C3C": 0,  # Vermell = falta → primer
        "#F39C12": 1,  # Groc = revisar
        "#BDC3C7": 2,  # Gris = sense dades
        "#27AE60": 3,  # Verd = OK → últim
    }

    def _create_semaphore_item(self, color, tooltip=""):
        """Crea un item semàfor amb forma + color per a la taula."""
        symbol = self._SEM_SHAPES.get(color, "\u25CF")
        item = _SortableImportItem(symbol)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        item.setForeground(QBrush(QColor(color)))
        font = item.font()
        font.setPointSize(14)
        item.setFont(font)
        # Guardar prioritat per ordenació
        priority = self._SEM_PRIORITY.get(color, 3)
        item.setData(Qt.UserRole, priority)
        if tooltip:
            item.setToolTip(tooltip)
        return item

    def _populate_row_estat(self, row, review_signals, missing_signals, needs_review,
                            direct_pts=0, uib_pts=0, dad_pts=0, sample_type="MOSTRA"):
        """Omple les columnes semàfor (DOC, UIB, DAD) d'una fila."""
        # Colors semàfor
        GREEN = "#27AE60"
        YELLOW = "#F39C12"
        RED = "#E74C3C"
        GREY = "#BDC3C7"

        uib_review = [s for s in review_signals if "UIB" in s]
        dad_review = [s for s in review_signals if "DAD" in s]
        uib_missing = "UIB" in missing_signals
        dad_missing = "DAD" in missing_signals

        is_optional_type = sample_type in self.TYPES_OPTIONAL_ASSIGNMENT

        # --- Semàfor DOC ---
        if direct_pts > 0:
            doc_item = self._create_semaphore_item(GREEN, f"DOC Direct: {direct_pts} punts")
        else:
            doc_item = self._create_semaphore_item(RED, "DOC Direct: sense dades")
        self.samples_table.setItem(row, self.COL_SEM_DOC, doc_item)

        # --- Semàfor UIB ---
        if self.COL_SEM_UIB is not None:
            if uib_review:
                uib_item = self._create_semaphore_item(YELLOW, "UIB: revisar assignació")
            elif uib_missing:
                uib_item = self._create_semaphore_item(RED, "UIB: falta fitxer")
            elif uib_pts > 0:
                uib_item = self._create_semaphore_item(GREEN, f"UIB: {uib_pts} punts")
            elif is_optional_type:
                uib_item = self._create_semaphore_item(GREY, "UIB: opcional")
            else:
                uib_item = self._create_semaphore_item(GREY, "UIB: sense dades")
            self.samples_table.setItem(row, self.COL_SEM_UIB, uib_item)

        # --- Semàfor DAD ---
        if dad_review:
            dad_item = self._create_semaphore_item(YELLOW, "DAD: revisar assignació")
        elif dad_missing:
            dad_item = self._create_semaphore_item(RED, "DAD: falta fitxer")
        elif dad_pts > 0:
            dad_item = self._create_semaphore_item(GREEN, f"DAD: {dad_pts} punts")
        elif is_optional_type:
            dad_item = self._create_semaphore_item(GREY, "DAD: opcional")
        else:
            dad_item = self._create_semaphore_item(GREY, "DAD: sense dades")
        self.samples_table.setItem(row, self.COL_SEM_DAD, dad_item)

        # Prioritat combinada al semàfor DOC (per ordenació per defecte)
        # Menor valor = més atenció necessària
        priorities = [doc_item.data(Qt.UserRole), dad_item.data(Qt.UserRole)]
        if self.COL_SEM_UIB is not None:
            uib_item_cur = self.samples_table.item(row, self.COL_SEM_UIB)
            if uib_item_cur:
                priorities.append(uib_item_cur.data(Qt.UserRole))
        worst = min(p for p in priorities if p is not None)
        doc_item.setData(Qt.UserRole, worst)

        if needs_review:
            self._unverified_fuzzy.add(row)

    # =========================================================================
    # _show_results PRINCIPAL (refactoritzat)
    # =========================================================================

    def _show_results(self, result):
        """Mostra els resultats d'importació a la taula."""
        # Inicialitzar estat
        self._init_results_state()

        # Guardar warnings d'importació per mostrar-los
        self._import_warnings = [w for w in result.get("warnings", []) if "⚠️" in w]

        # Processar manifest
        manifest = generate_import_manifest(result)
        samples = manifest.get("samples", [])
        seq_info = manifest.get("sequence", {})
        self._data_mode = seq_info.get("data_mode", "DUAL")

        # Processar fitxers orfes
        self._process_orphan_files(manifest, samples, result)

        # Configurar columnes segons mode
        self._setup_table_columns()

        # Construir llista d'injeccions ordenada
        all_injections = self._build_injection_list(samples)

        # Actualitzar barra d'informació
        self._update_info_bar(result, all_injections)

        # Omplir taula
        self._populate_table(all_injections)

        # Warnings
        self._update_warnings()

    def _populate_table(self, all_injections):
        """Omple la taula amb les injeccions."""
        self.samples_table.setSortingEnabled(False)
        self.samples_table.blockSignals(True)
        self.samples_table.setRowCount(0)
        self._sample_data = []

        for injection_num, inj in enumerate(all_injections, 1):
            row = self.samples_table.rowCount()
            self.samples_table.insertRow(row)

            sample_name = inj["sample_name"]
            sample_type = inj["sample_type"]
            rep = inj["rep"]

            # Columnes bàsiques (Inj, Mostra, Tipus, Rep, Vol, Direct)
            direct_pts = self._populate_row_basic(row, injection_num, inj)

            # Determinar si requereix assignació
            requires_assignment = sample_type in self.TYPES_REQUIRE_ASSIGNMENT
            optional_can_assign = sample_type in self.TYPES_OPTIONAL_ASSIGNMENT

            # Acumuladors per estat
            all_review_signals = []
            all_missing_signals = []
            needs_review = False
            uib_pts = 0
            dad_pts = 0

            # UIB (només si mode DUAL o UIB)
            if self._data_mode in ["DUAL", "UIB"]:
                review_uib, missing_uib, review_uib_flag, uib_pts = self._populate_row_uib(
                    row, rep, sample_name, sample_type, requires_assignment, optional_can_assign
                )
                all_review_signals.extend(review_uib)
                all_missing_signals.extend(missing_uib)
                needs_review = needs_review or review_uib_flag

            # DAD
            review_dad, missing_dad, review_dad_flag, dad_pts = self._populate_row_dad(
                row, rep, sample_name, sample_type, requires_assignment, optional_can_assign
            )
            all_review_signals.extend(review_dad)
            all_missing_signals.extend(missing_dad)
            needs_review = needs_review or review_dad_flag

            # Semàfors
            self._populate_row_estat(
                row, all_review_signals, all_missing_signals, needs_review,
                direct_pts=direct_pts, uib_pts=uib_pts, dad_pts=dad_pts,
                sample_type=sample_type
            )

            # Guardar per preview i lògica
            u = rep.get("uib", {})
            self._sample_data.append({
                "name": sample_name,
                "type": sample_type,
                "replica": rep.get("replica"),
                "direct_pts": direct_pts,
                "uib_pts": u.get("n_points", 0) if u else uib_pts,
                "dad_pts": dad_pts,
            })

        self.samples_table.blockSignals(False)
        self.samples_table.setSortingEnabled(True)
        # Per defecte, ordenar per atenció requerida (semàfor DOC conté prioritat combinada)
        self.samples_table.sortByColumn(self.COL_SEM_DOC, Qt.AscendingOrder)
        self.samples_table.setVisible(True)
        self.table_help.setVisible(True)

    def _add_simple_cell(self, row, col, text):
        """Afegeix una cel·la simple no editable."""
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        self.samples_table.setItem(row, col, item)

    def _add_file_cell(self, row, col, text, editable=False):
        """Afegeix una cel·la de fitxer, potencialment editable."""
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        else:
            item.setBackground(QBrush(QColor("#FADBD8")))  # Rosa per indicar que cal assignar
        self.samples_table.setItem(row, col, item)

    def _add_suggestion_cell(self, row, col, filename, confidence, display_name=None):
        """Afegeix una cel·la amb suggeriment de matching (editable per confirmar/canviar).

        Args:
            row: Fila de la taula
            col: Columna de la taula
            filename: Nom real del fitxer orfe
            confidence: Percentatge de confiança del match
            display_name: Nom a mostrar (format llistat injeccions). Si None, usa filename.
        """
        show_name = display_name if display_name else filename
        item = QTableWidgetItem(show_name)
        item.setTextAlignment(Qt.AlignCenter)
        # Editable per permetre canviar el suggeriment
        item.setBackground(QBrush(QColor("#FCF3CF")))  # Groc per indicar revisar
        # Tooltip mostra el nom real del fitxer orfe
        item.setToolTip(f"Fitxer: {filename}\nConfiança: {confidence:.0f}%\nDoble-clic per canviar.")
        # Guardar el path real del fitxer com a data
        item.setData(Qt.UserRole, filename)
        self.samples_table.setItem(row, col, item)
        # Marcar com a suggeriment
        self._match_types[(row, col)] = "SUGGESTED"

    def _setup_table_columns(self):
        """Configura les columnes segons el mode de dades."""
        if self._data_mode == "DIRECT":
            # Sense UIB
            self.samples_table.setColumnCount(11)
            headers = ["Inj", "Mostra", "Tipus", "Rep", "Vol (µL)", "Pts DOC", "Fila TOC", "Pts DAD", "Fitxer DAD", "DOC", "DAD"]
            self.COL_DAD_PTS_ACTUAL = 7
            self.COL_DAD_FILE_ACTUAL = 8
            self.COL_SEM_DOC = 9
            self.COL_SEM_UIB = None  # No UIB en mode DIRECT
            self.COL_SEM_DAD = 10
        else:
            # DUAL o UIB
            self.samples_table.setColumnCount(14)
            headers = ["Inj", "Mostra", "Tipus", "Rep", "Vol (µL)", "Pts DOC", "Fila TOC", "Pts UIB", "Fitxer UIB", "Pts DAD", "Fitxer DAD", "DOC", "UIB", "DAD"]
            self.COL_UIB_PTS_ACTUAL = 7
            self.COL_UIB_FILE_ACTUAL = 8
            self.COL_DAD_PTS_ACTUAL = 9
            self.COL_DAD_FILE_ACTUAL = 10
            self.COL_SEM_DOC = 11
            self.COL_SEM_UIB = 12
            self.COL_SEM_DAD = 13

        self.samples_table.setHorizontalHeaderLabels(headers)

        # Tooltips capçaleres
        header_tooltips = {
            self.COL_INJ: "Número d'injecció al MasterFile",
            self.COL_MOSTRA: "Nom de la mostra (normalitzat)",
            self.COL_TIPUS: "Tipus: MOSTRA, KHP, BLANC, NETEJA, PR. Doble-clic per editar",
            self.COL_REP: "Número de rèplica",
            self.COL_INJ_VOL: "Volum d'injecció (µL). Font: capçalera MasterFile o heurístic col-13",
            self.COL_DIRECT_PTS: "Punts del cromatograma DOC Direct",
            self.COL_DIRECT_FILE: "Rang de files TOC assignades (row_start – row_end)",
        }
        if self._data_mode != "DIRECT":
            header_tooltips[self.COL_UIB_PTS_ACTUAL] = "Punts UIB (després de downsample a cadència DOC)"
            header_tooltips[self.COL_UIB_FILE_ACTUAL] = "Fitxer UIB assignat. Doble-clic per canviar"
        header_tooltips[self.COL_DAD_PTS_ACTUAL] = "Punts DAD (Export3D o CSV)"
        header_tooltips[self.COL_DAD_FILE_ACTUAL] = "Fitxer DAD assignat. Doble-clic per canviar"
        header_tooltips[self.COL_SEM_DOC] = "Estat DOC: ✔ dades OK, ✖ sense dades"
        if self.COL_SEM_UIB is not None:
            header_tooltips[self.COL_SEM_UIB] = "Estat UIB: ✔ dades OK, ▲ revisar, ✖ falta fitxer"
        header_tooltips[self.COL_SEM_DAD] = "Estat DAD: ✔ dades OK, ▲ revisar, ✖ falta fitxer"

        h = self.samples_table.horizontalHeader()
        for col_idx, tip in header_tooltips.items():
            if col_idx is not None:
                # QHeaderView tooltips via model
                self.samples_table.horizontalHeaderItem(col_idx).setToolTip(tip)

        # Configurar delegates
        type_delegate = ComboBoxDelegate(list(self.sample_types_config.keys()), self)
        self.samples_table.setItemDelegateForColumn(self.COL_TIPUS, type_delegate)

        if self._data_mode == "DIRECT":
            dad_delegate = FileAssignmentDelegate(self._get_file_options, self)
            self.samples_table.setItemDelegateForColumn(self.COL_DAD_FILE_ACTUAL, dad_delegate)
        else:
            uib_delegate = FileAssignmentDelegate(self._get_file_options, self)
            dad_delegate = FileAssignmentDelegate(self._get_file_options, self)
            self.samples_table.setItemDelegateForColumn(self.COL_UIB_FILE_ACTUAL, uib_delegate)
            self.samples_table.setItemDelegateForColumn(self.COL_DAD_FILE_ACTUAL, dad_delegate)

        # Configurar mides
        header = self.samples_table.horizontalHeader()
        for col in range(self.samples_table.columnCount()):
            if col == COL_MOSTRA:
                header.setSectionResizeMode(col, QHeaderView.Stretch)
            else:
                header.setSectionResizeMode(col, QHeaderView.ResizeToContents)

        # Semàfors amb amplada fixa compacta
        sem_cols = [self.COL_SEM_DOC, self.COL_SEM_UIB, self.COL_SEM_DAD]
        for col in sem_cols:
            if col is not None:
                header.setSectionResizeMode(col, QHeaderView.Fixed)
                self.samples_table.setColumnWidth(col, 50)

    def _add_data_cell(self, row, col, text, match_type, editable=False):
        """Afegeix una cel·la amb color segons match_type."""
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)

        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)

        # Color de fons
        color = MATCH_COLORS.get(match_type, MATCH_COLORS["NORMAL"])
        item.setBackground(QBrush(color))

        # Guardar match type
        self._match_types[(row, col)] = match_type
        if match_type == "FUZZY":
            self._unverified_fuzzy.add((row, col))

        self.samples_table.setItem(row, col, item)

    def _on_cell_changed(self, row, col):
        """Handler quan canvia una cel·la."""
        # Determinar si és una columna de fitxer
        is_file_col = False
        pts_col = None
        file_type = None

        if self._data_mode == "DIRECT":
            if col == self.COL_DAD_FILE_ACTUAL:
                is_file_col = True
                pts_col = self.COL_DAD_PTS_ACTUAL
                file_type = "dad"
        else:
            if col == self.COL_UIB_FILE_ACTUAL:
                is_file_col = True
                pts_col = self.COL_UIB_PTS_ACTUAL
                file_type = "uib"
            elif col == self.COL_DAD_FILE_ACTUAL:
                is_file_col = True
                pts_col = self.COL_DAD_PTS_ACTUAL
                file_type = "dad"

        if is_file_col:
            # Marcar com a MANUAL i actualitzar color
            item = self.samples_table.item(row, col)
            if item:
                new_value = item.text()
                # Obtenir nom real del fitxer (pot ser diferent del display name)
                actual_filename = item.data(Qt.UserRole)
                if not actual_filename:
                    actual_filename = new_value  # Fallback al text visible

                # Obtenir nom i rèplica de la taula (funciona amb taula ordenada)
                name_item = self.samples_table.item(row, self.COL_MOSTRA)
                rep_item = self.samples_table.item(row, self.COL_REP)
                sample_name = name_item.data(Qt.UserRole) if name_item else ""
                try:
                    replica = int(rep_item.text()) if rep_item else 1
                except:
                    replica = 1

                # Si l'usuari ha seleccionat un fitxer del dropdown (nom real),
                # convertir a format llistat d'injeccions
                if new_value and new_value not in ["-", "(cap)"]:
                    # Si el text és un nom de fitxer (conté extensió o és diferent del format esperat)
                    # convertir-lo al format del llistat d'injeccions
                    expected_display = f"{sample_name}_R{replica}"
                    if new_value != expected_display and (
                        "." in new_value or "_R" not in new_value or not new_value.startswith(sample_name)
                    ):
                        # Guardar el fitxer real seleccionat
                        actual_filename = new_value
                        item.setData(Qt.UserRole, actual_filename)
                        # Mostrar el nom segons llistat d'injeccions
                        self.samples_table.blockSignals(True)
                        item.setText(expected_display)
                        item.setToolTip(f"Fitxer: {actual_filename}")
                        self.samples_table.blockSignals(False)

                    # Guardar assignació manual per al manifest (amb el path real)
                    # Usar clau (sample_name, replica) en lloc de row per suportar taula ordenada
                    if not hasattr(self, '_manual_assignments'):
                        self._manual_assignments = {}
                    key = (sample_name, replica)
                    self._manual_assignments.setdefault(key, {})[col] = actual_filename
                    # Marcar que hi ha canvis sense guardar
                    self.main_window.mark_unsaved_changes()
                    # Si era un suggeriment, marcar com CONFIRMED (verd)
                    # Sinó, marcar com MANUAL (blau)
                    prev_type = self._match_types.get((row, col), "")
                    if prev_type == "SUGGESTED":
                        item.setBackground(QBrush(MATCH_COLORS["EXACT"]))  # Verd clar = confirmat
                        self._match_types[(row, col)] = "CONFIRMED"
                    elif prev_type not in ["CONFIRMED", "EXACT"]:
                        item.setBackground(QBrush(MATCH_COLORS["MANUAL"]))  # Blau = manual
                        self._match_types[(row, col)] = "MANUAL"

                    # Actualitzar nombre de punts (usant el nom real del fitxer)
                    if pts_col is not None:
                        n_points = self._count_file_points(actual_filename, file_type)
                        pts_item = self.samples_table.item(row, pts_col)
                        if pts_item and n_points > 0:
                            pts_item.setText(str(n_points))
                            # Actualitzar _sample_data (buscar per nom/replica)
                            name_item = self.samples_table.item(row, self.COL_MOSTRA)
                            rep_item = self.samples_table.item(row, self.COL_REP)
                            if name_item and rep_item:
                                s_name = name_item.data(Qt.UserRole)
                                s_rep = rep_item.text()
                                for data in self._sample_data:
                                    if (data.get("name") == s_name and
                                        str(data.get("replica", "")) == s_rep):
                                        if file_type == "uib":
                                            data["uib_pts"] = n_points
                                        elif file_type == "dad":
                                            data["dad_pts"] = n_points
                                        break

                    # Carregar dades del fitxer per a la gràfica
                    self._load_and_store_file_data(actual_filename, file_type, sample_name, replica)
                    # Auto-guardar manifest per persistir assignació (I10)
                    try:
                        save_import_manifest(self.imported_data)
                        self.main_window.mark_manifest_saved()
                    except Exception as e:
                        logger.warning(f"No s'ha pogut auto-guardar manifest: {e}")
                else:
                    # Si es tria "(cap)", la cel·la queda sense assignació
                    item.setBackground(QBrush(MATCH_COLORS["NORMAL"]))
                    self._match_types[(row, col)] = "NONE"
                    # Restaurar punts a "-"
                    if pts_col is not None:
                        pts_item = self.samples_table.item(row, pts_col)
                        if pts_item:
                            pts_item.setText("-")

                # Recalcular l'estat de la fila
                self._update_row_state(row)

                self._update_next_button_state()

    def _count_file_points(self, filename, file_type):
        """Compta el nombre de punts d'un fitxer (comptatge ràpid per línies)."""
        if not filename or filename in ["-", "(cap)"]:
            return 0

        # Buscar el path complet del fitxer (usant llista amb paths complets)
        orphan_list = self._orphan_files_full.get(file_type, [])
        full_path = None
        for f in orphan_list:
            if Path(f).name == filename:
                full_path = f
                break

        if not full_path or not os.path.exists(full_path):
            return 0

        try:
            # Comptatge ràpid: llegir en binari i comptar línies (encoding-agnostic)
            with open(full_path, 'rb') as f:
                # Detectar UTF-16 BOM per ajustar comptatge
                header = f.read(2)
                is_utf16 = header in (b'\xff\xfe', b'\xfe\xff')
                f.seek(0)
                n_lines = sum(1 for _ in f)

            if is_utf16:
                # UTF-16 duplica cada caràcter, però line endings es detecten bé
                # Restar header (1 línia) → punts de dades
                return max(0, n_lines - 1)
            else:
                # CSV normal: restar header
                return max(0, n_lines - 1)

        except Exception as e:
            logger.warning(f"No s'ha pogut comptar punts de {filename}: {e}")

        return 0

    def _load_and_store_file_data(self, filename, file_type, sample_name, replica):
        """
        Carrega les dades d'un fitxer orfe assignat manualment i les guarda a imported_data.

        Args:
            filename: Nom del fitxer (sense path)
            file_type: "uib" o "dad"
            sample_name: Nom de la mostra
            replica: Número de rèplica
        """
        if not filename or filename in ["-", "(cap)"]:
            return

        QApplication.setOverrideCursor(QCursor(Qt.WaitCursor))
        try:
            self._load_and_store_file_data_impl(filename, file_type, sample_name, replica)
        finally:
            QApplication.restoreOverrideCursor()

    def _load_and_store_file_data_impl(self, filename, file_type, sample_name, replica):
        """Implementació interna de _load_and_store_file_data (sense cursor)."""
        logger.debug(f"_load_and_store: Intentant carregar {filename} per {sample_name} R{replica}")

        # Obtenir referència a les dades de la rèplica primer
        samples = self.imported_data.get("samples", {})
        if sample_name not in samples:
            logger.debug(f"_load_and_store: Mostra '{sample_name}' no trobada a imported_data")
            logger.debug(f"_load_and_store: Samples disponibles: {list(samples.keys())[:5]}...")
            return

        rep_data = samples[sample_name].get("replicas", {}).get(str(replica))
        if rep_data is None:
            logger.debug(f"_load_and_store: Rèplica {replica} no trobada per '{sample_name}'")
            logger.debug(f"_load_and_store: Rèpliques disponibles: {list(samples[sample_name].get('replicas', {}).keys())}")
            return

        # Buscar el path complet del fitxer
        full_path = None

        # 1. Primer buscar a la llista d'orfes completa
        orphan_list = self._orphan_files_full.get(file_type, [])
        for f in orphan_list:
            if Path(f).name == filename:
                full_path = f
                break

        # 2. Si no es troba, mirar si hi ha un suggeriment amb el path
        if not full_path:
            suggestion_key = f"{file_type}_suggestion"
            suggestion = rep_data.get(suggestion_key, {})
            if suggestion:
                # El suggeriment pot tenir el path complet a "file"
                sugg_file = suggestion.get("file", "")
                if sugg_file and Path(sugg_file).name == filename:
                    full_path = sugg_file
                elif sugg_file:
                    # Construir path a partir del seq_path
                    if hasattr(self, 'seq_path') and self.seq_path:
                        if file_type == "uib":
                            test_path = os.path.join(self.seq_path, "CSV", sugg_file)
                        else:
                            test_path = os.path.join(self.seq_path, "Export3d", sugg_file)
                        if os.path.exists(test_path):
                            full_path = test_path

        # 3. Construir path manualment si encara no es troba
        if not full_path and hasattr(self, 'seq_path') and self.seq_path:
            if file_type == "uib":
                possible_paths = [
                    os.path.join(self.seq_path, "CSV", filename),
                    os.path.join(self.seq_path, "csv", filename),
                ]
            else:
                possible_paths = [
                    os.path.join(self.seq_path, "Export3d", filename),
                    os.path.join(self.seq_path, "Export3D", filename),
                    os.path.join(self.seq_path, "CSV", filename),
                    os.path.join(self.seq_path, "csv", filename),
                ]
            for p in possible_paths:
                if os.path.exists(p):
                    full_path = p
                    break

        if not full_path or not os.path.exists(full_path):
            logger.warning(f"No s'ha trobat el fitxer {filename} (file_type={file_type})")
            return

        try:
            if file_type == "uib":
                # Carregar UIB
                df, status = llegir_doc_uib(full_path)
                if not df.empty and "OK" in status:
                    t = df["time (min)"].values
                    y = df["DOC"].values

                    # Downsample UIB a cadència DOC Direct
                    if len(t) > 10:
                        t, y = downsample_to_cadence(t, y)

                    # Calcular baseline i y_net (CRÍTIC per areas_uib)
                    baseline = None
                    y_net = None
                    if len(t) > 10:
                        # Determinar mode (BP o COLUMN)
                        method = self.imported_data.get("method", "COLUMN")
                        mode = "BP" if method == "BP" else "COLUMN"
                        baseline = get_baseline_value(t, y, mode=mode)
                        y_net = np.array(y) - baseline

                    rep_data["uib"] = {
                        "t": t,
                        "y": y,
                        "y_raw": y,
                        "y_net": y_net,
                        "baseline": baseline,
                        "file": filename,
                        "n_points": len(df),
                        "manual_assignment": True,
                    }
                    bl_val = baseline if baseline is not None else 0
                    logger.debug(f"_load_and_store: Carregat UIB: {filename} per {sample_name} R{replica}")
                    logger.debug(f"_load_and_store: rep_data['uib'] keys: {list(rep_data['uib'].keys())}")
                    logger.debug(f"_load_and_store: t is not None: {t is not None}, len(t)={len(t) if t is not None else 0}")
                    # Treure de la llista d'orfes
                    if "orphan_files" in self.imported_data:
                        uib_orphans = self.imported_data["orphan_files"].get("uib", [])
                        self.imported_data["orphan_files"]["uib"] = [
                            f for f in uib_orphans if Path(f).name != filename
                        ]

            elif file_type == "dad":
                # Provar primer Export3D, després DAD1A
                from hpsec_config import get_config
                _wl_keep = get_config().get("wavelengths", "selected")
                df, status = llegir_dad_export3d(full_path, wavelengths_to_keep=_wl_keep)
                if df.empty or "Error" in status:
                    df, status = llegir_dad_1a(full_path)

                if not df.empty and "OK" in status:
                    # Assegurar que tenim una còpia del DataFrame
                    df = df.copy()
                    rep_data["dad"] = {
                        "df": df,
                        "t": df["time (min)"].values if "time (min)" in df.columns else None,
                        "file": filename,
                        "n_points": len(df),
                        "manual_assignment": True,
                        "source": "manual",
                    }
                    logger.debug(f"Carregat DAD: {filename} ({len(df)} punts, columnes: {list(df.columns)[:5]})")
                    # Treure de la llista d'orfes
                    if "orphan_files" in self.imported_data:
                        dad_orphans = self.imported_data["orphan_files"].get("dad", [])
                        self.imported_data["orphan_files"]["dad"] = [
                            f for f in dad_orphans if Path(f).name != filename
                        ]
                else:
                    logger.warning(f"No s'han pogut llegir dades DAD de {filename}: {status}")

        except Exception as e:
            logger.error(f"Error carregant {filename}: {e}")

    def _update_row_state(self, row):
        """Actualitza l'estat (semàfors) d'una fila específica."""
        GREEN = "#27AE60"
        YELLOW = "#F39C12"
        RED = "#E74C3C"
        GREY = "#BDC3C7"

        # Obtenir tipus de mostra i dades (buscar per nom/replica per suportar taula ordenada)
        sample_type = "MOSTRA"
        direct_pts = 0
        uib_pts = 0
        dad_pts = 0
        name_item = self.samples_table.item(row, self.COL_MOSTRA)
        rep_item = self.samples_table.item(row, self.COL_REP)
        if name_item and rep_item:
            s_name = name_item.data(Qt.UserRole)
            s_rep = rep_item.text()
            for data in self._sample_data:
                if (data.get("name") == s_name and
                    str(data.get("replica", "")) == s_rep):
                    sample_type = data.get("type", "MOSTRA")
                    direct_pts = data.get("direct_pts", 0)
                    uib_pts = data.get("uib_pts", 0)
                    dad_pts = data.get("dad_pts", 0)
                    break

        requires_assignment = sample_type in self.TYPES_REQUIRE_ASSIGNMENT
        is_optional_type = sample_type in self.TYPES_OPTIONAL_ASSIGNMENT

        missing = []
        pending_review = []

        if requires_assignment:
            if self._data_mode in ["DUAL", "UIB"]:
                uib_item = self.samples_table.item(row, self.COL_UIB_FILE_ACTUAL)
                if uib_item:
                    val = uib_item.text()
                    match_type = self._match_types.get((row, self.COL_UIB_FILE_ACTUAL), "")
                    if val == "-" and self._orphan_files.get("uib"):
                        missing.append("UIB")
                    elif match_type == "SUGGESTED":
                        pending_review.append("UIB")

            dad_item = self.samples_table.item(row, self.COL_DAD_FILE_ACTUAL)
            if dad_item:
                val = dad_item.text()
                match_type = self._match_types.get((row, self.COL_DAD_FILE_ACTUAL), "")
                if val == "-" and self._orphan_files.get("dad"):
                    missing.append("DAD")
                elif match_type == "SUGGESTED":
                    pending_review.append("DAD")

        uib_missing = "UIB" in missing
        dad_missing = "DAD" in missing
        uib_review = "UIB" in pending_review
        dad_review = "DAD" in pending_review

        # --- Semàfor DOC ---
        if direct_pts > 0:
            self.samples_table.setItem(row, self.COL_SEM_DOC,
                self._create_semaphore_item(GREEN, f"DOC Direct: {direct_pts} punts"))
        else:
            self.samples_table.setItem(row, self.COL_SEM_DOC,
                self._create_semaphore_item(RED, "DOC Direct: sense dades"))

        # --- Semàfor UIB ---
        if self.COL_SEM_UIB is not None:
            if uib_review:
                self.samples_table.setItem(row, self.COL_SEM_UIB,
                    self._create_semaphore_item(YELLOW, "UIB: revisar assignació"))
            elif uib_missing:
                self.samples_table.setItem(row, self.COL_SEM_UIB,
                    self._create_semaphore_item(RED, "UIB: falta fitxer"))
            elif uib_pts > 0:
                self.samples_table.setItem(row, self.COL_SEM_UIB,
                    self._create_semaphore_item(GREEN, f"UIB: {uib_pts} punts"))
            elif is_optional_type:
                self.samples_table.setItem(row, self.COL_SEM_UIB,
                    self._create_semaphore_item(GREY, "UIB: opcional"))
            else:
                self.samples_table.setItem(row, self.COL_SEM_UIB,
                    self._create_semaphore_item(GREY, "UIB: sense dades"))

        # --- Semàfor DAD ---
        if dad_review:
            self.samples_table.setItem(row, self.COL_SEM_DAD,
                self._create_semaphore_item(YELLOW, "DAD: revisar assignació"))
        elif dad_missing:
            self.samples_table.setItem(row, self.COL_SEM_DAD,
                self._create_semaphore_item(RED, "DAD: falta fitxer"))
        elif dad_pts > 0:
            self.samples_table.setItem(row, self.COL_SEM_DAD,
                self._create_semaphore_item(GREEN, f"DAD: {dad_pts} punts"))
        elif is_optional_type:
            self.samples_table.setItem(row, self.COL_SEM_DAD,
                self._create_semaphore_item(GREY, "DAD: opcional"))
        else:
            self.samples_table.setItem(row, self.COL_SEM_DAD,
                self._create_semaphore_item(GREY, "DAD: sense dades"))

        # Marcar fila com no verificada si hi ha pendents
        if missing or pending_review:
            self._unverified_fuzzy.add(row)
        else:
            self._unverified_fuzzy.discard(row)

    def _on_cell_double_clicked(self, row, col):
        """Handler de doble clic."""
        # Determinar columnes editables
        editable_cols = [self.COL_TIPUS]
        if self._data_mode == "DIRECT":
            editable_cols.append(self.COL_DAD_FILE_ACTUAL)
        else:
            editable_cols.extend([self.COL_UIB_FILE_ACTUAL, self.COL_DAD_FILE_ACTUAL])

        # Si és columna editable, deixar que el delegate s'encarregui
        if col in editable_cols:
            return

        # Altrament, mostrar preview
        # Obtenir nom i rèplica de les cel·les (per funcionar amb taula ordenada)
        name_item = self.samples_table.item(row, self.COL_MOSTRA)
        rep_item = self.samples_table.item(row, self.COL_REP)

        if name_item and rep_item:
            sample_name = name_item.data(Qt.UserRole)
            rep_text = rep_item.text()

            # Buscar les dades corresponents a _sample_data
            sample_data = None
            for data in self._sample_data:
                # Comparar amb conversió a string per evitar problemes de tipus
                if (data.get("name") == sample_name and
                    str(data.get("replica", "")) == rep_text):
                    sample_data = data
                    break

            if sample_data:
                # Assegurar que les dades crues estan carregades per mostrar cromatogrames
                if self.imported_data and self.imported_data.get("data_deferred"):
                    from hpsec_import import ensure_data_loaded
                    self.main_window.set_status("Carregant cromatogrames...")
                    ensure_data_loaded(self.imported_data)
                    self.main_window.set_status("Dades carregades", 3000)

                try:
                    replica = int(rep_text) if rep_text.isdigit() else 1
                except:
                    replica = 1
                dialog = ChromatogramPreviewDialog(
                    self,
                    sample_name,
                    replica,
                    sample_data,
                    self.imported_data
                )
                dialog.exec()

    def _update_warnings(self):
        """Actualitza avisos al CommonToolbar (G01-G06: estructura unificada)."""
        # Recollir TOTS els avisos en una llista
        warnings_list = []

        # 1. Warnings d'importació (injeccions faltants, etc.)
        import_warnings = getattr(self, '_import_warnings', [])
        for w in import_warnings:
            clean_w = w.replace("⚠️ ", "").replace("⚠", "").strip()
            if clean_w:
                warnings_list.append(clean_w)

        # 2. Suggeriments pendents
        pending_suggestions = 0
        suggestion_samples = []
        for (r, c), mt in self._match_types.items():
            if mt == "SUGGESTED":
                pending_suggestions += 1
                name_item = self.samples_table.item(r, self.COL_MOSTRA)
                if name_item and name_item.text() not in suggestion_samples:
                    suggestion_samples.append(name_item.text())

        if pending_suggestions:
            samples_preview = ", ".join(suggestion_samples[:3])
            if len(suggestion_samples) > 3:
                samples_preview += f"... (+{len(suggestion_samples)-3})"
            warnings_list.append(f"{pending_suggestions} suggeriments FUZZY: {samples_preview}")

        # 3. Fitxers orfes
        unassigned_uib, unassigned_dad = self._count_unassigned_orphans()
        if unassigned_uib > 0:
            warnings_list.append(f"{unassigned_uib} fitxers UIB sense assignar")
        if unassigned_dad > 0:
            warnings_list.append(f"{unassigned_dad} fitxers DAD sense assignar")

    def _get_assigned_files_from_table(self, include_path_variants=False):
        """Obté els fitxers assignats des de la taula.

        Args:
            include_path_variants: Si True, afegeix també Path(val).name per matching més flexible.

        Returns:
            tuple: (assigned_uib: set, assigned_dad: set)
        """
        assigned_uib = set()
        assigned_dad = set()

        for row in range(self.samples_table.rowCount()):
            if self._data_mode in ["DUAL", "UIB"]:
                uib_item = self.samples_table.item(row, self.COL_UIB_FILE_ACTUAL)
                if uib_item:
                    val = uib_item.data(Qt.UserRole) or uib_item.text()
                    if val and val not in ["-", "(cap)"]:
                        assigned_uib.add(val)
                        if include_path_variants:
                            assigned_uib.add(Path(val).name)

            dad_item = self.samples_table.item(row, self.COL_DAD_FILE_ACTUAL)
            if dad_item:
                val = dad_item.data(Qt.UserRole) or dad_item.text()
                if val and val not in ["-", "(cap)", "[MasterFile]", "[export3d]", "[csv]"]:
                    assigned_dad.add(val)
                    if include_path_variants:
                        assigned_dad.add(Path(val).name)

        return assigned_uib, assigned_dad

    def _count_unassigned_orphans(self):
        """Compta quants orfes encara no estan assignats."""
        assigned_uib, assigned_dad = self._get_assigned_files_from_table()

        orphan_uib = self._orphan_files.get("uib", [])
        orphan_dad = self._orphan_files.get("dad", [])

        unassigned_uib = sum(1 for f in orphan_uib if Path(f).name not in assigned_uib)
        unassigned_dad = sum(1 for f in orphan_dad if Path(f).name not in assigned_dad)

        return unassigned_uib, unassigned_dad

    def _confirm_all_suggestions(self):
        """Confirma tots els suggeriments automàtics i carrega les dades."""
        confirmed = 0
        self.samples_table.blockSignals(True)
        QApplication.setOverrideCursor(QCursor(Qt.WaitCursor))

        try:
            self._confirm_all_suggestions_impl()
        finally:
            QApplication.restoreOverrideCursor()

    def _confirm_all_suggestions_impl(self):
        """Implementació interna de _confirm_all_suggestions."""
        confirmed = 0

        for row in range(self.samples_table.rowCount()):
            # Obtenir nom i rèplica de la fila
            name_item = self.samples_table.item(row, self.COL_MOSTRA)
            rep_item = self.samples_table.item(row, self.COL_REP)
            if not name_item or not rep_item:
                continue
            sample_name = name_item.data(Qt.UserRole)
            try:
                replica = int(rep_item.text())
            except:
                replica = 1

            # Comprovar UIB
            if self._data_mode in ["DUAL", "UIB"]:
                if self._match_types.get((row, self.COL_UIB_FILE_ACTUAL)) == "SUGGESTED":
                    item = self.samples_table.item(row, self.COL_UIB_FILE_ACTUAL)
                    if item:
                        item.setBackground(QBrush(MATCH_COLORS["EXACT"]))
                        self._match_types[(row, self.COL_UIB_FILE_ACTUAL)] = "CONFIRMED"
                        confirmed += 1
                        # Carregar dades del fitxer (sense cursor extra — ja el tenim)
                        filename = item.data(Qt.UserRole)
                        if filename:
                            self._load_and_store_file_data_impl(filename, "uib", sample_name, replica)
                            # Guardar a _manual_assignments per persistència
                            key = (sample_name, replica)
                            self._manual_assignments.setdefault(key, {})[self.COL_UIB_FILE_ACTUAL] = filename

            # Comprovar DAD
            if self._match_types.get((row, self.COL_DAD_FILE_ACTUAL)) == "SUGGESTED":
                item = self.samples_table.item(row, self.COL_DAD_FILE_ACTUAL)
                if item:
                    item.setBackground(QBrush(MATCH_COLORS["EXACT"]))
                    self._match_types[(row, self.COL_DAD_FILE_ACTUAL)] = "CONFIRMED"
                    confirmed += 1
                    # Carregar dades del fitxer (sense cursor extra — ja el tenim)
                    filename = item.data(Qt.UserRole)
                    if filename:
                        self._load_and_store_file_data_impl(filename, "dad", sample_name, replica)
                        # Guardar a _manual_assignments per persistència
                        key = (sample_name, replica)
                        self._manual_assignments.setdefault(key, {})[self.COL_DAD_FILE_ACTUAL] = filename

            # Processar events cada 5 files per mantenir UI responsiva
            if row % 5 == 0:
                QApplication.processEvents()

        self.samples_table.blockSignals(False)

        # Marcar canvis sense guardar perquè es persisteixin
        if confirmed > 0:
            self.main_window.mark_unsaved_changes()

        # Recalcular estats
        self._recalculate_row_states()
        self._update_warnings()
        self._update_next_button_state()

        if confirmed > 0:
            # Marcar warnings com a confirmats per evitar que reapareguin
            self._warnings_confirmed = True
            if self.imported_data:
                self.imported_data["warnings_confirmed"] = True

            # Guardar manifest immediatament
            try:
                logger.debug(f"confirm: Guardant manifest amb {confirmed} confirmacions...")
                self._apply_manual_assignments()
                manifest_path = save_import_manifest(self.imported_data)
                logger.debug(f"confirm: Manifest guardat a: {manifest_path}")
                self.main_window.mark_manifest_saved()
                QMessageBox.information(self, "Confirmat", f"S'han confirmat {confirmed} suggeriments i s'han guardat.")
            except Exception as e:
                import traceback
                logger.debug(f"confirm: ERROR: {e}")
                traceback.print_exc()
                QMessageBox.warning(self, "Avís", f"S'han confirmat {confirmed} suggeriments però no s'han pogut guardar: {e}")


    def _recalculate_row_states(self):
        """Recalcula l'estat (semàfors) de cada fila basant-se en assignacions actuals."""
        self.samples_table.blockSignals(True)
        self._unverified_fuzzy.clear()

        for row in range(self.samples_table.rowCount()):
            self._update_row_state(row)

        self.samples_table.blockSignals(False)

    def _update_next_button_state(self):
        """Actualitza l'estat intern de verificació FUZZY."""
        pass  # El wizard controla la navegació

    def _refresh_orphan_count(self):
        """Actualitza el comptador d'orfes i la llista després d'assignacions manuals."""
        assigned_uib, assigned_dad = self._get_assigned_files_from_table(include_path_variants=True)

        # Actualitzar llista d'orfes (treure els assignats)
        orig_uib = self._orphan_files.get("uib", [])
        orig_dad = self._orphan_files.get("dad", [])

        self._orphan_files["uib"] = [f for f in orig_uib if Path(f).name not in assigned_uib]
        self._orphan_files["dad"] = [f for f in orig_dad if Path(f).name not in assigned_dad]

        # Actualitzar warnings
        self._update_warnings()

    def _dismiss_orphan_warning(self):
        """Marca l'avís d'orfes com a revisat i amaga la barra d'avisos."""
        self._orphan_warning_dismissed = True
        # Guardar al manifest que l'avís ha estat revisat
        if self.imported_data:
            self.imported_data["orphan_warning_dismissed"] = True
            self.imported_data["warnings_confirmed"] = True  # Marcar warnings com confirmats
            try:
                save_import_manifest(self.imported_data)
                self.main_window.set_status("Avís marcat com a revisat", 3000)
            except Exception as e:
                logger.warning(f"No s'ha pogut guardar estat revisat: {e}")

        # Notificar al wizard que els warnings s'han descartat
        self.warnings_dismissed.emit()

    def _show_orphans(self):
        # Preparar dades amb punts per cada fitxer orfe
        orphans_with_info = {"uib": [], "dad": []}

        QApplication.setOverrideCursor(QCursor(Qt.WaitCursor))
        try:
            for file_type in ["uib", "dad"]:
                files = self._orphan_files.get(file_type, [])
                for f in sorted(files):
                    n_points = self._count_file_points(f, file_type)
                    orphans_with_info[file_type].append({
                        "file": f,
                        "n_points": n_points
                    })
        finally:
            QApplication.restoreOverrideCursor()

        dialog = OrphanFilesDialog(self, orphans_with_info)
        dialog.exec()

    def _on_warnings_confirmed(self, initials: str):
        """Handler quan es confirmen avisos via CommonToolbar."""
        self._warnings_confirmed = True
        self._warnings_confirmed_by = initials
        self._orphan_warning_dismissed = True

        # Guardar al manifest
        if self.imported_data:
            self.imported_data["warnings_confirmed"] = True
            self.imported_data["warnings_confirmed_by"] = initials
            self.imported_data["orphan_warning_dismissed"] = True
            try:
                save_import_manifest(self.imported_data)
                self.main_window.set_status(f"Avisos confirmats per {initials}", 3000)
            except Exception as e:
                logger.warning(f"No s'ha pogut guardar: {e}")

        # Notificar wizard
        self.warnings_dismissed.emit()
        self._update_next_button_state()

    def _on_notes_changed(self, notes: str):
        """Handler quan canvien les notes via CommonToolbar."""
        if self.imported_data:
            self.imported_data["notes"] = notes
            try:
                save_import_manifest(self.imported_data)
            except Exception as e:
                logger.warning(f"No s'ha pogut guardar notes: {e}")

    def _save_manifest(self):
        """Guarda el manifest amb les assignacions actuals."""
        if not self.imported_data:
            QMessageBox.warning(self, "Sense dades", "No hi ha dades per guardar.")
            return

        try:
            # Aplicar assignacions manuals
            self._apply_manual_assignments()
            # Guardar manifest
            save_import_manifest(self.imported_data)
            self.main_window.mark_manifest_saved()
            self.main_window.set_status("Manifest guardat correctament", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error guardant manifest:\n{e}")

    def _go_next(self):
        # Verificar si hi ha orfes sense assignar
        unassigned_uib, unassigned_dad = self._count_unassigned_orphans()
        has_orphans = unassigned_uib > 0 or unassigned_dad > 0
        has_unsaved = self.main_window.has_unsaved_changes

        # Si s'ha carregat des de manifest sense fer canvis, passar directament
        # (els orfes ja eren coneguts quan es va guardar el manifest)
        if self._loaded_from_manifest and not has_unsaved:
            return

        # Si no hi ha orfes ni canvis sense guardar, passar directament
        if not has_orphans and not has_unsaved:
            return

        # Construir missatge de confirmació
        msg_parts = []
        if has_orphans:
            msg_parts.append("Hi ha fitxers orfes sense assignar:")
            if unassigned_uib:
                msg_parts.append(f"  • {unassigned_uib} fitxers UIB")
            if unassigned_dad:
                msg_parts.append(f"  • {unassigned_dad} fitxers DAD")
            msg_parts.append("")

        if has_unsaved:
            msg_parts.append("Es guardaran els canvis i es passarà a la fase de calibració.")
        else:
            msg_parts.append("Es passarà a la fase de calibració.")

        msg_parts.append("\nVols continuar?")

        reply = QMessageBox.question(
            self, "Continuar",
            "\n".join(msg_parts),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply != QMessageBox.Yes:
            return

        # Aplicar assignacions manuals i guardar si hi ha canvis
        if has_unsaved:
            self._apply_manual_assignments()
            try:
                save_import_manifest(self.imported_data)
                self.main_window.mark_manifest_saved()
            except Exception as e:
                logger.warning(f"No s'ha pogut guardar manifest: {e}")

    def _apply_manual_assignments(self):
        """Aplica les assignacions manuals a imported_data."""
        if not hasattr(self, '_manual_assignments') or not self._manual_assignments:
            return

        # Iterar per les assignacions guardades amb clau (sample_name, replica)
        for key, assignments in self._manual_assignments.items():
            sample_name, replica = key

            # Obtenir dades de la mostra
            samples = self.imported_data.get("samples", {})
            if sample_name not in samples:
                continue

            rep_data = samples[sample_name].get("replicas", {}).get(str(replica))
            if not rep_data:
                continue

            # Aplicar assignacions
            for col, filename in assignments.items():
                if filename in ["-", "(cap)", ""]:
                    continue

                # Determinar tipus de senyal
                if col == self.COL_UIB_FILE_ACTUAL:
                    signal_type = "uib"
                elif col == self.COL_DAD_FILE_ACTUAL:
                    signal_type = "dad"
                else:
                    continue

                # Marcar com a assignació manual
                # Assegurar que el dict existeix (pot ser None o absent)
                if signal_type not in rep_data or rep_data[signal_type] is None:
                    rep_data[signal_type] = {}
                rep_data[signal_type]["manual_assignment"] = True
                rep_data[signal_type]["manual_file"] = filename
                rep_data[signal_type]["file"] = filename  # També guardar com a file per compatibilitat

        # Actualitzar manifest
        try:
            save_import_manifest(self.imported_data)
        except Exception as e:
            logger.warning(f"No s'ha pogut actualitzar manifest: {e}")
