# -*- coding: utf-8 -*-
"""
HPSEC Suite - Dashboard Panel
==============================

Vista general de totes les seqüències.
Disseny minimalista amb informació clara per columnes.
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFrame,
    QProgressBar, QComboBox, QLineEdit, QMessageBox, QInputDialog,
    QStyledItemDelegate, QMenu
)
from PySide6.QtCore import Qt, Signal, QThread
from PySide6.QtGui import QFont, QColor, QBrush

import os
import re
import sys
import json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from gui.models.sequence_state import SequenceState, Phase, get_all_sequences
from gui.widgets.styles import (
    COLOR_SUCCESS, COLOR_WARNING, COLOR_ERROR,
    COLOR_PENDING, COLOR_CURRENT, COLOR_CAL_BG, COLOR_CAL_TEXT
)
from hpsec_config import get_config, get_data_folders
# NOTA: hpsec_import, hpsec_calibrate, hpsec_analyze, hpsec_reports
# s'importen lazy dins les funcions run_*() per accelerar l'arrencada.

# Contrasenya per operacions batch i reset
BATCH_PASSWORD = "LEQUIA"


class SortableTableItem(QTableWidgetItem):
    """Item que ordena per UserRole si existeix, sinó per text."""
    def __lt__(self, other):
        my_data = self.data(Qt.UserRole)
        other_data = other.data(Qt.UserRole) if other else None

        # Si tots dos tenen UserRole numèric, ordenar per això
        if my_data is not None and other_data is not None:
            try:
                return float(my_data) < float(other_data)
            except (TypeError, ValueError):
                pass

        # Altrament, ordenar per text
        return self.text() < (other.text() if other else "")

# Constants de columna
COL_CHECK = 0
COL_NAME = 1
COL_DATE = 2
COL_MOSTRES = 3    # Nombre de mostres
COL_REF = 4        # Patrons de referència (KHP, PR)
COL_CTRL = 5       # Blancs i controls (MQ, NaOH, etc.)
COL_IMPORT = 6
COL_CAL = 7
COL_ANA = 8
COL_REVIEW = 9
COL_WARNINGS = 10  # Avisos de processament (recompte + tooltip)
COL_NOTES = 11     # Notes manuals
NUM_COLS = 12


# =============================================================================
# FUNCIONS CORE PER FASES INDIVIDUALS
# Cada funció executa UNA sola fase - cridem les funcions de hpsec_*.py
# =============================================================================

def run_import(seq_path, default_uib_sensitivity=None, siblings=None):
    """
    Executa IMPORT per una seqüència. Retorna (success, message, data).

    Args:
        seq_path: Path de la seqüència principal
        default_uib_sensitivity: Sensibilitat UIB per defecte (opcional)
        siblings: Llista de paths de siblings (282B_SEQ, 282C_SEQ...) o None
    """
    from hpsec_import import import_sequence, import_sequence_pack, save_import_manifest
    from hpsec_reports import generate_import_plots
    try:
        # Si hi ha siblings, usar import_sequence_pack
        if siblings:
            all_paths = [seq_path] + siblings
            result = import_sequence_pack(all_paths)
        else:
            result = import_sequence(seq_path)

        if result and result.get('success'):
            # Aplicar sensibilitat UIB per defecte si cal
            data_mode = result.get("data_mode", "")
            current_uib_sens = result.get("uib_sensitivity")
            if data_mode in ["DUAL", "UIB"] and not current_uib_sens and default_uib_sensitivity:
                result["uib_sensitivity"] = default_uib_sensitivity
                # Actualitzar MasterFile si existeix
                try:
                    master_file = result.get("master_file")
                    if master_file and os.path.exists(master_file):
                        import openpyxl
                        wb = openpyxl.load_workbook(master_file)
                        if "0-INFO" in wb.sheetnames:
                            ws = wb["0-INFO"]
                            ws["B5"] = default_uib_sensitivity
                            wb.save(master_file)
                except Exception:
                    pass  # Continuar sense actualitzar MasterFile

            save_import_manifest(result)
            # Generar gràfics
            try:
                generate_import_plots(seq_path, result)
            except Exception:
                pass  # Continuar sense gràfics
            return True, "OK", result
        errors = result.get('errors', ['?']) if result else ['?']
        return False, f"Error: {errors[0]}", None
    except Exception as e:
        return False, str(e), None


def run_calibrate(seq_path):
    """Executa CALIBRATE per una seqüència. Retorna (success, message, data)."""
    from hpsec_import import import_from_manifest
    from hpsec_calibrate import calibrate_from_import
    from hpsec_reports import generate_calibration_plots
    try:
        # IMPORTANT: El manifest JSON només conté metadades, no les dades reals.
        # Cal usar import_from_manifest per carregar les dades des dels fitxers.
        manifest_path = Path(seq_path) / "CHECK" / "data" / "import_manifest.json"
        if not manifest_path.exists():
            return False, "No importat", None

        # Reimportar dades usant el manifest
        imported = import_from_manifest(seq_path)
        if not imported or not imported.get("success"):
            errors = imported.get("errors", ["Error desconegut"]) if imported else ["Error importació"]
            return False, f"Error importació: {errors[0]}", None

        result = calibrate_from_import(imported)
        if result and result.get('success'):
            # Generar gràfics
            try:
                generate_calibration_plots(seq_path, result, imported)
            except Exception:
                pass  # Continuar sense gràfics
            return True, "OK", result
        return False, "Sense KHP", None
    except Exception as e:
        return False, str(e), None


def run_analyze(seq_path):
    """Executa ANALYZE per una seqüència. Retorna (success, message, data)."""
    from hpsec_import import import_from_manifest
    from hpsec_analyze import analyze_sequence, save_analysis_result
    from hpsec_reports import generate_analysis_plots
    try:
        import json
        data_path = Path(seq_path) / "CHECK" / "data"

        # Llegir manifest i reimportar les dades completes
        # IMPORTANT: El manifest JSON només conté metadades, no les dades reals.
        # Cal usar import_from_manifest per carregar les dades des dels fitxers.
        manifest_path = data_path / "import_manifest.json"
        if not manifest_path.exists():
            return False, "No importat", None

        # Reimportar dades usant el manifest
        imported = import_from_manifest(seq_path)
        if not imported or not imported.get("success"):
            errors = imported.get("errors", ["Error desconegut"]) if imported else ["Error importació"]
            return False, f"Error importació: {errors[0]}", None

        # Llegir calibració (opcional)
        cal_path = data_path / "calibration_result.json"
        if cal_path.exists():
            with open(cal_path, "r", encoding="utf-8") as f:
                calibrated = json.load(f)
        else:
            calibrated = {"factor": 1.0, "shift_uib": 0, "shift_direct": 0}

        result = analyze_sequence(imported, calibrated)
        if result and result.get('success'):
            save_analysis_result(result)
            # Generar gràfics
            try:
                generate_analysis_plots(seq_path, result)
            except Exception:
                pass  # Continuar sense gràfics
            return True, "OK", result
        errors = result.get('errors', ['?']) if result else ['?']
        return False, f"Error: {errors[0]}", None
    except Exception as e:
        return False, str(e), None




class SingleSeqWorker(QThread):
    """
    Worker per processar UNA seqüència (cas habitual).
    Executa totes les fases pendents: Import → Calibrate → Analyze.
    """
    progress = Signal(str)
    finished = Signal(bool, str)

    def __init__(self, seq_state):
        super().__init__()
        self.seq = seq_state

    def run(self):
        """Processa una seqüència executant les fases pendents."""
        seq_path = self.seq.seq_path
        siblings = self.seq.siblings if hasattr(self.seq, 'siblings') else []
        errors = []

        # IMPORT (si pendent)
        if not self.seq.import_status.completed:
            if siblings:
                self.progress.emit(f"Importar pack [{len(siblings)+1} carpetes]...")
            else:
                self.progress.emit("Importar...")
            ok, msg, _ = run_import(seq_path, siblings=siblings)
            if not ok:
                self.finished.emit(False, f"Import: {msg}")
                return

        # CALIBRATE (si pendent)
        if not self.seq.calibrate_status.completed:
            self.progress.emit("Verificar...")
            ok, msg, _ = run_calibrate(seq_path)
            # Verificar pot fallar sense KHP, continuem

        # ANALYZE (si pendent)
        if not self.seq.analyze_status.completed:
            self.progress.emit("Analitzar...")
            ok, msg, _ = run_analyze(seq_path)
            if not ok:
                self.finished.emit(False, f"Analyze: {msg}")
                return

        self.finished.emit(True, "Completat")


class BatchWorker(QThread):
    """
    Worker per processar múltiples seqüències.

    EXECUCIÓ HORITZONTAL: per cada seqüència, executa TOTES les fases.
    Això és més robust — si una SEQ falla, les següents continuen.
    """
    progress = Signal(int, int, str)  # current, total, message
    seq_completed = Signal(str, bool, str)  # seq_name, success, message
    finished = Signal(int, int)  # success_count, fail_count

    def __init__(self, sequences, phases, default_uib_sensitivity=None):
        super().__init__()
        self.sequences = sequences
        self.phases = phases
        self.default_uib_sensitivity = default_uib_sensitivity
        self._stop_requested = False

    def stop(self):
        self._stop_requested = True

    def run(self):
        """Execució HORITZONTAL: cada seqüència, totes les fases."""
        total_ok, total_fail = 0, 0
        n_seqs = len(self.sequences)

        phase_names = {
            Phase.IMPORT: "Importar",
            Phase.CALIBRATE: "Verificar",
            Phase.ANALYZE: "Analitzar",
        }

        for i, seq in enumerate(self.sequences):
            if self._stop_requested:
                break

            seq_ok = True
            for phase in self.phases:
                if self._stop_requested:
                    break

                phase_name = phase_names.get(phase, "?")

                # Display name (pack info per imports)
                if phase == Phase.IMPORT and hasattr(seq, 'siblings') and seq.siblings:
                    display_name = f"{seq.seq_name} [pack {len(seq.siblings)+1}]"
                else:
                    display_name = seq.seq_name
                self.progress.emit(i + 1, n_seqs, f"{display_name}: {phase_name}")

                # Executar fase
                if phase == Phase.IMPORT:
                    siblings = seq.siblings if hasattr(seq, 'siblings') else []
                    ok, msg, _ = run_import(seq.seq_path, self.default_uib_sensitivity, siblings)
                elif phase == Phase.CALIBRATE:
                    ok, msg, _ = run_calibrate(seq.seq_path)
                elif phase == Phase.ANALYZE:
                    ok, msg, _ = run_analyze(seq.seq_path)
                else:
                    continue

                self.seq_completed.emit(seq.seq_name, ok, f"{phase_name}: {msg}")

                if not ok and phase != Phase.CALIBRATE:
                    # Verificar pot fallar sense KHP — no és blocker
                    seq_ok = False
                    break  # Saltar fases restants per aquesta SEQ

            if seq_ok:
                total_ok += 1
            else:
                total_fail += 1

        self.finished.emit(total_ok, total_fail)


class DashboardPanel(QWidget):
    """Dashboard - Vista general de seqüències."""

    sequence_selected = Signal(str, str)

    # Noms de les etapes
    STAGE_NAMES = ["Importar", "Verificar", "Analitzar", "Exportar"]

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.sequences = []
        self.filtered_sequences = []
        self.batch_worker = None
        self.single_worker = None

        self._loading_overlay = None
        self._is_loading = False  # Evita doble-clic mentre carrega
        self._setup_ui()
        # Defer: carregar seqüències DESPRÉS que la finestra sigui visible
        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, self.refresh_sequences)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # === HEADER: Títol ===
        header = QHBoxLayout()
        header.setSpacing(16)

        self.lbl_title = QLabel("Seqüències")
        self.lbl_title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        header.addWidget(self.lbl_title)

        header.addStretch()
        layout.addLayout(header)

        # === FILA 1: Estadístiques + Botons ===
        stats_row = QHBoxLayout()
        stats_row.setSpacing(16)

        # Estadístiques per etapa
        self.lbl_stats = QLabel()
        self.lbl_stats.setFont(QFont("Segoe UI", 10))
        stats_row.addWidget(self.lbl_stats)

        stats_row.addStretch()

        # Botons
        self.refresh_btn = QPushButton("Actualitzar")
        self.refresh_btn.clicked.connect(self.refresh_sequences)
        stats_row.addWidget(self.refresh_btn)

        layout.addLayout(stats_row)

        # === FILA 2: Filtres ===
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(12)

        # Filtre Tipus (Column/BP)
        filter_layout.addWidget(QLabel("Tipus:"))
        self.filter_tipus = QComboBox()
        self.filter_tipus.addItems(["Tots", "Column", "BP"])
        self.filter_tipus.setMinimumWidth(80)
        self.filter_tipus.currentTextChanged.connect(self._apply_filter)
        filter_layout.addWidget(self.filter_tipus)

        # Filtre Mode (Dual/Direct/UIB)
        filter_layout.addWidget(QLabel("Mode:"))
        self.filter_mode = QComboBox()
        self.filter_mode.addItems(["Tots", "DUAL", "DIRECT", "UIB"])
        self.filter_mode.setMinimumWidth(80)
        self.filter_mode.currentTextChanged.connect(self._apply_filter)
        filter_layout.addWidget(self.filter_mode)

        # Filtre Estat
        filter_layout.addWidget(QLabel("Estat:"))
        self.filter_estat = QComboBox()
        self.filter_estat.addItems(["Tots", "Pendent", "En curs", "Complet", "Error", "CAL"])
        self.filter_estat.setMinimumWidth(90)
        self.filter_estat.currentTextChanged.connect(self._apply_filter)
        filter_layout.addWidget(self.filter_estat)

        # Filtre Carpeta (visible només amb >1 carpeta)
        self.lbl_filter_carpeta = QLabel("Carpeta:")
        filter_layout.addWidget(self.lbl_filter_carpeta)
        self.filter_carpeta = QComboBox()
        self.filter_carpeta.addItem("Totes")
        self.filter_carpeta.setMinimumWidth(90)
        self.filter_carpeta.currentTextChanged.connect(self._apply_filter)
        filter_layout.addWidget(self.filter_carpeta)
        # Amagar per defecte — es mostra si >1 carpeta
        self.lbl_filter_carpeta.setVisible(False)
        self.filter_carpeta.setVisible(False)

        filter_layout.addSpacing(20)

        # Cerca
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Cercar...")
        self.search_edit.setMaximumWidth(150)
        self.search_edit.textChanged.connect(self._apply_filter)
        filter_layout.addWidget(self.search_edit)

        filter_layout.addStretch()

        layout.addLayout(filter_layout)

        # === FILA 3: Barra d'accions batch ===
        batch_bar = QHBoxLayout()
        batch_bar.setSpacing(8)

        self.btn_select_filtered = QPushButton("✓ Sel. filtrades")
        self.btn_select_filtered.clicked.connect(self._select_filtered)
        batch_bar.addWidget(self.btn_select_filtered)

        self.btn_deselect_all = QPushButton("✗ Deseleccionar")
        self.btn_deselect_all.clicked.connect(self._deselect_all)
        batch_bar.addWidget(self.btn_deselect_all)

        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setStyleSheet("color: #ccc;")
        batch_bar.addWidget(sep)

        self.lbl_selected = QLabel("Seleccionades: 0")
        self.lbl_selected.setFont(QFont("Segoe UI", 9))
        self.lbl_selected.setStyleSheet("color: #555;")
        batch_bar.addWidget(self.lbl_selected)

        batch_bar.addStretch()

        # Botó Processar amb menú per etapa + reset
        self.btn_batch_process = QPushButton("▶ Processar ▾")
        process_menu = QMenu(self)
        process_menu.addAction("Importar seleccionades", lambda: self._batch_process_stage(Phase.IMPORT))
        process_menu.addAction("Verificar seleccionades", lambda: self._batch_process_stage(Phase.CALIBRATE))
        process_menu.addAction("Analitzar seleccionades", lambda: self._batch_process_stage(Phase.ANALYZE))
        process_menu.addSeparator()
        process_menu.addAction("Pipeline complet", lambda: self._batch_process_stage(None))
        process_menu.addSeparator()
        process_menu.addAction("↺ Reset des d'Importar", lambda: self._batch_reset_stage(0))
        process_menu.addAction("↺ Reset des de Verificar", lambda: self._batch_reset_stage(1))
        process_menu.addAction("↺ Reset des d'Analitzar", lambda: self._batch_reset_stage(2))
        process_menu.addAction("↺ Reset Resultats", lambda: self._batch_reset_stage(3))
        self.btn_batch_process.setMenu(process_menu)
        batch_bar.addWidget(self.btn_batch_process)

        layout.addLayout(batch_bar)

        # === TAULA DE SEQÜÈNCIES ===
        self.table = QTableWidget()
        self.table.setColumnCount(NUM_COLS)
        self.table.setHorizontalHeaderLabels([
            "", "Seqüència", "Data", "Mostres", "Ref", "Ctrl",
            "Importar", "Verificar", "Analitzar", "Exportar",
            "Avisos", "Notes"
        ])

        # Tooltips per capçaleres
        self.table.horizontalHeaderItem(COL_MOSTRES).setToolTip("Nombre de mostres")
        self.table.horizontalHeaderItem(COL_REF).setToolTip("Patrons de referència (KHP, PR)")
        self.table.horizontalHeaderItem(COL_CTRL).setToolTip("Blancs i controls (MQ, NaOH, etc.)")
        self.table.horizontalHeaderItem(COL_WARNINGS).setToolTip("Avisos de processament")
        self.table.horizontalHeaderItem(COL_NOTES).setToolTip("Doble-clic per afegir notes")

        # Configurar columnes - autoajust amb mínims per capçaleres
        h = self.table.horizontalHeader()

        # Primer: ResizeToContents per totes (ajusta a contingut)
        for i in range(self.table.columnCount()):
            if i not in (COL_NOTES,):
                h.setSectionResizeMode(i, QHeaderView.ResizeToContents)

        # Notes expandeix per omplir espai restant
        h.setSectionResizeMode(COL_NOTES, QHeaderView.Stretch)

        # Checkbox columna fixa
        h.setSectionResizeMode(COL_CHECK, QHeaderView.Fixed)
        h.resizeSection(COL_CHECK, 28)

        # Mínims per assegurar que capçaleres es veuen
        self._header_min_widths = {
            COL_NAME: 110,
            COL_DATE: 70,
            COL_MOSTRES: 55,
            COL_REF: 35,
            COL_CTRL: 35,
            COL_IMPORT: 60,
            COL_CAL: 60,
            COL_ANA: 60,
            COL_REVIEW: 55,
            COL_WARNINGS: 50,
        }

        # Estil per mantenir colors dels punts en selecció
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
            }
            QTableWidget::item:selected {
                background-color: #d5e8f7;
            }
            QTableWidget::item:hover {
                background-color: #ecf0f1;
            }
        """)

        # Permetre ordenació (per data descendent — SEQs recents primer)
        self.table.setSortingEnabled(True)
        self.table.sortByColumn(COL_NAME, Qt.DescendingOrder)

        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(False)  # Treure ombrejat alternatiu
        self.table.cellDoubleClicked.connect(self._on_double_click)
        self.table.itemChanged.connect(self._on_item_changed)

        # Menú contextual (clic dret)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)

        layout.addWidget(self.table)

    def refresh_sequences(self):
        cfg = get_config()
        cfg.reload()
        data_folders = get_data_folders()
        self.sequences = get_all_sequences(data_folders)

        # Actualitzar títol
        if len(data_folders) == 1:
            folder_short = os.path.basename(data_folders[0]) or data_folders[0]
            self.lbl_title.setText(f"{len(self.sequences)} Seqüències ({folder_short})")
            self.lbl_title.setToolTip(data_folders[0])
        elif data_folders:
            self.lbl_title.setText(f"{len(self.sequences)} Seqüències")
            self.lbl_title.setToolTip("\n".join(data_folders))
        else:
            self.lbl_title.setText("Seqüències")
            self.lbl_title.setToolTip("Cap carpeta configurada")

        # Actualitzar combo carpeta
        folder_names = sorted(set(s.source_folder for s in self.sequences if s.source_folder))
        multi_folder = len(folder_names) > 1
        self.lbl_filter_carpeta.setVisible(multi_folder)
        self.filter_carpeta.setVisible(multi_folder)

        self.filter_carpeta.blockSignals(True)
        current = self.filter_carpeta.currentText()
        self.filter_carpeta.clear()
        self.filter_carpeta.addItem("Totes")
        self.filter_carpeta.addItems(folder_names)
        idx = self.filter_carpeta.findText(current)
        self.filter_carpeta.setCurrentIndex(max(0, idx))
        self.filter_carpeta.blockSignals(False)

        self._apply_filter()
        self._update_stats()

    def _apply_filter(self):
        filter_tipus = self.filter_tipus.currentText()
        filter_mode = self.filter_mode.currentText()
        filter_estat = self.filter_estat.currentText()
        filter_carpeta = self.filter_carpeta.currentText()
        search_text = self.search_edit.text().lower()

        self.filtered_sequences = []

        for seq in self.sequences:
            # Filtre cerca
            if search_text and search_text not in seq.seq_name.lower():
                continue

            # Filtre carpeta
            if filter_carpeta != "Totes" and seq.source_folder != filter_carpeta:
                continue

            # Filtre tipus (Column/BP)
            if filter_tipus == "Column" and "_BP" in seq.seq_name.upper():
                continue
            elif filter_tipus == "BP" and "_BP" not in seq.seq_name.upper():
                continue

            # Filtre mode (Dual/Direct/UIB)
            if filter_mode != "Tots":
                if seq.data_mode.upper() != filter_mode.upper():
                    continue

            # Filtre estat
            if filter_estat == "Pendent" and seq.progress_pct > 0:
                continue
            elif filter_estat == "En curs" and (seq.progress_pct == 0 or seq.progress_pct == 100):
                continue
            elif filter_estat == "Complet" and seq.progress_pct < 100:
                continue
            elif filter_estat == "Error":
                # Filtrar per seqüències amb errors
                has_error = (
                    seq.import_state == 'error' or
                    seq.calibrate_state == 'error' or
                    seq.analyze_state == 'error' or
                    seq.review_state == 'error'
                )
                if not has_error:
                    continue
            elif filter_estat == "CAL":
                if "_CAL" not in seq.seq_name.upper():
                    continue

            self.filtered_sequences.append(seq)

        self._update_table()

    def _build_name_tooltip(self, seq):
        """Construeix tooltip enriquit pel nom (Tipus + Mode + origen + siblings)."""
        parts = []
        # Tipus i mode
        method = seq.method if seq.method else "?"
        data_mode = seq.data_mode if seq.data_mode else "?"
        parts.append(f"{method} · {data_mode}")

        # Origen (si multi-folder)
        if seq.source_folder:
            parts.append(f"Carpeta: {seq.source_folder}")

        # Siblings
        if seq.siblings:
            sibling_names = [os.path.basename(s) for s in seq.siblings]
            parts.append(f"Pack amb {len(seq.siblings)} siblings:")
            parts.extend(f"  {s}" for s in sibling_names)

        return "\n".join(parts)

    def _update_table(self):
        # Bloquejar signals i sorting mentre actualitzem
        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)

        for seq in self.filtered_sequences:
            row = self.table.rowCount()
            self.table.insertRow(row)

            is_cal = "_CAL" in seq.seq_name.upper()
            cal_bg = QColor(COLOR_CAL_BG) if is_cal else None

            # Col CHECK: checkbox
            item_check = QTableWidgetItem()
            item_check.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            item_check.setCheckState(Qt.Unchecked)
            if cal_bg:
                item_check.setBackground(cal_bg)
            self.table.setItem(row, COL_CHECK, item_check)

            # Col NAME: Nom (amb indicador de siblings, sense prefix [CAL])
            display_name = seq.seq_name
            if seq.siblings:
                display_name = f"{seq.seq_name} [+{len(seq.siblings)}]"
            item_name = SortableTableItem(display_name)
            # UserRole = número SEQ per ordenació numèrica descendent
            seq_num_match = re.match(r'(\d+)', seq.seq_name)
            item_name.setData(Qt.UserRole, int(seq_num_match.group(1)) if seq_num_match else 0)
            item_name.setData(Qt.UserRole + 1, seq.seq_path)
            item_name.setFlags(item_name.flags() & ~Qt.ItemIsEditable)
            if is_cal:
                item_name.setForeground(QColor(COLOR_CAL_TEXT))
                font = item_name.font()
                font.setBold(True)
                item_name.setFont(font)
                item_name.setBackground(cal_bg)
            # Tooltip enriquit: Tipus + Mode + siblings
            item_name.setToolTip(self._build_name_tooltip(seq))
            self.table.setItem(row, COL_NAME, item_name)

            # Col DATE: Data (amb valor ordenable)
            date_display = seq.seq_date if seq.seq_date else "-"
            item_date = SortableTableItem(date_display)
            item_date.setTextAlignment(Qt.AlignCenter)
            item_date.setForeground(QColor("#666"))
            item_date.setFlags(item_date.flags() & ~Qt.ItemIsEditable)
            if cal_bg:
                item_date.setBackground(cal_bg)
            # Guardar data en format ordenable (YYYYMMDD) — sempre DD/MM/YY normalitzat
            if seq.seq_date and seq.seq_date != "-":
                try:
                    d, m, y = seq.seq_date.split('/')
                    y = int(y)
                    y = 2000 + y if y < 100 else y
                    item_date.setData(Qt.UserRole, y * 10000 + int(m) * 100 + int(d))
                except Exception:
                    item_date.setData(Qt.UserRole, 0)
            else:
                item_date.setData(Qt.UserRole, 0)
            self.table.setItem(row, COL_DATE, item_date)

            # Col MOSTRES: nombre de mostres
            n_samples = seq.n_samples or 0
            n_imp = seq.n_inj_imported
            n_mst = seq.n_inj_master
            incomplete = n_mst > 0 and n_imp < n_mst and seq.import_status.completed

            if seq.import_status.completed:
                mostres_text = str(n_samples) if n_samples else "-"
                mostres_tooltip = f"{n_samples} mostres"
                if incomplete:
                    mostres_tooltip += f"\nIMPORTACIÓ INCOMPLETA: {n_imp}/{n_mst}"
            else:
                mostres_text = "-"
                mostres_tooltip = "Pendent d'importar"

            item_mostres = SortableTableItem(mostres_text)
            item_mostres.setData(Qt.UserRole, n_samples)
            item_mostres.setTextAlignment(Qt.AlignCenter)
            item_mostres.setForeground(QColor(COLOR_ERROR) if incomplete else QColor("#666"))
            item_mostres.setFlags(item_mostres.flags() & ~Qt.ItemIsEditable)
            item_mostres.setToolTip(mostres_tooltip)
            if cal_bg:
                item_mostres.setBackground(cal_bg)
            self.table.setItem(row, COL_MOSTRES, item_mostres)

            # Col REF: patrons de referència (KHP, PR)
            n_khp = seq.n_khp or 0
            n_pr = seq.n_pr or 0
            n_ref = n_khp + n_pr

            if seq.import_status.completed:
                ref_text = str(n_ref) if n_ref else "-"
                ref_parts = []
                if n_khp:
                    ref_parts.append(f"{n_khp} KHP")
                if n_pr:
                    ref_parts.append(f"{n_pr} PR")
                ref_tooltip = " · ".join(ref_parts) if ref_parts else "Cap patró"
            else:
                ref_text = "-"
                ref_tooltip = "Pendent d'importar"

            item_ref = SortableTableItem(ref_text)
            item_ref.setData(Qt.UserRole, n_ref)
            item_ref.setTextAlignment(Qt.AlignCenter)
            item_ref.setForeground(QColor("#666"))
            item_ref.setFlags(item_ref.flags() & ~Qt.ItemIsEditable)
            item_ref.setToolTip(ref_tooltip)
            if cal_bg:
                item_ref.setBackground(cal_bg)
            self.table.setItem(row, COL_REF, item_ref)

            # Col CTRL: blancs i controls (MQ, NaOH, etc.)
            n_blank = seq.n_blank or 0
            n_control = seq.n_control or 0
            n_ctrl = n_blank + n_control

            if seq.import_status.completed:
                ctrl_text = str(n_ctrl) if n_ctrl else "-"
                ctrl_parts = []
                if n_blank:
                    ctrl_parts.append(f"{n_blank} Blanc")
                if n_control:
                    ctrl_parts.append(f"{n_control} Control")
                ctrl_tooltip = " · ".join(ctrl_parts) if ctrl_parts else "Cap blanc/control"
            else:
                ctrl_text = "-"
                ctrl_tooltip = "Pendent d'importar"

            item_ctrl = SortableTableItem(ctrl_text)
            item_ctrl.setData(Qt.UserRole, n_ctrl)
            item_ctrl.setTextAlignment(Qt.AlignCenter)
            item_ctrl.setForeground(QColor("#666"))
            item_ctrl.setFlags(item_ctrl.flags() & ~Qt.ItemIsEditable)
            item_ctrl.setToolTip(ctrl_tooltip)
            if cal_bg:
                item_ctrl.setBackground(cal_bg)
            self.table.setItem(row, COL_CTRL, item_ctrl)

            # Fases (Importar, Verificar, Analitzar, Exportar)
            phases_data = [
                (seq.import_status, seq.import_state, "Importar", seq.import_warnings),
                (seq.calibrate_status, seq.calibrate_state, "Verificar", seq.calibrate_warnings),
                (seq.analyze_status, seq.analyze_state, "Analitzar", seq.analyze_warnings),
                (seq.review_status, seq.review_state, "Exportar", seq.review_warnings),
            ]

            current_phase_idx = None
            for i, (status, _, _, _) in enumerate(phases_data):
                if not status.completed:
                    current_phase_idx = i
                    break

            phase_cols = [COL_IMPORT, COL_CAL, COL_ANA, COL_REVIEW]
            for col_offset, (status, state, phase_name, phase_warnings) in enumerate(phases_data):
                col = phase_cols[col_offset]
                item = QTableWidgetItem()
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                font = item.font()
                font.setPointSize(11)
                item.setFont(font)

                if is_cal and col_offset > 0:
                    item.setText("—")
                    item.setForeground(QColor(COLOR_PENDING))
                    item.setToolTip("Flux calibració (sense wizard)")
                    if cal_bg:
                        item.setBackground(cal_bg)
                elif state == 'ok':
                    item.setText("✔")
                    item.setForeground(QColor(COLOR_SUCCESS))
                    timestamp = status.timestamp[:16] if status.timestamp else ""
                    tooltip = f"{phase_name}: Completat"
                    if timestamp:
                        tooltip += f"\n{timestamp}"
                    item.setToolTip(tooltip)
                    if cal_bg:
                        item.setBackground(cal_bg)
                elif state == 'warning':
                    if phase_name == "Exportar" and seq.is_bp_stale:
                        item.setText("⟳")
                    else:
                        item.setText("⚠")
                    item.setForeground(QColor(COLOR_WARNING))
                    tooltip_parts = []
                    if phase_name == "Importar" and seq.import_incomplete:
                        tooltip_parts.append(f"{seq.n_inj_imported}/{seq.n_inj_master} injeccions importades")
                    if phase_warnings:
                        tooltip_parts.extend(phase_warnings[:5])
                    elif phase_name == "Verificar":
                        tooltip_parts.append(f"KHP sibling ({seq.khp_source})")
                    if tooltip_parts:
                        tooltip = f"{phase_name}:\n" + "\n".join(tooltip_parts)
                    else:
                        tooltip = f"{phase_name}: Avisos"
                    item.setToolTip(tooltip)
                    if cal_bg:
                        item.setBackground(cal_bg)
                elif state == 'error':
                    item.setText("×")
                    item.setForeground(QColor(COLOR_ERROR))
                    if phase_name == "Importar":
                        tooltip = f"{phase_name}: Error MasterFile"
                    elif phase_name == "Verificar" and not seq.has_khp:
                        tooltip = f"{phase_name}: Només històric!"
                    elif status.errors:
                        tooltip = f"{phase_name}: Error\n" + "\n".join(status.errors)
                    else:
                        tooltip = f"{phase_name}: Error"
                    item.setToolTip(tooltip)
                    if cal_bg:
                        item.setBackground(cal_bg)
                else:  # pending
                    item.setText("○")
                    if current_phase_idx == col_offset:
                        item.setForeground(QColor(COLOR_CURRENT))
                        tooltip = f"{phase_name}: Pendent (següent)"
                    else:
                        item.setForeground(QColor(COLOR_PENDING))
                        tooltip = f"{phase_name}: Pendent"
                    item.setToolTip(tooltip)
                    if cal_bg:
                        item.setBackground(cal_bg)

                self.table.setItem(row, col, item)

            # Col WARNINGS: avisos de processament (recompte + tooltip)
            warn_preview, warn_tooltip, warn_color = self._format_warnings_cell(seq)
            item_warn = QTableWidgetItem(warn_preview)
            item_warn.setToolTip(warn_tooltip)
            item_warn.setForeground(warn_color)
            item_warn.setFlags(item_warn.flags() & ~Qt.ItemIsEditable)
            if cal_bg:
                item_warn.setBackground(cal_bg)
            self.table.setItem(row, COL_WARNINGS, item_warn)

            # Col NOTES: notes manuals (doble-clic per editar)
            note_preview, note_tooltip, note_color = self._format_notes_only_cell(seq)
            item_notes = QTableWidgetItem(note_preview)
            item_notes.setToolTip(note_tooltip)
            item_notes.setForeground(note_color)
            item_notes.setFlags(item_notes.flags() & ~Qt.ItemIsEditable)
            if cal_bg:
                item_notes.setBackground(cal_bg)
            self.table.setItem(row, COL_NOTES, item_notes)

        # Reactivar sorting i signals
        self.table.setSortingEnabled(True)
        self.table.sortByColumn(COL_NAME, Qt.DescendingOrder)
        self.table.blockSignals(False)

        # Aplicar mínims de capçalera
        self._apply_min_widths()
        self._update_selection_count()

    def _format_warnings_cell(self, seq: SequenceState):
        """Genera text, tooltip i color per la cel·la Avisos (processament).

        Mostra recompte d'avisos amb tooltip detallat per etapa.
        """
        json_notes = seq.dashboard_notes
        n_warn = 0
        has_blocker = False
        tooltip_parts = []

        for jn in json_notes:
            ntype = jn.get("type", "")
            sev = jn.get("severity", "")
            stage = jn.get("stage", "?")
            content = jn.get("content", "")

            if ntype in ("ANOM", "WARN", "QUAL"):
                n_warn += 1
                if sev == "blocker":
                    has_blocker = True
                tooltip_parts.append(f"[{stage}] {content}")

        if n_warn:
            preview = str(n_warn)
        else:
            preview = ""

        if has_blocker:
            color = QColor("#C62828")
        elif n_warn:
            color = QColor("#E65100")
        else:
            color = QColor("#999")

        tooltip = "\n".join(tooltip_parts) if tooltip_parts else ""

        return preview, tooltip, color

    def _format_notes_only_cell(self, seq: SequenceState):
        """Genera text, tooltip i color per la cel·la Notes (manuals + info).

        Mostra preview de notes manuals i notes informatives del processament.
        """
        json_notes = seq.dashboard_notes
        manual_notes = seq.notes.strip() if seq.notes else ""

        tooltip_parts = []
        n_info = 0

        for jn in json_notes:
            ntype = jn.get("type", "")
            stage = jn.get("stage", "?")
            content = jn.get("content", "")

            if ntype in ("NOTE", "USR"):
                n_info += 1
                tooltip_parts.append(f"[{stage}] {content}")

        if manual_notes:
            tooltip_parts.insert(0, manual_notes)

        # Preview: primera línia de notes manuals o recompte info
        if manual_notes:
            preview = manual_notes.split('\n')[0][:50]
            if len(manual_notes) > 50 or '\n' in manual_notes:
                preview += "..."
        elif n_info:
            preview = f"{n_info} {'nota' if n_info == 1 else 'notes'}"
        else:
            preview = ""

        color = QColor("#1565C0") if (manual_notes or n_info) else QColor("#999")
        tooltip = "\n".join(tooltip_parts) if tooltip_parts else "Doble-clic per afegir notes"

        return preview, tooltip, color

    def _apply_min_widths(self):
        """Aplica amplades mínimes per assegurar capçaleres visibles."""
        h = self.table.horizontalHeader()
        for col, min_width in self._header_min_widths.items():
            if h.sectionSize(col) < min_width:
                h.resizeSection(col, min_width)

    def _update_stats(self):
        """Actualitza estadístiques per etapa (format compacte)."""
        total = len(self.sequences)
        if total == 0:
            self.lbl_stats.setText("Cap seqüència")
            self.lbl_stats.setToolTip("")
            return

        imported = sum(1 for s in self.sequences if s.import_status.completed)
        calibrated = sum(1 for s in self.sequences if s.calibrate_status.completed)
        analyzed = sum(1 for s in self.sequences if s.analyze_status.completed)
        reviewed = sum(1 for s in self.sequences if s.review_status.completed)

        errors = sum(1 for s in self.sequences if (
            s.import_state == 'error' or
            s.calibrate_state == 'error' or
            s.analyze_state == 'error'
        ))

        stats = f"Importar:{imported}  Verificar:{calibrated}  Analitzar:{analyzed}  Exportar:{reviewed}  /{total}"
        if errors:
            stats += f" · {errors} errors"

        self.lbl_stats.setText(stats)
        self.lbl_stats.setToolTip(
            f"Importades: {imported}/{total}\n"
            f"Verificades: {calibrated}/{total}\n"
            f"Analitzades: {analyzed}/{total}\n"
            f"Revisades: {reviewed}/{total}"
        )

    def _show_context_menu(self, pos):
        """Mostra menú contextual amb opcions per la seqüència."""
        row = self.table.rowAt(pos.y())
        if row < 0:
            return

        item_name = self.table.item(row, COL_NAME)
        if not item_name:
            return

        seq_path = item_name.data(Qt.UserRole + 1)
        seq = None
        for s in self.filtered_sequences:
            if s.seq_path == seq_path:
                seq = s
                break

        if not seq:
            return

        menu = QMenu(self)

        # Opció: Processar (totes les etapes pendents)
        siblings = seq.siblings if hasattr(seq, 'siblings') else []
        if siblings:
            action_text = f"▶ Processar {seq.seq_name} [pack {len(siblings)+1}]"
        else:
            action_text = f"▶ Processar {seq.seq_name}"
        action_process = menu.addAction(action_text)
        action_process.triggered.connect(lambda: self._process_single(seq))

        menu.addSeparator()

        # Opcions individuals per etapa
        if not seq.import_status.completed:
            action_import = menu.addAction("  → Importar")
            action_import.triggered.connect(lambda: self._run_single_phase(seq, "import"))
        else:
            action_reimport = menu.addAction("  → Re-importar (des de MasterFile)")
            action_reimport.triggered.connect(lambda: self._run_single_phase(seq, "import"))

        if seq.import_status.completed and not seq.calibrate_status.completed:
            action_cal = menu.addAction("  → Verificar")
            action_cal.triggered.connect(lambda: self._run_single_phase(seq, "calibrate"))

        if seq.import_status.completed and not seq.analyze_status.completed:
            action_analyze = menu.addAction("  → Analitzar")
            action_analyze.triggered.connect(lambda: self._run_single_phase(seq, "analyze"))

        menu.addSeparator()

        # Obrir al wizard
        action_wizard = menu.addAction("Obrir al Wizard...")
        action_wizard.triggered.connect(lambda: self._open_in_wizard(seq))

        # Submenu Reset per SEQ individual
        menu.addSeparator()
        reset_submenu = menu.addMenu("↺ Reset...")
        reset_submenu.addAction("Des d'Importar", lambda s=seq: self._reset_single(s, 0))
        reset_submenu.addAction("Des de Verificar", lambda s=seq: self._reset_single(s, 1))
        reset_submenu.addAction("Des d'Analitzar", lambda s=seq: self._reset_single(s, 2))
        reset_submenu.addAction("Resultats", lambda s=seq: self._reset_single(s, 3))

        menu.exec(self.table.mapToGlobal(pos))

    def _run_single_phase(self, seq, phase_name):
        """Executa una sola fase per una seqüència."""
        self._set_controls_enabled(False)

        # Mostrar si és pack
        siblings = seq.siblings if hasattr(seq, 'siblings') else []
        if phase_name == "import" and siblings:
            status_msg = f"{seq.seq_name} [pack {len(siblings)+1}]: {phase_name}..."
        else:
            status_msg = f"{seq.seq_name}: {phase_name}..."
        self.main_window.set_status(status_msg)

        if phase_name == "import":
            ok, msg, _ = run_import(seq.seq_path, siblings=siblings)
        elif phase_name == "calibrate":
            ok, msg, _ = run_calibrate(seq.seq_path)
        elif phase_name == "analyze":
            ok, msg, _ = run_analyze(seq.seq_path)
        else:
            ok, msg = False, "Fase desconeguda"

        self._set_controls_enabled(True)
        self.main_window.set_status(
            f"{seq.seq_name}: {msg}" if ok else f"{seq.seq_name}: ERROR - {msg}",
            5000
        )
        self.refresh_sequences()

    def _on_double_click(self, row, col):
        """Doble-clic obre directament al wizard o edita notes."""
        # Bloquejar doble-clic mentre s'està carregant
        if self._is_loading:
            return

        item_name = self.table.item(row, COL_NAME)
        if not item_name:
            return

        seq_path = item_name.data(Qt.UserRole + 1)
        seq = None
        for s in self.filtered_sequences:
            if s.seq_path == seq_path:
                seq = s
                break

        if not seq:
            return

        # Si és la columna Notes o Avisos, obrir popup per editar/veure
        if col == COL_NOTES or col == COL_WARNINGS:
            self._edit_notes_popup(row, seq)
            return

        # Altrament, obrir al wizard
        self._open_in_wizard(seq)

    def _edit_notes_popup(self, row, seq: SequenceState):
        """Obre un diàleg per editar les notes i veure observacions dels JSON."""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QTextEdit, QDialogButtonBox,
            QLabel, QGroupBox, QScrollArea
        )

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Notes i Observacions - {seq.seq_name}")
        dialog.setMinimumSize(550, 400)

        layout = QVBoxLayout(dialog)

        # === SECCIÓ 1: Notes dels JSON (warnings, anomalies, etc.) ===
        json_notes = seq.dashboard_notes

        if json_notes:
            obs_group = QGroupBox("Observacions de processament")
            obs_layout = QVBoxLayout(obs_group)
            obs_layout.setSpacing(4)

            for note in json_notes:
                note_frame = QLabel()
                stage = note.get("stage", "?")
                reviewer = note.get("reviewer", "")
                content = note.get("content", "")
                date = note.get("date", "")[:10] if note.get("date") else ""

                html = f"<b>[{stage}]</b> "
                if reviewer:
                    html += f"<span style='color:#666;'>({reviewer} {date})</span><br>"
                html += f"<span style='color:#333;'>{content}</span>"
                note_frame.setText(html)
                note_frame.setWordWrap(True)
                note_frame.setStyleSheet("""
                    background-color: #FFF8E1;
                    border: 1px solid #FFE082;
                    border-radius: 4px;
                    padding: 8px;
                    margin: 2px;
                """)
                obs_layout.addWidget(note_frame)

            layout.addWidget(obs_group)
        else:
            no_obs = QLabel("<i style='color:#888;'>Sense observacions de processament</i>")
            layout.addWidget(no_obs)

        # === SECCIÓ 2: Notes manuals (editables) ===
        notes_group = QGroupBox("Notes manuals")
        notes_layout = QVBoxLayout(notes_group)

        text_edit = QTextEdit()
        text_edit.setPlaceholderText("Escriu notes sobre aquesta seqüència...")
        text_edit.setText(seq.notes if seq.notes else "")
        text_edit.setMinimumHeight(100)
        notes_layout.addWidget(text_edit)

        layout.addWidget(notes_group)

        # Botons
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec() == QDialog.Accepted:
            new_notes = text_edit.toPlainText().strip()
            if seq.save_notes(new_notes):
                # Actualitzar la cel·la de la taula
                self.table.blockSignals(True)
                item_notes = self.table.item(row, COL_NOTES)
                if item_notes:
                    # Mostrar només primera línia o resum
                    preview = new_notes.split('\n')[0][:50]
                    if len(new_notes) > 50 or '\n' in new_notes:
                        preview += "..."
                    item_notes.setText(preview)
                self.table.blockSignals(False)
                self.main_window.set_status(f"Notes guardades: {seq.seq_name}", 3000)
            else:
                QMessageBox.warning(
                    self, "Error",
                    "No s'han pogut guardar les notes.\n"
                    "Cal importar la seqüència primer."
                )

    def _open_in_wizard(self, seq: SequenceState):
        """Obre la seqüència al wizard per processar/revisar."""
        self._is_loading = True
        self._show_loading_overlay(seq.seq_name)
        from PySide6.QtWidgets import QApplication
        QApplication.setOverrideCursor(Qt.WaitCursor)
        QApplication.processEvents()
        self.sequence_selected.emit(seq.seq_path, seq.current_phase.value)

    def _show_loading_overlay(self, seq_name: str):
        """Mostra overlay 'Carregant...' sobre la taula."""
        if self._loading_overlay is None:
            self._loading_overlay = QLabel(self.table)
            self._loading_overlay.setAlignment(Qt.AlignCenter)
            self._loading_overlay.setStyleSheet(
                "background-color: rgba(255, 255, 255, 200);"
                "color: #2E86AB; font-size: 16px; font-weight: bold;"
                "border-radius: 8px;"
            )
        self._loading_overlay.setText(f"⏳ Carregant {seq_name}...")
        self._loading_overlay.setGeometry(self.table.rect())
        self._loading_overlay.raise_()
        self._loading_overlay.show()

    def hide_loading_overlay(self):
        """Amaga l'overlay de càrrega i restaura cursor."""
        self._is_loading = False
        if self._loading_overlay is not None:
            self._loading_overlay.hide()
        from PySide6.QtWidgets import QApplication
        QApplication.restoreOverrideCursor()

    def _process_single(self, seq: SequenceState):
        # Construir missatge amb info de siblings
        siblings = seq.siblings if hasattr(seq, 'siblings') else []
        if siblings:
            sibling_names = [os.path.basename(s) for s in siblings]
            sibling_info = f"\n\nPack amb {len(siblings)} siblings:\n• " + "\n• ".join(sibling_names)
        else:
            sibling_info = ""

        reply = QMessageBox.question(
            self, "Processar",
            f"Processar {seq.seq_name}?{sibling_info}\n\n"
            f"Executarà: {seq.next_action} i següents",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        self._set_controls_enabled(False)
        self.main_window.set_status(f"Processant {seq.seq_name}...")

        self.single_worker = SingleSeqWorker(seq)
        self.single_worker.progress.connect(
            lambda msg: self.main_window.set_status(f"{seq.seq_name}: {msg}")
        )
        self.single_worker.finished.connect(
            lambda ok, msg: self._on_single_finished(seq.seq_name, ok, msg)
        )
        self.single_worker.start()

    def _on_single_finished(self, seq_name, success, message):
        self._set_controls_enabled(True)
        self.main_window.set_status(
            f"{seq_name}: {message}" if success else f"{seq_name}: ERROR - {message}",
            5000
        )
        self.refresh_sequences()

    def _get_checked_sequences(self):
        """Retorna les SequenceState de les files amb checkbox marcat."""
        checked = []
        for row in range(self.table.rowCount()):
            item_check = self.table.item(row, COL_CHECK)
            if item_check and item_check.checkState() == Qt.Checked:
                item_name = self.table.item(row, COL_NAME)
                if item_name:
                    seq_path = item_name.data(Qt.UserRole + 1)
                    for seq in self.filtered_sequences:
                        if seq.seq_path == seq_path:
                            checked.append(seq)
                            break
        return checked

    def _update_selection_count(self):
        """Actualitza el comptador de seqüències seleccionades."""
        count = len(self._get_checked_sequences())
        self.lbl_selected.setText(f"Seleccionades: {count}")

    def _on_item_changed(self, item):
        """Actualitza comptador quan canvia un checkbox."""
        if item and item.column() == COL_CHECK:
            self._update_selection_count()

    def _select_filtered(self):
        """Marca checkboxes de totes les files visibles."""
        self.table.blockSignals(True)
        for row in range(self.table.rowCount()):
            item_check = self.table.item(row, COL_CHECK)
            if item_check:
                item_check.setCheckState(Qt.Checked)
        self.table.blockSignals(False)
        self._update_selection_count()

    def _deselect_all(self):
        """Desmarca totes les checkboxes."""
        self.table.blockSignals(True)
        for row in range(self.table.rowCount()):
            item_check = self.table.item(row, COL_CHECK)
            if item_check:
                item_check.setCheckState(Qt.Unchecked)
        self.table.blockSignals(False)
        self._update_selection_count()

    def _batch_process_stage(self, phase):
        """Processa les seqüències seleccionades per una etapa (o pipeline complet).

        Args:
            phase: Phase enum (IMPORT, CALIBRATE, ANALYZE) o None per pipeline complet
        """
        target_seqs = self._get_checked_sequences()
        if not target_seqs:
            QMessageBox.information(self, "Info", "Cap seqüència seleccionada.\n\nMarca les seqüències amb el checkbox.")
            return

        if phase is None:
            phases = [Phase.IMPORT, Phase.CALIBRATE, Phase.ANALYZE]
            op_name = "Pipeline complet"
        else:
            phases = [phase]
            op_name = {Phase.IMPORT: "Importar", Phase.CALIBRATE: "Verificar", Phase.ANALYZE: "Analitzar"}.get(phase, str(phase))

        # Confirmació
        seq_names = [s.seq_name for s in target_seqs[:8]]
        if len(target_seqs) > 8:
            seq_names.append(f"... i {len(target_seqs) - 8} més")

        reply = QMessageBox.question(
            self, f"Processar: {op_name}",
            f"{op_name}: {len(target_seqs)} seqüències\n\n"
            + "\n".join(f"  • {n}" for n in seq_names)
            + "\n\nVols continuar?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        # Verificar sensibilitat UIB si importem
        default_uib_sensitivity = None
        if Phase.IMPORT in phases:
            seqs_need_uib = self._get_seqs_needing_uib_sensitivity(target_seqs)
            if seqs_need_uib:
                sens, ok = QInputDialog.getText(
                    self,
                    "Sensibilitat UIB",
                    f"{len(seqs_need_uib)} seqüències DUAL/UIB sense sensibilitat UIB definida.\n"
                    "Indica la sensibilitat UIB per defecte (ex: 700, 1000):\n\n"
                    "Seqüències: " + ", ".join([s.seq_name for s in seqs_need_uib[:5]]) +
                    ("..." if len(seqs_need_uib) > 5 else ""),
                    text="1000"
                )
                if ok and sens.strip():
                    default_uib_sensitivity = sens.strip()

        self._set_controls_enabled(False)

        self.batch_worker = BatchWorker(target_seqs, phases, default_uib_sensitivity)
        self.batch_worker.progress.connect(self._on_batch_progress)
        self.batch_worker.seq_completed.connect(self._on_seq_completed)
        self.batch_worker.finished.connect(self._on_batch_finished)
        self.batch_worker.start()

    def _batch_reset_stage(self, from_stage):
        """Reset per etapa de les seqüències seleccionades (cascade).

        Args:
            from_stage: 0=Importar, 1=Verificar, 2=Analitzar, 3=Exportar
        """
        from hpsec_reset import reset_batch, STAGE_NAMES

        target_seqs = self._get_checked_sequences()
        if not target_seqs:
            QMessageBox.information(self, "Info", "Cap seqüència seleccionada.\n\nMarca les seqüències amb el checkbox.")
            return

        stages_affected = [STAGE_NAMES[s] for s in range(from_stage, 4)]

        seq_names = [s.seq_name for s in target_seqs[:8]]
        if len(target_seqs) > 8:
            seq_names.append(f"... i {len(target_seqs) - 8} més")

        reply = QMessageBox.warning(
            self, "Confirmar Reset",
            f"Reset des de '{STAGE_NAMES[from_stage]}'\n"
            f"Cascade: {' → '.join(stages_affected)}\n\n"
            f"{len(target_seqs)} seqüències:\n"
            + "\n".join(f"  • {n}" for n in seq_names)
            + "\n\nAixò esborrarà JSONs i outputs. Continuar?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        seq_paths = [s.seq_path for s in target_seqs]
        result = reset_batch(seq_paths, from_stage)

        QMessageBox.information(
            self, "Reset completat",
            f"Resetejades: {result['ok']}\nErrors: {result['fail']}"
        )

        self.refresh_sequences()

    def _reset_single(self, seq, from_stage):
        """Reset d'una sola seqüència des d'una etapa determinada."""
        from hpsec_reset import reset_batch, STAGE_NAMES

        stages_affected = [STAGE_NAMES[s] for s in range(from_stage, 4)]

        reply = QMessageBox.warning(
            self, "Confirmar Reset",
            f"Reset {seq.seq_name} des de '{STAGE_NAMES[from_stage]}'\n"
            f"Cascade: {' → '.join(stages_affected)}\n\n"
            "Això esborrarà JSONs i outputs. Continuar?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        result = reset_batch([seq.seq_path], from_stage)

        self.main_window.set_status(
            f"{seq.seq_name}: Reset completat" if result['ok'] else f"{seq.seq_name}: Reset error",
            3000
        )
        self.refresh_sequences()

    def _on_batch_progress(self, current, total, message):
        pct = int(100 * current / total) if total > 0 else 0
        self.main_window.show_progress(pct)
        self.main_window.set_status(f"[{current}/{total}] {message}")

    def _on_seq_completed(self, seq_name, success, message):
        """Actualitza la fila de la seqüència completada a la taula."""
        for row in range(self.table.rowCount()):
            item = self.table.item(row, COL_NAME)
            if not item:
                continue
            seq_path = item.data(Qt.UserRole + 1)
            # Buscar per seq_path (robust, independent del text de display)
            for seq in self.sequences:
                if seq.seq_path == seq_path and seq.seq_name == seq_name:
                    seq.refresh()
                    self._update_table_row(row, seq)
                    break
            else:
                continue
            break
        self._update_stats()

    def _update_table_row(self, row, seq: SequenceState):
        """Actualitza una sola fila de la taula amb l'estat actual de la seqüència."""
        is_cal = "_CAL" in seq.seq_name.upper()
        cal_bg = QColor(COLOR_CAL_BG) if is_cal else None

        # Actualitzar Mostres + Ref + Ctrl
        n_samples = seq.n_samples or 0
        item_mostres = self.table.item(row, COL_MOSTRES)
        if item_mostres:
            item_mostres.setText(str(n_samples) if n_samples else "-")
            item_mostres.setData(Qt.UserRole, n_samples)

        n_khp = seq.n_khp or 0
        n_pr = seq.n_pr or 0
        n_ref = n_khp + n_pr
        item_ref = self.table.item(row, COL_REF)
        if item_ref:
            item_ref.setText(str(n_ref) if n_ref else "-")
            ref_parts = []
            if n_khp:
                ref_parts.append(f"{n_khp} KHP")
            if n_pr:
                ref_parts.append(f"{n_pr} PR")
            item_ref.setToolTip(" · ".join(ref_parts) if ref_parts else "Cap patró")

        n_blank = seq.n_blank or 0
        n_control = seq.n_control or 0
        n_ctrl = n_blank + n_control
        item_ctrl = self.table.item(row, COL_CTRL)
        if item_ctrl:
            item_ctrl.setText(str(n_ctrl) if n_ctrl else "-")
            ctrl_parts = []
            if n_blank:
                ctrl_parts.append(f"{n_blank} Blanc")
            if n_control:
                ctrl_parts.append(f"{n_control} Control")
            item_ctrl.setToolTip(" · ".join(ctrl_parts) if ctrl_parts else "Cap blanc/control")

        phases_data = [
            (seq.import_status, seq.import_state, "Importar", seq.import_warnings),
            (seq.calibrate_status, seq.calibrate_state, "Verificar", seq.calibrate_warnings),
            (seq.analyze_status, seq.analyze_state, "Analitzar", seq.analyze_warnings),
            (seq.review_status, seq.review_state, "Exportar", seq.review_warnings),
        ]

        current_phase_idx = None
        for i, (status, _, _, _) in enumerate(phases_data):
            if not status.completed:
                current_phase_idx = i
                break

        phase_cols = [COL_IMPORT, COL_CAL, COL_ANA, COL_REVIEW]
        for col_offset, (status, state, phase_name, phase_warnings) in enumerate(phases_data):
            col = phase_cols[col_offset]
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)

            font = item.font()
            font.setPointSize(11)
            item.setFont(font)

            if is_cal and col_offset > 0:
                item.setText("—")
                item.setForeground(QColor(COLOR_PENDING))
                item.setToolTip("Flux calibració (sense wizard)")
            elif state == 'ok':
                item.setText("✔")
                item.setForeground(QColor(COLOR_SUCCESS))
                item.setToolTip(f"{phase_name}: Completat")
            elif state == 'warning':
                item.setText("⚠")
                item.setForeground(QColor(COLOR_WARNING))
                warns = phase_warnings[:3] if phase_warnings else []
                tooltip = f"{phase_name}: Avisos"
                if warns:
                    tooltip += "\n" + "\n".join(warns)
                item.setToolTip(tooltip)
            elif state == 'error':
                item.setText("×")
                item.setForeground(QColor(COLOR_ERROR))
                if phase_name == "Importar":
                    item.setToolTip(f"{phase_name}: Error MasterFile")
                else:
                    item.setToolTip(f"{phase_name}: Error")
            elif col_offset == current_phase_idx:
                item.setText("○")
                item.setForeground(QColor(COLOR_CURRENT))
                item.setToolTip(f"{phase_name}: En curs...")
            else:
                item.setText("○")
                item.setForeground(QColor(COLOR_PENDING))
                item.setToolTip(f"{phase_name}: Pendent")

            if cal_bg:
                item.setBackground(cal_bg)
            self.table.setItem(row, col, item)

        # Actualitzar avisos
        warn_preview, warn_tooltip, warn_color = self._format_warnings_cell(seq)
        current_warn = self.table.item(row, COL_WARNINGS)
        if current_warn:
            current_warn.setText(warn_preview)
            current_warn.setToolTip(warn_tooltip)
            current_warn.setForeground(warn_color)

        # Actualitzar notes
        note_preview, note_tooltip, note_color = self._format_notes_only_cell(seq)
        current_notes = self.table.item(row, COL_NOTES)
        if current_notes:
            current_notes.setText(note_preview)
            current_notes.setToolTip(note_tooltip)
            current_notes.setForeground(note_color)

    def _on_batch_finished(self, success, fail):
        self.main_window.show_progress(-1)
        self._set_controls_enabled(True)
        self.refresh_sequences()

        QMessageBox.information(
            self, "Completat",
            f"Correctes: {success}\nErrors: {fail}"
        )

    def _get_seqs_needing_uib_sensitivity(self, sequences):
        """
        Retorna les seqüències DUAL/UIB que no tenen sensibilitat UIB definida.
        Només verifica seqüències que encara no han estat importades.
        """
        need_uib = []
        for seq in sequences:
            # Només seqüències pendents d'importar
            if seq.import_status.completed:
                continue

            # Detectar mode pel nom del directori o estimació
            # Les seqüències entre 269-274 són DUAL (100µL)
            # Les seqüències >= 275 poden ser DUAL (400µL)
            try:
                seq_num = int(seq.seq_name.rstrip("ABCDEF_SEQ").rstrip("_BP"))
            except ValueError:
                continue

            # Heurística: seqüències modernes (>=269) poden ser DUAL
            if seq_num >= 269:
                # Verificar si té MasterFile amb sensibilitat UIB
                seq_path = Path(seq.seq_path)
                master_files = list(seq_path.glob("*MasterFile*.xlsx"))
                if master_files:
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(master_files[0], read_only=True, data_only=True)
                        if "0-INFO" in wb.sheetnames:
                            ws = wb["0-INFO"]
                            uib_sens = ws["B5"].value
                            if uib_sens:
                                continue  # Ja té sensibilitat definida
                    except Exception:
                        pass

                # Si arribem aquí, potencialment necessita sensibilitat UIB
                need_uib.append(seq)

        return need_uib

    def _set_controls_enabled(self, enabled):
        self.refresh_btn.setEnabled(enabled)
        self.btn_batch_process.setEnabled(enabled)
        self.btn_select_filtered.setEnabled(enabled)
        self.btn_deselect_all.setEnabled(enabled)
        self.filter_tipus.setEnabled(enabled)
        self.filter_mode.setEnabled(enabled)
        self.filter_estat.setEnabled(enabled)
        self.search_edit.setEnabled(enabled)
        self.table.setEnabled(enabled)
