"""
HPSEC Suite - Import Panel v3.0
================================

Panel per a la fase 1: Importació de seqüències.
- Columnes separades per punts i fitxers
- Colors segons tipus de match (EXACT/FUZZY/MANUAL/NONE)
- Dropdown per assignació manual d'orfes
- Verificació obligatòria per FUZZY
"""

import os
import json
from pathlib import Path

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QLineEdit, QFrame, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QRadioButton, QButtonGroup, QComboBox,
    QDialog, QDialogButtonBox, QTextEdit, QStyledItemDelegate
)
from PySide6.QtCore import Qt, Signal, QThread
from PySide6.QtGui import QFont, QColor, QBrush

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from hpsec_import import (
    import_sequence, load_manifest, import_from_manifest,
    generate_import_manifest, save_import_manifest,
    llegir_doc_uib, llegir_dad_export3d, llegir_dad_1a,
    get_baseline_value
)
import numpy as np

CONFIG_PATH = Path(__file__).parent.parent.parent / "hpsec_config.json"

# Colors per tipus de match
MATCH_COLORS = {
    "EXACT": QColor("#D5F5E3"),    # Verd clar
    "FUZZY": QColor("#FCF3CF"),    # Groc/taronja clar
    "MANUAL": QColor("#D6EAF8"),   # Blau clar
    "NONE": QColor("#FADBD8"),     # Vermell/rosa clar
    "NORMAL": QColor("#FFFFFF"),   # Blanc
}


def load_sample_types_config():
    """Carrega configuració de tipus de mostra."""
    default_types = {
        "MOSTRA": {"label": "MOSTRA", "color": "#2E86AB", "patterns": []},
        "PATRÓ_CAL": {"label": "PATRÓ_CAL", "color": "#2A9D8F", "patterns": ["KHP"]},
        "PATRÓ_REF": {"label": "PATRÓ_REF", "color": "#9B59B6", "patterns": ["REF", "QC"]},
        "CONTROL": {"label": "CONTROL", "color": "#F6AE2D", "patterns": ["NaOH", "CONTROL"]},
        "BLANC": {"label": "BLANC", "color": "#888888", "patterns": ["MQ", "BLANK", "BLK"]},
    }
    try:
        if CONFIG_PATH.exists():
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config.get("sample_types", default_types)
    except:
        pass
    return default_types


def detect_sample_type(sample_name, original_type, config):
    """Detecta el tipus de mostra basant-se en patrons del config."""
    name_upper = sample_name.upper()

    for type_key, type_info in config.items():
        patterns = type_info.get("patterns", [])
        for pattern in patterns:
            if pattern.upper() in name_upper:
                return type_key

    translations = {
        "SAMPLE": "MOSTRA",
        "KHP": "PATRÓ_CAL",
        "CONTROL": "CONTROL",
        "BLANK": "BLANC",
    }
    return translations.get(original_type.upper(), "MOSTRA")


class ComboBoxDelegate(QStyledItemDelegate):
    """Delegate per editar amb ComboBox al fer doble-clic."""

    def __init__(self, items, parent=None):
        super().__init__(parent)
        self.items = items

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(self.items)
        return combo

    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.DisplayRole)
        idx = editor.findText(value)
        if idx >= 0:
            editor.setCurrentIndex(idx)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class FileAssignmentDelegate(QStyledItemDelegate):
    """Delegate per assignar fitxers amb ComboBox."""

    def __init__(self, get_options_func, parent=None):
        super().__init__(parent)
        self.get_options_func = get_options_func

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        options = self.get_options_func(index.row(), index.column())
        combo.addItems(options)
        return combo

    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.DisplayRole)
        idx = editor.findText(value)
        if idx >= 0:
            editor.setCurrentIndex(idx)
        else:
            editor.setCurrentIndex(0)

    def setModelData(self, editor, model, index):
        new_value = editor.currentText()
        old_value = model.data(index, Qt.DisplayRole)
        # Guardar el fitxer seleccionat (nom real)
        # UserRole ja pot contenir el nom del fitxer original, preservar-lo si no canvia
        current_file = model.data(index, Qt.UserRole)

        # Si l'usuari selecciona un nou fitxer del dropdown, actualitzar UserRole
        if new_value and new_value not in ["-", "(cap)"]:
            # Si és un fitxer diferent al que hi havia, guardar-lo
            if new_value != current_file:
                model.setData(index, new_value, Qt.UserRole)

        # Actualitzar el text mostrat
        model.setData(index, new_value, Qt.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class ImportWorker(QThread):
    """Worker thread per importació asíncrona."""
    progress = Signal(int, str)
    finished = Signal(dict)
    error = Signal(str)

    def __init__(self, seq_path, use_manifest=False, manifest=None):
        super().__init__()
        self.seq_path = seq_path
        self.use_manifest = use_manifest
        self.manifest = manifest

    def run(self):
        try:
            def progress_cb(pct, msg):
                self.progress.emit(int(pct), msg)

            if self.use_manifest and self.manifest:
                result = import_from_manifest(
                    self.seq_path,
                    manifest=self.manifest,
                    progress_callback=progress_cb
                )
            else:
                result = import_sequence(
                    self.seq_path,
                    progress_callback=progress_cb
                )
            self.finished.emit(result)
        except Exception as e:
            import traceback
            self.error.emit(f"{str(e)}\n\n{traceback.format_exc()}")


class OrphanFilesDialog(QDialog):
    """Diàleg per mostrar fitxers orfes."""

    def __init__(self, parent, orphan_files):
        super().__init__(parent)
        self.setWindowTitle("Fitxers Orfes - Revisar Noms")
        self.setMinimumSize(500, 350)

        layout = QVBoxLayout(self)

        info = QLabel(
            "<b>Fitxers no assignats automàticament:</b><br>"
            "Els noms no coincideixen amb el MasterFile. Opcions:<br>"
            "• <b>Revisar noms i reimportar</b>: Verificar que la seqüència del cromatògraf coincideixi amb el MasterFile, després tornar a importar<br>"
            "• <b>Assignar manualment</b>: Doble-clic a la cel·la '-' de la mostra corresponent i seleccionar el fitxer"
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        text = QTextEdit()
        text.setReadOnly(True)

        content = ""
        uib_files = orphan_files.get("uib", [])
        dad_files = orphan_files.get("dad", [])

        if uib_files:
            content += "=== UIB (DOC) ===\n"
            for f in uib_files:
                content += f"  • {Path(f).name}\n"
            content += "\n"

        if dad_files:
            content += "=== DAD (254nm) ===\n"
            for f in dad_files:
                content += f"  • {Path(f).name}\n"

        text.setText(content or "Cap fitxer orfe.")
        layout.addWidget(text)

        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(self.close)
        layout.addWidget(buttons)


class ChromatogramPreviewDialog(QDialog):
    """Diàleg per mostrar preview del cromatograma."""

    def __init__(self, parent, sample_name, replica, sample_data, imported_data):
        super().__init__(parent)
        self.setWindowTitle(f"Preview: {sample_name} (Rep {replica})")
        self.setMinimumSize(900, 600)

        layout = QVBoxLayout(self)

        try:
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
            from matplotlib.figure import Figure

            fig = Figure(figsize=(10, 6), dpi=100)
            canvas = FigureCanvasQTAgg(fig)

            ax1 = fig.add_subplot(111)
            ax2 = ax1.twinx()

            self._plot_data(ax1, ax2, sample_name, replica, imported_data)

            ax1.set_xlabel("Temps (min)")
            ax1.set_ylabel("DOC (mAU)", color="#2E86AB")
            ax2.set_ylabel("DAD 254nm (mAU)", color="#E67E22")
            ax1.set_title(f"{sample_name} - Rèplica {replica}")
            # Només mostrar llegenda si hi ha dades
            if ax1.get_legend_handles_labels()[0]:
                ax1.legend(loc="upper left")
            if ax2.get_legend_handles_labels()[0]:
                ax2.legend(loc="upper right")
            ax1.grid(True, alpha=0.3)

            fig.tight_layout()
            layout.addWidget(canvas)

        except ImportError:
            info = QLabel(f"""
            <h3>{sample_name} - Rèplica {replica}</h3>
            <p><i>Instal·la matplotlib per veure el gràfic:</i></p>
            <code>pip install matplotlib</code>
            """)
            info.setWordWrap(True)
            layout.addWidget(info)

        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(self.close)
        layout.addWidget(buttons)

    def _plot_data(self, ax1, ax2, sample_name, replica, imported_data):
        samples = imported_data.get("samples", {})
        sample_info = samples.get(sample_name)
        if not sample_info:
            ax1.text(0.5, 0.5, "Dades no disponibles", ha='center', va='center')
            return

        rep_data = sample_info.get("replicas", {}).get(str(replica))
        if not rep_data:
            ax1.text(0.5, 0.5, f"Rèplica {replica} no trobada", ha='center', va='center')
            return

        direct = rep_data.get("direct", {})
        if direct and direct.get("t") is not None:
            t = direct["t"]
            y = direct.get("y") if direct.get("y") is not None else direct.get("y_raw")
            if y is not None:
                ax1.plot(t, y, color="#2E86AB", label="DOC Direct", linewidth=1)

        uib = rep_data.get("uib", {})
        if uib and uib.get("t") is not None:
            t = uib["t"]
            y = uib.get("y") if uib.get("y") is not None else uib.get("y_raw")
            if y is not None and len(t) > 0 and len(y) > 0:
                ax1.plot(t, y, color="#27AE60", label="DOC UIB", linewidth=1, alpha=0.8)

        dad = rep_data.get("dad", {})
        if dad:
            # Intentar obtenir dades del DAD (pot ser df o arrays separats)
            t_dad = None
            y254 = None

            # Format 1: DataFrame amb columnes "time (min)" i "254"
            df_dad = dad.get("df")
            try:
                import pandas as pd
                if df_dad is not None and isinstance(df_dad, pd.DataFrame) and len(df_dad) > 0:
                    if "time (min)" in df_dad.columns:
                        t_dad = df_dad["time (min)"].values
                    # Buscar columna 254nm (pot ser "254", "254nm", 254, o similar)
                    for col in df_dad.columns:
                        col_str = str(col)
                        if "254" in col_str or col_str == "254":
                            y254 = df_dad[col].values
                            break
            except Exception as e:
                print(f"Warning: Error llegint DataFrame DAD: {e}")

            # Format 2: Arrays separats (t, wavelengths dict)
            if t_dad is None and dad.get("t") is not None:
                t_dad = dad["t"]
                wavelengths = dad.get("wavelengths", {})
                if 254 in wavelengths or "254" in wavelengths:
                    y254 = wavelengths.get(254) or wavelengths.get("254")

            # Plotar si tenim dades
            if t_dad is not None and y254 is not None:
                ax2.plot(t_dad, y254, color="#E67E22", label="DAD 254nm",
                            linewidth=1, linestyle="--", alpha=0.7)


class ImportPanel(QWidget):
    """Panel d'importació de seqüències."""

    import_completed = Signal(dict)

    # Columnes base (s'ajusten segons mode a _setup_table_columns)
    COL_INJ = 0
    COL_MOSTRA = 1
    COL_TIPUS = 2
    COL_REP = 3
    COL_INJ_VOL = 4  # Volum d'injecció (µL)
    COL_DIRECT_PTS = 5
    COL_DIRECT_FILE = 6
    # Columnes dinàmiques (s'ajusten segons mode de dades)
    # Per DUAL: UIB=7,8, DAD=9,10, Estat=11
    # Per DIRECT: DAD=7,8, Estat=9
    COL_UIB_PTS_ACTUAL = 7
    COL_UIB_FILE_ACTUAL = 8
    COL_DAD_PTS_ACTUAL = 9
    COL_DAD_FILE_ACTUAL = 10
    COL_ESTAT = 11

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.seq_path = None
        self.existing_manifest = None
        self.imported_data = None
        self.worker = None
        self.sample_types_config = load_sample_types_config()
        self._sample_data = []
        self._orphan_files = {"uib": [], "dad": []}
        self._match_types = {}  # (row, col) -> match_type
        self._unverified_fuzzy = set()  # Set of rows needing verification
        self._manual_assignments = {}  # (sample_name, replica) -> {col: filename}
        self._data_mode = "DUAL"  # DUAL, DIRECT, UIB
        self._import_warnings = []  # Warnings d'importació
        self._loaded_from_manifest = False  # Si s'ha carregat des de manifest existent
        self._orphan_warning_dismissed = False  # Si l'usuari ha marcat l'avís d'orfes com revisat

        self._setup_ui()

    def _setup_ui(self):
        """Configura la interfície del panel."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 12)
        layout.setSpacing(8)

        # Camp ocult per compatibilitat (usat internament)
        self.path_input = QLineEdit()
        self.path_input.setVisible(False)
        layout.addWidget(self.path_input)

        # Botons ocults per compatibilitat (accions ara al header del wizard)
        self.import_btn = QPushButton()
        self.import_btn.setVisible(False)
        self.save_btn = QPushButton()
        self.save_btn.setVisible(False)
        self.next_btn = QPushButton()
        self.next_btn.setVisible(False)

        # === MANIFEST INFO (eliminat - redundant amb header del wizard) ===
        # El nom SEQ ja apareix al header i el botó "Executar" gestiona reimportació
        self.manifest_frame = QFrame()
        self.manifest_frame.setVisible(False)  # Mai visible - mantingut per compatibilitat
        self.manifest_info = QLabel()
        self.use_manifest_radio = QRadioButton()
        self.full_import_radio = QRadioButton()
        self.import_mode_group = QButtonGroup(self)

        # === BARRA D'AVISOS (consistent per tots els panels) ===
        self.warnings_bar = QFrame()
        self.warnings_bar.setVisible(False)
        self.warnings_bar.setStyleSheet("""
            QFrame {
                background-color: #fff3cd;
                border: 1px solid #ffc107;
                border-radius: 6px;
            }
        """)
        warnings_bar_layout = QHBoxLayout(self.warnings_bar)
        warnings_bar_layout.setContentsMargins(12, 8, 12, 8)

        warnings_icon = QLabel("⚠")
        warnings_icon.setStyleSheet("font-size: 16px;")
        warnings_bar_layout.addWidget(warnings_icon)

        self.warnings_bar_text = QLabel()
        self.warnings_bar_text.setStyleSheet("color: #856404;")
        self.warnings_bar_text.setWordWrap(True)
        warnings_bar_layout.addWidget(self.warnings_bar_text, 1)

        layout.addWidget(self.warnings_bar)

        # === INFO BARRA (resum injeccions) ===
        self.info_frame = QFrame()
        self.info_frame.setVisible(False)
        self.info_frame.setFixedHeight(28)
        info_layout = QHBoxLayout(self.info_frame)
        info_layout.setContentsMargins(0, 2, 0, 2)
        info_layout.setSpacing(16)

        self.total_label = QLabel()
        self.total_label.setStyleSheet("font-weight: bold; color: #2E86AB;")
        info_layout.addWidget(self.total_label)

        info_layout.addStretch()

        # Comptador de fitxers UIB i DAD
        self.files_label = QLabel()
        self.files_label.setStyleSheet("color: #666;")
        info_layout.addWidget(self.files_label)

        layout.addWidget(self.info_frame)

        # === TAULA DE MOSTRES ===
        # Nota d'ajuda
        self.table_help = QLabel("💡 Doble-clic a una fila per veure la gràfica")
        self.table_help.setStyleSheet("color: #666; font-size: 11px; padding: 2px;")
        self.table_help.setVisible(False)
        layout.addWidget(self.table_help)

        self.samples_table = QTableWidget()
        self.samples_table.setToolTip("Doble-clic per veure gràfica DOC + DAD 254nm")

        # Amagar numeració automàtica de files
        self.samples_table.verticalHeader().setVisible(False)

        # NO usar colors alternats (interfereixen amb colors de match)
        self.samples_table.setAlternatingRowColors(False)

        # Permetre ordenar
        self.samples_table.setSortingEnabled(True)

        self.samples_table.cellDoubleClicked.connect(self._on_cell_double_clicked)
        self.samples_table.cellChanged.connect(self._on_cell_changed)
        self.samples_table.setEditTriggers(
            QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed
        )
        self.samples_table.setVisible(False)

        layout.addWidget(self.samples_table, 1)

        # === WARNINGS / ORFES ===
        self.warnings_frame = QFrame()
        self.warnings_frame.setVisible(False)
        self.warnings_frame.setFixedHeight(48)  # Espai suficient pels botons
        warnings_layout = QHBoxLayout(self.warnings_frame)
        warnings_layout.setContentsMargins(8, 6, 8, 6)

        self.warnings_label = QLabel()
        self.warnings_label.setStyleSheet("color: #E67E22;")
        warnings_layout.addWidget(self.warnings_label, 1)

        self.confirm_btn = QPushButton("Confirmar suggeriments")
        self.confirm_btn.setToolTip("Confirma tots els suggeriments automàtics")
        self.confirm_btn.clicked.connect(self._confirm_all_suggestions)
        self.confirm_btn.setVisible(False)
        warnings_layout.addWidget(self.confirm_btn)

        self.orphans_btn = QPushButton("Veure orfes")
        self.orphans_btn.clicked.connect(self._show_orphans)
        self.orphans_btn.setVisible(False)
        warnings_layout.addWidget(self.orphans_btn)

        self.refresh_btn = QPushButton("Actualitzar")
        self.refresh_btn.setToolTip("Actualitza el comptador d'orfes després d'assignacions manuals")
        self.refresh_btn.clicked.connect(self._refresh_orphan_count)
        self.refresh_btn.setVisible(False)
        warnings_layout.addWidget(self.refresh_btn)

        self.dismiss_btn = QPushButton("Marcar revisat")
        self.dismiss_btn.setToolTip("Marca l'avís com a revisat (els fitxers orfes no s'assignaran)")
        self.dismiss_btn.clicked.connect(self._dismiss_orphan_warning)
        self.dismiss_btn.setVisible(False)
        warnings_layout.addWidget(self.dismiss_btn)

        layout.addWidget(self.warnings_frame)

        # === PLACEHOLDER ===
        self.placeholder = QLabel("Selecciona una carpeta SEQ i prem 'Importar'")
        self.placeholder.setAlignment(Qt.AlignCenter)
        self.placeholder.setStyleSheet("color: #888; font-size: 14px;")
        layout.addWidget(self.placeholder, 1)

    def _get_file_options(self, row, col):
        """Retorna opcions pel dropdown de fitxers."""
        options = ["(cap)"]

        # Determinar si és UIB o DAD segons la columna
        if self._data_mode == "DIRECT":
            if col == self.COL_DAD_FILE_ACTUAL:
                orphans = self._orphan_files.get("dad", [])
            else:
                return options
        else:
            if col == self.COL_UIB_FILE_ACTUAL:
                orphans = self._orphan_files.get("uib", [])
            elif col == self.COL_DAD_FILE_ACTUAL:
                orphans = self._orphan_files.get("dad", [])
            else:
                return options

        # Afegir valor actual si existeix
        current = self.samples_table.item(row, col)
        if current and current.text() and current.text() != "-" and current.text() != "(cap)":
            options.append(current.text())

        # Afegir orfes
        for f in orphans:
            fname = Path(f).name if "/" in f or "\\" in f else f
            if fname not in options:
                options.append(fname)

        return options

    def set_sequence_path(self, path):
        self.seq_path = path
        self.main_window.seq_path = path  # Actualitzar també el main_window
        self.path_input.setText(path)
        self._check_manifest()

    def load_from_dashboard(self, seq_path):
        """Carrega una seqüència des del Dashboard - auto-carrega si hi ha manifest."""
        self.set_sequence_path(seq_path)

        # Si hi ha manifest existent, carregar automàticament
        if self.existing_manifest:
            # Amagar placeholder immediatament - es mostrarà la taula quan acabi
            self.placeholder.setVisible(False)
            self._auto_load_from_manifest()

    def _go_to_dashboard(self):
        """Torna al Dashboard per seleccionar una seqüència."""
        self.main_window.tab_widget.setCurrentIndex(0)

    def _check_manifest(self):
        self.existing_manifest = load_manifest(self.seq_path)

        if self.existing_manifest:
            self._show_manifest_info()
        else:
            self.manifest_frame.setVisible(False)

        self.import_btn.setEnabled(True)
        self.samples_table.setVisible(False)
        self.table_help.setVisible(False)
        self.info_frame.setVisible(False)
        self.warnings_frame.setVisible(False)
        self.placeholder.setVisible(True)
        self.next_btn.setEnabled(False)
        self.save_btn.setEnabled(False)

    def _show_manifest_info(self):
        """Registra info del manifest (frame eliminat - info ja al header wizard)."""
        # Mantingut per compatibilitat però ja no mostra res
        # La info del SEQ es mostra al header del ProcessWizardPanel
        pass

    def _auto_load_from_manifest(self):
        """Carrega automàticament des del manifest existent."""
        self._loaded_from_manifest = True
        self.main_window.show_progress(0)

        self.worker = ImportWorker(
            self.seq_path,
            use_manifest=True,
            manifest=self.existing_manifest
        )
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_import_finished)
        self.worker.error.connect(self._on_import_error)
        self.worker.start()

    def _run_import(self, force_reimport=False):
        """Executa importació. Si force_reimport=True, reimporta tot."""
        if not self.seq_path:
            return

        self.main_window.show_progress(0)

        # Si ja hi ha manifest i no forcem reimportació, usar-lo
        use_manifest = self.existing_manifest and not force_reimport
        self._loaded_from_manifest = use_manifest

        self.worker = ImportWorker(
            self.seq_path,
            use_manifest=use_manifest,
            manifest=self.existing_manifest if use_manifest else None
        )
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_import_finished)
        self.worker.error.connect(self._on_import_error)
        self.worker.start()

    def _on_progress(self, pct, msg):
        self.main_window.show_progress(pct)
        self.main_window.set_status(msg)

    def _on_import_finished(self, result):
        self.main_window.show_progress(-1)
        self.import_btn.setEnabled(True)

        if not result.get("success"):
            errors = result.get("errors", ["Error desconegut"])
            QMessageBox.critical(self, "Error d'Importació", f"Error: {errors[0]}")
            return

        # Verificar si cal preguntar la sensibilitat UIB
        self._check_uib_sensitivity(result)

        self.imported_data = result
        self.main_window.imported_data = result
        # Restaurar estat "revisat" d'orfes si es va marcar anteriorment
        self._orphan_warning_dismissed = result.get("orphan_warning_dismissed", False)
        self._show_results(result)

        # Mostrar avisos si n'hi ha
        self._show_warnings_bar(result)

        try:
            save_import_manifest(result)
            self.main_window.mark_manifest_saved()
        except Exception as e:
            print(f"Warning: No s'ha pogut guardar manifest: {e}")

        self._update_next_button_state()
        self.save_btn.setEnabled(True)
        self.main_window.enable_tab(1)
        self.main_window.set_status("Importació completada", 5000)

    def _show_warnings_bar(self, result):
        """Mostra avisos a la barra superior si n'hi ha."""
        warnings = result.get("warnings", [])
        orphans_uib = len(result.get("orphan_files", {}).get("uib", []))
        orphans_dad = len(result.get("orphan_files", {}).get("dad", []))

        warning_parts = []
        if orphans_uib > 0:
            warning_parts.append(f"{orphans_uib} UIB orfes")
        if orphans_dad > 0:
            warning_parts.append(f"{orphans_dad} DAD orfes")

        # Afegir warnings generals
        for w in warnings:
            if "⚠" in w or "FUZZY" in w:
                warning_parts.append(w.replace("⚠️", "").strip()[:40])

        if warning_parts:
            self.warnings_bar.setVisible(True)
            n = len(warning_parts)
            display = warning_parts[:3]
            self.warnings_bar_text.setText(
                f"<b>{n} avisos:</b> " + " · ".join(display)
            )
        else:
            self.warnings_bar.setVisible(False)

    def _check_uib_sensitivity(self, result):
        """Verifica si cal preguntar la sensibilitat UIB i actualitza el MasterFile."""
        # Només per mode DUAL o UIB
        data_mode = result.get("data_mode", "")
        if data_mode not in ["DUAL", "UIB"]:
            return

        # Verificar si la sensibilitat UIB està definida
        uib_sens = result.get("uib_sensitivity")
        if uib_sens is not None and uib_sens not in ["None", "", None]:
            return

        # Preguntar a l'usuari (camp lliure)
        from PySide6.QtWidgets import QInputDialog
        text, ok = QInputDialog.getText(
            self,
            "Sensibilitat UIB",
            "Indica la sensibilitat UIB (ex: 700, 1000, o deixa buit si no aplica):",
            text=""
        )

        if ok and text.strip():
            sens_value = text.strip()

            # Actualitzar el resultat
            result["uib_sensitivity"] = sens_value

            # Actualitzar el MasterFile si existeix
            master_file = result.get("master_file")
            if master_file and os.path.exists(master_file):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(master_file)
                    if "0-INFO" in wb.sheetnames:
                        ws = wb["0-INFO"]
                        ws["B5"] = sens_value
                        wb.save(master_file)
                        print(f"Actualitzat UIB sensitivity a MasterFile: {sens_value}")
                except Exception as e:
                    print(f"Warning: No s'ha pogut actualitzar MasterFile: {e}")

    def _on_import_error(self, error_msg):
        self.main_window.show_progress(-1)
        self.import_btn.setEnabled(True)
        QMessageBox.critical(self, "Error", f"Error durant la importació:\n{error_msg}")

    def _show_results(self, result):
        self.placeholder.setVisible(False)
        self._match_types = {}
        self._unverified_fuzzy = set()
        self._manual_assignments = {}
        # Restaurar estat d'avís d'orfes (si carregat des de manifest)
        if not self._loaded_from_manifest:
            self._orphan_warning_dismissed = False

        # Guardar warnings d'importació per mostrar-los
        self._import_warnings = [w for w in result.get("warnings", []) if "⚠️" in w]

        manifest = generate_import_manifest(result)
        samples = manifest.get("samples", [])
        seq_info = manifest.get("sequence", {})

        # Detectar mode de dades
        self._data_mode = seq_info.get("data_mode", "DUAL")

        # Orfes del manifest (noms) per mostrar - només els no suggerits
        self._orphan_files = {
            "uib": manifest.get("orphan_files", {}).get("uib", []),
            "dad": manifest.get("orphan_files", {}).get("dad", []),
        }
        # TOTS els orfes amb paths complets (per comptar punts, inclou suggerits)
        self._orphan_files_full = {
            "uib": result.get("all_orphan_files", result.get("orphan_files", {})).get("uib", []),
            "dad": result.get("all_orphan_files", result.get("orphan_files", {})).get("dad", []),
        }

        # Si carregat des de manifest, filtrar orfes que ja estan assignats a mostres
        if self._loaded_from_manifest:
            assigned_uib = set()
            assigned_dad = set()
            for sample in samples:
                for rep in sample.get("replicas", []):
                    uib_info = rep.get("uib", {})
                    dad_info = rep.get("dad", {})
                    if uib_info and uib_info.get("file"):
                        assigned_uib.add(Path(uib_info["file"]).name)
                    if dad_info and dad_info.get("file") and not dad_info.get("file", "").startswith("["):
                        assigned_dad.add(Path(dad_info["file"]).name)
            # Treure fitxers assignats de la llista d'orfes
            self._orphan_files["uib"] = [f for f in self._orphan_files["uib"] if Path(f).name not in assigned_uib]
            self._orphan_files["dad"] = [f for f in self._orphan_files["dad"] if Path(f).name not in assigned_dad]


        # Configurar columnes segons mode
        self._setup_table_columns()

        # Crear llista plana de totes les injeccions i ordenar per line_num
        all_injections = []
        for sample in samples:
            original_type = sample.get("type", "SAMPLE")
            sample_type = detect_sample_type(
                sample["name"], original_type, self.sample_types_config
            )
            # Obtenir nom original (pot diferir del nom únic per controls repetits)
            original_name = sample.get("original_name", sample["name"])

            for rep in sample.get("replicas", []):
                # Usar line_num (ordre real d'injecció al MasterFile)
                # Nota: import_sequence usa "injection_info", manifest usa "injection"
                inj_info = rep.get("injection_info") or rep.get("injection") or {}
                line_num = inj_info.get("line_num")
                if line_num is None:
                    # Fallback a row_start si line_num no disponible
                    d = rep.get("direct", {})
                    line_num = d.get("row_start") if d else 999999
                if line_num is None:
                    line_num = 999999

                all_injections.append({
                    "sample_name": sample["name"],
                    "original_name": original_name,  # Nom original del MasterFile
                    "sample_type": sample_type,
                    "rep": rep,
                    "line_num": line_num,
                })

        # Ordenar per line_num (ordre real d'injecció al MasterFile)
        all_injections.sort(key=lambda x: (x["line_num"], x["sample_name"]))

        total_injections = len(all_injections)

        # Comparar amb el nombre de línies del MasterFile (Line#)
        stats = result.get("stats", {})
        master_line_count = stats.get("master_line_count", result.get("master_line_count", total_injections))

        # Recollir volums d'injecció
        volumes = []
        for inj in all_injections:
            rep = inj.get("rep", {})
            # Nota: import_sequence usa "injection_info", manifest usa "injection"
            inj_info = rep.get("injection_info") or rep.get("injection") or {}
            vol = inj_info.get("inj_volume")
            if vol is not None:
                volumes.append(vol)

        # Construir resum de la seqüència
        method = result.get("method", "COLUMN")  # COLUMN o BP
        info_parts = []

        # Injeccions (amb warning si no coincideixen)
        if master_line_count > total_injections:
            info_parts.append(f"⚠️ {total_injections}/{master_line_count} inj")
            has_warning = True
        else:
            info_parts.append(f"{total_injections} inj")
            has_warning = False

        # Mètode
        info_parts.append(method)

        # Mode dades
        info_parts.append(self._data_mode)

        # Volum d'injecció
        if volumes:
            vol_min = min(volumes)
            vol_max = max(volumes)
            if vol_min == vol_max:
                info_parts.append(f"{int(vol_min)}µL")
            else:
                info_parts.append(f"{int(vol_min)}-{int(vol_max)}µL")

        # Sensibilitat UIB (si disponible i mode DUAL o UIB)
        if self._data_mode in ["DUAL", "UIB"]:
            uib_sens = result.get("uib_sensitivity")
            if uib_sens is not None:
                # Mostrar com "700ppb" o "1000ppb"
                try:
                    sens_val = int(float(uib_sens))
                    info_parts.append(f"UIB:{sens_val}ppb")
                except (ValueError, TypeError):
                    # Si no és numèric, mostrar tal qual
                    info_parts.append(f"UIB:{uib_sens}")

        self.total_label.setText(" · ".join(info_parts))
        if has_warning:
            self.total_label.setStyleSheet("font-weight: bold; color: #E74C3C;")  # Vermell
        else:
            self.total_label.setStyleSheet("font-weight: bold; color: #2E86AB;")  # Blau normal

        # Comptar fitxers UIB i DAD
        stats = result.get("stats", {})
        uib_used = stats.get("uib_files_used", 0)
        uib_orphan = stats.get("orphan_uib", 0)
        dad_used = stats.get("dad_files_used", 0)
        dad_orphan = stats.get("orphan_dad", 0)

        uib_total = uib_used + uib_orphan
        dad_total = dad_used + dad_orphan

        files_parts = []
        if self._data_mode in ["DUAL", "UIB"]:
            files_parts.append(f"UIB: {uib_total}")
        files_parts.append(f"DAD: {dad_total}")
        self.files_label.setText(" · ".join(files_parts))

        self.info_frame.setVisible(True)

        # Bloquejar signals i sorting durant l'ompliment
        self.samples_table.setSortingEnabled(False)
        self.samples_table.blockSignals(True)
        self.samples_table.setRowCount(0)
        self._sample_data = []

        for injection_num, inj in enumerate(all_injections, 1):
            row = self.samples_table.rowCount()
            self.samples_table.insertRow(row)

            sample_name = inj["sample_name"]
            original_name = inj.get("original_name", sample_name)
            sample_type = inj["sample_type"]
            rep = inj["rep"]

            # Inj (no editable)
            inj_item = QTableWidgetItem()
            inj_item.setData(Qt.DisplayRole, injection_num)
            inj_item.setTextAlignment(Qt.AlignCenter)
            inj_item.setFlags(inj_item.flags() & ~Qt.ItemIsEditable)
            self.samples_table.setItem(row, self.COL_INJ, inj_item)

            # Mostra (no editable) - mostrar nom original si diferent
            if original_name != sample_name:
                # Nom transformat (controls repetits): mostrar original
                display_name = original_name
                name_item = QTableWidgetItem(display_name)
                name_item.setToolTip(f"Nom únic: {sample_name}\nNom MasterFile: {original_name}")
                # Color diferent per indicar transformació
                name_item.setForeground(QBrush(QColor("#2E86AB")))
            else:
                name_item = QTableWidgetItem(sample_name)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            # Guardar nom únic al UserRole per ús intern
            name_item.setData(Qt.UserRole, sample_name)
            self.samples_table.setItem(row, self.COL_MOSTRA, name_item)

            # Tipus (editable)
            type_item = QTableWidgetItem(sample_type)
            type_item.setTextAlignment(Qt.AlignCenter)
            self.samples_table.setItem(row, self.COL_TIPUS, type_item)

            # Rep (no editable)
            rep_item = QTableWidgetItem(str(rep.get("replica", "?")))
            rep_item.setTextAlignment(Qt.AlignCenter)
            rep_item.setFlags(rep_item.flags() & ~Qt.ItemIsEditable)
            self.samples_table.setItem(row, self.COL_REP, rep_item)

            # Volum d'injecció (no editable)
            # Nota: import_sequence usa "injection_info", manifest usa "injection"
            inj_info = rep.get("injection_info") or rep.get("injection") or {}
            inj_vol = inj_info.get("inj_volume") if inj_info else None
            vol_text = f"{int(inj_vol)}" if inj_vol else "-"
            vol_item = QTableWidgetItem(vol_text)
            vol_item.setTextAlignment(Qt.AlignCenter)
            vol_item.setFlags(vol_item.flags() & ~Qt.ItemIsEditable)
            self.samples_table.setItem(row, self.COL_INJ_VOL, vol_item)

            # Dades per cada senyal
            d = rep.get("direct", {})
            u = rep.get("uib", {})
            dad = rep.get("dad", {})

            # Direct
            direct_pts = d.get("n_points", 0) if d else 0
            row_start = d.get("row_start", "") if d else ""
            row_end = d.get("row_end", "") if d else ""
            direct_file = f"{row_start}-{row_end}" if row_start and row_end else "-"

            self._add_simple_cell(row, self.COL_DIRECT_PTS, str(direct_pts) if direct_pts else "-")
            self._add_simple_cell(row, self.COL_DIRECT_FILE, direct_file)

            # MOSTRA i PATRÓ sempre requereixen assignació (obligatori)
            # CONTROL i BLANC poden assignar manualment si hi ha orfes (opcional)
            requires_assignment = sample_type in ["MOSTRA", "PATRÓ_CAL", "PATRÓ_REF"]
            optional_can_assign = sample_type in ["CONTROL", "BLANC"]  # Permeten assignació manual opcional

            # Controlar què falta per assignar
            missing_signals = []
            review_signals = []  # Senyals amb suggeriment per revisar
            needs_review = False

            # UIB (només si mode DUAL o UIB)
            if self._data_mode in ["DUAL", "UIB"]:
                uib_pts = u.get("n_points", 0) if u else 0
                uib_file = u.get("file", "") if u else ""
                uib_suggestion = rep.get("uib_suggestion")

                if uib_file:
                    uib_file = Path(uib_file).name

                self._add_simple_cell(row, self.COL_UIB_PTS_ACTUAL, str(uib_pts) if uib_pts else "-")

                if uib_suggestion and (requires_assignment or optional_can_assign):
                    # Hi ha un suggeriment de matching (mostra/patró/control)
                    suggested_file = uib_suggestion.get("file", "")
                    confidence = uib_suggestion.get("confidence", 0)
                    # Nom segons llistat d'injeccions
                    replica_num = rep.get("replica", 1)
                    display_name = f"{sample_name}_R{replica_num}"

                    # Si carregat des de manifest, el suggeriment ja estava confirmat
                    if self._loaded_from_manifest:
                        self._add_simple_cell(row, self.COL_UIB_FILE_ACTUAL, display_name)
                        # No cal revisar - ja estava confirmat
                    else:
                        self._add_suggestion_cell(row, self.COL_UIB_FILE_ACTUAL, suggested_file, confidence, display_name)
                        if requires_assignment:
                            review_signals.append(f"UIB {int(confidence)}%")
                        needs_review = True

                    # Comptar punts del fitxer suggerit
                    n_points = self._count_file_points(suggested_file, "uib")
                    if n_points > 0:
                        self.samples_table.item(row, self.COL_UIB_PTS_ACTUAL).setText(str(n_points))
                        uib_pts = n_points
                elif uib_file and self._loaded_from_manifest:
                    # FITXER JA ASSIGNAT (carregat des de manifest) - mostrar com a OK
                    replica_num = rep.get("replica", 1)
                    display_name = f"{sample_name}_R{replica_num}"
                    self._add_simple_cell(row, self.COL_UIB_FILE_ACTUAL, display_name)
                    # Intentar comptar punts si no tenim
                    if not uib_pts:
                        n_points = self._count_file_points(uib_file, "uib")
                        if n_points > 0:
                            self.samples_table.item(row, self.COL_UIB_PTS_ACTUAL).setText(str(n_points))
                elif not uib_pts and not uib_file and self._orphan_files.get("uib") and requires_assignment:
                    # NO hi ha fitxer assignat i hi ha orfes UIB disponibles
                    self._add_file_cell(row, self.COL_UIB_FILE_ACTUAL, "-", editable=True)
                    missing_signals.append("UIB")
                    needs_review = True
                elif optional_can_assign and self._orphan_files.get("uib"):
                    # CONTROL: sempre permetre editar si hi ha orfes (per corregir assignacions)
                    display_val = uib_file if uib_file else "-"
                    self._add_file_cell(row, self.COL_UIB_FILE_ACTUAL, display_val, editable=True)
                else:
                    self._add_simple_cell(row, self.COL_UIB_FILE_ACTUAL, uib_file if uib_file else "-")

            # DAD
            dad_pts = dad.get("n_points", 0) if dad else 0
            dad_suggestion = rep.get("dad_suggestion")

            # Obtenir el fitxer DAD
            dad_file = ""
            if dad:
                dad_file = dad.get("file", "")
                # Si no hi ha fitxer però hi ha dades, indicar la font
                if not dad_file and dad_pts > 0:
                    source = dad.get("source", "")
                    if source == "masterfile":
                        dad_file = "[MasterFile]"
                    elif source in ["export3d", "csv"]:
                        dad_file = f"[{source}]"
                elif dad_file:
                    # Assegurar que només mostrem el nom del fitxer
                    dad_file = Path(dad_file).name if "/" in dad_file or "\\" in dad_file else dad_file

            self._add_simple_cell(row, self.COL_DAD_PTS_ACTUAL, str(dad_pts) if dad_pts else "-")

            if dad_suggestion and (requires_assignment or optional_can_assign):
                # Hi ha un suggeriment de matching (mostra/patró/control)
                suggested_file = dad_suggestion.get("file", "")
                confidence = dad_suggestion.get("confidence", 0)
                # Nom segons llistat d'injeccions
                replica_num = rep.get("replica", 1)
                display_name = f"{sample_name}_R{replica_num}"

                # Si carregat des de manifest, el suggeriment ja estava confirmat
                if self._loaded_from_manifest:
                    self._add_simple_cell(row, self.COL_DAD_FILE_ACTUAL, display_name)
                    # No cal revisar - ja estava confirmat
                else:
                    self._add_suggestion_cell(row, self.COL_DAD_FILE_ACTUAL, suggested_file, confidence, display_name)
                    if requires_assignment:
                        review_signals.append(f"DAD {int(confidence)}%")
                    needs_review = True

                # Comptar punts del fitxer suggerit
                n_points = self._count_file_points(suggested_file, "dad")
                if n_points > 0:
                    self.samples_table.item(row, self.COL_DAD_PTS_ACTUAL).setText(str(n_points))
                    dad_pts = n_points
            elif dad_file and self._loaded_from_manifest and not dad_file.startswith("["):
                # FITXER JA ASSIGNAT (carregat des de manifest) - mostrar com a OK
                replica_num = rep.get("replica", 1)
                display_name = f"{sample_name}_R{replica_num}"
                self._add_simple_cell(row, self.COL_DAD_FILE_ACTUAL, display_name)
                # Intentar comptar punts si no tenim
                if not dad_pts:
                    n_points = self._count_file_points(dad_file, "dad")
                    if n_points > 0:
                        self.samples_table.item(row, self.COL_DAD_PTS_ACTUAL).setText(str(n_points))
            elif not dad_pts and not dad_file and self._orphan_files.get("dad") and requires_assignment:
                # NO hi ha fitxer assignat i hi ha orfes DAD disponibles
                self._add_file_cell(row, self.COL_DAD_FILE_ACTUAL, "-", editable=True)
                missing_signals.append("DAD")
                needs_review = True
            elif optional_can_assign and self._orphan_files.get("dad"):
                # CONTROL: sempre permetre editar si hi ha orfes (per corregir assignacions)
                display_val = dad_file if dad_file else "-"
                self._add_file_cell(row, self.COL_DAD_FILE_ACTUAL, display_val, editable=True)
            else:
                self._add_simple_cell(row, self.COL_DAD_FILE_ACTUAL, dad_file if dad_file else "-")

            # Columna Estat - específic segons què falta o cal revisar
            if review_signals:
                estat = "Revisar " + ", ".join(review_signals)
                color = QColor("#FCF3CF")  # Groc
            elif missing_signals:
                estat = "Assignar " + "+".join(missing_signals)
                color = QColor("#FADBD8")  # Rosa
            else:
                estat = "OK"
                color = QColor("#D5F5E3")  # Verd

            estat_item = QTableWidgetItem(estat)
            estat_item.setTextAlignment(Qt.AlignCenter)
            estat_item.setFlags(estat_item.flags() & ~Qt.ItemIsEditable)
            estat_item.setBackground(QBrush(color))
            self.samples_table.setItem(row, self.COL_ESTAT, estat_item)

            if needs_review:
                self._unverified_fuzzy.add(row)

            # Guardar per preview i lògica
            self._sample_data.append({
                "name": sample_name,
                "type": sample_type,  # Guardar tipus per lògica d'assignació
                "replica": rep.get("replica"),
                "direct_pts": direct_pts,
                "uib_pts": u.get("n_points", 0) if u else 0,
                "dad_pts": dad_pts,
            })

        self.samples_table.blockSignals(False)
        self.samples_table.setSortingEnabled(True)
        self.samples_table.setVisible(True)
        self.table_help.setVisible(True)

        # Warnings
        self._update_warnings()

    def _add_simple_cell(self, row, col, text):
        """Afegeix una cel·la simple no editable."""
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        self.samples_table.setItem(row, col, item)

    def _add_file_cell(self, row, col, text, editable=False):
        """Afegeix una cel·la de fitxer, potencialment editable."""
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        else:
            item.setBackground(QBrush(QColor("#FADBD8")))  # Rosa per indicar que cal assignar
        self.samples_table.setItem(row, col, item)

    def _add_suggestion_cell(self, row, col, filename, confidence, display_name=None):
        """Afegeix una cel·la amb suggeriment de matching (editable per confirmar/canviar).

        Args:
            row: Fila de la taula
            col: Columna de la taula
            filename: Nom real del fitxer orfe
            confidence: Percentatge de confiança del match
            display_name: Nom a mostrar (format llistat injeccions). Si None, usa filename.
        """
        show_name = display_name if display_name else filename
        item = QTableWidgetItem(show_name)
        item.setTextAlignment(Qt.AlignCenter)
        # Editable per permetre canviar el suggeriment
        item.setBackground(QBrush(QColor("#FCF3CF")))  # Groc per indicar revisar
        # Tooltip mostra el nom real del fitxer orfe
        item.setToolTip(f"Fitxer: {filename}\nConfiança: {confidence:.0f}%\nDoble-clic per canviar.")
        # Guardar el path real del fitxer com a data
        item.setData(Qt.UserRole, filename)
        self.samples_table.setItem(row, col, item)
        # Marcar com a suggeriment
        self._match_types[(row, col)] = "SUGGESTED"

    def _setup_table_columns(self):
        """Configura les columnes segons el mode de dades."""
        if self._data_mode == "DIRECT":
            # Sense UIB: Inj, Mostra, Tipus, Rep, Vol, Direct, Fila, DAD, Fitxer DAD, Estat
            self.samples_table.setColumnCount(10)
            headers = ["Inj", "Mostra", "Tipus", "Rep", "Inj Vol", "Direct", "Fila", "DAD", "Fitxer DAD", "Estat"]
            self.COL_DAD_PTS_ACTUAL = 7
            self.COL_DAD_FILE_ACTUAL = 8
            self.COL_ESTAT = 9
        else:
            # DUAL o UIB: Inj, Mostra, Tipus, Rep, Vol, Direct, Fila, UIB, Fitxer UIB, DAD, Fitxer DAD, Estat
            self.samples_table.setColumnCount(12)
            headers = ["Inj", "Mostra", "Tipus", "Rep", "Inj Vol", "Direct", "Fila", "UIB", "Fitxer UIB", "DAD", "Fitxer DAD", "Estat"]
            self.COL_UIB_PTS_ACTUAL = 7
            self.COL_UIB_FILE_ACTUAL = 8
            self.COL_DAD_PTS_ACTUAL = 9
            self.COL_DAD_FILE_ACTUAL = 10
            self.COL_ESTAT = 11

        self.samples_table.setHorizontalHeaderLabels(headers)

        # Configurar delegates
        type_delegate = ComboBoxDelegate(list(self.sample_types_config.keys()), self)
        self.samples_table.setItemDelegateForColumn(self.COL_TIPUS, type_delegate)

        if self._data_mode == "DIRECT":
            dad_delegate = FileAssignmentDelegate(self._get_file_options, self)
            self.samples_table.setItemDelegateForColumn(self.COL_DAD_FILE_ACTUAL, dad_delegate)
        else:
            uib_delegate = FileAssignmentDelegate(self._get_file_options, self)
            dad_delegate = FileAssignmentDelegate(self._get_file_options, self)
            self.samples_table.setItemDelegateForColumn(self.COL_UIB_FILE_ACTUAL, uib_delegate)
            self.samples_table.setItemDelegateForColumn(self.COL_DAD_FILE_ACTUAL, dad_delegate)

        # Configurar mides
        header = self.samples_table.horizontalHeader()
        for col in range(self.samples_table.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.ResizeToContents)

    def _add_data_cell(self, row, col, text, match_type, editable=False):
        """Afegeix una cel·la amb color segons match_type."""
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)

        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)

        # Color de fons
        color = MATCH_COLORS.get(match_type, MATCH_COLORS["NORMAL"])
        item.setBackground(QBrush(color))

        # Guardar match type
        self._match_types[(row, col)] = match_type
        if match_type == "FUZZY":
            self._unverified_fuzzy.add((row, col))

        self.samples_table.setItem(row, col, item)

    def _on_cell_changed(self, row, col):
        """Handler quan canvia una cel·la."""
        # Determinar si és una columna de fitxer
        is_file_col = False
        pts_col = None
        file_type = None

        if self._data_mode == "DIRECT":
            if col == self.COL_DAD_FILE_ACTUAL:
                is_file_col = True
                pts_col = self.COL_DAD_PTS_ACTUAL
                file_type = "dad"
        else:
            if col == self.COL_UIB_FILE_ACTUAL:
                is_file_col = True
                pts_col = self.COL_UIB_PTS_ACTUAL
                file_type = "uib"
            elif col == self.COL_DAD_FILE_ACTUAL:
                is_file_col = True
                pts_col = self.COL_DAD_PTS_ACTUAL
                file_type = "dad"

        if is_file_col:
            # Marcar com a MANUAL i actualitzar color
            item = self.samples_table.item(row, col)
            if item:
                new_value = item.text()
                # Obtenir nom real del fitxer (pot ser diferent del display name)
                actual_filename = item.data(Qt.UserRole)
                if not actual_filename:
                    actual_filename = new_value  # Fallback al text visible

                # Obtenir nom i rèplica de la taula (funciona amb taula ordenada)
                name_item = self.samples_table.item(row, self.COL_MOSTRA)
                rep_item = self.samples_table.item(row, self.COL_REP)
                sample_name = name_item.data(Qt.UserRole) if name_item else ""
                try:
                    replica = int(rep_item.text()) if rep_item else 1
                except:
                    replica = 1

                # Si l'usuari ha seleccionat un fitxer del dropdown (nom real),
                # convertir a format llistat d'injeccions
                if new_value and new_value not in ["-", "(cap)"]:
                    # Si el text és un nom de fitxer (conté extensió o és diferent del format esperat)
                    # convertir-lo al format del llistat d'injeccions
                    expected_display = f"{sample_name}_R{replica}"
                    if new_value != expected_display and (
                        "." in new_value or "_R" not in new_value or not new_value.startswith(sample_name)
                    ):
                        # Guardar el fitxer real seleccionat
                        actual_filename = new_value
                        item.setData(Qt.UserRole, actual_filename)
                        # Mostrar el nom segons llistat d'injeccions
                        self.samples_table.blockSignals(True)
                        item.setText(expected_display)
                        item.setToolTip(f"Fitxer: {actual_filename}")
                        self.samples_table.blockSignals(False)

                    # Guardar assignació manual per al manifest (amb el path real)
                    # Usar clau (sample_name, replica) en lloc de row per suportar taula ordenada
                    if not hasattr(self, '_manual_assignments'):
                        self._manual_assignments = {}
                    key = (sample_name, replica)
                    self._manual_assignments.setdefault(key, {})[col] = actual_filename
                    # Marcar que hi ha canvis sense guardar
                    self.main_window.mark_unsaved_changes()
                    # Si era un suggeriment, marcar com CONFIRMED (verd)
                    # Sinó, marcar com MANUAL (blau)
                    prev_type = self._match_types.get((row, col), "")
                    if prev_type == "SUGGESTED":
                        item.setBackground(QBrush(MATCH_COLORS["EXACT"]))  # Verd clar = confirmat
                        self._match_types[(row, col)] = "CONFIRMED"
                    elif prev_type not in ["CONFIRMED", "EXACT"]:
                        item.setBackground(QBrush(MATCH_COLORS["MANUAL"]))  # Blau = manual
                        self._match_types[(row, col)] = "MANUAL"

                    # Actualitzar nombre de punts (usant el nom real del fitxer)
                    if pts_col is not None:
                        n_points = self._count_file_points(actual_filename, file_type)
                        pts_item = self.samples_table.item(row, pts_col)
                        if pts_item and n_points > 0:
                            pts_item.setText(str(n_points))
                            # Actualitzar _sample_data (buscar per nom/replica)
                            name_item = self.samples_table.item(row, self.COL_MOSTRA)
                            rep_item = self.samples_table.item(row, self.COL_REP)
                            if name_item and rep_item:
                                s_name = name_item.data(Qt.UserRole)
                                s_rep = rep_item.text()
                                for data in self._sample_data:
                                    if (data.get("name") == s_name and
                                        str(data.get("replica", "")) == s_rep):
                                        if file_type == "uib":
                                            data["uib_pts"] = n_points
                                        elif file_type == "dad":
                                            data["dad_pts"] = n_points
                                        break

                    # Carregar dades del fitxer per a la gràfica
                    self._load_and_store_file_data(actual_filename, file_type, sample_name, replica)
                else:
                    # Si es tria "(cap)", la cel·la queda sense assignació
                    item.setBackground(QBrush(MATCH_COLORS["NORMAL"]))
                    self._match_types[(row, col)] = "NONE"
                    # Restaurar punts a "-"
                    if pts_col is not None:
                        pts_item = self.samples_table.item(row, pts_col)
                        if pts_item:
                            pts_item.setText("-")

                # Recalcular l'estat de la fila
                self._update_row_state(row)

                self._update_next_button_state()

    def _count_file_points(self, filename, file_type):
        """Compta el nombre de punts d'un fitxer orfe."""
        if not filename or filename in ["-", "(cap)"]:
            return 0

        # Buscar el path complet del fitxer (usant llista amb paths complets)
        orphan_list = self._orphan_files_full.get(file_type, [])
        full_path = None
        for f in orphan_list:
            if Path(f).name == filename:
                full_path = f
                break

        if not full_path or not os.path.exists(full_path):
            return 0

        try:
            import pandas as pd
            # DAD files sovint són UTF-16, provar primer
            encodings = ['utf-16', 'utf-8', 'latin-1', 'cp1252']
            best_count = 0

            for encoding in encodings:
                try:
                    count = 0
                    if file_type == "uib":
                        df = pd.read_csv(full_path, sep=None, engine='python',
                                        encoding=encoding)
                        count = len(df)
                    elif file_type == "dad":
                        if full_path.lower().endswith('.csv'):
                            df = pd.read_csv(full_path, sep=None, engine='python',
                                            encoding=encoding)
                            count = len(df)
                        else:
                            with open(full_path, 'r', encoding=encoding) as f:
                                lines = f.readlines()
                            data_lines = [l for l in lines if l.strip() and not l.startswith('#')]
                            count = max(0, len(data_lines) - 1)

                    # Guardar el millor resultat (més files = encoding correcte)
                    if count > best_count:
                        best_count = count
                        # Si tenim un bon nombre de punts, retornar
                        if count > 100:
                            return count
                except (UnicodeDecodeError, UnicodeError):
                    continue
                except Exception:
                    continue

            return best_count

        except Exception as e:
            print(f"Warning: No s'ha pogut comptar punts de {filename}: {e}")

        return 0

    def _load_and_store_file_data(self, filename, file_type, sample_name, replica):
        """
        Carrega les dades d'un fitxer orfe assignat manualment i les guarda a imported_data.

        Args:
            filename: Nom del fitxer (sense path)
            file_type: "uib" o "dad"
            sample_name: Nom de la mostra
            replica: Número de rèplica
        """
        if not filename or filename in ["-", "(cap)"]:
            return

        print(f"[DEBUG _load_and_store] Intentant carregar {filename} per {sample_name} R{replica}")

        # Obtenir referència a les dades de la rèplica primer
        samples = self.imported_data.get("samples", {})
        if sample_name not in samples:
            print(f"[DEBUG _load_and_store] ERROR: Mostra '{sample_name}' no trobada a imported_data")
            print(f"[DEBUG _load_and_store] Samples disponibles: {list(samples.keys())[:5]}...")
            return

        rep_data = samples[sample_name].get("replicas", {}).get(str(replica))
        if rep_data is None:
            print(f"[DEBUG _load_and_store] ERROR: Rèplica {replica} no trobada per '{sample_name}'")
            print(f"[DEBUG _load_and_store] Rèpliques disponibles: {list(samples[sample_name].get('replicas', {}).keys())}")
            return

        # Buscar el path complet del fitxer
        full_path = None

        # 1. Primer buscar a la llista d'orfes completa
        orphan_list = self._orphan_files_full.get(file_type, [])
        for f in orphan_list:
            if Path(f).name == filename:
                full_path = f
                break

        # 2. Si no es troba, mirar si hi ha un suggeriment amb el path
        if not full_path:
            suggestion_key = f"{file_type}_suggestion"
            suggestion = rep_data.get(suggestion_key, {})
            if suggestion:
                # El suggeriment pot tenir el path complet a "file"
                sugg_file = suggestion.get("file", "")
                if sugg_file and Path(sugg_file).name == filename:
                    full_path = sugg_file
                elif sugg_file:
                    # Construir path a partir del seq_path
                    if hasattr(self, 'seq_path') and self.seq_path:
                        if file_type == "uib":
                            test_path = os.path.join(self.seq_path, "CSV", sugg_file)
                        else:
                            test_path = os.path.join(self.seq_path, "Export3d", sugg_file)
                        if os.path.exists(test_path):
                            full_path = test_path

        # 3. Construir path manualment si encara no es troba
        if not full_path and hasattr(self, 'seq_path') and self.seq_path:
            if file_type == "uib":
                possible_paths = [
                    os.path.join(self.seq_path, "CSV", filename),
                    os.path.join(self.seq_path, "csv", filename),
                ]
            else:
                possible_paths = [
                    os.path.join(self.seq_path, "Export3d", filename),
                    os.path.join(self.seq_path, "Export3D", filename),
                    os.path.join(self.seq_path, "CSV", filename),
                    os.path.join(self.seq_path, "csv", filename),
                ]
            for p in possible_paths:
                if os.path.exists(p):
                    full_path = p
                    break

        if not full_path or not os.path.exists(full_path):
            print(f"Warning: No s'ha trobat el fitxer {filename} (file_type={file_type})")
            return

        try:
            if file_type == "uib":
                # Carregar UIB
                df, status = llegir_doc_uib(full_path)
                if not df.empty and "OK" in status:
                    t = df["time (min)"].values
                    y = df["DOC"].values

                    # Calcular baseline i y_net (CRÍTIC per areas_uib)
                    baseline = None
                    y_net = None
                    if len(t) > 10:
                        # Determinar mode (BP o COLUMN)
                        method = self.imported_data.get("method", "COLUMN")
                        mode = "BP" if method == "BP" else "COLUMN"
                        baseline = get_baseline_value(t, y, mode=mode)
                        y_net = np.array(y) - baseline

                    rep_data["uib"] = {
                        "t": t,
                        "y": y,
                        "y_raw": y,
                        "y_net": y_net,
                        "baseline": baseline,
                        "file": filename,
                        "n_points": len(df),
                        "manual_assignment": True,
                    }
                    bl_val = baseline if baseline is not None else 0
                    print(f"[DEBUG _load_and_store] Carregat UIB: {filename} per {sample_name} R{replica}")
                    print(f"[DEBUG _load_and_store] rep_data['uib'] keys: {list(rep_data['uib'].keys())}")
                    print(f"[DEBUG _load_and_store] t is not None: {t is not None}, len(t)={len(t) if t is not None else 0}")
                    # Treure de la llista d'orfes
                    if "orphan_files" in self.imported_data:
                        uib_orphans = self.imported_data["orphan_files"].get("uib", [])
                        self.imported_data["orphan_files"]["uib"] = [
                            f for f in uib_orphans if Path(f).name != filename
                        ]

            elif file_type == "dad":
                # Provar primer Export3D, després DAD1A
                df, status = llegir_dad_export3d(full_path)
                if df.empty or "Error" in status:
                    df, status = llegir_dad_1a(full_path)

                if not df.empty and "OK" in status:
                    # Assegurar que tenim una còpia del DataFrame
                    df = df.copy()
                    rep_data["dad"] = {
                        "df": df,
                        "t": df["time (min)"].values if "time (min)" in df.columns else None,
                        "file": filename,
                        "n_points": len(df),
                        "manual_assignment": True,
                        "source": "manual",
                    }
                    print(f"Carregat DAD: {filename} ({len(df)} punts, columnes: {list(df.columns)[:5]})")
                    # Treure de la llista d'orfes
                    if "orphan_files" in self.imported_data:
                        dad_orphans = self.imported_data["orphan_files"].get("dad", [])
                        self.imported_data["orphan_files"]["dad"] = [
                            f for f in dad_orphans if Path(f).name != filename
                        ]
                else:
                    print(f"Warning: No s'han pogut llegir dades DAD de {filename}: {status}")

        except Exception as e:
            print(f"Error carregant {filename}: {e}")

    def _update_row_state(self, row):
        """Actualitza l'estat d'una fila específica."""
        # Obtenir tipus de mostra (buscar per nom/replica per suportar taula ordenada)
        sample_type = "MOSTRA"
        name_item = self.samples_table.item(row, self.COL_MOSTRA)
        rep_item = self.samples_table.item(row, self.COL_REP)
        if name_item and rep_item:
            s_name = name_item.data(Qt.UserRole)
            s_rep = rep_item.text()
            for data in self._sample_data:
                if (data.get("name") == s_name and
                    str(data.get("replica", "")) == s_rep):
                    sample_type = data.get("type", "MOSTRA")
                    break

        # BLANC i CONTROL no requereixen assignació de DAD/UIB
        # Només MOSTRA, PATRÓ_CAL, PATRÓ_REF necessiten verificació
        requires_assignment = sample_type in ["MOSTRA", "PATRÓ_CAL", "PATRÓ_REF"]

        missing = []
        pending_review = []

        if requires_assignment:
            # Comprovar UIB
            if self._data_mode in ["DUAL", "UIB"]:
                uib_item = self.samples_table.item(row, self.COL_UIB_FILE_ACTUAL)
                if uib_item:
                    val = uib_item.text()
                    match_type = self._match_types.get((row, self.COL_UIB_FILE_ACTUAL), "")
                    if val == "-" and self._orphan_files.get("uib"):
                        missing.append("UIB")
                    elif match_type == "SUGGESTED":
                        pending_review.append("UIB")

            # Comprovar DAD
            dad_item = self.samples_table.item(row, self.COL_DAD_FILE_ACTUAL)
            if dad_item:
                val = dad_item.text()
                match_type = self._match_types.get((row, self.COL_DAD_FILE_ACTUAL), "")
                if val == "-" and self._orphan_files.get("dad"):
                    missing.append("DAD")
                elif match_type == "SUGGESTED":
                    pending_review.append("DAD")

        # Actualitzar Estat
        estat_item = self.samples_table.item(row, self.COL_ESTAT)
        if estat_item:
            if missing:
                estat_text = "Assignar " + "+".join(missing)
                estat_item.setText(estat_text)
                estat_item.setBackground(QBrush(QColor("#FADBD8")))  # Rosa
                self._unverified_fuzzy.add(row)
            elif pending_review:
                estat_text = "Revisar " + "+".join(pending_review)
                estat_item.setText(estat_text)
                estat_item.setBackground(QBrush(QColor("#FCF3CF")))  # Groc
                self._unverified_fuzzy.add(row)
            else:
                estat_item.setText("OK")
                estat_item.setBackground(QBrush(QColor("#D5F5E3")))  # Verd
                self._unverified_fuzzy.discard(row)

    def _on_cell_double_clicked(self, row, col):
        """Handler de doble clic."""
        # Determinar columnes editables
        editable_cols = [self.COL_TIPUS]
        if self._data_mode == "DIRECT":
            editable_cols.append(self.COL_DAD_FILE_ACTUAL)
        else:
            editable_cols.extend([self.COL_UIB_FILE_ACTUAL, self.COL_DAD_FILE_ACTUAL])

        # Si és columna editable, deixar que el delegate s'encarregui
        if col in editable_cols:
            return

        # Altrament, mostrar preview
        # Obtenir nom i rèplica de les cel·les (per funcionar amb taula ordenada)
        name_item = self.samples_table.item(row, self.COL_MOSTRA)
        rep_item = self.samples_table.item(row, self.COL_REP)

        if name_item and rep_item:
            sample_name = name_item.data(Qt.UserRole)
            rep_text = rep_item.text()

            # Buscar les dades corresponents a _sample_data
            sample_data = None
            for data in self._sample_data:
                # Comparar amb conversió a string per evitar problemes de tipus
                if (data.get("name") == sample_name and
                    str(data.get("replica", "")) == rep_text):
                    sample_data = data
                    break

            if sample_data:
                try:
                    replica = int(rep_text) if rep_text.isdigit() else 1
                except:
                    replica = 1
                dialog = ChromatogramPreviewDialog(
                    self,
                    sample_name,
                    replica,
                    sample_data,
                    self.imported_data
                )
                dialog.exec()

    def _update_warnings(self):
        """Actualitza warnings i orfes."""
        # Si carregat des de manifest, no mostrar avisos d'orfes/suggeriments
        # (ja estaven gestionats quan es va guardar el manifest)
        if self._loaded_from_manifest:
            self.warnings_frame.setVisible(False)
            self.confirm_btn.setVisible(False)
            self.orphans_btn.setVisible(False)
            self.refresh_btn.setVisible(False)
            self.dismiss_btn.setVisible(False)
            return

        # Si l'avís d'orfes ha estat marcat com revisat, no mostrar
        if self._orphan_warning_dismissed:
            self.warnings_frame.setVisible(False)
            self.confirm_btn.setVisible(False)
            self.orphans_btn.setVisible(False)
            self.refresh_btn.setVisible(False)
            self.dismiss_btn.setVisible(False)
            return

        # Comptar orfes que encara no estan assignats
        unassigned_uib, unassigned_dad = self._count_unassigned_orphans()

        # Comptar suggeriments pendents de confirmar
        pending_suggestions = sum(
            1 for (r, c), mt in self._match_types.items()
            if mt == "SUGGESTED"
        )

        has_orphans = unassigned_uib > 0 or unassigned_dad > 0
        has_suggestions = pending_suggestions > 0

        # Warnings d'importació (injeccions faltants, etc.)
        import_warnings = getattr(self, '_import_warnings', [])
        has_import_warnings = len(import_warnings) > 0

        if has_orphans or has_suggestions or has_import_warnings:
            msg_parts = []

            # Primer els warnings crítics d'importació
            if has_import_warnings:
                # Mostrar el primer warning crític (normalment el de injeccions faltants)
                for w in import_warnings:
                    if "ATENCIÓ" in w:
                        msg_parts.append(w.replace("⚠️ ", ""))
                        break

            if pending_suggestions:
                msg_parts.append(f"{pending_suggestions} suggeriments per revisar")
            if unassigned_uib:
                msg_parts.append(f"{unassigned_uib} UIB orfes (revisar noms i reimportar, o assignar manualment)")
            if unassigned_dad:
                msg_parts.append(f"{unassigned_dad} DAD orfes (revisar noms i reimportar, o assignar manualment)")

            self.warnings_label.setText("⚠️ " + " · ".join(msg_parts))
            self.confirm_btn.setVisible(has_suggestions)
            self.orphans_btn.setVisible(has_orphans)
            self.refresh_btn.setVisible(has_orphans)
            # Mostrar botó "Marcar revisat" si hi ha orfes (permet continuar sense assignar)
            self.dismiss_btn.setVisible(has_orphans and not has_suggestions)
            self.warnings_frame.setVisible(True)
        else:
            self.warnings_frame.setVisible(False)
            self.confirm_btn.setVisible(False)
            self.orphans_btn.setVisible(False)
            self.refresh_btn.setVisible(False)
            self.dismiss_btn.setVisible(False)

    def _count_unassigned_orphans(self):
        """Compta quants orfes encara no estan assignats."""
        assigned_uib = set()
        assigned_dad = set()

        # Recórrer la taula per veure què s'ha assignat manualment
        for row in range(self.samples_table.rowCount()):
            if self._data_mode in ["DUAL", "UIB"]:
                uib_item = self.samples_table.item(row, self.COL_UIB_FILE_ACTUAL)
                if uib_item:
                    # Usar el nom real del fitxer (UserRole) si disponible
                    val = uib_item.data(Qt.UserRole) or uib_item.text()
                    if val and val not in ["-", "(cap)"]:
                        assigned_uib.add(val)

            dad_item = self.samples_table.item(row, self.COL_DAD_FILE_ACTUAL)
            if dad_item:
                # Usar el nom real del fitxer (UserRole) si disponible
                val = dad_item.data(Qt.UserRole) or dad_item.text()
                if val and val not in ["-", "(cap)", "[MasterFile]", "[export3d]", "[csv]"]:
                    assigned_dad.add(val)

        # Comptar orfes no assignats
        orphan_uib = self._orphan_files.get("uib", [])
        orphan_dad = self._orphan_files.get("dad", [])

        unassigned_uib = sum(1 for f in orphan_uib if Path(f).name not in assigned_uib)
        unassigned_dad = sum(1 for f in orphan_dad if Path(f).name not in assigned_dad)

        return unassigned_uib, unassigned_dad

    def _confirm_all_suggestions(self):
        """Confirma tots els suggeriments automàtics i carrega les dades."""
        confirmed = 0
        self.samples_table.blockSignals(True)

        for row in range(self.samples_table.rowCount()):
            # Obtenir nom i rèplica de la fila
            name_item = self.samples_table.item(row, self.COL_MOSTRA)
            rep_item = self.samples_table.item(row, self.COL_REP)
            if not name_item or not rep_item:
                continue
            sample_name = name_item.data(Qt.UserRole)
            try:
                replica = int(rep_item.text())
            except:
                replica = 1

            # Comprovar UIB
            if self._data_mode in ["DUAL", "UIB"]:
                if self._match_types.get((row, self.COL_UIB_FILE_ACTUAL)) == "SUGGESTED":
                    item = self.samples_table.item(row, self.COL_UIB_FILE_ACTUAL)
                    if item:
                        item.setBackground(QBrush(MATCH_COLORS["EXACT"]))
                        self._match_types[(row, self.COL_UIB_FILE_ACTUAL)] = "CONFIRMED"
                        confirmed += 1
                        # Carregar dades del fitxer
                        filename = item.data(Qt.UserRole)
                        if filename:
                            self._load_and_store_file_data(filename, "uib", sample_name, replica)
                            # Guardar a _manual_assignments per persistència
                            key = (sample_name, replica)
                            self._manual_assignments.setdefault(key, {})[self.COL_UIB_FILE_ACTUAL] = filename

            # Comprovar DAD
            if self._match_types.get((row, self.COL_DAD_FILE_ACTUAL)) == "SUGGESTED":
                item = self.samples_table.item(row, self.COL_DAD_FILE_ACTUAL)
                if item:
                    item.setBackground(QBrush(MATCH_COLORS["EXACT"]))
                    self._match_types[(row, self.COL_DAD_FILE_ACTUAL)] = "CONFIRMED"
                    confirmed += 1
                    # Carregar dades del fitxer
                    filename = item.data(Qt.UserRole)
                    if filename:
                        self._load_and_store_file_data(filename, "dad", sample_name, replica)
                        # Guardar a _manual_assignments per persistència
                        key = (sample_name, replica)
                        self._manual_assignments.setdefault(key, {})[self.COL_DAD_FILE_ACTUAL] = filename

        self.samples_table.blockSignals(False)

        # Marcar canvis sense guardar perquè es persisteixin
        if confirmed > 0:
            self.main_window.mark_unsaved_changes()

        # Recalcular estats
        self._recalculate_row_states()
        self._update_warnings()
        self._update_next_button_state()

        if confirmed > 0:
            # Guardar manifest immediatament
            try:
                print(f"[DEBUG confirm] Guardant manifest amb {confirmed} confirmacions...")
                self._apply_manual_assignments()
                manifest_path = save_import_manifest(self.imported_data)
                print(f"[DEBUG confirm] Manifest guardat a: {manifest_path}")
                self.main_window.mark_manifest_saved()
                QMessageBox.information(self, "Confirmat", f"S'han confirmat {confirmed} suggeriments i s'han guardat.")
            except Exception as e:
                import traceback
                print(f"[DEBUG confirm] ERROR: {e}")
                traceback.print_exc()
                QMessageBox.warning(self, "Avís", f"S'han confirmat {confirmed} suggeriments però no s'han pogut guardar: {e}")


    def _recalculate_row_states(self):
        """Recalcula l'estat de cada fila basant-se en assignacions actuals."""
        self.samples_table.blockSignals(True)
        self._unverified_fuzzy.clear()

        for row in range(self.samples_table.rowCount()):
            # Obtenir tipus de mostra (buscar per nom/replica)
            sample_type = "MOSTRA"
            name_item = self.samples_table.item(row, self.COL_MOSTRA)
            rep_item = self.samples_table.item(row, self.COL_REP)
            if name_item and rep_item:
                s_name = name_item.data(Qt.UserRole)
                s_rep = rep_item.text()
                for data in self._sample_data:
                    if (data.get("name") == s_name and
                        str(data.get("replica", "")) == s_rep):
                        sample_type = data.get("type", "MOSTRA")
                        break

            # BLANC i CONTROL no requereixen assignació
            requires_assignment = sample_type in ["MOSTRA", "PATRÓ_CAL", "PATRÓ_REF"]

            missing = []

            if requires_assignment:
                # Comprovar UIB
                if self._data_mode in ["DUAL", "UIB"]:
                    uib_item = self.samples_table.item(row, self.COL_UIB_FILE_ACTUAL)
                    if uib_item:
                        val = uib_item.text()
                        if val == "-" and self._orphan_files.get("uib"):
                            missing.append("UIB")

                # Comprovar DAD
                dad_item = self.samples_table.item(row, self.COL_DAD_FILE_ACTUAL)
                if dad_item:
                    val = dad_item.text()
                    if val == "-" and self._orphan_files.get("dad"):
                        missing.append("DAD")

            # Actualitzar Estat
            estat_item = self.samples_table.item(row, self.COL_ESTAT)
            if estat_item:
                if missing:
                    estat_text = "Assignar " + "+".join(missing)
                    estat_item.setText(estat_text)
                    estat_item.setBackground(QBrush(QColor("#FADBD8")))
                    self._unverified_fuzzy.add(row)
                else:
                    estat_item.setText("OK")
                    estat_item.setBackground(QBrush(QColor("#D5F5E3")))

        self.samples_table.blockSignals(False)

    def _update_next_button_state(self):
        """Actualitza l'estat del botó Següent."""
        # Verificar si hi ha FUZZY sense verificar
        if self._unverified_fuzzy:
            self.next_btn.setEnabled(False)
            self.next_btn.setToolTip(
                f"Cal verificar {len(self._unverified_fuzzy)} assignacions (taronja)"
            )
        else:
            self.next_btn.setEnabled(True)
            self.next_btn.setToolTip("")

    def _refresh_orphan_count(self):
        """Actualitza el comptador d'orfes i la llista després d'assignacions manuals."""
        # Obtenir fitxers assignats
        assigned_uib = set()
        assigned_dad = set()

        for row in range(self.samples_table.rowCount()):
            if self._data_mode in ["DUAL", "UIB"]:
                uib_item = self.samples_table.item(row, self.COL_UIB_FILE_ACTUAL)
                if uib_item:
                    val = uib_item.data(Qt.UserRole) or uib_item.text()
                    if val and val not in ["-", "(cap)"]:
                        assigned_uib.add(val)
                        assigned_uib.add(Path(val).name)  # També el nom sense path

            dad_item = self.samples_table.item(row, self.COL_DAD_FILE_ACTUAL)
            if dad_item:
                val = dad_item.data(Qt.UserRole) or dad_item.text()
                if val and val not in ["-", "(cap)", "[MasterFile]", "[export3d]", "[csv]"]:
                    assigned_dad.add(val)
                    assigned_dad.add(Path(val).name)

        # Actualitzar llista d'orfes (treure els assignats)
        orig_uib = self._orphan_files.get("uib", [])
        orig_dad = self._orphan_files.get("dad", [])

        self._orphan_files["uib"] = [f for f in orig_uib if Path(f).name not in assigned_uib]
        self._orphan_files["dad"] = [f for f in orig_dad if Path(f).name not in assigned_dad]

        # Actualitzar warnings
        self._update_warnings()

    def _dismiss_orphan_warning(self):
        """Marca l'avís d'orfes com a revisat i amaga la barra d'avisos."""
        self._orphan_warning_dismissed = True
        self.warnings_frame.setVisible(False)
        # Guardar al manifest que l'avís ha estat revisat
        if self.imported_data:
            self.imported_data["orphan_warning_dismissed"] = True
            try:
                save_import_manifest(self.imported_data)
                self.main_window.set_status("Avís marcat com a revisat", 3000)
            except Exception as e:
                print(f"Warning: No s'ha pogut guardar estat revisat: {e}")

    def _show_orphans(self):
        dialog = OrphanFilesDialog(self, self._orphan_files)
        dialog.exec()

    def _save_manifest(self):
        """Guarda el manifest amb les assignacions actuals."""
        if not self.imported_data:
            QMessageBox.warning(self, "Sense dades", "No hi ha dades per guardar.")
            return

        try:
            # Aplicar assignacions manuals
            self._apply_manual_assignments()
            # Guardar manifest
            save_import_manifest(self.imported_data)
            self.main_window.mark_manifest_saved()
            self.main_window.set_status("Manifest guardat correctament", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error guardant manifest:\n{e}")

    def _go_next(self):
        # Verificar si hi ha orfes sense assignar
        unassigned_uib, unassigned_dad = self._count_unassigned_orphans()
        has_orphans = unassigned_uib > 0 or unassigned_dad > 0
        has_unsaved = self.main_window.has_unsaved_changes

        # Si s'ha carregat des de manifest sense fer canvis, passar directament
        # (els orfes ja eren coneguts quan es va guardar el manifest)
        if self._loaded_from_manifest and not has_unsaved:
            self.main_window.go_to_tab(1)
            return

        # Si no hi ha orfes ni canvis sense guardar, passar directament
        if not has_orphans and not has_unsaved:
            self.main_window.go_to_tab(1)
            return

        # Construir missatge de confirmació
        msg_parts = []
        if has_orphans:
            msg_parts.append("Hi ha fitxers orfes sense assignar:")
            if unassigned_uib:
                msg_parts.append(f"  • {unassigned_uib} fitxers UIB")
            if unassigned_dad:
                msg_parts.append(f"  • {unassigned_dad} fitxers DAD")
            msg_parts.append("")

        if has_unsaved:
            msg_parts.append("Es guardaran els canvis i es passarà a la fase de calibració.")
        else:
            msg_parts.append("Es passarà a la fase de calibració.")

        msg_parts.append("\nVols continuar?")

        reply = QMessageBox.question(
            self, "Continuar",
            "\n".join(msg_parts),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply != QMessageBox.Yes:
            return

        # Aplicar assignacions manuals i guardar si hi ha canvis
        if has_unsaved:
            self._apply_manual_assignments()
            try:
                save_import_manifest(self.imported_data)
                self.main_window.mark_manifest_saved()
            except Exception as e:
                print(f"Warning: No s'ha pogut guardar manifest: {e}")

        self.main_window.go_to_tab(1)

    def _apply_manual_assignments(self):
        """Aplica les assignacions manuals a imported_data."""
        if not hasattr(self, '_manual_assignments') or not self._manual_assignments:
            return

        # Iterar per les assignacions guardades amb clau (sample_name, replica)
        for key, assignments in self._manual_assignments.items():
            sample_name, replica = key

            # Obtenir dades de la mostra
            samples = self.imported_data.get("samples", {})
            if sample_name not in samples:
                continue

            rep_data = samples[sample_name].get("replicas", {}).get(str(replica))
            if not rep_data:
                continue

            # Aplicar assignacions
            for col, filename in assignments.items():
                if filename in ["-", "(cap)", ""]:
                    continue

                # Determinar tipus de senyal
                if col == self.COL_UIB_FILE_ACTUAL:
                    signal_type = "uib"
                elif col == self.COL_DAD_FILE_ACTUAL:
                    signal_type = "dad"
                else:
                    continue

                # Marcar com a assignació manual
                # Assegurar que el dict existeix (pot ser None o absent)
                if signal_type not in rep_data or rep_data[signal_type] is None:
                    rep_data[signal_type] = {}
                rep_data[signal_type]["manual_assignment"] = True
                rep_data[signal_type]["manual_file"] = filename

        # Actualitzar manifest
        try:
            save_import_manifest(self.imported_data)
        except Exception as e:
            print(f"Warning: No s'ha pogut actualitzar manifest: {e}")
