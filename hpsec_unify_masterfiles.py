"""
hpsec_unify_masterfiles.py - Unifica estructura de tots els MasterFiles

Estructura canònica:
  0-INFO:       Metadata + sincronització
  1-HPLC-SEQ:   Injeccions + Inj_Index + Sample_Rep
  2-TOC:        Dades DOC raw (intocable)
  3-DAD_KHP:    Dades KHP (opcional)
  4-TOC_CALC:   Càlculs (sempre al final, és l'única fulla calculada)

Accions:
  1. Reordenar fulls segons estructura canònica
  2. Renombrar fulls (TOC_CALC -> 4-TOC_CALC, etc.)
  3. Eliminar 3-SEQ_DATA (redundant amb 1-HPLC-SEQ)
  4. Afegir Sample_Rep a 1-HPLC-SEQ si falta

Autor: HPSEC Suite
Data: 2025-01
"""

import os
import shutil
import glob
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd


# Estructura canònica
CANONICAL_SHEETS = ['0-INFO', '1-HPLC-SEQ', '2-TOC', '3-DAD_KHP', '4-TOC_CALC']


def analyze_masterfile(filepath):
    """Analitza un MasterFile i retorna info sobre la seva estructura."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        issues = []

        # Detectar problemes
        has_toc_calc = any('TOC_CALC' in s.upper() for s in sheets)
        has_dad_khp = any('DAD_KHP' in s.upper() for s in sheets)
        has_seq_data = any('SEQ_DATA' in s.upper() for s in sheets)

        # Noms incorrectes
        for s in sheets:
            if 'TOC_CALC' in s.upper() and s != '4-TOC_CALC':
                issues.append(f"Rename: '{s}' -> '4-TOC_CALC'")
            if 'DAD_KHP' in s.upper() and s != '3-DAD_KHP':
                issues.append(f"Rename: '{s}' -> '3-DAD_KHP'")

        # SEQ_DATA redundant
        if has_seq_data:
            issues.append("Remove: '3-SEQ_DATA' (redundant)")

        # Ordre incorrecte
        toc_calc_idx = next((i for i, s in enumerate(sheets) if 'TOC_CALC' in s.upper()), -1)
        dad_khp_idx = next((i for i, s in enumerate(sheets) if 'DAD_KHP' in s.upper()), -1)

        if has_toc_calc and has_dad_khp and toc_calc_idx < dad_khp_idx:
            issues.append("Reorder: TOC_CALC ha d'anar després de DAD_KHP")

        return {
            'sheets': sheets,
            'has_toc_calc': has_toc_calc,
            'has_dad_khp': has_dad_khp,
            'has_seq_data': has_seq_data,
            'issues': issues,
            'needs_fix': len(issues) > 0
        }

    except Exception as e:
        return {'error': str(e), 'needs_fix': False}


def check_hplc_columns(filepath):
    """Verifica si 1-HPLC-SEQ té les columnes necessàries."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['1-HPLC-SEQ']

        # Llegir capçaleres
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val) if val else '')

        wb.close()

        has_inj_index = any('Inj_Index' in h for h in headers)
        has_sample_rep = any('Sample_Rep' in h for h in headers)

        return {
            'headers': headers,
            'has_inj_index': has_inj_index,
            'has_sample_rep': has_sample_rep,
            'needs_sample_rep': has_inj_index and not has_sample_rep
        }

    except Exception as e:
        return {'error': str(e)}


def backup_file(filepath):
    """Crea backup del fitxer."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base, ext = os.path.splitext(filepath)
    backup_path = f"{base}_backup_{timestamp}{ext}"
    shutil.copy2(filepath, backup_path)
    return backup_path


def unify_masterfile(filepath, dry_run=True):
    """
    Unifica un MasterFile a l'estructura canònica.

    Args:
        filepath: Ruta al fitxer
        dry_run: Si True, només mostra què faria sense modificar

    Returns:
        dict amb resultat de l'operació
    """
    folder = os.path.basename(os.path.dirname(filepath))
    filename = os.path.basename(filepath)

    print(f"\n{'='*60}")
    print(f"{folder}/{filename}")
    print('='*60)

    # Analitzar
    analysis = analyze_masterfile(filepath)

    if 'error' in analysis:
        print(f"  ERROR: {analysis['error']}")
        return {'status': 'error', 'message': analysis['error']}

    if not analysis['needs_fix']:
        print("  [OK] Estructura correcta")

        # Verificar columnes HPLC
        hplc = check_hplc_columns(filepath)
        if hplc.get('needs_sample_rep'):
            print("  ! Falta Sample_Rep a 1-HPLC-SEQ")
            analysis['issues'].append("Add: Sample_Rep a 1-HPLC-SEQ")
            analysis['needs_fix'] = True

        if not analysis['needs_fix']:
            return {'status': 'ok', 'message': 'Ja unificat'}

    print(f"  Fulls actuals: {analysis['sheets']}")
    print(f"  Problemes:")
    for issue in analysis['issues']:
        print(f"    - {issue}")

    if dry_run:
        print("  [DRY RUN - no s'han fet canvis]")
        return {'status': 'needs_fix', 'issues': analysis['issues']}

    # Crear backup
    backup_path = backup_file(filepath)
    print(f"  Backup: {os.path.basename(backup_path)}")

    # Carregar workbook
    wb = openpyxl.load_workbook(filepath)

    # 1. Renombrar fulls
    sheet_mapping = {}
    for sheet_name in wb.sheetnames:
        new_name = sheet_name
        upper = sheet_name.upper()

        if 'TOC_CALC' in upper and sheet_name != '4-TOC_CALC':
            new_name = '4-TOC_CALC'
        elif 'DAD_KHP' in upper and sheet_name != '3-DAD_KHP':
            new_name = '3-DAD_KHP'

        if new_name != sheet_name:
            sheet_mapping[sheet_name] = new_name
            wb[sheet_name].title = new_name
            print(f"  Renamed: '{sheet_name}' -> '{new_name}'")

    # 2. Eliminar SEQ_DATA si existeix
    for sheet_name in list(wb.sheetnames):
        if 'SEQ_DATA' in sheet_name.upper():
            del wb[sheet_name]
            print(f"  Removed: '{sheet_name}'")

    # 3. Reordenar fulls
    current_sheets = wb.sheetnames

    # Ordre desitjat
    ordered = []
    for canonical in CANONICAL_SHEETS:
        for sheet in current_sheets:
            if sheet == canonical or (canonical == '3-DAD_KHP' and 'DAD_KHP' in sheet.upper()) or \
               (canonical == '4-TOC_CALC' and 'TOC_CALC' in sheet.upper()):
                if sheet not in ordered:
                    ordered.append(sheet)
                break

    # Afegir fulls que no estan a la llista canònica
    for sheet in current_sheets:
        if sheet not in ordered:
            ordered.append(sheet)

    # Aplicar ordre
    for i, sheet_name in enumerate(ordered):
        wb.move_sheet(sheet_name, offset=i - wb.sheetnames.index(sheet_name))

    print(f"  Nou ordre: {wb.sheetnames}")

    # 4. Afegir Sample_Rep si falta
    hplc = check_hplc_columns(filepath)
    if hplc.get('needs_sample_rep'):
        add_sample_rep(wb['1-HPLC-SEQ'])
        print("  Added: Sample_Rep a 1-HPLC-SEQ")

    # Guardar
    wb.save(filepath)
    wb.close()

    print("  [OK] Unificat correctament")
    return {'status': 'fixed', 'issues': analysis['issues']}


def add_sample_rep(ws):
    """Afegeix columna Sample_Rep a 1-HPLC-SEQ."""
    # Trobar columnes necessàries
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col

    sample_col = headers.get('Sample  Name') or headers.get('Sample Name')
    inj_num_col = headers.get('Inj#')
    inj_index_col = headers.get('Inj_Index')

    if not sample_col:
        return False

    # Nova columna al final
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value='Sample_Rep')

    # Generar Sample_Rep per cada fila
    # Lògica de blocs per mostres repetides
    sample_inj_counts = {}  # {sample: {inj_num: count}}

    # Primera passada: comptar repeticions
    for row in range(2, ws.max_row + 1):
        sample = ws.cell(row=row, column=sample_col).value
        inj_num = ws.cell(row=row, column=inj_num_col).value if inj_num_col else 1

        if sample:
            sample = str(sample)
            inj_num = int(inj_num) if inj_num else 1

            if sample not in sample_inj_counts:
                sample_inj_counts[sample] = {}
            if inj_num not in sample_inj_counts[sample]:
                sample_inj_counts[sample][inj_num] = 0
            sample_inj_counts[sample][inj_num] += 1

    # Detectar mostres que necessiten blocs
    samples_need_blocks = set()
    for sample, inj_counts in sample_inj_counts.items():
        for inj_num, count in inj_counts.items():
            if count > 1:
                samples_need_blocks.add(sample)
                break

    # Segona passada: generar Sample_Rep
    sample_block_counter = {}
    sample_last_inj = {}

    for row in range(2, ws.max_row + 1):
        sample = ws.cell(row=row, column=sample_col).value
        inj_num = ws.cell(row=row, column=inj_num_col).value if inj_num_col else 1

        if sample:
            sample = str(sample)
            inj_num = int(inj_num) if inj_num else 1

            if sample in samples_need_blocks:
                # Mostra amb repeticions: usar blocs
                if sample not in sample_block_counter:
                    sample_block_counter[sample] = 0
                    sample_last_inj[sample] = 0

                # Nou bloc si inj_num <= últim inj_num
                if inj_num <= sample_last_inj.get(sample, 0):
                    sample_block_counter[sample] += 1

                sample_last_inj[sample] = inj_num
                block = sample_block_counter[sample] + 1
                sample_rep = f"{sample}_B{block}_R{inj_num}"
            else:
                # Mostra única: només _R1/_R2
                sample_rep = f"{sample}_R{inj_num}"

            ws.cell(row=row, column=new_col, value=sample_rep)

    return True


def scan_all_masterfiles(base_path, dry_run=True):
    """Escaneja i unifica tots els MasterFiles."""

    files = glob.glob(f'{base_path}/**/*MasterFile*.xlsx', recursive=True)
    files = [f for f in files if 'backup' not in f.lower()]

    print(f"\n{'#'*60}")
    print(f"# ESCANEIG MASTERFILES")
    print(f"# Base: {base_path}")
    print(f"# Fitxers: {len(files)}")
    print(f"# Mode: {'DRY RUN' if dry_run else 'APLICAR CANVIS'}")
    print('#'*60)

    results = {'ok': 0, 'needs_fix': 0, 'fixed': 0, 'error': 0}

    for f in sorted(files):
        result = unify_masterfile(f, dry_run=dry_run)
        results[result['status']] += 1

    print(f"\n{'='*60}")
    print("RESUM")
    print('='*60)
    print(f"  Correctes:        {results['ok']}")
    print(f"  Necessiten fix:   {results['needs_fix']}")
    print(f"  Arreglats:        {results['fixed']}")
    print(f"  Errors:           {results['error']}")

    return results


# --- CLI ---
if __name__ == '__main__':
    import sys

    print("="*60)
    print("HPSEC MasterFile Unifier")
    print("="*60)

    if len(sys.argv) < 2:
        print("""
Ús:
  python hpsec_unify_masterfiles.py --scan              Escanejar tots (dry run)
  python hpsec_unify_masterfiles.py --apply             Aplicar canvis a tots
  python hpsec_unify_masterfiles.py <fitxer.xlsx>       Unificar un fitxer (dry run)
  python hpsec_unify_masterfiles.py <fitxer.xlsx> --apply   Aplicar canvis
        """)
        sys.exit(1)

    arg = sys.argv[1]
    apply_changes = '--apply' in sys.argv

    base_path = 'C:/Users/Lequia/Desktop/Dades2'

    if arg == '--scan':
        scan_all_masterfiles(base_path, dry_run=True)
    elif arg == '--apply':
        scan_all_masterfiles(base_path, dry_run=False)
    elif os.path.isfile(arg):
        unify_masterfile(arg, dry_run=not apply_changes)
    else:
        print(f"ERROR: No trobat: {arg}")
        sys.exit(1)
