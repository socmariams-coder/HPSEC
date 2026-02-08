# -*- coding: utf-8 -*-
"""
Script per actualitzar la sensibilitat UIB als MasterFiles.

Regles:
- SEQ 269-274: 700 ppb
- SEQ 275+: 1000 ppb
- Les dues últimes seqüències: 700 ppb
"""

import os
import re
from pathlib import Path
import openpyxl

# Configuració
DATA_FOLDER = "C:/Users/Lequia/Desktop/Dades3"

def get_seq_number(folder_name):
    """Extreu el número de seqüència del nom de carpeta."""
    # Patró: 269B_SEQ, 275_SEQ, 270_SEQ_BP, etc.
    match = re.match(r'^(\d+)', folder_name)
    if match:
        return int(match.group(1))
    return None

def get_uib_sensitivity(seq_num, is_last_two=False):
    """Determina la sensibilitat UIB segons el número de seqüència."""
    if is_last_two:
        return 700
    if 269 <= seq_num <= 274:
        return 700
    elif seq_num >= 275:
        return 1000
    return None  # Seqüències antigues no tenen UIB

def find_masterfile(seq_path):
    """Troba el MasterFile d'una seqüència."""
    seq_path = Path(seq_path)
    masterfiles = list(seq_path.glob("*MasterFile*.xlsx"))
    if masterfiles:
        # Preferir el més recent o el que no sigui backup
        for mf in masterfiles:
            if "backup" not in mf.name.lower():
                return mf
        return masterfiles[0]
    return None

def update_masterfile(masterfile_path, uib_sensitivity):
    """Actualitza la sensibilitat UIB al MasterFile (B5 de 0-INFO)."""
    try:
        wb = openpyxl.load_workbook(masterfile_path)

        if "0-INFO" not in wb.sheetnames:
            print(f"  [!] No te fulla 0-INFO: {masterfile_path.name}")
            return False, "No 0-INFO"

        ws = wb["0-INFO"]
        old_value = ws["B5"].value
        ws["B5"] = uib_sensitivity
        wb.save(masterfile_path)

        return True, f"{old_value} -> {uib_sensitivity}"
    except Exception as e:
        return False, str(e)

def main():
    print("=" * 60)
    print("ACTUALITZACIÓ SENSIBILITAT UIB ALS MASTERFILES")
    print("=" * 60)
    print()

    # Llistar totes les seqüències
    data_path = Path(DATA_FOLDER)
    if not data_path.exists():
        print(f"ERROR: No existeix {DATA_FOLDER}")
        return

    seq_folders = sorted([
        f for f in data_path.iterdir()
        if f.is_dir() and "_SEQ" in f.name
    ], key=lambda x: get_seq_number(x.name) or 0)

    # Filtrar seqüències >= 269
    seq_folders = [f for f in seq_folders if (get_seq_number(f.name) or 0) >= 269]

    if not seq_folders:
        print("No s'han trobat seqüències >= 269")
        return

    # Identificar les dues últimes
    last_two_nums = set()
    if len(seq_folders) >= 2:
        # Obtenir els dos números més alts
        all_nums = sorted(set(get_seq_number(f.name) for f in seq_folders if get_seq_number(f.name)))
        if len(all_nums) >= 2:
            last_two_nums = set(all_nums[-2:])
            print(f"Les dues últimes seqüències (700 ppb): {sorted(last_two_nums)}")

    print(f"\nTrobades {len(seq_folders)} seqüències >= 269")
    print()

    # Processar
    updated = 0
    skipped = 0
    errors = 0

    print(f"{'SEQ':<20} {'Sensibilitat':<12} {'Resultat':<30}")
    print("-" * 60)

    for seq_folder in seq_folders:
        seq_num = get_seq_number(seq_folder.name)
        if seq_num is None:
            continue

        is_last_two = seq_num in last_two_nums
        uib_sens = get_uib_sensitivity(seq_num, is_last_two)

        if uib_sens is None:
            print(f"{seq_folder.name:<20} {'N/A':<12} Skipped (no UIB)")
            skipped += 1
            continue

        masterfile = find_masterfile(seq_folder)
        if not masterfile:
            print(f"{seq_folder.name:<20} {uib_sens:<12} [!] No MasterFile trobat")
            skipped += 1
            continue

        success, msg = update_masterfile(masterfile, uib_sens)

        if success:
            marker = "*" if is_last_two else ""
            print(f"{seq_folder.name:<20} {uib_sens:<12} [OK] {msg} {marker}")
            updated += 1
        else:
            print(f"{seq_folder.name:<20} {uib_sens:<12} [ERR] {msg}")
            errors += 1

    print()
    print("=" * 60)
    print(f"RESUM: Actualitzats={updated}, Omesos={skipped}, Errors={errors}")
    print("=" * 60)
    print()
    print("* = Últimes dues seqüències (forçat a 700 ppb)")

if __name__ == "__main__":
    main()
