"""
Delay tool v3 — TOC + DAD 254 superposat amb drift analysis.
=============================================================
- TOC continu amb pics marcats
- DAD 254 reconstruït (Export3D) posicionat als timestamps HPLC
- Per cada injecció: delay real = t_max_TOC - t_max_DAD
- Gràfic drift: delay per injecció al llarg de la seqüència

Usage:
    python _tool_delay_v3.py 231
"""
import sys, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button
from scipy.ndimage import uniform_filter1d
from scipy.signal import find_peaks

from hpsec_config import get_data_folders
from hpsec_consolidate import extract_seq_number, detect_seq_type

FLUSH_TIME_MIN = 3.637


def find_seq_path(arg):
    if os.path.isdir(arg):
        return arg
    folders = get_data_folders()
    for folder in folders:
        full = os.path.join(folder, arg)
        if os.path.isdir(full):
            return full
        try:
            num = int(arg)
            for d in os.listdir(folder):
                if extract_seq_number(d) == num:
                    return os.path.join(folder, d)
        except ValueError:
            pass
    return None


def read_export3d_254(filepath):
    """Read one Export3D CSV, return (time, value_254) arrays."""
    try:
        with open(filepath, encoding='utf-16') as fh:
            lines = fh.readlines()
    except:
        return None, None

    header = lines[0].strip().split(',')
    idx254 = None
    for i, h in enumerate(header):
        if '254' in h:
            idx254 = i
            break
    if idx254 is None:
        return None, None

    times, vals = [], []
    for line in lines[1:]:
        fields = line.strip().split(',')
        if len(fields) > idx254:
            try:
                times.append(float(fields[0]))
                vals.append(float(fields[idx254]))
            except ValueError:
                pass

    return np.array(times), np.array(vals)


def main():
    if len(sys.argv) < 2:
        print("Usage: python _tool_delay_v3.py <SEQ>")
        sys.exit(1)

    seq_path = find_seq_path(sys.argv[1])
    if not seq_path:
        print(f"ERROR: '{sys.argv[1]}' no trobat")
        sys.exit(1)

    seq_name = os.path.basename(seq_path)
    method = detect_seq_type(seq_name)
    print(f"Loading {seq_name} ({method})...")

    # MasterFile
    mfs = [f for f in os.listdir(seq_path)
           if 'MasterFile' in f and f.endswith('.xlsx')
           and '~' not in f and 'pre_vol' not in f and 'backup' not in f]
    if not mfs:
        mfs = [f for f in os.listdir(seq_path)
               if 'MasterFile' in f and f.endswith('.xlsx')
               and '~' not in f and 'pre_vol' not in f]
    mf_path = os.path.join(seq_path, mfs[0])

    # === TOC ===
    df_toc = pd.read_excel(mf_path, sheet_name='2-TOC', header=6)
    toc_col = [c for c in df_toc.columns if 'toc' in str(c).lower()][0]
    date_col = [c for c in df_toc.columns if 'started' in str(c).lower()]
    y_toc = pd.to_numeric(df_toc[toc_col], errors='coerce').values
    dt = 4.0 / 60.0
    t_toc = np.arange(len(y_toc)) * dt

    t0_toc = None
    if date_col:
        dates = pd.to_datetime(df_toc[date_col[0]], errors='coerce')
        valid = dates.dropna()
        if len(valid) > 0:
            t0_toc = valid.iloc[0]

    y_clean = np.nan_to_num(y_toc, nan=np.nanmedian(y_toc[~np.isnan(y_toc)]))
    y_smooth = uniform_filter1d(y_clean, size=max(3, int(0.3/dt)))

    # TOC peaks
    baseline = np.percentile(y_smooth, 15)
    pk_dist = int(6.0/dt) if method == 'BP' else int(40.0/dt)
    toc_peaks, _ = find_peaks(y_smooth, distance=pk_dist,
                               prominence=max(5, (y_smooth.max()-baseline)*0.02),
                               height=baseline+10)
    print(f"TOC: {len(y_toc)} pts, {len(toc_peaks)} pics")

    # === HPLC injections ===
    df_hplc = pd.read_excel(mf_path, sheet_name='1-HPLC-SEQ')
    name_col = [c for c in df_hplc.columns if 'sample' in str(c).lower() and 'name' in str(c).lower()][0]
    hplc_date_col = [c for c in df_hplc.columns if 'acquired' in str(c).lower()][0]

    injs = []
    for _, row in df_hplc.iterrows():
        try:
            name = str(row[name_col]).strip()
            dt_val = pd.to_datetime(row[hplc_date_col])
            t_min = (dt_val - t0_toc).total_seconds() / 60.0 if t0_toc else 0
            injs.append({
                'name': name, 't_min': t_min,
                'is_khp': 'khp' in name.lower(),
                'is_control': any(x in name.lower() for x in ['mq', 'naoh', 'blanc', 'blnc']),
            })
        except:
            pass

    # === DAD 254 from Export3D ===
    e3d = os.path.join(seq_path, 'Export3d')
    e3d_files = sorted(os.listdir(e3d)) if os.path.isdir(e3d) else []

    # Map Export3D files to injections by name + replica count
    name_counter = {}
    dad_segments = []

    for j, inj in enumerate(injs):
        name_key = inj['name'].lower().replace(' ', '').replace('-', '')
        name_counter[name_key] = name_counter.get(name_key, 0) + 1
        rep_num = name_counter[name_key]

        # Find matching file
        matched = None
        for f in e3d_files:
            f_base = f.lower().replace('.csv', '').replace('-', '').replace(' ', '')
            # Check name match
            name_variants = [
                name_key,
                inj['name'].lower().replace(' ', '_'),
                inj['name'].lower().replace(' ', ''),
            ]
            for nv in name_variants:
                if nv in f_base and f'r{rep_num}' in f_base:
                    matched = f
                    break
                # Also try without special chars
                nv_clean = ''.join(c for c in nv if c.isalnum())
                f_clean = ''.join(c for c in f_base if c.isalnum())
                if nv_clean in f_clean and f'r{rep_num}' in f_base:
                    matched = f
                    break
            if matched:
                break

        if not matched:
            continue

        t_dad, y_dad = read_export3d_254(os.path.join(e3d, matched))
        if t_dad is None or len(t_dad) < 5:
            continue

        # Peak in DAD
        pk_idx = np.argmax(y_dad)
        t_max_dad = t_dad[pk_idx]  # relative to injection start
        y_max_dad = y_dad[pk_idx]

        # Absolute time = HPLC injection time + relative time
        t_dad_abs = t_dad + inj['t_min']

        dad_segments.append({
            't': t_dad_abs, 'y': y_dad,
            't_max_rel': t_max_dad,
            't_max_abs': inj['t_min'] + t_max_dad,
            'y_max': y_max_dad,
            'name': inj['name'],
            'inj_idx': j,
            'is_khp': inj['is_khp'],
            'is_control': inj['is_control'],
            'inj_t': inj['t_min'],
        })

    print(f"DAD 254: {len(dad_segments)} segments carregats")

    # === Calculate per-injection delay ===
    # For each DAD segment, find closest TOC peak
    delay_per_inj = []
    for seg in dad_segments:
        if seg['is_control']:
            continue
        t_dad_max = seg['t_max_abs']
        # Find closest TOC peak
        if len(toc_peaks) > 0:
            dists = t_toc[toc_peaks] - t_dad_max
            # Only look at positive delays (TOC should be AFTER DAD)
            pos_mask = dists > -2  # allow small negative (noise)
            if pos_mask.any():
                pos_dists = dists[pos_mask]
                best = np.argmin(np.abs(pos_dists))
                delay_real = pos_dists[best]
                toc_pk_t = t_toc[toc_peaks[np.where(pos_mask)[0][best]]]
            else:
                delay_real = None
                toc_pk_t = None
        else:
            delay_real = None
            toc_pk_t = None

        delay_per_inj.append({
            'name': seg['name'],
            'inj_idx': seg['inj_idx'],
            't_dad': t_dad_max,
            't_toc': toc_pk_t,
            'delay': delay_real,
            'y_dad': seg['y_max'],
        })

    valid_delays = [d for d in delay_per_inj if d['delay'] is not None]
    if valid_delays:
        delays = np.array([d['delay'] for d in valid_delays])
        print(f"\nDelay per injecció:")
        print(f"  Mediana: {np.median(delays):.2f} min")
        print(f"  Rang: {delays.min():.2f} - {delays.max():.2f} min")
        print(f"  Drift: {delays[-1] - delays[0]:.2f} min")

    # === Current delay from MasterFile ===
    current_delay = FLUSH_TIME_MIN
    df_info = pd.read_excel(mf_path, sheet_name='0-INFO', header=None)
    hora_hplc = hora_toc = None
    for _, row in df_info.iterrows():
        key = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        val = row.iloc[1] if len(row) > 1 else None
        if 'hora hplc' in key: hora_hplc = val
        elif 'hora toc' in key: hora_toc = val
        elif 'net delay' in key and 'suite' not in key:
            try:
                v = float(val)
                if not np.isnan(v): current_delay = v
            except: pass

    if current_delay == FLUSH_TIME_MIN and hora_hplc and hora_toc:
        try:
            def to_min(t):
                if hasattr(t, 'hour'):
                    return t.hour * 60 + t.minute + t.second / 60
                return sum(int(p) * m for p, m in zip(str(t).split(':'), [60, 1]))
            current_delay = FLUSH_TIME_MIN - (to_min(hora_hplc) - to_min(hora_toc))
        except: pass

    # === PLOT ===
    fig = plt.figure(figsize=(20, 12))
    gs = fig.add_gridspec(3, 1, height_ratios=[3, 2, 1], hspace=0.15)
    ax_main = fig.add_subplot(gs[0])
    ax_dad = fig.add_subplot(gs[1], sharex=ax_main)
    ax_drift = fig.add_subplot(gs[2])
    plt.subplots_adjust(bottom=0.12, top=0.93)

    # --- Main: TOC continu ---
    ax_main.plot(t_toc, y_clean, 'b-', lw=0.15, alpha=0.2)
    ax_main.plot(t_toc, y_smooth, 'b-', lw=0.5, alpha=0.8)
    ax_main.plot(t_toc[toc_peaks], y_smooth[toc_peaks], 'rv', ms=5, alpha=0.7,
                 label=f'{len(toc_peaks)} pics TOC', zorder=5)
    for i, pk in enumerate(toc_peaks):
        ax_main.annotate(f'{i+1}', (t_toc[pk], y_smooth[pk]+5),
                         fontsize=5, ha='center', color='red')
    ax_main.set_ylabel('TOC (ppb)')
    ax_main.legend(fontsize=7)

    # --- DAD 254 ---
    for seg in dad_segments:
        if seg['is_khp']:
            color, alpha = '#E74C3C', 0.8
        elif seg['is_control']:
            color, alpha = '#95A5A6', 0.3
        else:
            color, alpha = '#27AE60', 0.6
        ax_dad.plot(seg['t'], seg['y'], color=color, lw=0.5, alpha=alpha)
        # Mark peak
        ax_dad.plot(seg['t_max_abs'], seg['y_max'], '^', color=color, ms=4, zorder=5)

    ax_dad.set_ylabel('DAD 254 (mAU)')
    ax_dad.set_xlabel('min (des de inici TOC)')

    # --- Drift plot ---
    if valid_delays:
        inj_nums = [d['inj_idx']+1 for d in valid_delays]
        delays_arr = [d['delay'] for d in valid_delays]
        colors = ['#E74C3C' if d['name'] and 'khp' in d['name'].lower() else '#2E86AB'
                  for d in valid_delays]
        ax_drift.scatter(inj_nums, delays_arr, c=colors, s=20, alpha=0.7, zorder=5)
        ax_drift.plot(inj_nums, delays_arr, 'b-', lw=0.5, alpha=0.3)
        ax_drift.axhline(np.median(delays), color='orange', ls='--', lw=1,
                         label=f'Mediana: {np.median(delays):.2f} min')
        ax_drift.set_xlabel('Injeccio #')
        ax_drift.set_ylabel('Delay real (min)')
        ax_drift.set_title(f'Drift delay: {delays[0]:.1f} -> {delays[-1]:.1f} min '
                           f'(canvi {delays[-1]-delays[0]:+.1f})', fontsize=9)
        ax_drift.legend(fontsize=7)

        # Label some points
        for d in valid_delays:
            if d['name'] and ('khp' in d['name'].lower()):
                ax_drift.annotate(d['name'][:6], (d['inj_idx']+1, d['delay']),
                                  fontsize=6, color='red')

    # --- Injection markers on TOC (movable with slider) ---
    vlines = []
    title_text = [None]

    def draw_markers(delay_val):
        for vl in vlines:
            vl.remove()
        vlines.clear()

        n_match = 0
        n_samples = 0
        for j, inj in enumerate(injs):
            t_exp = inj['t_min'] + delay_val
            if inj['is_khp']:
                color, lw = '#E74C3C', 1.0
            elif inj['is_control']:
                color, lw = '#95A5A6', 0.3
            else:
                color, lw = '#2E86AB', 0.5

            vl = ax_main.axvline(t_exp, color=color, ls=':', lw=lw, alpha=0.4)
            vlines.append(vl)

            if not inj['is_control']:
                n_samples += 1
                if len(toc_peaks) > 0:
                    cadence = np.median(np.diff([i['t_min'] for i in injs])) if len(injs) >= 2 else 11
                    if np.min(np.abs(t_toc[toc_peaks] - t_exp)) < cadence * 0.3:
                        n_match += 1

        pct = n_match / n_samples * 100 if n_samples > 0 else 0
        ttl = (f"{seq_name} ({method}) — delay slider={delay_val:.1f} min — "
               f"Alineats: {n_match}/{n_samples} ({pct:.0f}%)")
        if valid_delays:
            ttl += f" | Delay real mediana: {np.median(delays):.2f} min"
        fig.suptitle(ttl, fontsize=11, fontweight='bold')

    draw_markers(current_delay)

    # --- Slider ---
    ax_slider = plt.axes([0.15, 0.04, 0.50, 0.02])
    cadence = np.median(np.diff([i['t_min'] for i in injs])) if len(injs) >= 2 else 11
    slider = Slider(ax_slider, 'Delay', current_delay - 30, current_delay + 30,
                    valinit=current_delay, valstep=0.1)
    slider.on_changed(lambda val: (draw_markers(val), fig.canvas.draw_idle()))

    # --- Buttons ---
    ax_auto = plt.axes([0.70, 0.04, 0.07, 0.02])
    btn_auto = Button(ax_auto, 'Auto')
    def on_auto(e):
        if valid_delays:
            slider.set_val(np.median(delays))
    btn_auto.on_clicked(on_auto)

    ax_save = plt.axes([0.78, 0.04, 0.07, 0.02])
    btn_save = Button(ax_save, 'GUARDAR')
    def on_save(e):
        import openpyxl
        val = slider.val
        wb = openpyxl.load_workbook(mf_path)
        ws = wb['0-INFO']
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value and 'net delay' in str(row[0].value).lower() and 'suite' not in str(row[0].value).lower():
                ws.cell(row=row[0].row, column=2, value=round(val, 3))
                wb.save(mf_path)
                print(f"GUARDAT: {val:.3f} min")
                fig.suptitle(fig._suptitle.get_text() + " GUARDAT!", color='green')
                fig.canvas.draw_idle()
                return
        lr = ws.max_row + 1
        ws.cell(row=lr, column=1, value='Net delay (Suite)')
        ws.cell(row=lr, column=2, value=round(val, 3))
        wb.save(mf_path)
        print(f"GUARDAT: {val:.3f} min")
    btn_save.on_clicked(on_save)

    ax_reset = plt.axes([0.86, 0.04, 0.07, 0.02])
    btn_reset = Button(ax_reset, 'Reset')
    btn_reset.on_clicked(lambda e: slider.set_val(current_delay))

    fig.text(0.15, 0.01,
             f"MF delay: {current_delay:.2f} | Cadencia: {cadence:.1f} min | "
             f"Inj: {len(injs)} | Pics TOC: {len(toc_peaks)} | DAD: {len(dad_segments)} seg",
             fontsize=8, color='#666')

    print("Obert. Auto = delay mediana des de DAD↔TOC matching.")
    plt.show()


if __name__ == "__main__":
    main()
