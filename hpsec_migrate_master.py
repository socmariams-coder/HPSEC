"""
hpsec_migrate_master.py - Creació de MasterFiles des de rawdata

Estructura canònica:
  0-INFO:       Metadata + sincronització
  1-HPLC-SEQ:   Injeccions ordenades + Inj_Index + Sample_Rep
  2-TOC:        Dades DOC raw (intocable)
  3-DAD_KHP:    Dades KHP (opcional, si n'hi ha)
  4-TOC_CALC:   Càlculs Sample mapping (sempre al final)

API per GUI:
  migrate_single(seq_path, progress_callback) -> dict
  migrate_batch(seq_paths, progress_callback) -> list[dict]

Autor: HPSEC Suite
Data: 2025-01
"""

import os
import glob
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Optional, Callable, List, Dict, Any


# =============================================================================
# CONSTANTS
# =============================================================================
FLUSH_TIME_MIN = 3.637  # Delay físic DOC-DAD en minuts
TOC_DATA_START_ROW = 8  # Les dades TOC comencen a la fila 8
CANONICAL_SHEETS = ['0-INFO', '1-HPLC-SEQ', '2-TOC', '3-DAD_KHP', '4-TOC_CALC']


# =============================================================================
# PUBLIC API
# =============================================================================

def migrate_single(seq_path: str, progress_callback: Optional[Callable] = None,
                   force: bool = False) -> Dict[str, Any]:
    """
    Migra una sola carpeta SEQ a MasterFile.

    Args:
        seq_path: Ruta a la carpeta SEQ
        progress_callback: Opcional, funció(step, total, message) per reportar progrés
        force: Si True, sobreescriu MasterFile existent

    Returns:
        dict amb keys: status ('ok'|'error'|'skip'), file, rows, message
    """
    folder_name = os.path.basename(seq_path)

    def report(step, total, msg):
        if progress_callback:
            progress_callback(step, total, msg)

    report(0, 5, f"Cercant rawdata a {folder_name}...")

    # 1. Trobar rawdata
    rawdata_path = _find_rawdata(seq_path)
    if not rawdata_path:
        return {'status': 'skip', 'message': 'No rawdata (v11/v12) trobat'}

    # 2. Verificar si ja existeix MasterFile
    existing = glob.glob(os.path.join(seq_path, '*_MasterFile.xlsx'))
    existing = [f for f in existing if 'backup' not in f.lower()]
    if existing and not force:
        return {'status': 'skip', 'message': 'MasterFile ja existeix', 'file': existing[0]}
    elif existing and force:
        # Eliminar existents per regenerar
        for f in existing:
            os.remove(f)

    report(1, 5, "Llegint dades...")

    # 3. Llegir dades
    try:
        data = _read_rawdata(rawdata_path)
    except Exception as e:
        return {'status': 'error', 'message': f'Error llegint rawdata: {e}'}

    report(2, 5, "Extraient metadata...")

    # 4. Extreure info
    info = _extract_info(data, seq_path)

    report(3, 5, "Creant MasterFile...")

    # 5. Crear MasterFile
    try:
        output_path, n_rows = _create_masterfile(data, info, seq_path)
    except Exception as e:
        return {'status': 'error', 'message': f'Error creant MasterFile: {e}'}

    report(5, 5, "Completat")

    return {
        'status': 'ok',
        'file': output_path,
        'rows': n_rows,
        'seq_id': info.get('seq_id'),
        'method': info.get('method'),
        'has_khp': data.get('khp') is not None
    }


def migrate_batch(seq_paths: List[str], progress_callback: Optional[Callable] = None,
                  force: bool = False) -> List[Dict[str, Any]]:
    """
    Migra múltiples carpetes SEQ.

    Args:
        seq_paths: Llista de rutes a carpetes SEQ
        progress_callback: Opcional, funció(current, total, seq_path) per reportar progrés
        force: Si True, sobreescriu MasterFiles existents

    Returns:
        Llista de resultats (un dict per cada SEQ)
    """
    results = []
    total = len(seq_paths)

    for i, path in enumerate(seq_paths):
        if progress_callback:
            progress_callback(i, total, path)

        result = migrate_single(path, force=force)
        result['path'] = path
        results.append(result)

    if progress_callback:
        progress_callback(total, total, "Completat")

    return results


def find_seq_folders(base_path: str) -> List[str]:
    """
    Troba totes les carpetes SEQ a un directori base.

    Args:
        base_path: Ruta base (ex: Dades2)

    Returns:
        Llista de rutes a carpetes SEQ
    """
    folders = []
    for item in os.listdir(base_path):
        item_path = os.path.join(base_path, item)
        if os.path.isdir(item_path) and 'SEQ' in item.upper():
            folders.append(item_path)
    return sorted(folders)


# =============================================================================
# PRIVATE FUNCTIONS
# =============================================================================

def _find_rawdata(seq_path: str) -> Optional[str]:
    """Troba el fitxer rawdata (v11 o v12) a la carpeta."""
    patterns = [
        '*_v11*.xlsx', '*_v12*.xlsx',
        '*RAWDATA*.xlsx', '*rawdata*.xlsx'
    ]

    for pattern in patterns:
        matches = glob.glob(os.path.join(seq_path, pattern))
        # Filtrar backups i MasterFiles
        matches = [m for m in matches if 'backup' not in m.lower() and 'MasterFile' not in m]
        if matches:
            return matches[0]

    return None


def _read_rawdata(filepath: str) -> Dict[str, Any]:
    """Llegeix les dades del fitxer rawdata."""
    import sys
    print(f"  Llegint: {os.path.basename(filepath)}...", end='', flush=True)
    data = {}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(" obert", end='', flush=True)

    # 0-CHECK (metadata)
    if '0-CHECK' in wb.sheetnames:
        print(" CHECK", end='', flush=True)
        data['check'] = pd.read_excel(filepath, sheet_name='0-CHECK', header=None, engine='openpyxl')

    # 1-HPLC-SEQ_RAW
    if '1-HPLC-SEQ_RAW' in wb.sheetnames:
        print(" HPLC", end='', flush=True)
        data['hplc'] = pd.read_excel(filepath, sheet_name='1-HPLC-SEQ_RAW', engine='openpyxl')

    # 2-TOC
    if '2-TOC' in wb.sheetnames:
        print(" TOC", end='', flush=True)
        data['toc'] = pd.read_excel(filepath, sheet_name='2-TOC', header=None, engine='openpyxl')

    # 3-DAD_data (per extreure KHP)
    if '3-DAD_data' in wb.sheetnames:
        print(" DAD", end='', flush=True)
        df_dad = pd.read_excel(filepath, sheet_name='3-DAD_data', header=None, engine='openpyxl')
        data['khp'] = _extract_khp_from_dad(df_dad)

    wb.close()
    print(" OK", flush=True)
    return data


def _extract_info(data: Dict, seq_path: str) -> Dict[str, Any]:
    """Extreu metadata del rawdata i carpeta."""
    info = {
        'seq_id': None,
        'date': None,
        'method': 'COLUMN',
        'hora_hplc': None,
        'hora_toc': None,
        'uib_range': None
    }

    # Extreure de 0-CHECK
    df_check = data.get('check')
    if df_check is not None:
        for i, row in df_check.iterrows():
            label = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
            value = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None

            if 'SEQ-ID' in label or 'SEQ_ID' in label:
                info['seq_id'] = value
            elif label == 'Date':
                info['date'] = value
            elif 'HPLC' in label and 'TOC' not in label and 'delay' not in label.lower():
                info['hora_hplc'] = value
            elif 'TOC' in label and 'HPLC' not in label and 'Flush' not in label:
                info['hora_toc'] = value

    # Detectar method pel nom del folder
    folder_name = os.path.basename(seq_path).upper()
    if '_BP' in folder_name or folder_name.endswith('BP'):
        info['method'] = 'BP'

    # Detectar UIB range
    seq_num = _extract_seq_number(info['seq_id'])
    if seq_num:
        if seq_num >= 275:
            info['uib_range'] = '1000 ppb'
        elif seq_num >= 269:
            info['uib_range'] = '700 ppb'

    return info


def _extract_seq_number(seq_id) -> Optional[int]:
    """Extreu número de SEQ del ID (ex: '283_SEQ' -> 283)."""
    if not seq_id:
        return None
    import re
    match = re.match(r'^(\d+)', str(seq_id))
    return int(match.group(1)) if match else None


def _extract_khp_from_dad(df_dad: pd.DataFrame) -> Optional[Dict[str, pd.DataFrame]]:
    """Extreu dades KHP del full DAD."""
    if df_dad is None or df_dad.empty:
        return None

    khp_data = {}
    khp_count = 0

    for col in range(df_dad.shape[1]):
        name_val = df_dad.iloc[1, col] if df_dad.shape[0] > 1 and pd.notna(df_dad.iloc[1, col]) else None

        if name_val and 'KHP' in str(name_val).upper():
            khp_count += 1
            khp_name = f"{name_val}_R{khp_count}"

            # Trobar columnes amb dades
            time_col = val_col = None
            for c in range(col, min(col + 10, df_dad.shape[1])):
                col_data = df_dad.iloc[2:, c].dropna()
                if len(col_data) > 100:
                    if time_col is None:
                        time_col = c
                    elif val_col is None:
                        val_col = c
                        break

            if time_col is not None and val_col is not None:
                times = pd.to_numeric(df_dad.iloc[2:, time_col], errors='coerce')
                vals = pd.to_numeric(df_dad.iloc[2:, val_col], errors='coerce')

                df_khp = pd.DataFrame({'time': times.values, 'value': vals.values}).dropna()
                if len(df_khp) > 0:
                    khp_data[khp_name] = df_khp

    return khp_data if khp_data else None


def _create_masterfile(data: Dict, info: Dict, seq_path: str) -> tuple:
    """Crea el MasterFile amb estructura canònica."""

    seq_id = info.get('seq_id', 'UNKNOWN')
    output_path = os.path.join(seq_path, f"{seq_id}_MasterFile.xlsx")

    wb = openpyxl.Workbook()

    # --- 0-INFO ---
    ws_info = wb.active
    ws_info.title = '0-INFO'

    ws_info['A1'] = 'SEQ-ID'
    ws_info['B1'] = info.get('seq_id', '')
    ws_info['A2'] = 'Date'
    ws_info['B2'] = info.get('date', '')
    if info.get('date'):
        ws_info['B2'].number_format = 'DD/MM/YYYY'
    ws_info['A3'] = 'Method'
    ws_info['B3'] = info.get('method', 'COLUMN')
    ws_info['A4'] = 'Inj_Volume (uL)'
    ws_info['B4'] = 100 if info.get('method') == 'BP' else 400
    ws_info['A5'] = 'UIB_range'
    ws_info['B5'] = info.get('uib_range') or 'None'

    ws_info['A7'] = 'SINCRONITZACIO HPLC-TOC'
    ws_info['A8'] = 'Hora HPLC'
    ws_info['B8'] = info.get('hora_hplc', '')
    ws_info['A9'] = 'Hora TOC'
    ws_info['B9'] = info.get('hora_toc', '')
    ws_info['A10'] = 'Desfase (min)'
    ws_info['B10'] = '=(B8-B9)*24*60'
    ws_info['A11'] = 'Flush time (min)'
    ws_info['B11'] = FLUSH_TIME_MIN
    ws_info['C11'] = '(constant)'
    ws_info['A12'] = 'Net delay (min)'
    ws_info['B12'] = '=-B10+B11'

    # --- 1-HPLC-SEQ ---
    ws_hplc = wb.create_sheet('1-HPLC-SEQ')
    hplc_times = []
    hplc_samples = []

    if data.get('hplc') is not None:
        df_hplc = data['hplc'].copy()

        # Trobar columnes
        date_col = sample_col = None
        for col in df_hplc.columns:
            col_lower = col.lower()
            if 'acquired date' in col_lower or ('injection' in col_lower and 'date' in col_lower):
                date_col = col
            elif 'sample' in col_lower and 'name' in col_lower:
                sample_col = col

        # Filtrar i ordenar
        if sample_col:
            df_hplc = df_hplc[df_hplc[sample_col].notna()].copy()

        if date_col:
            df_hplc[date_col] = pd.to_datetime(df_hplc[date_col])
            df_hplc = df_hplc.sort_values(date_col).reset_index(drop=True)
            df_hplc['Inj_Index'] = range(1, len(df_hplc) + 1)

            # Crear Sample_Rep
            df_hplc['Sample_Rep'] = _create_sample_rep(df_hplc, sample_col)

            # Guardar per TOC_CALC
            hplc_times = df_hplc[date_col].values
            hplc_samples = df_hplc['Sample_Rep'].values

        # Escriure
        for r_idx, row in enumerate(dataframe_to_rows(df_hplc, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    continue
                if isinstance(value, str) and value.startswith('='):
                    value = "'" + value
                ws_hplc.cell(row=r_idx, column=c_idx, value=value)

    # --- 2-TOC ---
    ws_toc = wb.create_sheet('2-TOC')
    toc_timestamps = []

    if data.get('toc') is not None:
        df_toc = data['toc']

        for r_idx, row in enumerate(dataframe_to_rows(df_toc, index=False, header=False), 1):
            for c_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    continue
                if isinstance(value, str) and value.startswith('='):
                    value = "'" + value
                ws_toc.cell(row=r_idx, column=c_idx, value=value)

        # Extreure timestamps per TOC_CALC (columna D = index 3)
        if df_toc.shape[1] > 3:
            for i in range(TOC_DATA_START_ROW - 1, len(df_toc)):
                val = df_toc.iloc[i, 3]
                if pd.notna(val):
                    try:
                        toc_timestamps.append((i + 1, pd.to_datetime(val)))  # (row_num, timestamp)
                    except:
                        pass

    # --- 3-DAD_KHP (opcional) ---
    if data.get('khp'):
        ws_khp = wb.create_sheet('3-DAD_KHP')
        col_offset = 1

        for khp_name, df_khp in data['khp'].items():
            ws_khp.cell(row=1, column=col_offset, value=khp_name)
            ws_khp.cell(row=2, column=col_offset, value='time (min)')
            ws_khp.cell(row=2, column=col_offset + 1, value='value (mAU)')

            for r_idx, row in df_khp.iterrows():
                ws_khp.cell(row=r_idx + 3, column=col_offset, value=row['time'])
                ws_khp.cell(row=r_idx + 3, column=col_offset + 1, value=row['value'])

            col_offset += 3

    # --- 4-TOC_CALC ---
    ws_calc = wb.create_sheet('4-TOC_CALC')
    ws_calc['A1'] = 'TOC_Row'
    ws_calc['B1'] = 'Sample'
    ws_calc['C1'] = 'Temps_Relatiu (min)'
    ws_calc['D1'] = 'Inj_Index'

    # Calcular net_delay a partir de 0-INFO (hores rellotge HPLC i TOC)
    # Net delay = FLUSH_TIME - desfase_rellotges
    # Desfase = (HPLC_clock - TOC_clock) en minuts
    hora_hplc_clock = info.get('hora_hplc')
    hora_toc_clock = info.get('hora_toc')

    if hora_hplc_clock and hora_toc_clock:
        try:
            # Convertir a minuts des de mitjanit
            if hasattr(hora_hplc_clock, 'hour'):
                hplc_min = hora_hplc_clock.hour * 60 + hora_hplc_clock.minute + hora_hplc_clock.second / 60
            else:
                parts = str(hora_hplc_clock).split(':')
                hplc_min = int(parts[0]) * 60 + int(parts[1])

            if hasattr(hora_toc_clock, 'hour'):
                toc_min = hora_toc_clock.hour * 60 + hora_toc_clock.minute + hora_toc_clock.second / 60
            else:
                parts = str(hora_toc_clock).split(':')
                toc_min = int(parts[0]) * 60 + int(parts[1])

            desfase_min = hplc_min - toc_min
            net_delay_min = FLUSH_TIME_MIN - desfase_min
        except:
            net_delay_min = FLUSH_TIME_MIN
    else:
        net_delay_min = FLUSH_TIME_MIN

    n_rows = 0
    if toc_timestamps and len(hplc_times) > 0:
        for i, (toc_row, toc_time) in enumerate(toc_timestamps):
            # hora_HPLC = toc_time - net_delay
            hora_hplc = toc_time - pd.Timedelta(minutes=net_delay_min)
            inj_index = int((hplc_times <= hora_hplc).sum())

            if 0 < inj_index <= len(hplc_samples):
                sample = hplc_samples[inj_index - 1]
                inj_start = pd.Timestamp(hplc_times[inj_index - 1])
                temps_rel = (hora_hplc - inj_start).total_seconds() / 60.0
            else:
                sample = ''
                temps_rel = None

            excel_row = i + 2
            ws_calc.cell(row=excel_row, column=1, value=toc_row)
            ws_calc.cell(row=excel_row, column=2, value=sample)
            if temps_rel is not None:
                ws_calc.cell(row=excel_row, column=3, value=round(temps_rel, 3))
            ws_calc.cell(row=excel_row, column=4, value=inj_index)
            n_rows += 1

    # Guardar
    wb.save(output_path)
    wb.close()

    return output_path, n_rows


def _create_sample_rep(df_hplc: pd.DataFrame, sample_col: str) -> List[str]:
    """Crea Sample_Rep amb lògica de blocs per duplicats."""

    inj_num_col = 'Inj#' if 'Inj#' in df_hplc.columns else None

    if not inj_num_col:
        return [f"{row[sample_col]}_R1" for _, row in df_hplc.iterrows()]

    df_hplc['_inj_num'] = df_hplc[inj_num_col].fillna(1).astype(int)

    # Detectar duplicats
    df_hplc['_key'] = df_hplc[sample_col].astype(str) + '_R' + df_hplc['_inj_num'].astype(str)
    dup_counts = df_hplc['_key'].value_counts()
    samples_need_blocks = set(df_hplc.loc[df_hplc['_key'].isin(dup_counts[dup_counts > 1].index), sample_col])

    # Generar Sample_Rep
    result = []
    block_counter = {}

    for _, row in df_hplc.iterrows():
        sample = row[sample_col]
        inj_num = row['_inj_num']

        if sample in samples_need_blocks:
            if sample not in block_counter:
                block_counter[sample] = 0
            if inj_num == 1:
                block_counter[sample] += 1
            result.append(f"{sample}_B{block_counter[sample]}_R{inj_num}")
        else:
            result.append(f"{sample}_R{inj_num}")

    return result


# =============================================================================
# CLI
# =============================================================================

if __name__ == '__main__':
    import sys

    print("=" * 60)
    print("HPSEC MasterFile Migration Tool")
    print("=" * 60)

    if len(sys.argv) < 2:
        print("""
Ús:
  python hpsec_migrate_master.py <carpeta_SEQ>    Migra una SEQ
  python hpsec_migrate_master.py --all            Migra totes a Dades2
  python hpsec_migrate_master.py --all --force    Regenera TOTES (sobreescriu)
  python hpsec_migrate_master.py --list           Llista SEQs disponibles
        """)
        sys.exit(1)

    args = sys.argv[1:]
    force = '--force' in args
    if force:
        args.remove('--force')

    arg = args[0] if args else '--list'
    base_path = 'C:/Users/Lequia/Desktop/Dades2'

    if arg == '--all':
        folders = find_seq_folders(base_path)
        print(f"\nTrobats {len(folders)} carpetes SEQ")
        if force:
            print("Mode --force: regenerant TOTS els MasterFiles\n")
        else:
            print("(usar --force per regenerar existents)\n")

        def progress(i, total, path):
            print(f"[{i+1}/{total}] {os.path.basename(path)}...", flush=True)

        results = migrate_batch(folders, progress, force=force)

        # Resum
        ok = sum(1 for r in results if r['status'] == 'ok')
        skip = sum(1 for r in results if r['status'] == 'skip')
        err = sum(1 for r in results if r['status'] == 'error')

        print(f"\n{'=' * 60}")
        print(f"RESUM: {ok} OK, {skip} SKIP, {err} ERROR")

    elif arg == '--list':
        folders = find_seq_folders(base_path)
        print(f"\nCarpetes SEQ a {base_path}:\n")
        for f in folders:
            has_master = bool(glob.glob(os.path.join(f, '*_MasterFile.xlsx')))
            status = '[OK]' if has_master else '[--]'
            print(f"  {status} {os.path.basename(f)}")
        print(f"\nTotal: {len(folders)}")

    elif os.path.isdir(arg):
        result = migrate_single(arg, force=force)
        print(f"\nResultat: {result}")

    else:
        print(f"ERROR: No trobat: {arg}")
        sys.exit(1)
