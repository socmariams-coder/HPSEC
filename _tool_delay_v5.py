"""
Delay tool v5 — dos continus superposats: TOC + DAD 254.
==========================================================
Construeix el continu DAD 254 de la mateixa manera que el TOC:
- Per cada injecció: temps_absolut = hora_injecció_HPLC + temps_relatiu_Export3D
- Concatena tot → línia contínua DAD 254
- Superposa amb TOC continu
- Els pics DOC haurien d'aparèixer "delay" minuts després dels pics DAD

Usage:
    python _tool_delay_v5.py 231
"""
import sys, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button
from scipy.ndimage import uniform_filter1d

from hpsec_config import get_data_folders
from hpsec_consolidate import extract_seq_number, detect_seq_type


def find_seq_path(arg):
    if os.path.isdir(arg):
        return arg
    for folder in get_data_folders():
        full = os.path.join(folder, arg)
        if os.path.isdir(full):
            return full
        try:
            num = int(arg)
            for d in os.listdir(folder):
                if extract_seq_number(d) == num:
                    return os.path.join(folder, d)
        except ValueError:
            pass
    return None


def read_export3d_254(filepath):
    """Read Export3D, return (time_relative, value_254)."""
    try:
        with open(filepath, encoding='utf-16') as fh:
            lines = fh.readlines()
        header = lines[0].strip().split(',')
        idx254 = next((i for i, h in enumerate(header) if '254' in h), None)
        if idx254 is None:
            return None, None
        times, vals = [], []
        for line in lines[1:]:
            fields = line.strip().split(',')
            if len(fields) > idx254:
                try:
                    times.append(float(fields[0]))
                    vals.append(float(fields[idx254]))
                except ValueError:
                    pass
        return np.array(times), np.array(vals)
    except:
        return None, None


def build_dad_continuous(seq_path, injs):
    """Build continuous DAD 254 signal.

    For each injection: absolute_time = hplc_injection_time + relative_time
    Concatenate all → continuous timeline.
    """
    e3d = os.path.join(seq_path, 'Export3d')
    if not os.path.isdir(e3d):
        return None, None

    e3d_files = sorted(os.listdir(e3d))

    # Map files to injections
    name_counter = {}
    all_t = []
    all_y = []
    matched_count = 0

    for j, inj in enumerate(injs):
        if pd.isna(inj['name']) or not inj['name'].strip():
            continue
        nk = inj['name'].lower().replace(' ', '').replace('-', '')
        name_counter[nk] = name_counter.get(nk, 0) + 1
        rep = name_counter[nk]

        # Find matching Export3D file
        # Patterns: NAME_R1.csv, NAME_1.CSV, NAME_R1_DAD1A.CSV
        matched = None
        nk_alnum = ''.join(c for c in nk if c.isalnum())

        rep_patterns = [f'_r{rep}.', f'_r{rep}_', f'_{rep}.', f'_{rep}_']

        for f in e3d_files:
            fb = f.lower()
            fb_noext = fb.replace('.csv', '')
            fb_alnum = ''.join(c for c in fb_noext if c.isalnum())

            # Skip UIB files
            if 'uib' in fb:
                continue

            # Check name match
            name_match = nk_alnum in fb_alnum
            if not name_match:
                orig = inj['name'].lower().replace(' ', '_')
                name_match = orig in fb_noext

            if not name_match:
                continue

            # Check replica match
            for pat in rep_patterns:
                if pat in fb:
                    matched = f
                    break
            if matched:
                break

        if matched is None:
            continue

        t_rel, y_254 = read_export3d_254(os.path.join(e3d, matched))
        if t_rel is None or len(t_rel) < 5:
            continue

        # Absolute time
        t_abs = t_rel + inj['t_min']
        all_t.append(t_abs)
        all_y.append(y_254)
        matched_count += 1

    if not all_t:
        return None, None

    # Concatenate and sort
    t_cat = np.concatenate(all_t)
    y_cat = np.concatenate(all_y)
    order = np.argsort(t_cat)

    return t_cat[order], y_cat[order], matched_count


def main():
    if len(sys.argv) < 2:
        print("Usage: python _tool_delay_v5.py <SEQ>")
        sys.exit(1)

    seq_path = find_seq_path(sys.argv[1])
    if not seq_path:
        print(f"ERROR: '{sys.argv[1]}' no trobat")
        sys.exit(1)

    seq_name = os.path.basename(seq_path)
    method = detect_seq_type(seq_name)
    print(f"Loading {seq_name} ({method})...")

    # MasterFile
    mfs = [f for f in os.listdir(seq_path)
           if 'MasterFile' in f and f.endswith('.xlsx')
           and '~' not in f and 'pre_vol' not in f and 'backup' not in f]
    if not mfs:
        mfs = [f for f in os.listdir(seq_path)
               if 'MasterFile' in f and f.endswith('.xlsx')
               and '~' not in f and 'pre_vol' not in f]
    mf_path = os.path.join(seq_path, mfs[0])

    # === TOC continu ===
    df_toc = pd.read_excel(mf_path, sheet_name='2-TOC', header=6)
    toc_col = [c for c in df_toc.columns if 'toc' in str(c).lower()][0]
    date_col_toc = [c for c in df_toc.columns if 'started' in str(c).lower()]
    y_toc_raw = pd.to_numeric(df_toc[toc_col], errors='coerce').values
    dt = 4.0 / 60.0
    t_toc = np.arange(len(y_toc_raw)) * dt
    y_toc = np.nan_to_num(y_toc_raw, nan=np.nanmedian(y_toc_raw[~np.isnan(y_toc_raw)]))
    y_toc_smooth = uniform_filter1d(y_toc, size=5)

    t0_toc = None
    if date_col_toc:
        dates = pd.to_datetime(df_toc[date_col_toc[0]], errors='coerce')
        valid = dates.dropna()
        if len(valid) > 0:
            t0_toc = valid.iloc[0]

    print(f"TOC: {len(y_toc)} pts, {t_toc[-1]:.0f} min")

    # === HPLC injections ===
    df_hplc = pd.read_excel(mf_path, sheet_name='1-HPLC-SEQ')
    name_col = [c for c in df_hplc.columns
                if 'sample' in str(c).lower() and 'name' in str(c).lower()][0]
    hplc_date_col = [c for c in df_hplc.columns if 'acquired' in str(c).lower()][0]

    injs = []
    for _, row in df_hplc.iterrows():
        try:
            name = str(row[name_col]).strip()
            dt_val = pd.to_datetime(row[hplc_date_col])
            t_min = (dt_val - t0_toc).total_seconds() / 60.0 if t0_toc else 0
            injs.append({'name': name, 't_min': t_min,
                         'is_khp': 'khp' in name.lower(),
                         'is_control': any(x in name.lower()
                                           for x in ['mq', 'naoh', 'blanc', 'blnc'])})
        except:
            pass

    cadence = np.median(np.diff([i['t_min'] for i in injs])) if len(injs) >= 2 else 11.0
    print(f"HPLC: {len(injs)} inj, cadencia={cadence:.1f} min")

    # === DAD 254 continu ===
    dad_result = build_dad_continuous(seq_path, injs)
    has_dad = dad_result is not None and dad_result[0] is not None
    if has_dad:
        t_dad, y_dad, n_dad = dad_result
        print(f"DAD 254: {len(t_dad)} pts, {n_dad}/{len(injs)} injeccions")
    else:
        print("DAD 254: no disponible")

    # === Current delay ===
    current_delay = 3.637
    df_info = pd.read_excel(mf_path, sheet_name='0-INFO', header=None)
    hora_hplc = hora_toc = None
    for _, row in df_info.iterrows():
        key = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        val = row.iloc[1] if len(row) > 1 else None
        if 'hora hplc' in key: hora_hplc = val
        elif 'hora toc' in key: hora_toc = val
        elif 'net delay' in key and 'suite' not in key:
            try:
                v = float(val)
                if not np.isnan(v): current_delay = v
            except: pass
    if hora_hplc and hora_toc:
        try:
            def to_min(t):
                if hasattr(t, 'hour'):
                    return t.hour * 60 + t.minute + t.second / 60
                return sum(int(p)*m for p, m in zip(str(t).split(':'), [60, 1]))
            calc = 3.637 - (to_min(hora_hplc) - to_min(hora_toc))
            if current_delay == 3.637:
                current_delay = calc
        except: pass

    # === Detect TOC timeouts ===
    # Timeouts: drops in TOC signal followed by recovery
    # The TOC cycle is ~77 min; timeout is ~74 sec gap
    # Detect as sudden drops > 30% of local signal
    timeout_positions = []
    window_pts = int(2.0 / dt)  # 2 min window
    for i in range(window_pts, len(y_toc_smooth) - window_pts):
        before = np.mean(y_toc_smooth[i-window_pts:i])
        after = np.mean(y_toc_smooth[i:i+window_pts])
        # Timeout: signal drops significantly then recovers
        val = y_toc_smooth[i]
        if before > 30 and val < before * 0.5 and after > val * 1.3:
            # Check not too close to previous timeout
            if not timeout_positions or (t_toc[i] - timeout_positions[-1]) > 10:
                timeout_positions.append(t_toc[i])

    # Also detect via gaps in timestamps (if timestamps have jumps)
    if date_col_toc:
        ts = pd.to_datetime(df_toc[date_col_toc[0]], errors='coerce')
        ts_diff = ts.diff().dt.total_seconds().values
        for i, d in enumerate(ts_diff):
            if d > 30 and not np.isnan(d):  # gap > 30 sec = timeout
                t_gap = t_toc[i]
                if not any(abs(t_gap - tp) < 5 for tp in timeout_positions):
                    timeout_positions.append(t_gap)

    timeout_positions.sort()
    print(f"Timeouts detectats: {len(timeout_positions)}")

    # === FIGURE ===
    fig, ax = plt.subplots(figsize=(22, 7))
    plt.subplots_adjust(bottom=0.15, top=0.90, left=0.05, right=0.92)

    # TOC — eix esquerre
    line_toc, = ax.plot(t_toc, y_toc_smooth, 'b-', lw=0.5, alpha=0.7, label='TOC (ppb)')
    ax.set_ylabel('TOC (ppb)', color='blue', fontsize=10)
    ax.tick_params(axis='y', labelcolor='blue')

    # Mark timeouts
    for tp in timeout_positions:
        ax.axvspan(tp - 0.5, tp + 1.5, alpha=0.15, color='red', zorder=0)
        ax.annotate('TO', (tp, ax.get_ylim()[1] * 0.95 if ax.get_ylim()[1] > 0 else 100),
                    fontsize=6, color='red', alpha=0.6, ha='center')

    # DAD 254 — eix dret
    if has_dad:
        ax2 = ax.twinx()
        line_dad, = ax2.plot(t_dad, y_dad, 'g-', lw=0.4, alpha=0.6, label='DAD 254 (mAU)')
        ax2.set_ylabel('DAD 254 (mAU)', color='green', fontsize=10)
        ax2.tick_params(axis='y', labelcolor='green')
    else:
        ax2 = None

    ax.set_xlabel('min (temps absolut des de inici TOC)', fontsize=10)

    # Injection name labels at bottom
    for j, inj in enumerate(injs):
        if pd.isna(inj.get('name', '')) or not str(inj['name']).strip():
            continue
        color = '#E74C3C' if inj['is_khp'] else ('#95A5A6' if inj['is_control'] else '#333')
        ax.annotate(f"{j+1}:{inj['name'][:6]}",
                    (inj['t_min'], ax.get_ylim()[0] if ax.get_ylim()[0] > 0 else 0),
                    fontsize=4, rotation=90, va='bottom', ha='center',
                    color=color, alpha=0.5)

    # Store dynamic elements
    shift_line = [None]

    def update_view(delay_val):
        """Shift DAD signal by delay to align with TOC."""
        if not has_dad or ax2 is None:
            return

        # Save current zoom
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        ylim2 = ax2.get_ylim()

        if shift_line[0] is not None:
            shift_line[0].remove()

        t_dad_shifted = t_dad + delay_val
        shift_line[0], = ax2.plot(t_dad_shifted, y_dad, color='#FF6B00',
                                   lw=0.5, alpha=0.5, ls='--')

        # Restore zoom
        ax.set_xlim(xlim)
        ax.set_ylim(ylim)
        ax2.set_ylim(ylim2)

        n_to = len(timeout_positions)
        fig.suptitle(
            f"{seq_name} ({method}) — delay={delay_val:.1f}min — "
            f"{n_to} timeouts\n"
            f"Blau=TOC | Verd=DAD254 | Taronja=DAD254+delay | "
            f"Vermell=timeout TOC",
            fontsize=10, fontweight='bold')

    update_view(current_delay)

    # Slider — force it to stay active during zoom
    ax_slider = plt.axes([0.15, 0.04, 0.50, 0.025])
    slider = Slider(ax_slider, 'Delay', current_delay - 20, current_delay + 20,
                    valinit=current_delay, valstep=0.1)
    def on_slider(val):
        update_view(val)
        fig.canvas.draw_idle()
    slider.on_changed(on_slider)

    # Keyboard shortcuts for fine adjustment
    def on_key(event):
        if event.key == 'left':
            slider.set_val(max(slider.val - 0.1, slider.valmin))
        elif event.key == 'right':
            slider.set_val(min(slider.val + 0.1, slider.valmax))
        elif event.key == 'shift+left':
            slider.set_val(max(slider.val - 1.0, slider.valmin))
        elif event.key == 'shift+right':
            slider.set_val(min(slider.val + 1.0, slider.valmax))
    fig.canvas.mpl_connect('key_press_event', on_key)

    # Auto: find delay that maximizes correlation between shifted DAD and TOC
    ax_auto = plt.axes([0.70, 0.04, 0.07, 0.025])
    btn_auto = Button(ax_auto, 'Auto')
    def on_auto(e):
        if not has_dad:
            return
        # Cross-correlation on a common time grid
        t_common = np.arange(0, min(t_toc[-1], t_dad[-1]+30), dt)
        y_toc_interp = np.interp(t_common, t_toc, y_toc_smooth,
                                  left=np.nan, right=np.nan)
        best_corr = -1
        best_d = current_delay
        for d_try in np.arange(current_delay - 15, current_delay + 15, 0.2):
            y_dad_interp = np.interp(t_common, t_dad + d_try, y_dad,
                                      left=np.nan, right=np.nan)
            valid = ~np.isnan(y_toc_interp) & ~np.isnan(y_dad_interp)
            if valid.sum() > 100:
                corr = np.corrcoef(y_toc_interp[valid], y_dad_interp[valid])[0, 1]
                if corr > best_corr:
                    best_corr = corr
                    best_d = d_try
        slider.set_val(best_d)
        print(f"Auto: delay={best_d:.1f} min (correlacio={best_corr:.3f})")
    btn_auto.on_clicked(on_auto)

    # Save
    ax_save = plt.axes([0.78, 0.04, 0.07, 0.025])
    btn_save = Button(ax_save, 'GUARDAR')
    def on_save(e):
        import openpyxl
        val = slider.val
        wb = openpyxl.load_workbook(mf_path)
        ws = wb['0-INFO']
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value and 'net delay' in str(row[0].value).lower() and 'suite' not in str(row[0].value).lower():
                ws.cell(row=row[0].row, column=2, value=round(val, 3))
                wb.save(mf_path)
                print(f"GUARDAT: {val:.3f} min")
                return
        lr = ws.max_row + 1
        ws.cell(row=lr, column=1, value='Net delay (Suite)')
        ws.cell(row=lr, column=2, value=round(val, 3))
        wb.save(mf_path)
        print(f"GUARDAT: {val:.3f} min")
    btn_save.on_clicked(on_save)

    # Reset
    ax_reset = plt.axes([0.86, 0.04, 0.07, 0.025])
    btn_reset = Button(ax_reset, 'Reset')
    btn_reset.on_clicked(lambda e: slider.set_val(current_delay))

    fig.text(0.05, 0.01,
             f"MF delay: {current_delay:.2f} min | "
             f"DAD: {n_dad if has_dad else 0}/{len(injs)} inj | "
             f"Cadencia: {cadence:.1f} min",
             fontsize=8, color='#666')

    print(f"Delay actual: {current_delay:.2f}")
    print("Verd = DAD directe (referencia)")
    print("Taronja = DAD desplaçat pel delay (hauria de coincidir amb blau)")
    print("Auto = correlació creuada TOC↔DAD")
    plt.show()


if __name__ == "__main__":
    main()
